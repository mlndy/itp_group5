"""Tests for unscheduled-assignment diagnostics."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from data.models import Assignment, Course, Room, TimeSlot
from engine.unscheduled_diagnostics import (
    REASON_DELIVERY_COMPATIBILITY,
    REASON_DURATION_WINDOW,
    REASON_NO_COMPATIBLE_ROOM_TYPE,
    REASON_NO_ROOM_LARGE_ENOUGH,
    REASON_NO_VALID_WEEK,
    diagnose_unscheduled_assignments,
    diagnose_unscheduled_assignment,
    export_unscheduled_diagnostics,
)
from generator.scheduler import MAX_CANDIDATE_PATTERN_LIMIT_REASON


def make_course(**overrides: object) -> Course:
    """Create a small test course."""
    data = {
        "module_code": "ENG1001",
        "activity": "Lecture",
        "prog_yr": "ENG/YR 1",
        "class_size": 30,
        "delivery_mode": "f2f",
        "teaching_weeks": [1],
        "week_pattern": "ALL",
        "staff_ids": ["S001"],
        "duration_hrs": 2,
        "is_common_module": False,
    }
    data.update(overrides)
    return Course(**data)


def test_no_compatible_room_reason() -> None:
    """Online classes without a virtual room should be diagnosed clearly."""
    assignment = Assignment(course=make_course(delivery_mode="Online - Synchronous"), room=None, timeslot=None)
    rooms = [Room("R1", 100, "physical")]

    diagnostic = diagnose_unscheduled_assignment(assignment, [], rooms)

    assert REASON_NO_COMPATIBLE_ROOM_TYPE in diagnostic.reasons
    assert REASON_DELIVERY_COMPATIBILITY in diagnostic.reasons


def test_no_room_large_enough_reason() -> None:
    """Assignments with only undersized rooms should report capacity shortage."""
    assignment = Assignment(course=make_course(class_size=120), room=None, timeslot=None)
    rooms = [Room("R1", 80, "physical"), Room("R2", 100, "physical")]

    diagnostic = diagnose_unscheduled_assignment(assignment, [], rooms)

    assert REASON_NO_ROOM_LARGE_ENOUGH in diagnostic.reasons


def test_blocked_week_reason(monkeypatch) -> None:
    """Blocked teaching weeks should be diagnosed separately."""
    from generator import scheduler

    monkeypatch.setattr(scheduler, "BLOCKED_WEEKS", {7})
    assignment = Assignment(course=make_course(teaching_weeks=[7]), room=None, timeslot=None)

    diagnostic = diagnose_unscheduled_assignment(assignment, [], [Room("R1", 100, "physical")])

    assert REASON_NO_VALID_WEEK in diagnostic.reasons


def test_duration_cannot_fit_valid_day_window(monkeypatch) -> None:
    """Long classes should report when the remaining day window is too short."""
    from generator import scheduler

    monkeypatch.setattr(scheduler, "VALID_DAYS", ["Monday"])
    monkeypatch.setattr(scheduler, "VALID_START_TIMES", ["16:00", "17:00"])

    assignment = Assignment(course=make_course(duration_hrs=3), room=None, timeslot=None)
    rooms = [Room("R1", 100, "physical")]

    diagnostic = diagnose_unscheduled_assignment(assignment, [], rooms)

    assert REASON_DURATION_WINDOW in diagnostic.reasons


def test_diagnostics_do_not_modify_schedule(monkeypatch) -> None:
    """Running diagnostics must not change the schedule objects."""
    from generator import scheduler

    monkeypatch.setattr(scheduler, "VALID_DAYS", ["Monday"])
    monkeypatch.setattr(scheduler, "VALID_START_TIMES", ["09:00", "10:00"])

    scheduled = Assignment(
        course=make_course(module_code="ENG2001", staff_ids=["S010"]),
        room=Room("R1", 100, "physical"),
        timeslot=TimeSlot("Monday", "09:00", 1),
    )
    unscheduled = Assignment(course=make_course(module_code="ENG2002", class_size=120), room=None, timeslot=None)
    schedule = [scheduled, unscheduled]
    snapshot = [(item.room.room_id if item.room else None, item.timeslot) for item in schedule]

    report = diagnose_unscheduled_assignments(schedule, [Room("R1", 100, "physical")])

    assert [(item.room.room_id if item.room else None, item.timeslot) for item in schedule] == snapshot
    assert report.scheduled_assignments == 1
    assert report.unscheduled_assignments == 1
    assert report.hard_violations_on_scheduled_assignments == 0


def test_diagnostics_can_be_limited_to_eligible_assignments() -> None:
    """Candidate-limit rows should be preserved without expensive diagnosis."""
    assignments = [
        Assignment(
            course=make_course(module_code="ENG2001"),
            room=None,
            timeslot=None,
            hard_violations=[MAX_CANDIDATE_PATTERN_LIMIT_REASON],
        ),
        Assignment(course=make_course(module_code="ENG2002", class_size=120), room=None, timeslot=None),
        Assignment(
            course=make_course(module_code="ENG2003", class_size=130),
            room=None,
            timeslot=None,
            hard_violations=["Could not find feasible weekly room/day/start pattern"],
        ),
    ]

    report = diagnose_unscheduled_assignments(assignments, [Room("R1", 100, "physical")], max_diagnostic_assignments=1)

    assert report.diagnosed_assignments == 1
    assert report.undiagnosed_assignments == 2
    assert MAX_CANDIDATE_PATTERN_LIMIT_REASON in report.reason_counts()
    assert "Could not find feasible weekly room/day/start pattern" in report.reason_counts()


def test_unscheduled_diagnostics_export(tmp_path: Path) -> None:
    """The diagnostics report should export the expected workbook sheets."""
    report = diagnose_unscheduled_assignments(
        [Assignment(course=make_course(class_size=120), room=None, timeslot=None)],
        [Room("R1", 80, "physical")],
    )
    output = tmp_path / "unscheduled_diagnostics.xlsx"

    export_unscheduled_diagnostics(report, output)

    workbook = load_workbook(output)
    assert workbook.sheetnames == ["Summary", "Unscheduled Assignments", "Reason Counts"]
    assert workbook["Summary"]["A1"].value == "Metric"
    assert workbook["Reason Counts"]["A1"].value == "Reason"
