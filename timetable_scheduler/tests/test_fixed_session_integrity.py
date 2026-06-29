"""Tests for fixed-session integrity evidence exports."""

from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

from data.models import Assignment, Course, FixedSession, Room, TimeSlot
from output.fixed_session_integrity import export_fixed_session_integrity_report, fixed_session_integrity_rows


def make_fixed_session(**overrides: object) -> FixedSession:
    """Create one compact fixed-session source row."""
    data = {
        "programme_year": "ENG/Y1",
        "module_code": "ENG1001",
        "group_id": "P1",
        "group_size": 30,
        "day": "Monday",
        "start_time": "09:00",
        "duration_hours": 2.0,
        "teaching_weeks": (1,),
        "locations": ("E2-07-01",),
        "staff_ids": ("TUTOR A",),
        "staff_names": ("Tutor A",),
        "source_file": "fixed.xlsx",
        "source_sheet": "ENG",
        "source_row": 2,
    }
    data.update(overrides)
    return FixedSession(**data)


def make_assignment(session: FixedSession, **overrides: object) -> Assignment:
    """Create a scheduled fixed assignment tied to a source row."""
    timeslot = overrides.pop("timeslot", TimeSlot(session.day, session.start_time, session.teaching_weeks[0]))
    course = Course(
        module_code="ENG1001",
        activity="Laboratory",
        prog_yr="ENG/Y1",
        class_size=30,
        delivery_mode="f2f",
        teaching_weeks=list(session.teaching_weeks),
        week_pattern="CUSTOM",
        staff_ids=["TUTOR A"],
        duration_hrs=session.duration_hours,
        staff_names=["Tutor A"],
        fixed_source=f"{session.source_file}:{session.source_sheet}:{session.source_row}",
        is_fixed_requirement=True,
    )
    return Assignment(
        course=course,
        room=Room("E2-07-01", 40, "physical"),
        timeslot=timeslot,
        is_fixed=True,
        fixed_source=course.fixed_source,
        **overrides,
    )


def test_fixed_session_integrity_passes_for_exact_anchored_source(tmp_path: Path) -> None:
    """Exact fixed source rows should pass when represented in the final schedule."""
    session = make_fixed_session()
    assignment = make_assignment(session)

    summary = export_fixed_session_integrity_report(
        [session],
        [assignment],
        tmp_path / "fixed_session_integrity_validation.xlsx",
    )

    assert summary["fixed-session integrity status"] == "PASS"
    workbook = load_workbook(tmp_path / "fixed_session_integrity_validation.xlsx", read_only=True, data_only=True)
    try:
        assert {"Summary", "Fixed Source Integrity", "Integrity Issues"} <= set(workbook.sheetnames)
        values = {row[0]: row[1] for row in workbook["Summary"].iter_rows(min_row=2, values_only=True)}
        assert values["anchored fixed teaching occurrences"] == 1
    finally:
        workbook.close()


def test_fixed_session_integrity_detects_moved_fixed_start() -> None:
    """Moved fixed sessions should be visible as integrity failures."""
    session = make_fixed_session()
    assignment = make_assignment(session, timeslot=TimeSlot("Monday", "10:00", 1))

    rows = fixed_session_integrity_rows([session], [assignment])

    assert rows[0]["Status"] == "FAIL"
    assert rows[0]["Start Match"] == "FAIL"


def test_fixed_session_integrity_marks_quarantined_source_separately() -> None:
    """Quarantined fixed rows should remain visible without implying movement."""
    session = make_fixed_session()
    quarantined = [SimpleNamespace(requirement_id="fixed.xlsx:ENG:2")]

    rows = fixed_session_integrity_rows([session], [], quarantined)

    assert rows[0]["Status"] == "QUARANTINED"
    assert "intentionally not scheduled" in rows[0]["Issue"]
