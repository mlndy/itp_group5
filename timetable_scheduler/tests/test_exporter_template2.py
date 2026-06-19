"""Tests for the Template 2 timetable exporter."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from data.models import Assignment, Course, Room, TimeSlot
from output import exporter
from output.exporter import assignment_to_row, export_schedule


def make_course(**overrides: object) -> Course:
    """Create a small test course."""
    data = {
        "module_code": "ENG1001",
        "activity": "Lecture",
        "prog_yr": "ENG/YR 1",
        "class_size": 30,
        "delivery_mode": "f2f",
        "teaching_weeks": [1],
        "week_pattern": "ALL",
        "staff_ids": ["S001"],
        "duration_hrs": 2,
        "is_common_module": False,
    }
    data.update(overrides)
    return Course(**data)


def _headers(ws) -> list[str | None]:
    return [cell.value for cell in ws[1]]


def test_template2_workbook_is_used_when_available(tmp_path: Path, monkeypatch) -> None:
    """Exporter should preserve the provided Template 2 workbook structure."""
    template_copy = tmp_path / "template2.xlsx"
    workbook = load_workbook(Path("input/Upload template_System (Template 2).xlsx"))
    workbook["Course Code"]["A2"] = "TEMPLATE_MARKER"
    workbook.save(template_copy)

    monkeypatch.setattr(exporter, "DEFAULT_TEMPLATE2_FILE", template_copy)
    output = tmp_path / "final_timetable.xlsx"
    export_schedule([], output)

    exported = load_workbook(output)
    assert "Timetable" in exported.sheetnames
    assert exported["Course Code"]["A2"].value == "TEMPLATE_MARKER"


def test_timetable_sheet_has_required_columns(tmp_path: Path, monkeypatch) -> None:
    """The Timetable sheet should expose the required Template 2 columns."""
    monkeypatch.setattr(exporter, "DEFAULT_TEMPLATE2_FILE", Path("input/Upload template_System (Template 2).xlsx"))
    output = tmp_path / "final_timetable.xlsx"
    export_schedule([Assignment(course=make_course(), room=None, timeslot=None)], output)

    workbook = load_workbook(output)
    headers = _headers(workbook["Timetable"])

    required = ["Module", "Class Type", "Group", "Day", "Start", "End", "Class Size", "Room1", "Staff1", "Staff2", "Tri Week", "Activity Type", "Duration", "Location Hostkey", "Remark"]
    for header in required:
        assert header in headers


def test_room1_and_location_hostkey_use_assigned_room_id(tmp_path: Path, monkeypatch) -> None:
    """Room1 and Location Hostkey should carry the assigned room ID."""
    monkeypatch.setattr(exporter, "DEFAULT_TEMPLATE2_FILE", Path("input/Upload template_System (Template 2).xlsx"))
    assignment = Assignment(
        course=make_course(staff_ids=["S001", "S002"], prog_yr="ENG/YR 2", group_ids=["ENG/YR 2"]),
        room=Room("PGB-LT-01", 120, "physical"),
        timeslot=TimeSlot("Monday", "09:00", 1),
    )
    output = tmp_path / "final_timetable.xlsx"
    export_schedule([assignment], output)

    workbook = load_workbook(output)
    headers = _headers(workbook["Timetable"])
    values = [workbook["Timetable"].cell(2, idx + 1).value for idx in range(len(headers))]
    row = dict(zip(headers, values, strict=False))

    assert row["Room1"] == "PGB-LT-01"
    assert row["Location Hostkey"] == "PGB-LT-01"


def test_room2_exports_second_assigned_room(tmp_path: Path, monkeypatch) -> None:
    """A two-room assignment should populate Room2 without changing Room1."""
    monkeypatch.setattr(exporter, "DEFAULT_TEMPLATE2_FILE", Path("input/Upload template_System (Template 2).xlsx"))
    assignment = Assignment(
        course=make_course(remarks="2 rooms", group_ids=["ENG/YR 1"]),
        room=Room("PGB-R1", 30, "physical"),
        timeslot=TimeSlot("Monday", "09:00", 1),
        additional_rooms=(Room("PGB-R2", 30, "physical"),),
    )
    output = tmp_path / "final_timetable.xlsx"
    export_schedule([assignment], output)

    workbook = load_workbook(output)
    headers = _headers(workbook["Timetable"])
    values = [workbook["Timetable"].cell(2, idx + 1).value for idx in range(len(headers))]
    row = dict(zip(headers, values, strict=False))

    assert row["Room1"] == "PGB-R1"
    assert row["Room2"] == "PGB-R2"


def test_online_assignment_exports_online_room(tmp_path: Path, monkeypatch) -> None:
    """Online assignments should export the synthetic ONLINE_ROOM placeholder."""
    monkeypatch.setattr(exporter, "DEFAULT_TEMPLATE2_FILE", Path("input/Upload template_System (Template 2).xlsx"))
    assignment = Assignment(
        course=make_course(delivery_mode="Online - Synchronous", group_ids=["ENG/YR 1"]),
        room=Room("ONLINE_ROOM", 9999, "virtual"),
        timeslot=TimeSlot("Monday", "09:00", 1),
    )
    output = tmp_path / "final_timetable.xlsx"
    export_schedule([assignment], output)

    workbook = load_workbook(output)
    headers = _headers(workbook["Timetable"])
    values = [workbook["Timetable"].cell(2, idx + 1).value for idx in range(len(headers))]
    row = dict(zip(headers, values, strict=False))

    assert row["Room1"] == "ONLINE_ROOM"
    assert row["Location Hostkey"] == "ONLINE_ROOM"


def test_unscheduled_assignment_keeps_room1_blank(tmp_path: Path, monkeypatch) -> None:
    """Unscheduled assignments should not invent a room ID."""
    monkeypatch.setattr(exporter, "DEFAULT_TEMPLATE2_FILE", Path("input/Upload template_System (Template 2).xlsx"))
    assignment = Assignment(course=make_course(class_size=80), room=None, timeslot=None)
    output = tmp_path / "final_timetable.xlsx"
    export_schedule([assignment], output)

    workbook = load_workbook(output)
    headers = _headers(workbook["Timetable"])
    values = [workbook["Timetable"].cell(2, idx + 1).value for idx in range(len(headers))]
    row = dict(zip(headers, values, strict=False))

    assert row["Room1"] in (None, "")
    assert row["Location Hostkey"] in (None, "")


def test_exporter_does_not_modify_scheduler_results(tmp_path: Path, monkeypatch) -> None:
    """Exporting should not mutate the input schedule objects."""
    monkeypatch.setattr(exporter, "DEFAULT_TEMPLATE2_FILE", Path("input/Upload template_System (Template 2).xlsx"))
    assignment = Assignment(course=make_course(), room=Room("PGB-LT-01", 120, "physical"), timeslot=TimeSlot("Monday", "09:00", 1))
    snapshot = (list(assignment.hard_violations), list(assignment.soft_violations))

    output = tmp_path / "final_timetable.xlsx"
    export_schedule([assignment], output)

    assert (assignment.hard_violations, assignment.soft_violations) == snapshot


def test_assignment_to_row_uses_room_id_in_room1() -> None:
    """The row mapper should expose the room ID in Room1 and Location Hostkey."""
    assignment = Assignment(
        course=make_course(group_ids=["ENG/YR 1"], prog_yr="ENG/YR 1"),
        room=Room("PGB-LT-01", 120, "physical"),
        timeslot=TimeSlot("Monday", "09:00", 1),
    )
    row = assignment_to_row(assignment)

    assert row["Room1"] == "PGB-LT-01"
    assert row["Location Hostkey"] == "PGB-LT-01"
