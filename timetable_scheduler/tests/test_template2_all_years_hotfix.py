"""Regression tests for Template 2 all-years validation evidence."""

from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

from config import DEFAULT_TEMPLATE2_FILE
from data.models import Assignment, Course, Room, TimeSlot
from engine.programme_year import canonical_programme_year, canonical_programme_year_from_source
from output.exporter import assignment_to_row, export_schedule
from output.submission_validator import (
    build_template2_exclusion_audit_rows,
    export_all_valid_scheduled_schedule,
    export_template2_exclusion_audit,
    export_template2_programme_year_reconciliation,
    export_template2_validation_report,
    normalise_template2_programme_year,
    saved_row_programme_year,
    validate_template2_submission,
)


def make_course(index: int = 1, **overrides: object) -> Course:
    """Create one scheduled test requirement."""
    data = {
        "module_code": f"ENG{index:04d}",
        "activity": "Lecture",
        "prog_yr": f"P{index:02d}/YR 1",
        "class_size": 30,
        "delivery_mode": "f2f",
        "teaching_weeks": [1],
        "week_pattern": "ALL",
        "staff_ids": [f"S{index:03d}"],
        "duration_hrs": 2,
        "group_ids": [f"P{index:02d}/YR 1"],
    }
    data.update(overrides)
    return Course(**data)


def make_assignment(index: int = 1, **course_overrides: object) -> Assignment:
    """Create one valid scheduled assignment."""
    return Assignment(
        make_course(index, **course_overrides),
        Room("R1", 100, "physical"),
        TimeSlot("Monday", "09:00", 1),
    )


def _summary(path: Path) -> dict[str, object]:
    """Read a Metric/Value summary sheet into a dictionary."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["Summary"]
        return {row[0]: row[1] for row in sheet.iter_rows(min_row=2, values_only=True)}
    finally:
        workbook.close()


def _timetable_rows(path: Path) -> list[dict[str, object]]:
    """Return populated Timetable rows from a workbook."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["Timetable"]
        headers = [cell.value for cell in sheet[1]]
        rows: list[dict[str, object]] = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            if not any(value not in (None, "") for value in values):
                continue
            rows.append(dict(zip(headers, values, strict=False)))
        return rows
    finally:
        workbook.close()


def test_template2_year_normalisation_preserves_later_years() -> None:
    """Year 2 and Year 3 labels should not collapse to Year 1."""
    assert canonical_programme_year("ASE Y1") == "ASE/Y1"
    assert canonical_programme_year("ASE/Y1") == "ASE/Y1"
    assert canonical_programme_year("ASE/1") == "ASE/Y1"
    assert canonical_programme_year("ASE YR 1") == "ASE/Y1"
    assert canonical_programme_year("MEC-3") == "MEC/Y3"
    assert normalise_template2_programme_year("DSC/YR 2") == "DSC/Y2"
    assert normalise_template2_programme_year("MEC / Year 3") == "MEC/Y3"
    assert normalise_template2_programme_year("ESE/Y4") == "ESE/Y4"
    assert normalise_template2_programme_year("EPE/2") == "EPE/Y2"
    assert canonical_programme_year("P1") == ""
    assert canonical_programme_year("YR1/2") == ""
    assert canonical_programme_year("Y2/Y3") == ""
    assert normalise_template2_programme_year("METS/2022") == ""
    assert normalise_template2_programme_year("EDE, EPE, ESE, SBE") == ""
    assert canonical_programme_year("CBWL ESE/Yr 4") == "ESE/Y4"
    assert canonical_programme_year("MEC Yr 2 -Design-") == "MEC/Y2"
    assert canonical_programme_year("EPE Y1 - 80 pax") == ""
    assert canonical_programme_year("EEE and ISE CBE/Yr 1") == ""


def test_source_file_evidence_recovers_mets_cohort_year() -> None:
    """METS cohort labels need the source workbook year to become countable."""
    assert canonical_programme_year("METS/2022") == ""
    assert canonical_programme_year_from_source("METS/2022", "METS_Year 4.xlsx") == "METS/Y4"
    assert canonical_programme_year_from_source("METS/2024", "METS_Year 2.xlsx") == "METS/Y2"
    assert canonical_programme_year_from_source("METS/2024", "Unrelated_Year 2.xlsx") == ""


def test_template2_row_uses_source_recovered_programme_year() -> None:
    """Template 2 rows should expose source-recovered programme-year identity."""
    assignment = make_assignment(1, prog_yr="METS/2024", group_ids=["METS/2024"], source_file="METS_Year 2.xlsx")

    row = assignment_to_row(assignment)

    assert row["Programme/Year"] == "METS/Y2"
    assert saved_row_programme_year(row) == "METS/Y2"


def test_saved_row_programme_year_uses_actual_template2_fields() -> None:
    """Programme-year evidence should come from the saved row, not source-only data."""
    assert saved_row_programme_year({"Group": "CVE/YR 2"}) == "CVE/Y2"
    assert saved_row_programme_year({"Activity Hostkey": "ENG1001-2510-ENG-UGRD-PU-LEC/RSE/YR 3"}) == "RSE/Y3"
    assert saved_row_programme_year({"Group": "EEE + ISE/P1"}) == ""


def test_validator_counts_programme_years_from_saved_workbook(tmp_path: Path) -> None:
    """Readiness should use actual saved Timetable rows as the source of truth."""
    assignments = [make_assignment(index) for index in range(1, 21)]
    workbook_path = tmp_path / "submission_ready.xlsx"
    export_schedule(assignments, workbook_path, template2_path=DEFAULT_TEMPLATE2_FILE)

    result = validate_template2_submission(
        workbook_path,
        [assignment.course for assignment in assignments],
        [],
        assignments,
        [Room("R1", 100, "physical")],
        DEFAULT_TEMPLATE2_FILE,
    )

    assert result.ready
    assert result.summary["actual saved programme-year schedules"] == 20
    assert result.summary["submission-ready programme-year schedules"] == 20
    assert result.summary["qualifying submission-ready programme-years"] == 20
    assert result.summary["minimum programme-year status"] == "PASS"


def test_validator_fails_when_saved_workbook_has_fewer_than_twenty_programme_years(tmp_path: Path) -> None:
    """Source coverage cannot pass readiness when saved rows are missing."""
    assignments = [make_assignment(index) for index in range(1, 21)]
    workbook_path = tmp_path / "submission_ready.xlsx"
    export_schedule(assignments[:19], workbook_path, template2_path=DEFAULT_TEMPLATE2_FILE)

    result = validate_template2_submission(
        workbook_path,
        [assignment.course for assignment in assignments],
        [],
        assignments,
        [Room("R1", 100, "physical")],
        DEFAULT_TEMPLATE2_FILE,
    )

    assert not result.ready
    assert result.summary["actual saved programme-year schedules"] == 19
    assert result.summary["submission-ready programme-year schedules"] == 19
    assert result.summary["minimum programme-year status"] == "FAIL"


def test_validator_passes_with_more_than_twenty_qualifying_programme_years(tmp_path: Path) -> None:
    """The minimum gate should allow counts above the threshold."""
    assignments = [make_assignment(index) for index in range(1, 22)]
    workbook_path = tmp_path / "submission_ready.xlsx"
    export_schedule(assignments, workbook_path, template2_path=DEFAULT_TEMPLATE2_FILE)

    result = validate_template2_submission(
        workbook_path,
        [assignment.course for assignment in assignments],
        [],
        assignments,
        [Room("R1", 100, "physical")],
        DEFAULT_TEMPLATE2_FILE,
    )

    assert result.ready
    assert result.summary["submission-ready programme-year schedules"] == 21


def test_represented_but_search_incomplete_programme_year_does_not_count(tmp_path: Path) -> None:
    """A saved row is not enough when required weeks remain unscheduled."""
    course = make_course(1, prog_yr="ENG/YR 2", group_ids=["ENG/YR 2"], teaching_weeks=[1, 2])
    assignment = Assignment(course, Room("R1", 100, "physical"), TimeSlot("Monday", "09:00", 1))
    workbook_path = tmp_path / "submission_ready.xlsx"
    export_schedule([assignment], workbook_path, template2_path=DEFAULT_TEMPLATE2_FILE)

    result = validate_template2_submission(
        workbook_path,
        [course],
        [],
        [assignment],
        [Room("R1", 100, "physical")],
        DEFAULT_TEMPLATE2_FILE,
        all_valid_workbook_path=workbook_path,
    )

    assert not result.ready
    assert result.summary["programme-years represented in submission workbook"] == 1
    assert result.summary["complete programme-year schedules"] == 0
    assert result.summary["submission-ready programme-year schedules"] == 0
    assert "Scheduler search-failure" in result.programme_rows[0]["Exclusion Reason"]


def test_quarantined_occurrence_makes_programme_year_incomplete(tmp_path: Path) -> None:
    """Quarantined demand must block strict submission readiness."""
    scheduled_course = make_course(1, prog_yr="ENG/YR 2", group_ids=["ENG/YR 2"])
    quarantined_course = make_course(2, prog_yr="ENG/YR 2", group_ids=["ENG/YR 2"])
    assignment = Assignment(scheduled_course, Room("R1", 100, "physical"), TimeSlot("Monday", "09:00", 1))
    workbook_path = tmp_path / "submission_ready.xlsx"
    export_schedule([assignment], workbook_path, template2_path=DEFAULT_TEMPLATE2_FILE)

    result = validate_template2_submission(
        workbook_path,
        [scheduled_course, quarantined_course],
        [],
        [assignment],
        [Room("R1", 100, "physical")],
        DEFAULT_TEMPLATE2_FILE,
        all_valid_workbook_path=workbook_path,
        quarantined_requirements=[SimpleNamespace(programme_year="ENG/YR 2", affected_occurrences=1)],
    )

    assert result.summary["submission-ready programme-year schedules"] == 0
    assert "Quarantined required occurrences remain" in result.programme_rows[0]["Exclusion Reason"]


def test_hard_violation_makes_programme_year_incomplete(tmp_path: Path) -> None:
    """A scheduled hard violation must prevent programme-year completeness."""
    assignment = make_assignment(1, prog_yr="ENG/YR 2", group_ids=["ENG/YR 2"])
    assignment.hard_violations = ["Room clash"]
    workbook_path = tmp_path / "submission_ready.xlsx"
    export_schedule([], workbook_path, template2_path=DEFAULT_TEMPLATE2_FILE)

    result = validate_template2_submission(
        workbook_path,
        [assignment.course],
        [],
        [assignment],
        [Room("R1", 100, "physical")],
        DEFAULT_TEMPLATE2_FILE,
    )

    assert result.summary["complete programme-year schedules"] == 0
    assert result.summary["submission-ready programme-year schedules"] == 0
    assert "Scheduled hard violations exist" in result.programme_rows[0]["Exclusion Reason"]


def test_all_valid_export_keeps_valid_rows_from_incomplete_programme_year(tmp_path: Path) -> None:
    """One unresolved requirement should not erase other valid rows from all-valid output."""
    scheduled = make_assignment(1, prog_yr="ENG/YR 2", group_ids=["ENG/YR 2"])
    unscheduled = Assignment(
        make_course(2, prog_yr="ENG/YR 2", group_ids=["ENG/YR 2"]),
        None,
        None,
        hard_violations=["No feasible placement"],
    )
    output = tmp_path / "all_valid.xlsx"

    export_all_valid_scheduled_schedule(
        [scheduled, unscheduled],
        output,
        template2_path=DEFAULT_TEMPLATE2_FILE,
        rooms=[Room("R1", 100, "physical")],
    )

    rows = _timetable_rows(output)
    assert len(rows) == 1
    assert rows[0]["Module"] == "ENG0001"
    assert rows[0]["Group"] == "ENG/YR 2"


def test_strict_export_excludes_valid_rows_from_incomplete_programme_year(tmp_path: Path) -> None:
    """Strict export should use the completeness filter while all-valid remains broad."""
    from output.submission_validator import export_submission_ready_schedule

    assignment = make_assignment(1, prog_yr="ENG/YR 2", group_ids=["ENG/YR 2"])
    output = tmp_path / "strict.xlsx"

    export_submission_ready_schedule(
        [assignment],
        output,
        template2_path=DEFAULT_TEMPLATE2_FILE,
        complete_programmes={"ENG/Y1"},
        rooms=[Room("R1", 100, "physical")],
    )

    assert _timetable_rows(output) == []


def test_exclusion_audit_reports_incomplete_programme_year_filter(tmp_path: Path) -> None:
    """Strict submission exclusions should explain dropped valid scheduled rows."""
    assignment = make_assignment(1, prog_yr="ENG/YR 2", group_ids=["ENG/YR 2"])
    rows = build_template2_exclusion_audit_rows(
        [assignment],
        complete_programmes={"ENG/Y1"},
        rooms=[Room("R1", 100, "physical")],
    )

    assert len(rows) == 1
    assert rows[0]["Exclusion Rule"] == "incomplete-programme-year-filter"
    assert rows[0]["Programme-Year"] == "ENG/Y2"

    output = tmp_path / "template2_exclusion_audit.xlsx"
    export_template2_exclusion_audit(
        [assignment],
        output,
        complete_programmes={"ENG/Y1"},
        rooms=[Room("R1", 100, "physical")],
    )
    workbook = load_workbook(output, read_only=True)
    try:
        assert "Exclusion Audit" in workbook.sheetnames
    finally:
        workbook.close()


def test_template2_reconciliation_workbook_exports_evidence_sheets(tmp_path: Path) -> None:
    """The new reconciliation workbook should expose summary and row-level evidence."""
    assignments = [make_assignment(index) for index in range(1, 21)]
    workbook_path = tmp_path / "submission_ready.xlsx"
    export_schedule(assignments, workbook_path, template2_path=DEFAULT_TEMPLATE2_FILE)
    result = validate_template2_submission(
        workbook_path,
        [assignment.course for assignment in assignments],
        [],
        assignments,
        [Room("R1", 100, "physical")],
        DEFAULT_TEMPLATE2_FILE,
    )

    validation_path = tmp_path / "template2_submission_validation.xlsx"
    reconciliation_path = tmp_path / "template2_programme_year_reconciliation.xlsx"
    export_template2_validation_report(result, validation_path)
    export_template2_programme_year_reconciliation(result, reconciliation_path)

    validation_summary = _summary(validation_path)
    reconciliation_summary = _summary(reconciliation_path)
    assert validation_summary["actual saved programme-year schedules"] == 20
    assert reconciliation_summary["qualifying submission-ready programme-years"] == 20

    workbook = load_workbook(reconciliation_path, read_only=True)
    try:
        assert {
            "Summary",
            "Programme-Year Reconciliation",
            "Missing Programme-Years",
            "Row-Level Reconciliation",
            "Identity Normalisation Audit",
            "Incomplete Programme-Years",
            "Saved Workbook Verification",
        } <= set(workbook.sheetnames)
    finally:
        workbook.close()


def test_all_valid_export_preserves_template2_sheet_structure(tmp_path: Path) -> None:
    """Hotfix workbooks must not alter the official Template 2 sheet structure."""
    output = tmp_path / "all_valid.xlsx"
    export_all_valid_scheduled_schedule(
        [make_assignment(1)],
        output,
        template2_path=DEFAULT_TEMPLATE2_FILE,
        rooms=[Room("R1", 100, "physical")],
    )

    source = load_workbook(DEFAULT_TEMPLATE2_FILE, read_only=True)
    exported = load_workbook(output, read_only=True)
    try:
        assert exported.sheetnames == source.sheetnames
        assert len([cell.value for cell in exported["Timetable"][1]]) == 31
    finally:
        source.close()
        exported.close()
