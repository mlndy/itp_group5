"""Tests for the read-only v1.1 release validator."""

from __future__ import annotations

import importlib
from pathlib import Path

from openpyxl import Workbook

import validate_release


def _replace_default_sheet(workbook: Workbook, title: str):
    """Rename the default worksheet and return it."""
    sheet = workbook.active
    sheet.title = title
    return sheet


def _add_metric_sheet(workbook: Workbook, title: str, rows: list[tuple[object, object]]) -> None:
    """Add a two-column metric sheet."""
    sheet = _replace_default_sheet(workbook, title) if workbook.sheetnames == ["Sheet"] else workbook.create_sheet(title)
    sheet.append(["Metric", "Value"])
    for row in rows:
        sheet.append(list(row))


def _add_validation_checks(
    workbook: Workbook,
    demand_status: str = "PASS",
    hard_status: str = "PASS",
    dsc_status: str = "PASS",
) -> None:
    """Add Validation Checks sheet."""
    sheet = workbook.create_sheet("Validation Checks")
    sheet.append(["Check", "Value", "Status", "Notes"])
    sheet.append(["Demand occurrence consistency", validate_release.EXPECTED_TOTAL_TEACHING_OCCURRENCES, demand_status, ""])
    sheet.append(["Hard-constraint safety status", "0 scheduled hard violations", hard_status, ""])
    sheet.append(["DSC inclusion status", "DSC rows found", dsc_status, ""])


def _add_programme_breakdown(workbook: Workbook, has_dsc: bool = True) -> None:
    """Add Programme Breakdown sheet."""
    sheet = workbook.create_sheet("Programme Breakdown")
    sheet.append(["Programme/Year", "Source File", "DSC Indicator"])
    sheet.append(["DSC/Y1" if has_dsc else "ENG/Y1", "2510_DSC.xlsx" if has_dsc else "ENG.xlsx", "Yes" if has_dsc else "No"])


def _create_run_summary(
    path: Path,
    *,
    required: int = validate_release.EXPECTED_TOTAL_TEACHING_OCCURRENCES,
    scheduled: int = validate_release.EXPECTED_SCHEDULED_OCCURRENCES,
    unscheduled: int = validate_release.EXPECTED_TOTAL_TEACHING_OCCURRENCES - validate_release.EXPECTED_SCHEDULED_OCCURRENCES,
    scheduled_hard: int = 0,
    has_dsc: bool = True,
) -> None:
    """Create a minimal valid run_summary workbook."""
    workbook = Workbook()
    _add_metric_sheet(
        workbook,
        "Summary",
        [
            ("Required teaching occurrences", required),
            ("Scheduled teaching occurrences", scheduled),
            ("Unscheduled teaching occurrences", unscheduled),
            ("Hard violations on scheduled assignments", scheduled_hard),
        ],
    )
    _add_validation_checks(workbook)
    _add_metric_sheet(workbook, "Run Metadata", [("scope", "eng")])
    _add_programme_breakdown(workbook, has_dsc=has_dsc)
    _add_metric_sheet(workbook, "Resource Audit", [("Loaded physical room count", 169)])
    _add_metric_sheet(workbook, "Virtual Room Detail", [("Room ID", "ONLINE_ROOM")])
    _add_metric_sheet(workbook, "Unscheduled Analysis", [("Original Reason", "Could not find feasible slot")])
    _add_metric_sheet(workbook, "Unscheduled Breakdown", [("Reason Category", "Physical room scarcity")])
    _add_metric_sheet(workbook, "Residual F2F Analysis", [("Module Code", "ENG1001")])
    _add_metric_sheet(workbook, "Optimisation Summary", [("Optimisation enabled", "No")])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _create_guarded_report(
    path: Path,
    *,
    scheduled: int = validate_release.EXPECTED_SCHEDULED_OCCURRENCES,
    search_failures: int = validate_release.EXPECTED_SEARCH_FAILURE_OCCURRENCES,
) -> None:
    """Create a minimal guarded-generation report."""
    workbook = Workbook()
    _add_metric_sheet(
        workbook,
        "Summary",
        [
            ("Total teaching occurrences", validate_release.EXPECTED_TOTAL_TEACHING_OCCURRENCES),
            ("Schedulable occurrences", validate_release.EXPECTED_SCHEDULABLE_OCCURRENCES),
            ("Quarantined occurrences", validate_release.EXPECTED_QUARANTINED_OCCURRENCES),
            ("Scheduled occurrences", scheduled),
            ("Unscheduled search failures", search_failures),
            ("Scheduled hard violations", validate_release.EXPECTED_SCHEDULED_HARD_VIOLATIONS),
            ("Template 2 complete programme-years", 21),
            ("Submission-ready programme-years", 21),
        ],
    )
    for sheet in validate_release.REQUIRED_GUARDED_SHEETS:
        if sheet not in workbook.sheetnames:
            workbook.create_sheet(sheet).append(["Value"])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _create_template2_validation(path: Path, *, readiness: str = "PASS", invalid_rows: int = 0, ready_count: int = 21) -> None:
    """Create a minimal Template 2 validation workbook."""
    workbook = Workbook()
    _add_metric_sheet(
        workbook,
        "Summary",
        [
            ("Template 2 output rows", validate_release.EXPECTED_TEMPLATE2_SUBMISSION_ROWS),
            ("Actual saved Template 2 rows", validate_release.EXPECTED_TEMPLATE2_SUBMISSION_ROWS),
            ("All-valid Template 2 rows", validate_release.EXPECTED_ALL_VALID_TEMPLATE2_ROWS),
            ("rows with missing required fields", invalid_rows),
            ("rows with mapping errors", invalid_rows),
            ("programme-years represented in submission workbook", ready_count),
            ("actual saved programme-year schedules", ready_count),
            ("complete programme-year schedules", ready_count),
            ("submission-ready programme-year schedules", ready_count),
            ("qualifying submission-ready programme-years", ready_count),
            ("minimum programme-year status", readiness),
            ("Template 2 readiness status", readiness),
        ],
    )
    for sheet in validate_release.REQUIRED_TEMPLATE2_VALIDATION_SHEETS:
        if sheet not in workbook.sheetnames:
            created = workbook.create_sheet(sheet)
            if sheet == "Submission Readiness":
                created.append(["Check", "Status", "Notes"])
                created.append(["Submission readiness", readiness, ""])
            elif sheet == "Programme Schedule Coverage":
                created.append(["Canonical programme-year", "Complete Schedule Status", "Submission-Ready Status", "Counts Toward Minimum 20"])
                for index in range(ready_count):
                    created.append([f"P{index:02d}/Y1", "PASS", "PASS", "Yes"])
            elif sheet == "Invalid Rows":
                created.append(["Row", "Issue"])
                for index in range(invalid_rows):
                    created.append([index + 2, "Invalid saved row"])
            else:
                created.append(["Value"])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _create_template2_submission(
    path: Path,
    *,
    row_count: int = validate_release.EXPECTED_TEMPLATE2_SUBMISSION_ROWS,
    ready_count: int = 21,
) -> None:
    """Create a saved Template 2 workbook with countable programme-year rows."""
    workbook = Workbook()
    sheet = _replace_default_sheet(workbook, validate_release.OFFICIAL_TEMPLATE2_SHEETS[0])
    for sheet_name in validate_release.OFFICIAL_TEMPLATE2_SHEETS[1:]:
        workbook.create_sheet(sheet_name)
    sheet.append(validate_release.OFFICIAL_TIMETABLE_COLUMNS)
    group_index = validate_release.OFFICIAL_TIMETABLE_COLUMNS.index("Group")
    for index in range(row_count):
        programme_year = f"P{index % ready_count:02d}/Y1"
        row = [""] * len(validate_release.OFFICIAL_TIMETABLE_COLUMNS)
        row[0] = "ENG1001"
        row[1] = "LEC"
        row[group_index] = programme_year
        sheet.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _create_template2_reconciliation(path: Path, *, readiness: str = "PASS", ready_count: int = 21) -> None:
    """Create a minimal programme-year reconciliation workbook."""
    workbook = Workbook()
    _add_metric_sheet(
        workbook,
        "Summary",
        [
            ("submission-ready programme-year schedules", ready_count),
            ("qualifying submission-ready programme-years", ready_count),
            ("actual saved programme-year schedules", ready_count),
            ("Template 2 output rows", validate_release.EXPECTED_TEMPLATE2_SUBMISSION_ROWS),
            ("All-valid Template 2 rows", validate_release.EXPECTED_ALL_VALID_TEMPLATE2_ROWS),
            ("minimum programme-year status", readiness),
            ("Template 2 readiness status", readiness),
        ],
    )
    for sheet in validate_release.REQUIRED_TEMPLATE2_RECONCILIATION_SHEETS:
        if sheet not in workbook.sheetnames:
            created = workbook.create_sheet(sheet)
            if sheet == "Programme-Year Reconciliation":
                created.append(["Canonical programme-year", "Counts Toward Minimum 20"])
                for index in range(ready_count):
                    created.append([f"P{index:02d}/Y1", "Yes"])
            else:
                created.append(["Value"])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _create_visual_validation(path: Path, *, status: str = "PASS", missing: int = 0) -> None:
    """Create a minimal visualisation validation workbook."""
    workbook = Workbook()
    _add_metric_sheet(
        workbook,
        "Summary",
        [
            ("scheduled assignments received", validate_release.EXPECTED_SCHEDULED_OCCURRENCES),
            ("programme visual entries", 680),
            ("tutor visual entries", 616),
            ("room visual entries", 535),
            ("programme sheets", 86),
            ("tutor sheets", 235),
            ("room sheets", 48),
            ("missing entries", missing),
            ("unexpected entries", 0),
            ("invalid overlaps", 0),
            ("visual export status", status),
        ],
    )
    for sheet in validate_release.REQUIRED_VISUAL_VALIDATION_SHEETS:
        if sheet not in workbook.sheetnames:
            created = workbook.create_sheet(sheet)
            if sheet == "Export Status":
                created.append(["Workbook", "Status", "Issue"])
                created.append(["Programme_Timetable_Visuals.xlsx", status, ""])
            else:
                created.append(["Value"])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _create_fixed_integrity(path: Path, *, status: str = "PASS", mismatches: int = 0) -> None:
    """Create a minimal fixed-session integrity workbook."""
    workbook = Workbook()
    _add_metric_sheet(
        workbook,
        "Summary",
        [
            ("fixed source rows", validate_release.EXPECTED_FIXED_SOURCE_ROWS),
            ("expected fixed teaching occurrences", validate_release.EXPECTED_FIXED_SOURCE_OCCURRENCES),
            ("anchored fixed teaching occurrences", validate_release.EXPECTED_ANCHORED_FIXED_SOURCE_OCCURRENCES),
            ("quarantined fixed teaching occurrences", validate_release.EXPECTED_QUARANTINED_FIXED_SOURCE_OCCURRENCES),
            ("missing fixed teaching occurrences", 0),
            ("placement mismatches", mismatches),
            ("scheduled hard violations on fixed assignments", 0),
            ("fixed-session integrity status", status),
        ],
    )
    for sheet in validate_release.REQUIRED_FIXED_INTEGRITY_SHEETS:
        if sheet not in workbook.sheetnames:
            created = workbook.create_sheet(sheet)
            created.append(["Value"])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _create_timetable(path: Path) -> None:
    """Create a minimal Template 2-style proposed timetable workbook."""
    workbook = Workbook()
    sheet = _replace_default_sheet(workbook, "Timetable")
    sheet.append(validate_release.REQUIRED_TIMETABLE_COLUMNS)
    row = ["ENG1001"] + [""] * (len(validate_release.REQUIRED_TIMETABLE_COLUMNS) - 1)
    for _ in range(3006):
        sheet.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _create_workbook(path: Path, sheet_name: str = "Sheet1") -> None:
    """Create a readable placeholder workbook."""
    workbook = Workbook()
    workbook.active.title = sheet_name
    workbook.active.append(["Value"])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _create_stakeholder_views(path: Path) -> None:
    """Create a minimal stakeholder views workbook."""
    workbook = Workbook()
    programme = _replace_default_sheet(workbook, "Programme Timetable")
    programme.append(["Programme/Year", "Week", "Day"])
    tutor = workbook.create_sheet("Tutor Timetable")
    tutor.append(["Tutor", "Week", "Day"])
    room = workbook.create_sheet("Room Timetable")
    room.append(["Room", "Week", "Day"])
    queue = workbook.create_sheet("Exception Queue")
    queue.append(["Original Reason", "Classification", "Recommended Operational Action", "Review Status"])
    queue.append(["Could not find feasible slot", "Physical room scarcity", "Review large-room availability", "Open"])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _create_run_manifest(path: Path, validation_status: str = "PASS", template_status: str = "PASS") -> None:
    """Create a minimal run manifest workbook."""
    workbook = Workbook()
    manifest = _replace_default_sheet(workbook, "Run Manifest")
    manifest.append(["Setting", "Value"])
    manifest.append(["validation_status", validation_status])
    manifest.append(["required_teaching_occurrences", validate_release.EXPECTED_TOTAL_TEACHING_OCCURRENCES])
    weights = workbook.create_sheet("Soft Constraint Weights")
    weights.append(["Soft Rule", "Weight"])
    baseline = workbook.create_sheet("Soft Rule Baseline")
    baseline.append(["Soft Rule", "Count"])
    template = workbook.create_sheet("Template Validation")
    template.append(["Check", "Status", "Notes"])
    template.append(["Template 2 output structure retained", template_status, ""])
    trace = workbook.create_sheet("Traceability")
    trace.append(["Status", "Source File", "Module Code"])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _create_release_workbooks(
    tmp_path: Path,
    *,
    timetable_filename: str = validate_release.LEGACY_TIMETABLE_FILENAME,
) -> dict[str, Path]:
    """Create all workbook paths needed by the release validator."""
    paths = {
        "run_summary": tmp_path / "run_summary.xlsx",
        "timetable": tmp_path / timetable_filename,
        "stakeholder_views": tmp_path / "stakeholder_views.xlsx",
        "run_manifest": tmp_path / "run_manifest.xlsx",
        "guarded_report": tmp_path / "guarded_generation_report.xlsx",
        "template2_validation": tmp_path / "template2_submission_validation.xlsx",
        "template2_submission": tmp_path / "Template2_Submission_Ready.xlsx",
        "template2_reconciliation": tmp_path / "template2_programme_year_reconciliation.xlsx",
        "visual_validation": tmp_path / "timetable_visualisation_validation.xlsx",
        "fixed_integrity": tmp_path / "fixed_session_integrity_validation.xlsx",
        "programme_visuals": tmp_path / "Programme_Timetable_Visuals.xlsx",
        "tutor_visuals": tmp_path / "Tutor_Timetable_Visuals.xlsx",
        "room_visuals": tmp_path / "Room_Timetable_Visuals.xlsx",
    }
    _create_run_summary(paths["run_summary"])
    _create_timetable(paths["timetable"])
    _create_stakeholder_views(paths["stakeholder_views"])
    _create_run_manifest(paths["run_manifest"])
    _create_guarded_report(paths["guarded_report"])
    _create_template2_validation(paths["template2_validation"])
    _create_template2_submission(paths["template2_submission"])
    _create_template2_reconciliation(paths["template2_reconciliation"])
    _create_visual_validation(paths["visual_validation"])
    _create_fixed_integrity(paths["fixed_integrity"])
    _create_workbook(paths["programme_visuals"])
    _create_workbook(paths["tutor_visuals"])
    _create_workbook(paths["room_visuals"])
    return paths


def _validate(paths: dict[str, Path], *, test_root: Path | None = None) -> validate_release.ReleaseValidationResult:
    """Run validation with explicit temporary paths."""
    return validate_release.validate_release(
        paths["run_summary"],
        paths["timetable"],
        paths["stakeholder_views"],
        paths["run_manifest"],
        paths["guarded_report"],
        paths["template2_validation"],
        paths["template2_submission"],
        paths["template2_reconciliation"],
        paths["visual_validation"],
        paths["fixed_integrity"],
        paths["programme_visuals"],
        paths["tutor_visuals"],
        paths["room_visuals"],
        test_root or Path(__file__).resolve().parent,
    )


def test_release_validator_passes_with_valid_temporary_workbooks(tmp_path: Path) -> None:
    """A minimal valid v1.1 release package should pass."""
    paths = _create_release_workbooks(tmp_path)

    result = _validate(paths)

    assert result.passed
    assert result.failures == []


def test_run_dir_resolution_uses_one_complete_evidence_folder(tmp_path: Path) -> None:
    """A run directory should resolve every validator workbook from that folder."""
    paths = _create_release_workbooks(tmp_path)

    resolution = validate_release.resolve_evidence_paths(run_dir=tmp_path, test_root=Path(__file__).resolve().parent)

    assert resolution.paths.run_id == tmp_path.name
    assert resolution.paths.run_summary == paths["run_summary"].resolve()
    assert resolution.paths.timetable == paths["timetable"].resolve()
    assert "Resolved run ID" in resolution.message
    assert f"Resolved timetable filename: {validate_release.LEGACY_TIMETABLE_FILENAME}" in resolution.message


def test_run_dir_resolution_prefers_current_timetable_filename(tmp_path: Path) -> None:
    """A current pipeline run should resolve Proposed_Timetable.xlsx."""
    paths = _create_release_workbooks(tmp_path, timetable_filename=validate_release.CURRENT_TIMETABLE_FILENAME)

    resolution = validate_release.resolve_evidence_paths(run_dir=tmp_path, test_root=Path(__file__).resolve().parent)

    assert resolution.paths.timetable == paths["timetable"].resolve()
    assert resolution.paths.timetable.name == validate_release.CURRENT_TIMETABLE_FILENAME
    assert f"Resolved timetable filename: {validate_release.CURRENT_TIMETABLE_FILENAME}" in resolution.message


def test_run_dir_resolution_accepts_legacy_timetable_filename(tmp_path: Path) -> None:
    """Legacy run folders should still resolve final_timetable_engineering_cluster.xlsx."""
    paths = _create_release_workbooks(tmp_path, timetable_filename=validate_release.LEGACY_TIMETABLE_FILENAME)

    resolution = validate_release.resolve_evidence_paths(run_dir=tmp_path, test_root=Path(__file__).resolve().parent)

    assert resolution.paths.timetable == paths["timetable"].resolve()
    assert resolution.paths.timetable.name == validate_release.LEGACY_TIMETABLE_FILENAME


def test_run_dir_resolution_rejects_missing_timetable_filename(tmp_path: Path) -> None:
    """A run folder must contain either the current or legacy timetable workbook."""
    paths = _create_release_workbooks(tmp_path)
    paths["timetable"].unlink()

    try:
        validate_release.resolve_evidence_paths(run_dir=tmp_path, test_root=Path(__file__).resolve().parent)
    except FileNotFoundError as exc:
        message = str(exc)
    else:  # pragma: no cover - defensive
        raise AssertionError("Expected FileNotFoundError")

    assert validate_release.CURRENT_TIMETABLE_FILENAME in message
    assert validate_release.LEGACY_TIMETABLE_FILENAME in message


def test_run_dir_resolution_accepts_identical_current_and_legacy_timetables(tmp_path: Path) -> None:
    """A copied legacy/current pair should resolve to the current filename."""
    paths = _create_release_workbooks(tmp_path, timetable_filename=validate_release.LEGACY_TIMETABLE_FILENAME)
    current = tmp_path / validate_release.CURRENT_TIMETABLE_FILENAME
    current.write_bytes(paths["timetable"].read_bytes())

    resolution = validate_release.resolve_evidence_paths(run_dir=tmp_path, test_root=Path(__file__).resolve().parent)

    assert resolution.paths.timetable == current.resolve()


def test_run_dir_resolution_rejects_inconsistent_current_and_legacy_timetables(tmp_path: Path) -> None:
    """Mixed timetable evidence should fail when current and legacy files differ."""
    _create_release_workbooks(tmp_path, timetable_filename=validate_release.LEGACY_TIMETABLE_FILENAME)
    _create_timetable(tmp_path / validate_release.CURRENT_TIMETABLE_FILENAME)
    # Add a byte-level difference while preserving a readable workbook.
    workbook = Workbook()
    sheet = _replace_default_sheet(workbook, "Timetable")
    sheet.append(validate_release.REQUIRED_TIMETABLE_COLUMNS)
    sheet.append(["DIFFERENT"] + [""] * (len(validate_release.REQUIRED_TIMETABLE_COLUMNS) - 1))
    workbook.save(tmp_path / validate_release.CURRENT_TIMETABLE_FILENAME)

    try:
        validate_release.resolve_evidence_paths(run_dir=tmp_path, test_root=Path(__file__).resolve().parent)
    except ValueError as exc:
        message = str(exc)
    else:  # pragma: no cover - defensive
        raise AssertionError("Expected ValueError")

    assert "ambiguous mixed timetable evidence" in message


def test_no_argument_latest_run_resolution_uses_newest_complete_run(tmp_path: Path, monkeypatch) -> None:
    """No-argument resolution should select the newest complete run folder."""
    older = tmp_path / "20260630_010000_old"
    newer = tmp_path / "20260630_020000_new"
    _create_release_workbooks(older)
    paths = _create_release_workbooks(newer, timetable_filename=validate_release.CURRENT_TIMETABLE_FILENAME)
    monkeypatch.setattr(validate_release, "RUNS_DIR", tmp_path)

    resolution = validate_release.resolve_evidence_paths(test_root=Path(__file__).resolve().parent)

    assert resolution.paths.run_id == newer.name
    assert resolution.paths.timetable == paths["timetable"].resolve()


def test_explicit_run_dir_cli_resolution_prints_timetable_filename(tmp_path: Path, capsys) -> None:
    """The CLI should print the resolved timetable filename for explicit run folders."""
    _create_release_workbooks(tmp_path, timetable_filename=validate_release.CURRENT_TIMETABLE_FILENAME)

    exit_code = validate_release.main(
        ["--run-dir", str(tmp_path), "--test-root", str(Path(__file__).resolve().parent)]
    )

    captured = capsys.readouterr()
    assert exit_code == 0
    assert f"Resolved timetable filename: {validate_release.CURRENT_TIMETABLE_FILENAME}" in captured.out
    assert "FINAL RELEASE VALIDATION: PASS" in captured.out


def test_saved_template2_programmes_must_match_reconciliation(tmp_path: Path) -> None:
    """The validator should not pass if saved rows and reconciliation count different programme-years."""
    paths = _create_release_workbooks(tmp_path)
    _create_template2_reconciliation(paths["template2_reconciliation"], ready_count=20)

    result = _validate(paths)

    assert not result.passed
    assert any("programme-years do not match reconciliation" in failure for failure in result.failures)


def test_missing_workbook_produces_fail(tmp_path: Path) -> None:
    """A missing workbook should fail validation."""
    paths = _create_release_workbooks(tmp_path)
    paths["stakeholder_views"] = tmp_path / "missing.xlsx"

    result = _validate(paths)

    assert not result.passed
    assert any("Missing workbook" in failure for failure in result.failures)


def test_missing_required_sheet_produces_fail(tmp_path: Path) -> None:
    """Missing run-summary sheets should fail validation."""
    paths = _create_release_workbooks(tmp_path)
    workbook = Workbook()
    _add_metric_sheet(workbook, "Summary", [("Required teaching occurrences", validate_release.EXPECTED_TOTAL_TEACHING_OCCURRENCES)])
    workbook.save(paths["run_summary"])

    result = _validate(paths)

    assert not result.passed
    assert any("missing required sheet" in failure for failure in result.failures)


def test_demand_inconsistency_produces_fail(tmp_path: Path) -> None:
    """Required occurrences must equal scheduled plus unscheduled occurrences."""
    paths = _create_release_workbooks(tmp_path)
    _create_run_summary(paths["run_summary"], scheduled=3000, unscheduled=10)

    result = _validate(paths)

    assert not result.passed
    assert any("Demand inconsistency" in failure for failure in result.failures)


def test_non_zero_scheduled_hard_violations_produces_fail(tmp_path: Path) -> None:
    """Scheduled hard violations must be zero."""
    paths = _create_release_workbooks(tmp_path)
    _create_run_summary(paths["run_summary"], scheduled_hard=1)

    result = _validate(paths)

    assert not result.passed
    assert any("Scheduled hard violations" in failure for failure in result.failures)


def test_missing_dsc_produces_fail(tmp_path: Path) -> None:
    """Programme Breakdown must contain DSC evidence."""
    paths = _create_release_workbooks(tmp_path)
    _create_run_summary(paths["run_summary"], has_dsc=False)

    result = _validate(paths)

    assert not result.passed
    assert any("does not contain DSC" in failure for failure in result.failures)


def test_guarded_schedulable_split_must_match_v1_1_metrics(tmp_path: Path) -> None:
    """Schedulable demand must equal scheduled occurrences plus search failures."""
    paths = _create_release_workbooks(tmp_path)
    _create_guarded_report(paths["guarded_report"], search_failures=89)

    result = _validate(paths)

    assert not result.passed
    assert any("search failures" in failure or "Schedulable split" in failure for failure in result.failures)


def test_template2_readiness_failure_produces_fail(tmp_path: Path) -> None:
    """Template 2 readiness must pass."""
    paths = _create_release_workbooks(tmp_path)
    _create_template2_validation(paths["template2_validation"], readiness="FAIL")

    result = _validate(paths)

    assert not result.passed
    assert any("Template 2 readiness status" in failure for failure in result.failures)


def test_visual_validation_failure_produces_fail(tmp_path: Path) -> None:
    """Visual export validation must pass with no missing entries."""
    paths = _create_release_workbooks(tmp_path)
    _create_visual_validation(paths["visual_validation"], status="FAIL", missing=1)

    result = _validate(paths)

    assert not result.passed
    assert any("visual export status" in failure or "missing entries" in failure for failure in result.failures)


def test_fixed_integrity_failure_produces_fail(tmp_path: Path) -> None:
    """Fixed-session integrity evidence must pass."""
    paths = _create_release_workbooks(tmp_path)
    _create_fixed_integrity(paths["fixed_integrity"], status="FAIL", mismatches=1)

    result = _validate(paths)

    assert not result.passed
    assert any("fixed-session integrity status" in failure or "placement mismatches" in failure for failure in result.failures)


def test_normal_test_suite_size_is_checked(tmp_path: Path) -> None:
    """The validator should flag a too-small test folder."""
    paths = _create_release_workbooks(tmp_path)
    small_tests = tmp_path / "tests"
    small_tests.mkdir()
    (small_tests / "test_one.py").write_text("def test_one():\n    assert True\n", encoding="utf-8")

    result = _validate(paths, test_root=small_tests)

    assert not result.passed
    assert any("expected at least" in failure for failure in result.failures)


def test_manifest_template_validation_failure_produces_fail(tmp_path: Path) -> None:
    """Template Validation rows must pass in the run manifest."""
    paths = _create_release_workbooks(tmp_path)
    _create_run_manifest(paths["run_manifest"], template_status="FAIL")

    result = _validate(paths)

    assert not result.passed
    assert any("Template Validation" in failure for failure in result.failures)


def test_importing_validate_release_does_not_execute_validation(capsys) -> None:
    """Importing the module should not print release validation output."""
    importlib.reload(validate_release)

    captured = capsys.readouterr()
    assert "FINAL RELEASE VALIDATION" not in captured.out
