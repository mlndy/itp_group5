"""Tests for baseline versus remarks-enabled coverage comparison."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from data.models import Assignment, Course, Room, TimeSlot
from engine.remarks_comparison import compare_remark_runs, export_remarks_coverage_comparison


def make_course(**overrides: object) -> Course:
    """Create a small course for comparison tests."""
    data = {
        "module_code": "ENG9001",
        "activity": "Lecture",
        "prog_yr": "ENG/YR 1",
        "class_size": 30,
        "delivery_mode": "f2f",
        "teaching_weeks": [1],
        "week_pattern": "ALL",
        "staff_ids": ["S001"],
        "duration_hrs": 1,
        "is_common_module": False,
        "group_ids": ["ENG/YR 1"],
        "source_file": "requirements.xlsx",
        "source_row": 4,
    }
    data.update(overrides)
    return Course(**data)


def test_baseline_and_enhanced_demand_totals_are_identical() -> None:
    """Comparison should use one shared demand pool."""
    course = make_course(remarks="Need 2 rooms")
    baseline = [Assignment(course, Room("R1", 100, "physical"), TimeSlot("Monday", "09:00", 1))]
    enhanced = [
        Assignment(
            course,
            None,
            None,
            hard_violations=[
                "Could not find feasible weekly room/day/start pattern",
                "Explicit two-room requirement could not be satisfied simultaneously.",
            ],
            remark_unscheduled_reason="Explicit two-room requirement could not be satisfied simultaneously.",
        )
    ]

    comparison = compare_remark_runs([course], baseline, enhanced, input_course_records=1)

    assert comparison.baseline_metrics.required_teaching_occurrences == 1
    assert comparison.enhanced_metrics.required_teaching_occurrences == 1
    assert comparison.baseline_metrics.scheduled_teaching_occurrences == 1
    assert comparison.enhanced_metrics.scheduled_teaching_occurrences == 0


def test_newly_unscheduled_occurrence_is_attributed_to_explicit_rule() -> None:
    """Explicit enforced rules may create attributed additional unscheduled rows."""
    course = make_course(remarks="Need 2 rooms")
    baseline = [Assignment(course, Room("R1", 100, "physical"), TimeSlot("Monday", "09:00", 1))]
    enhanced = [
        Assignment(
            course,
            None,
            None,
            hard_violations=[
                "Could not find feasible weekly room/day/start pattern",
                "Explicit two-room requirement could not be satisfied simultaneously.",
            ],
            remark_unscheduled_reason="Explicit two-room requirement could not be satisfied simultaneously.",
        )
    ]

    comparison = compare_remark_runs([course], baseline, enhanced)

    assert len(comparison.newly_unscheduled_rows) == 1
    assert comparison.rule_attribution["multiple-room requirement"] == 1
    assert comparison.attribution_reconciles is True


def test_unsupported_remark_difference_is_indirect_displacement() -> None:
    """Unsupported enhanced-only failures should be classified without suspected buckets."""
    course = make_course(remarks="Discuss with programme")
    baseline = [Assignment(course, Room("R1", 100, "physical"), TimeSlot("Monday", "09:00", 1))]
    enhanced = [Assignment(course, None, None, hard_violations=["Could not find feasible pattern"])]

    comparison = compare_remark_runs([course], baseline, enhanced)

    assert comparison.rule_attribution["indirect search displacement"] == 1
    assert comparison.remark_related_hard_violations == 0
    assert comparison.attribution_reconciles is True
    assert all("suspected" not in row["Attribution Category"] for row in comparison.newly_unscheduled_rows)


def test_preferences_create_zero_newly_unscheduled_when_scheduled() -> None:
    """Preferences should not create additional unscheduled occurrences."""
    course = make_course(remarks="Computer room preferred")
    baseline = [Assignment(course, Room("R1", 100, "physical"), TimeSlot("Monday", "09:00", 1))]
    enhanced = [Assignment(course, Room("R1", 100, "physical"), TimeSlot("Monday", "09:00", 1))]

    comparison = compare_remark_runs([course], baseline, enhanced)

    assert comparison.newly_unscheduled_rows == []
    assert comparison.attribution_total == 0


def test_export_remarks_coverage_comparison_creates_expected_workbook(tmp_path: Path) -> None:
    """The comparison report should export the requested sheets."""
    output = tmp_path / "remarks_coverage_comparison.xlsx"
    course = make_course(remarks="Need 2 rooms")
    baseline = [Assignment(course, Room("R1", 100, "physical"), TimeSlot("Monday", "09:00", 1))]
    enhanced = [
        Assignment(
            course,
            None,
            None,
            hard_violations=[
                "Could not find feasible weekly room/day/start pattern",
                "Explicit two-room requirement could not be satisfied simultaneously.",
            ],
            remark_unscheduled_reason="Explicit two-room requirement could not be satisfied simultaneously.",
        )
    ]

    export_remarks_coverage_comparison([course], baseline, enhanced, output)

    workbook = load_workbook(output)
    assert workbook.sheetnames == [
        "Run Fingerprints",
        "Overall Metrics",
        "Direct Remark Effects",
        "Indirect Remark Effects",
        "Unchanged Unscheduled",
        "Enhanced Improvements",
        "Attribution Diagnostics",
        "Attribution Reconciliation",
    ]
    reconciliation = {row[0]: row[2] for row in workbook["Attribution Reconciliation"].iter_rows(min_row=2, values_only=True)}
    assert reconciliation["Calculated enhanced unscheduled"] == "PASS"
    assert reconciliation["Attribution reconciliation"] == "PASS"


def test_comparison_tracks_recoveries_in_reconciliation() -> None:
    """Enhanced recoveries should subtract from the reconciliation equation."""
    course = make_course(remarks="Computer room preferred")
    baseline = [Assignment(course, None, None, hard_violations=["Could not find feasible pattern"])]
    enhanced = [Assignment(course, Room("R1", 100, "physical"), TimeSlot("Monday", "09:00", 1))]

    comparison = compare_remark_runs([course], baseline, enhanced)

    assert len(comparison.enhanced_improvement_rows) == 1
    assert comparison.attribution_reconciles is True


def test_comparison_synthesises_missing_unscheduled_placeholders(tmp_path: Path) -> None:
    """Demand-derived unscheduled occurrences should reconcile even without placeholders."""
    output = tmp_path / "remarks_coverage_comparison.xlsx"
    course = make_course(remarks="Need 2 rooms")
    baseline = [
        Assignment(course, Room("R1", 100, "physical"), TimeSlot("Monday", "09:00", 1)),
    ]
    enhanced: list[Assignment] = []

    comparison = export_remarks_coverage_comparison([course], baseline, enhanced, output)

    assert len(comparison.newly_unscheduled_rows) == 1
    assert comparison.attribution_reconciles is True
    assert len(comparison.attribution_diagnostic_rows) == 1
    assert comparison.attribution_diagnostic_rows[0]["Classification"] == "MISSING_ATTRIBUTION"
    workbook = load_workbook(output)
    diagnostics = list(workbook["Attribution Diagnostics"].iter_rows(min_row=2, values_only=True))
    assert len(diagnostics) == 1


def test_comparison_exports_run_fingerprints(tmp_path: Path) -> None:
    """The comparison workbook should include comparable run fingerprints."""
    output = tmp_path / "remarks_coverage_comparison.xlsx"
    course = make_course(remarks="")
    assignment = Assignment(course, Room("R1", 100, "physical"), TimeSlot("Monday", "09:00", 1))

    export_remarks_coverage_comparison(
        [course],
        [assignment],
        [assignment],
        output,
        input_course_records=1,
        scheduler_metadata={"max_candidate_patterns": 300},
    )

    workbook = load_workbook(output)
    rows = list(workbook["Run Fingerprints"].iter_rows(min_row=2, values_only=True))
    metrics = {(run, metric): value for run, metric, value in rows}
    assert metrics[("Baseline", "input_demand_fingerprint")] == metrics[("Enhanced", "input_demand_fingerprint")]
    assert metrics[("Baseline", "max_candidate_patterns")] == 300
    assert metrics[("Enhanced", "max_candidate_patterns")] == 300
