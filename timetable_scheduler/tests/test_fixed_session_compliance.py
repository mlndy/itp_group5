"""Tests for fixed-session compliance and submission-readiness helpers."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from data.fixed_sessions import FixedSessionLoaderReport, load_fixed_sessions
from data.models import Assignment, Course, FixedSession, Room, TimeSlot
from engine.constraint_checker import check_hard_constraints
from engine.fixed_issue_analysis import export_fixed_issue_workbooks
from engine.fixed_reconciliation import FixedReconciliationReport, normalise_programme_year, reconcile_fixed_sessions
from engine.guarded_generation import (
    build_guarded_generation_state,
    build_programme_completeness_rows,
    complete_programme_set,
    export_guarded_generation_report,
    quarantined_requirement_courses,
)
from engine.fixed_resolution import export_resolution_template, load_resolution_workbook
from engine.input_readiness import build_input_readiness_result
from engine.location_mapping import classify_location
from generator.fixed_scheduler import create_fixed_assignments, normalise_staff_name, validate_fixed_assignments
from generator.scheduler import generate_schedule
from optimiser.local_search import optimise_schedule_with_stats
from output.submission_validator import (
    Template2ValidationResult,
    export_template2_validation_report,
    submission_assignments,
)


def make_course(**overrides: object) -> Course:
    """Create a compact course for fixed-session tests."""
    data = {
        "module_code": "ENG1001",
        "activity": "Lecture",
        "prog_yr": "ENG/Y1",
        "class_size": 30,
        "delivery_mode": "f2f",
        "teaching_weeks": [1],
        "week_pattern": "ALL",
        "staff_ids": ["S001"],
        "duration_hrs": 1,
        "group_ids": ["ENG/Y1"],
    }
    data.update(overrides)
    return Course(**data)


def make_fixed_session(**overrides: object) -> FixedSession:
    """Create one fixed source row."""
    data = {
        "programme_year": "ENG/Y1",
        "module_code": "ENG1001",
        "group_id": "P1",
        "group_size": 30,
        "day": "Monday",
        "start_time": "09:00",
        "duration_hours": 1.0,
        "teaching_weeks": (1,),
        "locations": ("E2-07-01",),
        "staff_ids": ("Tutor A",),
        "staff_names": ("Tutor A",),
        "source_file": "fixed.xlsx",
        "source_sheet": "ENG",
        "source_row": 2,
    }
    data.update(overrides)
    return FixedSession(**data)


def write_fixed_workbook(path: Path, rows: list[list[object]]) -> None:
    """Write a minimal fixed-session workbook."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "ENG"
    worksheet.append(
        [
            "Prog/Yr",
            "Module code",
            "Group",
            "Group Size",
            "Day",
            "Start Time",
            "Duration (Hr)",
            "Weeks",
            "Location",
            "Staff 1",
        ]
    )
    for row in rows:
        worksheet.append(row)
    workbook.save(path)


def write_template2_locations(path: Path, rows: list[list[object]]) -> None:
    """Write a minimal Template 2-style Location support sheet."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Location"
    worksheet.append(["Name", "Host Key", "Capacity"])
    for row in rows:
        worksheet.append(row)
    workbook.save(path)


def write_resolution_workbook(path: Path, rows: list[list[object]]) -> None:
    """Write a supervisor resolution workbook fixture."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Resolution Decisions"
    worksheet.append(["Query ID", "Decision", "Approved Value", "Reason", "Approved By", "Approval Date", "Notes"])
    for row in rows:
        worksheet.append(row)
    workbook.save(path)


def test_fixed_loader_separates_warning_and_critical_rows(tmp_path: Path) -> None:
    """Incomplete placements should warn, while invalid fixed placements remain critical."""
    workbook_path = tmp_path / "fixed.xlsx"
    write_fixed_workbook(
        workbook_path,
        [
            ["ENG/Y1", "ENG1001", "P1", 30, "", "", 1, "1", "Seminar Room", "Tutor A"],
            ["ENG/Y1", "ENG1002", "P1", 30, "Mon", "9am", 1, "", "Seminar Room", "Tutor A"],
        ],
    )

    sessions, report = load_fixed_sessions(workbook_path)

    assert sessions == []
    assert report.source_rows == 2
    assert report.warnings == 1
    assert report.critical_errors == 1
    assert [row["loader status"] for row in report.audit_rows] == ["not loaded", "invalid"]


def test_input_readiness_preserves_warning_only_incomplete_rows(tmp_path: Path) -> None:
    """Warning-only incomplete rows should not be promoted into critical blockers."""
    workbook_path = tmp_path / "fixed.xlsx"
    write_fixed_workbook(
        workbook_path,
        [["ENG/Y1", "ENG1001", "P1", 30, "", "", 1, "1", "Seminar Room", "Tutor A"]],
    )
    fixed_sessions, loader_report = load_fixed_sessions(workbook_path)
    reconciliation = reconcile_fixed_sessions(fixed_sessions, [], loader_report)

    readiness = build_input_readiness_result(
        fixed_loader_report=loader_report,
        reconciliation_report=reconciliation,
        fixed_assignment_issues=[],
    )

    assert readiness.ready
    assert not readiness.critical_errors
    assert readiness.warnings
    assert all(issue["severity"] == "warning" for issue in readiness.warnings)


def test_fixed_assignments_resolve_exact_room_alias_and_remain_anchored() -> None:
    """Exact fixed room codes should map to official venue IDs without moving the slot."""
    rooms = [Room("E2-07-01-SR259", 40, "physical", "Seminar Room")]
    assignments, issues = create_fixed_assignments([make_fixed_session()], rooms)

    assert issues == []
    assert len(assignments) == 1
    assert assignments[0].is_fixed
    assert assignments[0].room is not None
    assert assignments[0].room.room_id == "E2-07-01-SR259"
    assert assignments[0].timeslot == TimeSlot("Monday", "09:00", 1)


def test_location_mapping_classifies_w3_exact_and_alias(tmp_path: Path) -> None:
    """W3 venue codes should resolve only when an exact authoritative room exists."""
    rooms = [Room("W3-01-03", 40, "physical", "Laboratory")]

    exact = classify_location("W3-01-03", rooms, tmp_path / "missing_template2.xlsx")
    alias = classify_location("Non-Destructive Testing Lab (W3 01 03)", rooms, tmp_path / "missing_template2.xlsx")

    assert exact.blocking_status == "non-blocking"
    assert exact.treatment == "exact internal venue fully validated"
    assert alias.blocking_status == "non-blocking"
    assert alias.candidate_venue_code == "W3-01-03"


def test_unknown_w3_venue_remains_blocking(tmp_path: Path) -> None:
    """Missing W3 rooms should not be inferred from unrelated internal venues."""
    evidence = classify_location("W3-01-99", [], tmp_path / "missing_template2.xlsx")

    assert evidence.blocking_status == "blocking"
    assert evidence.treatment == "unknown venue"


def test_exact_w3_fixed_room_is_retained_as_capacity_unverified() -> None:
    """Exact official W3 fixed codes should anchor without inventing capacity."""
    session = make_fixed_session(locations=("W3-01-03",))

    assignments, issues = create_fixed_assignments([session], [])

    assert len(assignments) == 1
    assert assignments[0].room is not None
    assert assignments[0].room.room_id == "W3-01-03"
    assert assignments[0].room.capacity == 0
    assert "capacity unavailable" in assignments[0].room.resource_type.casefold()
    assert all(issue["severity"] == "warning" for issue in issues)
    assert check_hard_constraints(assignments[0], [], enable_remark_interpretation=False) == []


def test_external_venue_does_not_reserve_internal_room() -> None:
    """Recognised external venues should anchor the row without consuming an internal room."""
    session = make_fixed_session(locations=("ENG External Venue",), source_row=12)

    assignments, issues = create_fixed_assignments([session], [])

    assert len(assignments) == 1
    assert assignments[0].room is not None
    assert assignments[0].room.room_type == "external"
    assert assignments[0].room.room_id == "ENG External Venue"
    assert all(issue["severity"] == "warning" for issue in issues)


def test_recognised_venue_without_capacity_is_classified(tmp_path: Path) -> None:
    """Template-supported venues missing CSV capacity should be visible but not fully validated."""
    template2 = tmp_path / "template2.xlsx"
    write_template2_locations(template2, [["W3-01-03", "W3-01-03", None]])

    evidence = classify_location("W3-01-03", [], template2)

    assert evidence.blocking_status == "warning"
    assert evidence.treatment == "recognised institutional venue missing capacity data"
    assert evidence.capacity is None


def test_unknown_location_remains_critical() -> None:
    """Unknown locations should continue to block fixed assignment creation."""
    assignments, issues = create_fixed_assignments([make_fixed_session(locations=("UNKNOWN ROOM",))], [])

    assert assignments == []
    assert any(issue["severity"] == "critical" for issue in issues)


def test_guarded_readiness_quarantines_incomplete_row_without_blocking(tmp_path: Path) -> None:
    """Record-level fixed errors should produce ready-with-exclusions rather than a global block."""
    workbook_path = tmp_path / "fixed.xlsx"
    write_fixed_workbook(
        workbook_path,
        [["RSE/Y1", "RSE1001", "P1", 30, "Mon", "09:00", 1, "", "R1", "Tutor A"]],
    )
    fixed_sessions, loader_report = load_fixed_sessions(workbook_path)
    reconciliation = reconcile_fixed_sessions(fixed_sessions, [], loader_report)
    guarded = build_guarded_generation_state(
        courses=[make_course()],
        rooms_loaded=1,
        fixed_sessions=fixed_sessions,
        fixed_loader_report=loader_report,
        reconciliation_report=reconciliation,
        fixed_assignments=[],
        fixed_assignment_issues=[],
    )
    readiness = build_input_readiness_result(
        fixed_loader_report=loader_report,
        reconciliation_report=reconciliation,
        fixed_assignment_issues=[],
        global_errors=guarded.global_errors,
        quarantined_requirements=guarded.quarantined_requirements,
    )

    assert readiness.ready
    assert "Ready with exclusions" in readiness.message
    assert len(readiness.quarantined_requirements) == 1
    assert readiness.quarantined_requirements[0].module_code == "RSE1001"


def test_programme_and_staff_alias_normalisation_is_conservative() -> None:
    """Safe aliases should normalise without changing source display values."""
    assert normalise_programme_year("DSC / Year 1") == "DSC/Y1"
    assert normalise_programme_year("DSC/Yr 1") == "DSC/Y1"
    assert normalise_staff_name("KHOO TECK PING .") == "KHOO TECK PING"


def test_shared_fixed_sessions_group_without_self_conflict() -> None:
    """Rows describing the same shared class should reserve resources once."""
    rooms = [Room("R1", 100, "physical", "Laboratory")]
    sessions = [
        make_fixed_session(group_id="P1", group_size=30, locations=("R1",), source_row=2),
        make_fixed_session(group_id="P2", group_size=25, locations=("R1",), source_row=3),
    ]

    assignments, issues = create_fixed_assignments(sessions, rooms)
    conflicts = validate_fixed_assignments(assignments)

    assert issues == []
    assert conflicts == []
    assert len(assignments) == 1
    assert assignments[0].course.class_size == 55
    assert assignments[0].course.group_ids == ["ENG/Y1/P1", "ENG/Y1/P2"]
    assert "fixed.xlsx:ENG:2" in (assignments[0].fixed_source or "")
    assert "fixed.xlsx:ENG:3" in (assignments[0].fixed_source or "")


def test_genuine_fixed_overlap_still_conflicts() -> None:
    """Different fixed classes sharing one physical room/time should remain blocked."""
    rooms = [Room("R1", 100, "physical", "Laboratory")]
    sessions = [
        make_fixed_session(module_code="ENG1001", locations=("R1",), staff_names=("Tutor A",), staff_ids=("Tutor A",), source_row=2),
        make_fixed_session(module_code="ENG1002", locations=("R1",), staff_names=("Tutor B",), staff_ids=("Tutor B",), source_row=3),
    ]

    assignments, issues = create_fixed_assignments(sessions, rooms)
    conflicts = validate_fixed_assignments(assignments)

    assert issues == []
    assert len(assignments) == 2
    assert any("Room clash" in issue["problem"] for issue in conflicts)


def test_authoritative_fixed_session_can_span_lunch() -> None:
    """Official fixed sessions may keep an exact lunch-spanning source placement."""
    rooms = [Room("R1", 100, "physical", "Laboratory")]
    session = make_fixed_session(start_time="09:00", duration_hours=9.0, locations=("R1",))

    assignments, mapping_issues = create_fixed_assignments([session], rooms)
    issues = validate_fixed_assignments(assignments)

    assert mapping_issues == []
    assert assignments[0].hard_violations == []
    assert not any(issue["severity"] == "critical" for issue in issues)
    assert any("AUTHORITATIVE_FIXED_LUNCH_SPAN" in issue["problem"] for issue in issues)


def test_authoritative_fixed_session_can_end_after_1800() -> None:
    """Official fixed sessions may keep exact source times beyond generated slots."""
    rooms = [Room("R1", 100, "physical", "Laboratory")]
    session = make_fixed_session(start_time="18:00", duration_hours=2.0, locations=("R1",))

    assignments, mapping_issues = create_fixed_assignments([session], rooms)
    issues = validate_fixed_assignments(assignments)

    assert mapping_issues == []
    assert assignments[0].timeslot == TimeSlot("Monday", "18:00", 1)
    assert assignments[0].course.duration_hrs == 2.0
    assert assignments[0].hard_violations == []
    assert not any(issue["severity"] == "critical" for issue in issues)
    assert any("AUTHORITATIVE_FIXED_AFTER_1800" in issue["problem"] for issue in issues)


def test_non_fixed_long_session_still_obeys_lunch_policy() -> None:
    """The fixed-source exception must not weaken non-fixed scheduling policy."""
    assignment = Assignment(
        course=make_course(duration_hrs=9.0, group_ids=["ENG/Y1/P1"]),
        room=Room("R1", 100, "physical", "Laboratory"),
        timeslot=TimeSlot("Monday", "09:00", 1),
    )

    violations = check_hard_constraints(assignment, [], enable_remark_interpretation=False)

    assert any("no free lunch block" in violation for violation in violations)


def test_fixed_tutor_and_group_clashes_still_fail() -> None:
    """Official fixed-window exceptions must not hide real resource clashes."""
    rooms = [Room("R1", 100, "physical", "Laboratory"), Room("R2", 100, "physical", "Laboratory")]
    tutor_sessions = [
        make_fixed_session(module_code="ENG1001", locations=("R1",), staff_ids=("Tutor A",), staff_names=("Tutor A",), source_row=2),
        make_fixed_session(module_code="ENG1002", locations=("R2",), staff_ids=("Tutor A",), staff_names=("Tutor A",), group_id="P2", source_row=3),
    ]
    group_sessions = [
        make_fixed_session(module_code="ENG1003", locations=("R1",), staff_ids=("Tutor A",), staff_names=("Tutor A",), source_row=4),
        make_fixed_session(module_code="ENG1004", locations=("R2",), staff_ids=("Tutor B",), staff_names=("Tutor B",), source_row=5),
    ]

    tutor_assignments, _mapping_issues = create_fixed_assignments(tutor_sessions, rooms)
    group_assignments, _mapping_issues = create_fixed_assignments(group_sessions, rooms)

    assert any("Staff clash" in issue["problem"] for issue in validate_fixed_assignments(tutor_assignments))
    assert any("Student group clash" in issue["problem"] for issue in validate_fixed_assignments(group_assignments))


def test_guarded_conflict_quarantines_both_linked_fixed_assignments() -> None:
    """Unresolved fixed conflicts should exclude every linked fixed assignment."""
    rooms = [Room("R1", 100, "physical", "Laboratory")]
    sessions = [
        make_fixed_session(module_code="ENG1001", locations=("R1",), staff_names=("Tutor A",), staff_ids=("Tutor A",), source_row=2),
        make_fixed_session(module_code="ENG1002", locations=("R1",), staff_names=("Tutor B",), staff_ids=("Tutor B",), source_row=3),
        make_fixed_session(module_code="ENG1003", locations=("R1",), day="Tuesday", source_row=4),
    ]
    assignments, mapping_issues = create_fixed_assignments(sessions, rooms)
    conflicts = validate_fixed_assignments(assignments)
    report = FixedSessionLoaderReport(workbook_path="fixed.xlsx")
    report.audit_rows = [
        {
            "source workbook": "fixed.xlsx",
            "source sheet": "ENG",
            "source row": session.source_row,
            "programme/year": session.programme_year,
            "module code": session.module_code,
            "group": session.group_id,
            "teaching weeks": "1",
            "loader status": "loaded",
            "severity": "info",
        }
        for session in sessions
    ]
    guarded = build_guarded_generation_state(
        courses=[make_course(module_code="ENG1001"), make_course(module_code="ENG1002"), make_course(module_code="ENG1003")],
        rooms_loaded=len(rooms),
        fixed_sessions=sessions,
        fixed_loader_report=report,
        reconciliation_report=FixedReconciliationReport(),
        fixed_assignments=assignments,
        fixed_assignment_issues=mapping_issues + conflicts,
    )

    anchored_sources = {assignment.fixed_source for assignment in guarded.anchored_fixed_assignments}
    assert len(guarded.quarantined_requirements) == 2
    assert len(guarded.quarantined_fixed_assignments) == 2
    assert any("ENG1003" in (source or "") or source == "fixed.xlsx:ENG:4" for source in anchored_sources)


def test_generate_schedule_reserves_initial_fixed_assignments() -> None:
    """Non-fixed scheduling should avoid room/time occupied by anchored sessions."""
    room = Room("R1", 100, "physical", "Lectorial")
    fixed = Assignment(
        make_course(module_code="ENG0001", staff_ids=["F001"], group_ids=["ENG/Y9"], prog_yr="ENG/Y9"),
        room,
        TimeSlot("Monday", "09:00", 1),
        is_fixed=True,
        fixed_source="fixed.xlsx:ENG:2",
    )

    result = generate_schedule(
        [make_course(module_code="ENG1001", staff_ids=["S002"], group_ids=["ENG/Y1"])],
        [room],
        initial_assignments=[fixed],
        allow_weekly_fallback=False,
        max_retry_assignments=0,
    )

    movable = [assignment for assignment in result if not assignment.is_fixed][0]
    assert result[0].is_fixed
    assert movable.room == room
    assert movable.timeslot != TimeSlot("Monday", "09:00", 1)


def test_quarantined_rows_do_not_block_unaffected_scheduling() -> None:
    """Guarded scheduling should continue with quarantined rows excluded from candidates."""
    room = Room("R1", 100, "physical", "Lectorial")
    session = make_fixed_session(locations=("UNKNOWN ROOM",), source_row=9)
    fixed_assignments, mapping_issues = create_fixed_assignments([session], [room])
    report = FixedSessionLoaderReport(workbook_path="fixed.xlsx")
    guarded = build_guarded_generation_state(
        courses=[make_course(module_code="ENG2001")],
        rooms_loaded=1,
        fixed_sessions=[session],
        fixed_loader_report=report,
        reconciliation_report=FixedReconciliationReport(),
        fixed_assignments=fixed_assignments,
        fixed_assignment_issues=mapping_issues,
    )

    result = generate_schedule(
        [make_course(module_code="ENG2001", staff_ids=["S002"], group_ids=["ENG/Y2"])],
        [room],
        initial_assignments=guarded.anchored_fixed_assignments,
        allow_weekly_fallback=False,
        max_retry_assignments=0,
    )

    assert guarded.quarantined_requirements
    assert all(assignment.course.module_code != session.module_code for assignment in result)
    assert any(assignment.room is not None for assignment in result)


def test_optimiser_preserves_fixed_assignment_on_early_return() -> None:
    """Early optimiser exits should still preserve fixed placements exactly."""
    room = Room("R1", 100, "physical", "Lectorial")
    fixed = Assignment(
        make_course(),
        room,
        TimeSlot("Monday", "09:00", 1),
        is_fixed=True,
        fixed_source="fixed.xlsx:ENG:2",
    )

    result = optimise_schedule_with_stats([fixed], [room], time_limit_seconds=0)

    assert result.status == "Time limit reached"
    assert result.assignments[0].is_fixed
    assert result.assignments[0].timeslot == TimeSlot("Monday", "09:00", 1)
    assert result.assignments[0].room == room


def test_submission_assignments_excludes_unresolved_or_invalid_rows() -> None:
    """Submission-ready exports should include only complete hard-valid rows."""
    scheduled = Assignment(make_course(), Room("R1", 100, "physical"), TimeSlot("Monday", "09:00", 1))
    unscheduled = Assignment(make_course(module_code="ENG1002"), None, None, hard_violations=["No room"])
    invalid = Assignment(
        make_course(module_code="ENG1003"),
        Room("R1", 100, "physical"),
        TimeSlot("Monday", "10:00", 1),
        hard_violations=["Room clash"],
    )

    assert submission_assignments([scheduled, unscheduled, invalid]) == [scheduled]


def test_submission_assignments_respects_complete_programme_and_room_mapping() -> None:
    """Submission-ready rows should be stricter than proposed timetable rows."""
    complete = Assignment(make_course(prog_yr="ENG/Y1"), Room("R1", 100, "physical"), TimeSlot("Monday", "09:00", 1))
    incomplete_programme = Assignment(make_course(module_code="ENG2002", prog_yr="ENG/Y2"), Room("R1", 100, "physical"), TimeSlot("Monday", "10:00", 1))
    missing_host_key = Assignment(make_course(module_code="ENG2003", prog_yr="ENG/Y1"), Room("W3-01-03", 0, "physical", "Recognised Venue (capacity unavailable)"), TimeSlot("Tuesday", "09:00", 1))

    rows = submission_assignments(
        [complete, incomplete_programme, missing_host_key],
        complete_programmes={"ENG/Y1"},
        rooms=[Room("R1", 100, "physical")],
    )

    assert rows == [complete]


def test_template2_validation_report_exports_required_sheets(tmp_path: Path) -> None:
    """The submission validation evidence workbook should expose all review sheets."""
    output_path = tmp_path / "template2_validation.xlsx"
    export_template2_validation_report(Template2ValidationResult(ready=False), output_path)

    workbook = load_workbook(output_path, read_only=True)
    try:
        assert {
            "Summary",
            "Programme Schedule Coverage",
            "Fixed Session Accuracy",
            "Required Field Validation",
            "Source-to-Output Reconciliation",
            "Duplicate Check",
            "Invalid Rows",
            "Submission Readiness",
        } <= set(workbook.sheetnames)
    finally:
        workbook.close()


def test_fixed_issue_workbooks_include_supervisor_queries(tmp_path: Path) -> None:
    """Root-cause exports should distinguish issue instances from affected rows."""
    session = make_fixed_session(locations=("UNKNOWN",))
    loader_report = FixedSessionLoaderReport(workbook_path="fixed.xlsx")
    reconciliation = FixedReconciliationReport()
    mapping_issue = {
        "severity": "critical",
        "source": "fixed.xlsx",
        "sheet": "ENG",
        "row": 2,
        "field": "location",
        "problem": "Location 'UNKNOWN' is neither an exact room nor a supported generic room type.",
        "recommendation": "Clarify the room.",
    }
    second_issue_same_row = {**mapping_issue, "problem": "Exact fixed room 'UNKNOWN' was not found in the venue data."}

    stats = export_fixed_issue_workbooks(
        fixed_sessions=[session],
        courses=[],
        assignments=[],
        rooms=[],
        loader_report=loader_report,
        reconciliation_report=reconciliation,
        mapping_issues=[mapping_issue, second_issue_same_row],
        conflict_issues=[],
        root_cause_path=tmp_path / "root.xlsx",
        conflict_triage_path=tmp_path / "conflicts.xlsx",
        supervisor_queries_path=tmp_path / "queries.xlsx",
    )

    assert stats["critical_issue_instances"] == 2
    assert stats["unique_affected_rows"] == 1
    workbook = load_workbook(tmp_path / "queries.xlsx", read_only=True)
    try:
        assert "Query Summary" in workbook.sheetnames
        assert "Affected Source Rows" in workbook.sheetnames
    finally:
        workbook.close()


def test_fixed_issue_workbook_uses_valid_sheet_names(tmp_path: Path) -> None:
    """Generated evidence workbooks should avoid over-length worksheet names."""
    session = make_fixed_session(locations=("UNKNOWN",))
    stats = export_fixed_issue_workbooks(
        fixed_sessions=[session],
        courses=[],
        assignments=[],
        rooms=[],
        loader_report=FixedSessionLoaderReport(workbook_path="fixed.xlsx"),
        reconciliation_report=FixedReconciliationReport(),
        mapping_issues=[
            {
                "severity": "critical",
                "source": "fixed.xlsx",
                "sheet": "ENG",
                "row": 2,
                "field": "location",
                "problem": "Location 'UNKNOWN' is neither an exact room nor a supported generic room type.",
                "recommendation": "Clarify the room.",
            }
        ],
        conflict_issues=[],
        root_cause_path=tmp_path / "root.xlsx",
        conflict_triage_path=tmp_path / "conflicts.xlsx",
        supervisor_queries_path=tmp_path / "queries.xlsx",
        supervisor_pack_path=tmp_path / "pack.xlsx",
        resolution_template_path=tmp_path / "resolution.xlsx",
        resolution_audit_path=tmp_path / "audit.xlsx",
    )

    assert stats["supervisor_decisions"] == 1
    workbook = load_workbook(tmp_path / "root.xlsx", read_only=True)
    try:
        assert all(len(name) <= 31 for name in workbook.sheetnames)
        assert "Supervisor Queries" in workbook.sheetnames
    finally:
        workbook.close()


def test_grouped_supervisor_queries_retain_all_source_rows(tmp_path: Path) -> None:
    """A grouped decision should still link every affected source row."""
    issues = [
        {
            "severity": "critical",
            "source": "fixed.xlsx",
            "sheet": "ENG",
            "row": row,
            "field": "location",
            "problem": "Exact fixed room 'W3-01-03' was not found in the venue data.",
            "recommendation": "Clarify the room.",
        }
        for row in [2, 3]
    ]
    export_fixed_issue_workbooks(
        fixed_sessions=[make_fixed_session(source_row=2), make_fixed_session(source_row=3)],
        courses=[],
        assignments=[],
        rooms=[],
        loader_report=FixedSessionLoaderReport(workbook_path="fixed.xlsx"),
        reconciliation_report=FixedReconciliationReport(),
        mapping_issues=issues,
        conflict_issues=[],
        root_cause_path=tmp_path / "root.xlsx",
        conflict_triage_path=tmp_path / "conflicts.xlsx",
        supervisor_queries_path=tmp_path / "queries.xlsx",
    )

    workbook = load_workbook(tmp_path / "queries.xlsx", read_only=True, data_only=True)
    try:
        assert workbook["Query Summary"].max_row == 2
        assert workbook["Affected Source Rows"].max_row == 3
    finally:
        workbook.close()


def test_resolution_workbook_validation(tmp_path: Path) -> None:
    """Resolution loader should reject unknown, unsupported, and unauthenticated decisions."""
    valid_path = tmp_path / "valid.xlsx"
    write_resolution_workbook(
        valid_path,
        [["Q001", "PROVIDE_TEACHING_WEEKS", "1-6,8-13", "Confirmed by owner", "Supervisor", "2026-06-26", ""]],
    )
    invalid_path = tmp_path / "invalid.xlsx"
    write_resolution_workbook(
        invalid_path,
        [
            ["Q999", "PROVIDE_TEACHING_WEEKS", "1-6", "Confirmed", "Supervisor", "2026-06-26", ""],
            ["Q001", "IGNORE_ERROR", "", "No", "Supervisor", "2026-06-26", ""],
            ["Q002", "CONFIRM_ROOM_ALIAS", "W3-01-03", "Confirmed", "", "2026-06-26", ""],
            ["Q003", "PROVIDE_TEACHING_WEEKS", "banana", "Confirmed", "Supervisor", "2026-06-26", ""],
        ],
    )

    valid = load_resolution_workbook(valid_path, {"Q001"})
    invalid = load_resolution_workbook(invalid_path, {"Q001", "Q002", "Q003"})

    assert valid.valid
    assert valid.decisions[0].approved_value == "1-6,8-13"
    assert not invalid.valid
    assert len(invalid.errors) >= 4


def test_resolution_template_lists_query_ids(tmp_path: Path) -> None:
    """Blank resolution templates should preserve linked query IDs."""
    output_path = tmp_path / "resolution_template.xlsx"

    export_resolution_template(output_path, ["Q001", "Q002"])

    workbook = load_workbook(output_path, read_only=True, data_only=True)
    try:
        assert workbook["Resolution Decisions"]["A2"].value == "Q001"
        assert workbook["Resolution Decisions"]["A3"].value == "Q002"
    finally:
        workbook.close()


def test_guarded_generation_report_exports_required_sheets(tmp_path: Path) -> None:
    """Guarded report should keep quarantine and search-failure buckets separate."""
    scheduled = Assignment(make_course(prog_yr="ENG/Y1"), Room("R1", 100, "physical"), TimeSlot("Monday", "09:00", 1))
    quarantined = [
        build_guarded_generation_state(
            courses=[make_course()],
            rooms_loaded=1,
            fixed_sessions=[make_fixed_session(source_row=9)],
            fixed_loader_report=FixedSessionLoaderReport(workbook_path="fixed.xlsx"),
            reconciliation_report=FixedReconciliationReport(),
            fixed_assignments=[],
            fixed_assignment_issues=[
                {
                    "severity": "critical",
                    "source": "fixed.xlsx",
                    "sheet": "ENG",
                    "row": 9,
                    "problem": "Invalid or missing fixed-session weeks.",
                }
            ],
        ).quarantined_requirements[0]
    ]
    demand_courses = [make_course(prog_yr="ENG/Y1"), *quarantined_requirement_courses(quarantined)]
    programme_rows = build_programme_completeness_rows(demand_courses, [scheduled], quarantined)
    output = tmp_path / "guarded.xlsx"

    export_guarded_generation_report(
        output_path=output,
        global_errors=[],
        quarantined=quarantined,
        fixed_conflict_issues=[],
        warnings=[],
        assignments=[scheduled],
        demand_courses=demand_courses,
        programme_rows=programme_rows,
        template2_summary={
            "Template 2 readiness status": "FAIL",
            "complete programme-year schedules": 7,
            "submission-ready programme-year schedules": 5,
        },
    )

    workbook = load_workbook(output, read_only=True)
    try:
        assert {
            "Summary",
            "Global Errors",
            "Quarantined Requirements",
            "Programme Completeness",
            "Submission Exclusions",
            "Resolution Guidance",
        } <= set(workbook.sheetnames)
        summary = {row[0]: row[1] for row in workbook["Summary"].iter_rows(min_row=2, values_only=True)}
        assert summary["Template 2 complete programme-years"] == 7
        assert summary["Submission-ready programme-years"] == 5
    finally:
        workbook.close()
