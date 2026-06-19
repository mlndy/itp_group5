"""Tests for stakeholder report exports."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from data.models import Assignment, Course, Room, TimeSlot
from generator.scheduler import MAX_CANDIDATE_PATTERN_LIMIT_REASON
from output.report_exporter import (
    categorise_unscheduled_reason,
    export_preflight_report,
    export_run_manifest,
    export_run_summary,
    export_stakeholder_views,
)


def make_course(**overrides: object) -> Course:
    """Create a small test course."""
    data = {
        "module_code": "ENG1001",
        "activity": "Lecture",
        "prog_yr": "ENG/YR 1",
        "class_size": 30,
        "delivery_mode": "f2f",
        "teaching_weeks": [1],
        "week_pattern": "ALL",
        "staff_ids": ["S001"],
        "duration_hrs": 2,
        "is_common_module": False,
    }
    data.update(overrides)
    return Course(**data)


def test_export_preflight_report_creates_workbook(tmp_path: Path) -> None:
    """The preflight report should export issue rows."""
    output = tmp_path / "preflight_report.xlsx"
    issues = [
        {
            "severity": "error",
            "entity_type": "course",
            "entity_id": "ENG1001",
            "issue": "Class size is not positive",
            "recommendation": "Enter a class size greater than 0.",
        }
    ]

    export_preflight_report(issues, output)

    workbook = load_workbook(output)
    assert workbook.sheetnames == ["Preflight Issues"]
    assert workbook["Preflight Issues"]["A1"].value == "severity"
    assert workbook["Preflight Issues"]["D2"].value == "Class size is not positive"


def test_export_run_summary_creates_expected_sheets(tmp_path: Path) -> None:
    """The run summary should include all stakeholder report sheets."""
    output = tmp_path / "run_summary.xlsx"
    assignments = [
        Assignment(
            course=make_course(),
            room=Room("R1", 100, "physical"),
            timeslot=TimeSlot("Monday", "09:00", 1),
        ),
        Assignment(
            course=make_course(module_code="ENG1002"),
            room=None,
            timeslot=None,
            hard_violations=["Could not find feasible weekly room/day/start pattern"],
        ),
    ]

    export_run_summary(assignments, output)

    workbook = load_workbook(output)
    assert workbook.sheetnames == [
        "Summary",
        "Hard Violations",
        "Soft Violations",
        "Unscheduled Reasons",
        "Unscheduled Analysis",
        "Unscheduled Breakdown",
        "Residual F2F Analysis",
        "Room Utilisation",
        "Resource Audit",
        "Virtual Room Detail",
        "Programme Breakdown",
        "Remarks Interpretation",
        "Validation Checks",
        "Optimisation Summary",
        "Run Metadata",
    ]
    assert workbook["Summary"]["A1"].value == "Metric"


def test_candidate_limit_reason_appears_in_run_summary(tmp_path: Path) -> None:
    """Original unscheduled reasons should remain visible in the run summary."""
    output = tmp_path / "run_summary.xlsx"
    assignments = [
        Assignment(
            course=make_course(),
            room=None,
            timeslot=None,
            hard_violations=[MAX_CANDIDATE_PATTERN_LIMIT_REASON],
        )
    ]

    export_run_summary(assignments, output)

    workbook = load_workbook(output)
    reasons = workbook["Unscheduled Reasons"]
    assert reasons["A2"].value == MAX_CANDIDATE_PATTERN_LIMIT_REASON
    assert reasons["B2"].value == 1


def test_unscheduled_reason_categorisation() -> None:
    """Candidate-limit reasons should be grouped for bottleneck analysis."""
    assignment = Assignment(course=make_course(), room=None, timeslot=None)

    category = categorise_unscheduled_reason(assignment, MAX_CANDIDATE_PATTERN_LIMIT_REASON)

    assert category == "Candidate pattern limit"


def test_unscheduled_analysis_preserves_original_reason(tmp_path: Path) -> None:
    """Unscheduled Analysis should keep the original scheduler reason visible."""
    output = tmp_path / "run_summary.xlsx"
    reason = "Could not find feasible slot for week 3"
    assignments = [Assignment(course=make_course(), room=None, timeslot=None, hard_violations=[reason])]

    export_run_summary(assignments, output)

    workbook = load_workbook(output)
    analysis = workbook["Unscheduled Analysis"]
    assert analysis["A2"].value == reason
    assert analysis["B2"].value == "No complete multi-week placement"
    assert analysis["I2"].value == 2
    assert analysis["J2"].value == 3


def test_unscheduled_analysis_includes_demand_week_counts(tmp_path: Path) -> None:
    """Unscheduled Analysis should show requirement-level week coverage."""
    output = tmp_path / "run_summary.xlsx"
    course = make_course(teaching_weeks=[1, 2, 3])
    assignments = [
        Assignment(course=course, room=Room("R1", 100, "physical"), timeslot=TimeSlot("Monday", "09:00", 1)),
        Assignment(course=course, room=None, timeslot=None, hard_violations=["Could not find feasible slot for week 2"]),
    ]

    export_run_summary(assignments, output, demand_courses=[course], input_course_records=1)

    workbook = load_workbook(output)
    analysis = workbook["Unscheduled Analysis"]
    headers = [cell.value for cell in analysis[1]]
    values = dict(zip(headers, [cell.value for cell in analysis[2]], strict=False))
    assert values["Required Week Count"] == 3
    assert values["Scheduled Week Count"] == 1
    assert values["Unscheduled Week Count"] == 2


def test_residual_f2f_analysis_classifies_physical_room_scarcity(tmp_path: Path) -> None:
    """Residual F2F Analysis should identify oversized classes with no suitable room."""
    output = tmp_path / "run_summary.xlsx"
    course = make_course(class_size=500, teaching_weeks=[1, 2], is_common_module=True)
    assignments = [
        Assignment(
            course=course,
            room=None,
            timeslot=None,
            hard_violations=["Could not find feasible slot for week 1"],
        )
    ]

    export_run_summary(assignments, output, rooms=[Room("SMALL", 100, "physical")])

    workbook = load_workbook(output)
    residual = workbook["Residual F2F Analysis"]
    headers = [cell.value for cell in residual[1]]
    values = dict(zip(headers, [cell.value for cell in residual[2]], strict=False))
    assert values["Residual Classification"] == "Physical room scarcity"
    assert values["Compatible Physical Room Count"] == 0
    assert values["Failed Reason"] == "Could not find feasible slot for week 1"


def test_residual_f2f_analysis_classifies_candidate_limit(tmp_path: Path) -> None:
    """Residual F2F Analysis should preserve candidate-limit evidence."""
    output = tmp_path / "run_summary.xlsx"
    course = make_course(class_size=30)
    assignments = [
        Assignment(
            course=course,
            room=None,
            timeslot=None,
            hard_violations=[MAX_CANDIDATE_PATTERN_LIMIT_REASON],
        )
    ]

    export_run_summary(assignments, output, rooms=[Room("R1", 100, "physical")])

    workbook = load_workbook(output)
    residual = workbook["Residual F2F Analysis"]
    headers = [cell.value for cell in residual[1]]
    values = dict(zip(headers, [cell.value for cell in residual[2]], strict=False))
    assert values["Residual Classification"] == "Search limitation"
    assert values["Candidate Limit"] == "Yes"
    assert values["Compatible Physical Room Count"] == 1
    assert values["Feasible Start Windows Before Clash Checking"] > 0


def test_programme_breakdown_marks_dsc_assignments(tmp_path: Path) -> None:
    """Engineering evidence should show whether DSC rows are included."""
    output = tmp_path / "run_summary.xlsx"
    assignments = [
        Assignment(
            course=make_course(module_code="DSC1001", prog_yr="DSC/YR 1", source_file="2510_DSC.xlsx"),
            room=Room("R1", 100, "physical"),
            timeslot=TimeSlot("Monday", "09:00", 1),
        )
    ]

    export_run_summary(assignments, output)

    workbook = load_workbook(output)
    breakdown = workbook["Programme Breakdown"]
    assert breakdown["A2"].value == "DSC/YR 1"
    assert breakdown["C2"].value == "Yes"


def _sheet_rows(workbook, sheet_name: str) -> list[tuple[object, ...]]:
    """Return worksheet rows as tuples."""
    return list(workbook[sheet_name].iter_rows(values_only=True))


def test_validation_checks_show_pass_for_scheduled_hard_safety(tmp_path: Path) -> None:
    """Validation Checks should pass when scheduled assignments have no hard violations."""
    output = tmp_path / "run_summary.xlsx"
    assignments = [
        Assignment(
            course=make_course(module_code="DSC1001", prog_yr="DSC/YR 1", source_file="2510_DSC.xlsx"),
            room=Room("R1", 100, "physical"),
            timeslot=TimeSlot("Monday", "09:00", 1),
        ),
        Assignment(
            course=make_course(module_code="ENG1002"),
            room=None,
            timeslot=None,
            hard_violations=["Could not find feasible slot for week 1"],
        ),
    ]

    export_run_summary(assignments, output)

    workbook = load_workbook(output)
    rows = _sheet_rows(workbook, "Validation Checks")
    safety = next(row for row in rows if row[0] == "Hard-constraint safety status")
    assert safety[2] == "PASS"


def test_validation_checks_detect_dsc_inclusion(tmp_path: Path) -> None:
    """Validation Checks should pass DSC inclusion when Programme Breakdown has DSC data."""
    output = tmp_path / "run_summary.xlsx"
    assignments = [
        Assignment(
            course=make_course(module_code="DSC1001", prog_yr="DSC/YR 1", source_file="2510_DSC.xlsx"),
            room=Room("R1", 100, "physical"),
            timeslot=TimeSlot("Monday", "09:00", 1),
        )
    ]

    export_run_summary(assignments, output)

    workbook = load_workbook(output)
    rows = _sheet_rows(workbook, "Validation Checks")
    dsc = next(row for row in rows if row[0] == "DSC inclusion status")
    assert dsc[2] == "PASS"


def test_summary_distinguishes_scheduled_and_all_hard_violations(tmp_path: Path) -> None:
    """Summary should separate scheduled hard violations from unscheduled feasibility failures."""
    output = tmp_path / "run_summary.xlsx"
    assignments = [
        Assignment(
            course=make_course(),
            room=Room("R1", 100, "physical"),
            timeslot=TimeSlot("Monday", "09:00", 1),
        ),
        Assignment(
            course=make_course(module_code="ENG1002"),
            room=None,
            timeslot=None,
            hard_violations=["Could not find feasible slot for week 1"],
        ),
    ]

    export_run_summary(assignments, output)

    workbook = load_workbook(output)
    summary = {row[0]: row[1] for row in _sheet_rows(workbook, "Summary")[1:]}
    assert summary["Hard violations on scheduled assignments"] == 0
    assert summary["Hard violations from unscheduled feasibility failures"] == 1
    assert summary["Hard violations on all assignments"] == 1


def test_summary_includes_invariant_demand_metrics(tmp_path: Path) -> None:
    """Summary should report stable teaching-occurrence demand metrics."""
    output = tmp_path / "run_summary.xlsx"
    course = make_course(teaching_weeks=[1, 2, 3])
    assignments = [
        Assignment(course=course, room=Room("R1", 100, "physical"), timeslot=TimeSlot("Monday", "09:00", 1)),
        Assignment(course=course, room=None, timeslot=None, hard_violations=["Could not find feasible slot for week 2"]),
    ]

    export_run_summary(assignments, output, demand_courses=[course], input_course_records=1)

    workbook = load_workbook(output)
    summary = {row[0]: row[1] for row in _sheet_rows(workbook, "Summary")[1:]}
    assert summary["Required teaching occurrences"] == 3
    assert summary["Scheduled teaching occurrences"] == 1
    assert summary["Unscheduled teaching occurrences"] == 2
    assert summary["Partially scheduled course requirements"] == 1


def test_validation_checks_pass_demand_occurrence_consistency(tmp_path: Path) -> None:
    """Validation Checks should pass when occurrence demand balances."""
    output = tmp_path / "run_summary.xlsx"
    course = make_course(teaching_weeks=[1, 2, 3])
    assignments = [Assignment(course=course, room=None, timeslot=None, hard_violations=["Could not find feasible pattern"])]

    export_run_summary(assignments, output, demand_courses=[course], input_course_records=1)

    workbook = load_workbook(output)
    rows = _sheet_rows(workbook, "Validation Checks")
    demand = next(row for row in rows if row[0] == "Demand occurrence consistency")
    assert demand[2] == "PASS"


def test_run_summary_records_metadata(tmp_path: Path) -> None:
    """Run Metadata should record Engineering command settings."""
    output = tmp_path / "run_summary.xlsx"
    assignments = [
        Assignment(
            course=make_course(),
            room=Room("R1", 100, "physical"),
            timeslot=TimeSlot("Monday", "09:00", 1),
        )
    ]

    export_run_summary(assignments, output, metadata={"scope": "eng", "max_candidate_patterns": 300})

    workbook = load_workbook(output)
    metadata = {row[0]: row[1] for row in _sheet_rows(workbook, "Run Metadata")[1:]}
    assert metadata["scope"] == "eng"
    assert metadata["max_candidate_patterns"] == 300


def test_optimisation_summary_sheet_exists(tmp_path: Path) -> None:
    """Run summary should record optimiser acceptance metrics."""
    output = tmp_path / "run_summary.xlsx"
    assignments = [
        Assignment(
            course=make_course(),
            room=Room("R1", 100, "physical"),
            timeslot=TimeSlot("Monday", "09:00", 1),
        )
    ]

    export_run_summary(
        assignments,
        output,
        optimisation_summary={
            "optimisation_enabled": "Yes",
            "status": "Accepted improved schedule",
            "soft_violations_before": 4,
            "soft_violations_after": 2,
            "coverage_unchanged_status": "PASS",
            "hard_safety_status": "PASS",
            "soft_score_not_worsened_status": "PASS",
        },
    )

    workbook = load_workbook(output)
    summary = {row[0]: row[1] for row in _sheet_rows(workbook, "Optimisation Summary")[1:]}
    assert summary["Optimisation enabled"] == "Yes"
    assert summary["Soft violations before optimisation"] == 4
    assert summary["Soft violations after optimisation"] == 2


def test_skipped_optimisation_is_reported(tmp_path: Path) -> None:
    """Run summary should show skipped status when no optimisation metrics are supplied."""
    output = tmp_path / "run_summary.xlsx"
    assignments = [
        Assignment(
            course=make_course(),
            room=Room("R1", 100, "physical"),
            timeslot=TimeSlot("Monday", "09:00", 1),
        )
    ]

    export_run_summary(assignments, output)

    workbook = load_workbook(output)
    summary = {row[0]: row[1] for row in _sheet_rows(workbook, "Optimisation Summary")[1:]}
    assert summary["Optimisation enabled"] == "No"
    assert summary["Status"] == "Skipped"


def test_resource_audit_sheet_exists(tmp_path: Path) -> None:
    """Run summary should include virtual-room supply and online demand evidence."""
    output = tmp_path / "run_summary.xlsx"
    course = make_course(delivery_mode="Online - Synchronous", teaching_weeks=[1, 2])
    assignments = [
        Assignment(course, Room("ONLINE", 9999, "virtual"), TimeSlot("Monday", "09:00", 1)),
    ]

    export_run_summary(
        assignments,
        output,
        demand_courses=[course],
        input_course_records=1,
        rooms=[Room("ONLINE", 9999, "virtual"), Room("R1", 100, "physical")],
    )

    workbook = load_workbook(output)
    audit = {row[0]: row[1] for row in _sheet_rows(workbook, "Resource Audit")[1:]}
    assert audit["Loaded virtual room count"] == 1
    assert audit["Virtual room policy"] == "Shared delivery-mode placeholder"
    assert "tutor or student-group clash checks" in audit["Virtual room policy note"]
    assert audit["Required online teaching occurrences"] == 2
    assert audit["Scheduled online teaching occurrences"] == 1
    assert "Virtual Room Detail" in workbook.sheetnames


def test_stakeholder_views_export_expected_sheets(tmp_path: Path) -> None:
    """Stakeholder views should provide timetable and exception-review tabs."""
    output = tmp_path / "stakeholder_views.xlsx"
    scheduled = Assignment(
        make_course(staff_names=["Tutor A"], group_ids=["ENG/YR 1"]),
        Room("R1", 100, "physical"),
        TimeSlot("Monday", "09:00", 1),
    )
    unscheduled = Assignment(
        make_course(module_code="ENG1002", class_size=500, source_file="input.xlsx"),
        None,
        None,
        hard_violations=["Could not find feasible slot for week 1"],
    )

    export_stakeholder_views([scheduled, unscheduled], [Room("R1", 100, "physical")], output)

    workbook = load_workbook(output)
    assert workbook.sheetnames == [
        "Programme Timetable",
        "Tutor Timetable",
        "Room Timetable",
        "Exception Queue",
        "Special Requests Review",
        "Special Requests Summary",
    ]
    queue_headers = [cell.value for cell in workbook["Exception Queue"][1]]
    assert "Recommended Operational Action" in queue_headers
    assert "Review Status" in queue_headers
    assert workbook["Exception Queue"]["J2"].value


def test_run_summary_includes_remarks_interpretation_sheet(tmp_path: Path) -> None:
    """Run summary should explain interpreted remark rules."""
    output = tmp_path / "run_summary.xlsx"
    course = make_course(remarks="2 rooms", source_file="requirements.xlsx", source_sheet="Module", source_row=4)
    assignments = [
        Assignment(
            course=course,
            room=Room("R1", 20, "physical"),
            timeslot=TimeSlot("Monday", "09:00", 1),
            additional_rooms=(Room("R2", 20, "physical"),),
        )
    ]

    export_run_summary(assignments, output)

    workbook = load_workbook(output)
    sheet = workbook["Remarks Interpretation"]
    headers = [cell.value for cell in sheet[1]]
    values = dict(zip(headers, [cell.value for cell in sheet[2]], strict=False))
    assert values["Raw Remark"] == "2 rooms"
    assert values["Detected Rule"] == "multiple_room_requirement"
    assert values["Applied Status"] == "Applied"
    assert values["Assigned Rooms"] == "R1, R2"


def test_stakeholder_views_include_special_requests_review(tmp_path: Path) -> None:
    """Stakeholder views should include plain-language special-request handling."""
    output = tmp_path / "stakeholder_views.xlsx"
    assignment = Assignment(
        make_course(remarks="Additional rooms for quizzes"),
        Room("R1", 100, "physical"),
        TimeSlot("Monday", "09:00", 1),
    )

    export_stakeholder_views([assignment], [Room("R1", 100, "physical")], output)

    workbook = load_workbook(output)
    headers = [cell.value for cell in workbook["Special Requests Review"][1]]
    values = dict(zip(headers, [cell.value for cell in workbook["Special Requests Review"][2]], strict=False))
    assert values["Special Request"] == "Additional rooms for quizzes"
    assert values["Needs Manual Review"] == "Yes"


def test_remarks_interpretation_handles_room_type_application(tmp_path: Path) -> None:
    """Room-type applied-status checks should not fail report generation."""
    output = tmp_path / "run_summary.xlsx"
    course = make_course(remarks="must use computer lab")
    assignment = Assignment(
        course,
        Room("COMP-LAB-01", 40, "physical", resource_type="Computer Lab"),
        TimeSlot("Monday", "09:00", 1),
    )

    export_run_summary([assignment], output)

    workbook = load_workbook(output)
    rows = _sheet_rows(workbook, "Remarks Interpretation")
    values = dict(zip(rows[0], rows[1], strict=False))
    assert values["Detected Rule"] == "room_type"
    assert values["Applied Status"] == "Applied"


def test_run_manifest_exports_template_validation_and_traceability(tmp_path: Path) -> None:
    """Run manifest should capture reproducibility and Template/traceability checks."""
    output = tmp_path / "run_manifest.xlsx"
    timetable = tmp_path / "timetable.xlsx"
    workbook = load_workbook(Path("input/Upload template_System (Template 2).xlsx"))
    workbook.save(timetable)
    course = make_course(source_file="requirements.xlsx")
    assignments = [Assignment(course, Room("R1", 100, "physical"), TimeSlot("Monday", "09:00", 1))]

    export_run_manifest(
        [course],
        assignments,
        output,
        metadata={"scope": "eng", "skip_optimisation": True, "input_workbook": "selected_template1.xlsx"},
        rooms=[Room("R1", 100, "physical")],
        output_files={"timetable": timetable},
        template2_path=Path("input/Upload template_System (Template 2).xlsx"),
    )

    exported = load_workbook(output)
    assert exported.sheetnames == [
        "Run Manifest",
        "Soft Constraint Weights",
        "Soft Rule Baseline",
        "Template Validation",
        "Traceability",
    ]
    manifest = {row[0]: row[1] for row in _sheet_rows(exported, "Run Manifest")[1:]}
    assert manifest["scope"] == "eng"
    assert manifest["Input role"] == "Template 1 - Scheduling Requirements"
    assert manifest["Input workbook"] == "selected_template1.xlsx"
    assert manifest["Output-template role"] == "Template 2 - Proposed Timetable"
    assert manifest["Output template"].endswith("Upload template_System (Template 2).xlsx")
    assert manifest["Generated timetable"] == str(timetable)
    assert manifest["validation_status"] == "PASS"
    validation = {row[0]: row[1] for row in _sheet_rows(exported, "Template Validation")[1:]}
    assert validation["Template 2 output structure retained"] == "PASS"
    traceability_headers = [cell.value for cell in exported["Traceability"][1]]
    assert "Source File" in traceability_headers


def test_optimisation_summary_includes_weighted_and_rule_breakdown(tmp_path: Path) -> None:
    """Optimisation Summary should support new weighted score and per-rule rows."""
    output = tmp_path / "run_summary.xlsx"
    assignments = [Assignment(make_course(), Room("R1", 100, "physical"), TimeSlot("Monday", "09:00", 1))]

    export_run_summary(
        assignments,
        output,
        optimisation_summary={
            "optimisation_enabled": "Yes",
            "status": "Improved",
            "weighted_soft_score_before": 10,
            "weighted_soft_score_after": 8,
            "Before soft rule - Short campus day": 1,
            "After soft rule - Short campus day": 0,
        },
    )

    workbook = load_workbook(output)
    summary = {row[0]: row[1] for row in _sheet_rows(workbook, "Optimisation Summary")[1:]}
    assert summary["Weighted soft score before optimisation"] == 10
    assert summary["Weighted soft score after optimisation"] == 8
    assert summary["Before soft rule - Short campus day"] == 1
