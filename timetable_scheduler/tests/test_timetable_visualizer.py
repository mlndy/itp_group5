"""Tests for calendar-style timetable visual exports."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from data.models import Assignment, Course, Room, TimeSlot
from output.timetable_visualizer import (
    allocate_lanes,
    build_programme_visual_entries,
    build_room_visual_entries,
    build_tutor_visual_entries,
    export_timetable_visuals,
    safe_sheet_name,
    scheduled_assignment_id,
    validate_visual_exports,
)


def make_course(**overrides: object) -> Course:
    """Create a compact course for visualisation tests."""
    data = {
        "module_code": "ENG1001",
        "activity": "Lecture",
        "prog_yr": "ENG/Y1",
        "class_size": 30,
        "delivery_mode": "f2f",
        "teaching_weeks": [1],
        "week_pattern": "ALL",
        "staff_ids": ["S001"],
        "duration_hrs": 1,
        "staff_names": ["Tutor A"],
        "source_file": "source.xlsx",
        "source_sheet": "Module",
        "source_row": 2,
        "group_ids": ["ENG/Y1"],
    }
    data.update(overrides)
    return Course(**data)


def make_assignment(
    *,
    course: Course | None = None,
    room: Room | None = None,
    day: str = "Monday",
    start: str = "09:00",
    week: int = 1,
    hard_violations: list[str] | None = None,
    is_fixed: bool = False,
    fixed_source: str | None = None,
) -> Assignment:
    """Create a scheduled assignment unless room is explicitly None."""
    course = course or make_course()
    return Assignment(
        course=course,
        room=room if room is not None else Room("R1", 40, "physical", "Classroom"),
        timeslot=TimeSlot(day, start, week),
        hard_violations=hard_violations or [],
        is_fixed=is_fixed,
        fixed_source=fixed_source,
    )


def programme_rows() -> list[dict[str, object]]:
    """Return programme completeness rows for tests."""
    return [
        {
            "Programme/Year": "ENG/Y1",
            "Total Required Occurrences": 1,
            "Valid Exported Rows": 1,
            "Quarantined Occurrences": 0,
            "Unscheduled Search Failures": 0,
            "Status": "COMPLETE",
            "Submission-Ready Status": "PASS",
        }
    ]


def test_visual_entries_use_scheduled_hard_valid_assignments_only() -> None:
    """Unscheduled and hard-invalid assignments should not become visual blocks."""
    scheduled = make_assignment()
    unscheduled = Assignment(make_course(module_code="ENG2001"), None, None)
    invalid = make_assignment(course=make_course(module_code="ENG3001"), hard_violations=["Room clash"])

    entries = build_programme_visual_entries([scheduled, unscheduled, invalid], programme_rows())

    assert [entry.assignment_id for entry in entries] == [scheduled_assignment_id(scheduled)]
    assert entries[0].module_code == "ENG1001"


def test_visual_entries_retain_assignment_ids_and_fixed_marker() -> None:
    """Visual rows should preserve traceable assignment IDs and fixed status."""
    assignment = make_assignment(is_fixed=True, fixed_source="fixed.xlsx|ENG|9")

    entry = build_programme_visual_entries([assignment], programme_rows())[0]

    assert entry.assignment_id.startswith("fixed.xlsx|ENG|9")
    assert entry.is_fixed


def test_lane_allocation_separates_overlaps_and_reuses_free_lanes() -> None:
    """Overlapping classes use separate lanes while later classes can reuse one."""
    first = build_programme_visual_entries([make_assignment(start="09:00")], programme_rows())[0]
    overlapping = build_programme_visual_entries(
        [make_assignment(course=make_course(module_code="ENG2001"), start="09:30")],
        programme_rows(),
    )[0]
    later = build_programme_visual_entries(
        [make_assignment(course=make_course(module_code="ENG3001"), start="11:00")],
        programme_rows(),
    )[0]

    lanes = allocate_lanes([later, overlapping, first])

    assert lanes[first.assignment_id] == lanes[later.assignment_id]
    assert lanes[overlapping.assignment_id] != lanes[first.assignment_id]
    assert lanes == allocate_lanes([first, overlapping, later])


def test_disjoint_week_same_time_does_not_overwrite() -> None:
    """Same-time entries in different weeks still receive deterministic lanes."""
    week_one = build_programme_visual_entries([make_assignment(week=1)], programme_rows())[0]
    week_two = build_programme_visual_entries(
        [make_assignment(course=make_course(module_code="ENG2001"), week=2)],
        programme_rows(),
    )[0]

    lanes = allocate_lanes([week_one, week_two])

    assert len(set(lanes.values())) == 2


def test_safe_sheet_name_handles_long_duplicates_invalid_and_external_names() -> None:
    """Worksheet names should be safe, short and deterministic."""
    used: set[str] = set()
    first = safe_sheet_name("Very Long Tutor / Name With [Invalid] Characters And Extra Text", used, "Tutor")
    second = safe_sheet_name("Very Long Tutor / Name With [Invalid] Characters And Extra Text", used, "Tutor")
    external = safe_sheet_name("ENG External Venue / Special:Room", used, "Room")

    assert len(first) <= 31
    assert len(second) <= 31
    assert first != second
    assert "/" not in external and ":" not in external


def test_shared_session_expands_to_programmes_but_dedupes_tutor_and_room() -> None:
    """Shared sessions should appear per programme and once per tutor/room occupancy."""
    course = make_course(
        is_common_module=True,
        group_ids=["DSC/Y1", "EDE/Y1"],
        staff_names=["Tutor A"],
    )
    assignment = make_assignment(course=course, fixed_source="shared|source")

    assert len(build_programme_visual_entries([assignment], programme_rows())) == 3
    assert len(build_tutor_visual_entries([assignment])) == 1
    assert len(build_room_visual_entries([assignment])) == 1


def test_room_visual_entries_exclude_online_and_include_external() -> None:
    """Online sessions should not consume room views, while external venues remain visible."""
    online = make_assignment(
        course=make_course(module_code="ONL1001", delivery_mode="online"),
        room=Room("ONLINE_ROOM", 9999, "virtual", "Online"),
    )
    external = make_assignment(
        course=make_course(module_code="EXT1001"),
        room=Room("ENG External Venue", 999, "external", "External Venue"),
    )

    entries = build_room_visual_entries([online, external])

    assert [entry.module_code for entry in entries] == ["EXT1001"]
    assert entries[0].is_external


def test_export_visual_workbooks_and_validation_pass(tmp_path: Path) -> None:
    """Visual exporter should create the three workbooks and PASS validation."""
    assignment = make_assignment()

    result = export_timetable_visuals(
        assignments=[assignment],
        rooms=[assignment.room],
        programme_rows=programme_rows(),
        programme_path=tmp_path / "Programme_Timetable_Visuals.xlsx",
        tutor_path=tmp_path / "Tutor_Timetable_Visuals.xlsx",
        room_path=tmp_path / "Room_Timetable_Visuals.xlsx",
        validation_path=tmp_path / "timetable_visualisation_validation.xlsx",
    )

    assert result.status == "PASS"
    assert result.programme_sheets == 1
    assert result.tutor_sheets == 1
    assert result.room_sheets == 1
    for path in [result.programme_path, result.tutor_path, result.room_path, result.validation_path]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            assert "Index" in workbook.sheetnames or "Summary" in workbook.sheetnames
        finally:
            workbook.close()


def test_programme_visual_sheet_shows_incomplete_banner(tmp_path: Path) -> None:
    """Incomplete programme-years should display a warning banner."""
    assignment = make_assignment()
    rows = [
        {
            "Programme/Year": "ENG/Y1",
            "Total Required Occurrences": 2,
            "Valid Exported Rows": 1,
            "Quarantined Occurrences": 1,
            "Unscheduled Search Failures": 0,
            "Status": "INCOMPLETE_QUARANTINED_INPUT",
            "Submission-Ready Status": "FAIL",
        }
    ]

    result = export_timetable_visuals(
        assignments=[assignment],
        rooms=[assignment.room],
        programme_rows=rows,
        programme_path=tmp_path / "Programme_Timetable_Visuals.xlsx",
        tutor_path=tmp_path / "Tutor_Timetable_Visuals.xlsx",
        room_path=tmp_path / "Room_Timetable_Visuals.xlsx",
        validation_path=tmp_path / "timetable_visualisation_validation.xlsx",
    )

    workbook = load_workbook(result.programme_path, read_only=True, data_only=True)
    try:
        programme_sheet = [name for name in workbook.sheetnames if name.startswith("Prog_")][0]
        assert "Incomplete: input requirements excluded" in workbook[programme_sheet]["A3"].value
    finally:
        workbook.close()


def test_validation_detects_tutor_overlap() -> None:
    """Tutor and room visual validation should flag overlapping entries."""
    left = make_assignment()
    right = make_assignment(course=make_course(module_code="ENG2001"), start="09:30")
    programme_entries = build_programme_visual_entries([left, right], programme_rows())
    tutor_entries = build_tutor_visual_entries([left, right])
    room_entries = build_room_visual_entries([left, right])

    validation = validate_visual_exports(
        assignments=[left, right],
        rooms=[left.room],
        programme_entries=programme_entries,
        tutor_entries=tutor_entries,
        room_entries=room_entries,
        programme_sheets=1,
        tutor_sheets=1,
        room_sheets=1,
        programme_rows=programme_rows(),
    )

    assert validation["status"] == "FAIL"
    assert validation["overlaps"]


def test_validation_does_not_treat_unknown_tutors_as_same_person() -> None:
    """Missing tutor identity should not create a synthetic tutor hard clash."""
    left = make_assignment(course=make_course(staff_ids=[], staff_names=["TBC"]))
    right = make_assignment(course=make_course(module_code="ENG2001", staff_ids=[], staff_names=["TBC"]), start="09:30")
    programme_entries = build_programme_visual_entries([left, right], programme_rows())
    tutor_entries = build_tutor_visual_entries([left, right])
    room_entries: list = []

    validation = validate_visual_exports(
        assignments=[left, right],
        rooms=[left.room],
        programme_entries=programme_entries,
        tutor_entries=tutor_entries,
        room_entries=room_entries,
        programme_sheets=1,
        tutor_sheets=1,
        room_sheets=0,
        programme_rows=programme_rows(),
    )

    assert validation["overlaps"] == []
