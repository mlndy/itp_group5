"""Unit tests for loader robustness and diagnostics."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from data.loader import (
    ConsolidatedScheduleValidationError,
    WorkbookRole,
    detect_workbook_role,
    export_loader_report,
    load_consolidated_schedule,
    load_courses_from_folder,
    load_courses_from_requirements,
    workbook_appears_template2_output,
)

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = PROJECT_ROOT / "Data"
DATA_TEMPLATE1 = DATA_DIR / "Requirements Template.xlsx"
DATA_TEMPLATE2 = DATA_DIR / "Upload template_System (Template 2).xlsx"


def _write_requirement_workbook(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    """Write a small workbook in the expected requirements format."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Module"
    sheet.append([None] * len(headers))
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def _write_template2_output_workbook(path: Path, sheet_name: str = "Timetable") -> None:
    """Write a small proposed-timetable output workbook."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(
        [
            "Module",
            "Class Type",
            "Template",
            "Group",
            "Day",
            "Start",
            "End",
            "Class Size",
            "Room1",
            "Staff1",
            "Tri Week",
            "Activity Hostkey",
        ]
    )
    sheet.append(["ENG1001", "Lecture", 1, "All", "Mon", "0900", "1100", 50, "R1", "Tutor", "1,2", "ENG1001-LEC"])
    workbook.save(path)


def test_dataset_template1_is_detected_as_requirements_input() -> None:
    """The confirmed dataset Template 1 workbook should be accepted as requirements input."""
    assert detect_workbook_role(DATA_TEMPLATE1) == WorkbookRole.TEMPLATE1_REQUIREMENTS
    courses = load_consolidated_schedule(DATA_TEMPLATE1)
    assert courses
    assert courses[0].module_code


def test_dataset_template2_is_detected_as_timetable_output() -> None:
    """The confirmed dataset Template 2 workbook should be rejected as UI input."""
    assert detect_workbook_role(DATA_TEMPLATE2) == WorkbookRole.TEMPLATE2_TIMETABLE
    try:
        load_consolidated_schedule(DATA_TEMPLATE2)
    except ConsolidatedScheduleValidationError as exc:
        assert "generated timetable" in str(exc)
    else:  # pragma: no cover - defensive branch
        raise AssertionError("Template 2 output workbook should be rejected")


def test_workbook_role_detection_ignores_filename_for_template1(tmp_path: Path) -> None:
    """A Template 1 workbook renamed to Template 2 should still be accepted as input."""
    renamed = tmp_path / "Template 2.xlsx"
    _write_requirement_workbook(
        renamed,
        ["Prog/Yr", "Class Size", "Module code", "Activity", "Delivery Mode", "Weeks", "Staff ID 1"],
        [["ENG/YR 1", 30, "ENG1001", "Lecture", "f2f", "1,2", "S001"]],
    )

    assert detect_workbook_role(renamed) == WorkbookRole.TEMPLATE1_REQUIREMENTS
    assert load_consolidated_schedule(renamed)[0].module_code == "ENG1001"


def test_workbook_role_detection_ignores_filename_for_template2(tmp_path: Path) -> None:
    """A Template 2 workbook renamed to Template 1 should still be rejected."""
    renamed = tmp_path / "Template 1.xlsx"
    _write_template2_output_workbook(renamed)

    assert detect_workbook_role(renamed) == WorkbookRole.TEMPLATE2_TIMETABLE


def test_successful_workbook_parsing_supports_column_variations(tmp_path: Path) -> None:
    """The loader should parse safe header variations without failing."""
    path = tmp_path / "eng_workbook.xlsx"
    _write_requirement_workbook(
        path,
        ["Prog Yr", "Enrolment", "Module", "Activity", "Mode", "Tri Week", "Staff 1", "Staff ID 1"],
        [["EPE/YR 1", 42, "UCM1001", "Lecture", "Online - Synchronous", "1,3-4", "Tan", "A1001"]],
    )

    courses, diagnostic = load_courses_from_requirements(path, common_modules={"UCM1001"})

    assert len(courses) == 1
    assert courses[0].module_code == "UCM1001"
    assert courses[0].is_common_module is True
    assert diagnostic.status == "parsed"
    assert diagnostic.rows_parsed == 1
    assert diagnostic.rows_skipped == 0
    assert diagnostic.reason == "Workbook parsed successfully"


def test_consolidated_schedule_loader_converts_template1_rows_to_courses(tmp_path: Path) -> None:
    """A valid consolidated Template 1 workbook should become Course objects."""
    path = tmp_path / "consolidated.xlsx"
    _write_requirement_workbook(
        path,
        ["Prog/Yr", "Class Size", "Module Code", "Activity", "Delivery Mode", "Teaching Weeks", "Staff 1", "Staff ID 1"],
        [["ENG/YR 1", "42 students", "eng1001", "Lecture", "online sync", "1,3-4", "Ada", "S001"]],
    )

    courses = load_consolidated_schedule(path)

    assert len(courses) == 1
    course = courses[0]
    assert course.prog_yr == "ENG/YR 1"
    assert course.module_code == "ENG1001"
    assert course.activity == "Lecture"
    assert course.class_size == 42
    assert course.delivery_mode == "Online - Synchronous"
    assert course.teaching_weeks == [1, 3, 4]
    assert course.staff_ids == ["S001"]
    assert course.staff_names == ["Ada"]


def test_consolidated_schedule_supports_header_row_one_and_inherited_values(tmp_path: Path) -> None:
    """Header row 1 and inherited programme, module, and class size should parse."""
    path = tmp_path / "header_row_one.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Module"
    sheet.append(["Prog/Yr", "Class Size", "Module Code", "Activity", "Delivery Mode", "Weeks", "Staff ID 1", "Staff ID 2", "Staff ID 3", "Staff ID 4"])
    sheet.append(["ENG/YR 1", 40, "ENG1001", "Lecture", "f2f", "1,2", "S001", "S002", None, None])
    sheet.append([None, None, None, "Tutorial", "f2f", 3, "S003", None, "S004", "S005"])
    workbook.save(path)

    courses = load_consolidated_schedule(path)

    assert [course.activity for course in courses] == ["Lecture", "Tutorial"]
    assert courses[1].prog_yr == "ENG/YR 1"
    assert courses[1].module_code == "ENG1001"
    assert courses[1].class_size == 40
    assert courses[1].teaching_weeks == [3]
    assert courses[1].staff_ids == ["S003", "S004", "S005"]


def test_template2_output_workbook_is_rejected_as_consolidated_input(tmp_path: Path) -> None:
    """Proposed-timetable output headers should be rejected as Template 1 input."""
    path = tmp_path / "looks_like_input.xlsx"
    _write_template2_output_workbook(path)

    assert workbook_appears_template2_output(path)
    try:
        load_consolidated_schedule(path)
    except ConsolidatedScheduleValidationError as exc:
        assert "generated timetable" in str(exc)
    else:  # pragma: no cover - defensive branch
        raise AssertionError("Template 2 output workbook should be rejected")


def test_template2_detection_uses_headers_not_filename(tmp_path: Path) -> None:
    """Template 2 rejection should be based on worksheet headers."""
    path = tmp_path / "Consolidated Schedule.xlsx"
    _write_template2_output_workbook(path, sheet_name="Anything")

    assert workbook_appears_template2_output(path)


def test_consolidated_schedule_missing_columns_has_concise_error(tmp_path: Path) -> None:
    """Missing Template 1 columns should return the role-specific error."""
    path = tmp_path / "missing.xlsx"
    _write_requirement_workbook(path, ["Prog/Yr", "Activity"], [["ENG/YR 1", "Lecture"]])

    try:
        load_consolidated_schedule(path)
    except ConsolidatedScheduleValidationError as exc:
        assert "consolidated schedule format" in str(exc)
    else:  # pragma: no cover - defensive branch
        raise AssertionError("Missing columns should fail")


def test_consolidated_schedule_empty_workbook_has_concise_error(tmp_path: Path) -> None:
    """A Template 1 workbook with no course rows should fail clearly."""
    path = tmp_path / "empty.xlsx"
    _write_requirement_workbook(path, ["Prog/Yr", "Class Size", "Module Code", "Activity", "Delivery Mode", "Teaching Weeks"], [])

    try:
        load_consolidated_schedule(path)
    except ConsolidatedScheduleValidationError as exc:
        assert "contains no scheduling records" in str(exc)
    else:  # pragma: no cover - defensive branch
        raise AssertionError("Empty Template 1 should fail")


def test_missing_required_columns_are_reported(tmp_path: Path) -> None:
    """Missing required columns should be returned as an explicit diagnostic."""
    path = tmp_path / "missing_columns.xlsx"
    _write_requirement_workbook(path, ["Prog Yr", "Activity"], [["EPE/YR 1", "Lecture"]])

    courses, diagnostic = load_courses_from_requirements(path)

    assert courses == []
    assert diagnostic.status == "skipped"
    assert diagnostic.file_path == str(path)
    assert diagnostic.sheet_name == "Module"
    assert diagnostic.missing_columns
    assert any("Class Size" in item or "Enrolment" in item for item in diagnostic.missing_columns)
    assert diagnostic.reason == "Missing required columns"


def test_skipped_workbook_is_reported_and_exported(tmp_path: Path) -> None:
    """The folder loader should expose every skipped workbook in its report."""
    good_path = tmp_path / "good.xlsx"
    bad_path = tmp_path / "bad.xlsx"
    report_path = tmp_path / "loader_report.xlsx"

    _write_requirement_workbook(
        good_path,
        ["Prog Yr", "Enrolment", "Module", "Activity", "Mode", "Teaching Weeks", "Staff ID 1"],
        [["METS/YR 3", 30, "MET3001", "Tutorial", "f2f", "1,2", "A2001"]],
    )
    _write_requirement_workbook(bad_path, ["Prog Yr", "Activity"], [["METS/YR 3", "Tutorial"]])

    courses, report = load_courses_from_folder(tmp_path)

    assert len(courses) == 1
    assert report.skipped_workbooks == 1
    assert report.partial_workbooks == 0
    assert len(report.workbooks) == 2

    skipped = next(item for item in report.workbooks if item.status == "skipped")
    assert skipped.file_path == str(bad_path)
    assert skipped.reason
    assert skipped.missing_columns

    export_loader_report(report, report_path)
    assert report_path.exists()

    workbook = load_workbook(report_path, read_only=True)
    assert workbook.sheetnames == ["Summary", "Diagnostics"]
    summary = workbook["Summary"]
    diagnostics = workbook["Diagnostics"]
    assert summary["A2"].value == "Workbooks scanned"
    assert summary["B5"].value == 1
    assert diagnostics["A2"].value is not None
