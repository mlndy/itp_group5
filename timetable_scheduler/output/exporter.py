"""Export timetable and violation reports to Excel."""

from __future__ import annotations

from collections import defaultdict
from copy import deepcopy
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import ACTIVITY_TYPE_CODES, DAY_ABBREVIATIONS, DEFAULT_TEMPLATE2_FILE, ENABLE_REMARK_INTERPRETATION
from data.models import Assignment
from engine.constraint_checker import annotate_schedule_violations, assignment_end_time
from engine.remarks_interpreter import assignment_room_ids, assignment_rooms


def _activity_type_code(activity: str) -> str:
    """Map activity text to Template 2 activity code."""
    activity_lower = activity.lower()
    for keyword, code in ACTIVITY_TYPE_CODES.items():
        if keyword in activity_lower:
            return code
    return activity[:3].upper() if activity else "ACT"


def _format_template_time(time_text: str | None) -> str | None:
    """Convert HH:MM to Template 2 HHMM text."""
    if not time_text:
        return None
    return time_text.replace(":", "")


def _staff_value(assignment: Assignment, index: int) -> str | None:
    """Return staff name when available, otherwise staff ID."""
    if index < len(assignment.course.staff_names):
        return assignment.course.staff_names[index]
    if index < len(assignment.course.staff_ids):
        return assignment.course.staff_ids[index]
    return None


def _annotated_snapshot(
    assignments: list[Assignment],
    enable_remark_interpretation: bool = ENABLE_REMARK_INTERPRETATION,
) -> list[Assignment]:
    """Return a deep-copied, annotated version of the schedule."""
    snapshot = deepcopy(assignments)
    annotate_schedule_violations(
        snapshot,
        enable_remark_interpretation=enable_remark_interpretation,
    )
    return snapshot


def _remark_text(assignment: Assignment) -> str | None:
    """Compose a readable remark for Template 2 output."""
    parts: list[str] = []
    if assignment.course.remarks:
        parts.append(assignment.course.remarks)
    if assignment.selected_delivery_mode == "hybrid":
        parts.append("Interpreted as hybrid delivery")
    if len(assignment_rooms(assignment)) > 1:
        parts.append(f"Assigned rooms: {assignment_room_ids(assignment)}")
    if assignment.room is None or assignment.timeslot is None:
        parts.append("Unscheduled assignment")
    if assignment.hard_violations:
        parts.append("Hard: " + "; ".join(assignment.hard_violations))
    if assignment.soft_violations:
        parts.append("Soft: " + "; ".join(assignment.soft_violations))
    return " | ".join(parts) if parts else None


def _room_id(assignment: Assignment) -> str | None:
    """Return the assigned room ID when available."""
    return assignment.room.room_id if assignment.room else None


def _room2_id(assignment: Assignment) -> str | None:
    """Return the second assigned room ID when available."""
    rooms = assignment_rooms(assignment)
    return rooms[1].room_id if len(rooms) > 1 else None


def _template_row_values(assignment: Assignment) -> dict[str, object]:
    """Convert one assignment to logical Template 2 values."""
    course = assignment.course
    timeslot = assignment.timeslot
    end_time = None
    if timeslot is not None:
        end_time = assignment_end_time(assignment)

    room_id = _room_id(assignment)
    room2_id = _room2_id(assignment)
    return {
        "Module": course.module_code,
        "Class Type": course.activity,
        "Template": 1,
        "Group": " / ".join(course.group_ids or [course.prog_yr]),
        "Day": DAY_ABBREVIATIONS.get(timeslot.day, timeslot.day) if timeslot else None,
        "Start": _format_template_time(timeslot.start_time) if timeslot else None,
        "End": _format_template_time(end_time) if end_time else None,
        "Class Size": course.class_size,
        "Sector": "PUNGGOL",
        "RoomGrouping": None,
        "Room1": room_id,
        "Room2": room2_id,
        "StaffGrouping": None,
        "Staff1": _staff_value(assignment, 0),
        "Staff2": _staff_value(assignment, 1),
        "Tri Week": str(timeslot.week) if timeslot else None,
        "Recording Mode": "A0",
        "Remark": _remark_text(assignment),
        "FMTS Tri Start Week": 1,
        "Activity Hostkey": f"{course.module_code}-2510-ENG-UGRD-PU-{_activity_type_code(course.activity)}/{course.prog_yr}",
        "SIS Module Code": f"{course.module_code}-2510-ENG-UGRD-PU",
        "Term": 2510,
        "Activity Type": _activity_type_code(course.activity),
        "Duration": course.duration_hrs,
        "Staff Suitability ID": "#N/A",
        "SIS Staff ID": course.staff_ids[0] if len(course.staff_ids) >= 1 else "#N/A",
        "SIS Staff ID.1": course.staff_ids[1] if len(course.staff_ids) >= 2 else "#N/A",
        "Zone Hoskey": "PUNGGOL",
        "Zone Hostkey": "PUNGGOL",
        "Location Suitability ID": "#N/A",
        "Location Hostkey": room_id,
        "Location Hostkey.1": room_id,
        "Programme/Year": course.prog_yr,
        "Delivery Mode": assignment.selected_delivery_mode or course.delivery_mode,
        "Status": assignment.status,
    }


def _template_cell_value(header: str, occurrence: int, row_values: dict[str, object]) -> object:
    """Return the workbook cell value for one Template 2 header cell."""
    if header == "SIS Staff ID":
        if occurrence == 0:
            return row_values[header]
        if occurrence == 1:
            return row_values.get("SIS Staff ID.1")
        return "#N/A"
    if header == "Location Hostkey":
        return row_values[header]
    if header == "Zone Hoskey":
        return row_values[header]
    return row_values.get(header)


def _write_template_timetable_sheet(workbook, assignments: list[Assignment]) -> None:
    """Populate the Template 2 Timetable sheet in-place."""
    if "Timetable" not in workbook.sheetnames:
        workbook.create_sheet("Timetable")
    sheet = workbook["Timetable"]
    headers = [cell.value for cell in sheet[1]]
    if not headers or all(header is None for header in headers):
        headers = list(_template_row_values(assignments[0]).keys()) if assignments else []
        for column, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=column, value=header)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            cell.value = None
    for row_index, assignment in enumerate(assignments, start=2):
        row_values = _template_row_values(assignment)
        header_counts: dict[str, int] = defaultdict(int)
        for column_index, header in enumerate(headers, start=1):
            if header is None:
                continue
            occurrence = header_counts[header]
            header_counts[header] += 1
            sheet.cell(row=row_index, column=column_index, value=_template_cell_value(header, occurrence, row_values))


def assignment_to_row(assignment: Assignment) -> dict[str, object]:
    """Convert one assignment to a Template 2-compatible row."""
    return _template_row_values(assignment)


def assignments_to_dataframe(
    assignments: list[Assignment],
    enable_remark_interpretation: bool = ENABLE_REMARK_INTERPRETATION,
) -> pd.DataFrame:
    """Convert assignments to a timetable DataFrame."""
    snapshot = _annotated_snapshot(
        assignments,
        enable_remark_interpretation=enable_remark_interpretation,
    )
    rows = [assignment_to_row(assignment) for assignment in snapshot]
    return pd.DataFrame(rows)


def violations_to_dataframe(assignments: list[Assignment], violation_type: str) -> pd.DataFrame:
    """Convert hard or soft violations to a DataFrame."""
    columns = [
        "Type",
        "Module",
        "Activity",
        "Programme/Year",
        "Week",
        "Day",
        "Start",
        "Room",
        "Violation",
    ]
    rows: list[dict[str, object]] = []
    for assignment in assignments:
        violations = assignment.hard_violations if violation_type == "hard" else assignment.soft_violations
        for violation in violations:
            rows.append(
                {
                    "Type": violation_type.upper(),
                    "Module": assignment.course.module_code,
                    "Activity": assignment.course.activity,
                    "Programme/Year": assignment.course.prog_yr,
                    "Week": assignment.timeslot.week if assignment.timeslot else None,
                    "Day": assignment.timeslot.day if assignment.timeslot else None,
                    "Start": assignment.timeslot.start_time if assignment.timeslot else None,
                    "Room": assignment_room_ids(assignment),
                    "Violation": violation,
                }
            )
    return pd.DataFrame(rows, columns=columns)


def summary_to_dataframe(
    assignments: list[Assignment],
    enable_remark_interpretation: bool = ENABLE_REMARK_INTERPRETATION,
) -> pd.DataFrame:
    """Create summary metrics for stakeholder reporting."""
    snapshot = _annotated_snapshot(
        assignments,
        enable_remark_interpretation=enable_remark_interpretation,
    )
    total = len(snapshot)
    hard = sum(len(item.hard_violations) for item in snapshot)
    soft = sum(len(item.soft_violations) for item in snapshot)
    scheduled = sum(1 for item in snapshot if item.room is not None and item.timeslot is not None)
    return pd.DataFrame(
        [
            {"Metric": "Assignments", "Value": total},
            {"Metric": "Scheduled assignments", "Value": scheduled},
            {"Metric": "Unscheduled assignments", "Value": total - scheduled},
            {"Metric": "Hard violations", "Value": hard},
            {"Metric": "Soft violations", "Value": soft},
            {"Metric": "Feasibility rate", "Value": f"{scheduled / total:.1%}" if total else "0.0%"},
        ]
    )


def _autosize_and_style(path: Path) -> None:
    """Apply simple readable styling to an exported workbook."""
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            width = min(max(max_length + 2, 10), 38)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(path)


def export_schedule(
    assignments: list[Assignment],
    output_path: str | Path,
    template2_path: str | Path | None = None,
    enable_remark_interpretation: bool = ENABLE_REMARK_INTERPRETATION,
) -> None:
    """Export the final timetable, preferring the provided Template 2 workbook."""
    output_path = Path(output_path)
    template2_path = Path(template2_path or DEFAULT_TEMPLATE2_FILE)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    snapshot = _annotated_snapshot(
        assignments,
        enable_remark_interpretation=enable_remark_interpretation,
    )

    if template2_path.exists():
        workbook = load_workbook(template2_path)
        _write_template_timetable_sheet(workbook, snapshot)
        workbook.save(output_path)
        return

    timetable_df = assignments_to_dataframe(
        snapshot,
        enable_remark_interpretation=enable_remark_interpretation,
    )
    hard_df = violations_to_dataframe(snapshot, "hard")
    soft_df = violations_to_dataframe(snapshot, "soft")
    summary_df = summary_to_dataframe(
        snapshot,
        enable_remark_interpretation=enable_remark_interpretation,
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        timetable_df.to_excel(writer, sheet_name="Timetable", index=False)
        hard_df.to_excel(writer, sheet_name="Hard Violations", index=False)
        soft_df.to_excel(writer, sheet_name="Soft Violations", index=False)
    _autosize_and_style(output_path)


def export_violations(
    assignments: list[Assignment],
    output_path: str | Path,
    enable_remark_interpretation: bool = ENABLE_REMARK_INTERPRETATION,
) -> None:
    """Export only hard and soft violation reports."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    snapshot = _annotated_snapshot(
        assignments,
        enable_remark_interpretation=enable_remark_interpretation,
    )
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        violations_to_dataframe(snapshot, "hard").to_excel(writer, sheet_name="Hard Violations", index=False)
        violations_to_dataframe(snapshot, "soft").to_excel(writer, sheet_name="Soft Violations", index=False)
    _autosize_and_style(output_path)
