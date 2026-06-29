"""Create and validate the submission-ready Template 2 workbook."""

from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
import re

import pandas as pd
from openpyxl import load_workbook

from data.models import Assignment, Course, FixedSession, Room
from engine.demand_metrics import build_demand_metrics, build_requirement_demands
from engine.programme_year import (
    canonical_programme_year,
    canonical_programme_year_from_source,
    clean_programme_year_text,
    identify_programme_year,
    normalise_programme_year,
    programme_year_report_value,
    programme_year_report_value_from_source,
)
from output.exporter import assignment_to_row, export_schedule

REQUIRED_SUBMISSION_COLUMNS = [
    "Module",
    "Class Type",
    "Group",
    "Day",
    "Start",
    "End",
    "Class Size",
    "Room1",
    "Staff1",
    "Tri Week",
    "Activity Type",
    "Duration",
    "Location Hostkey",
]
REQUIRED_MIN_PROGRAMME_YEARS = 20
UNKNOWN_YEAR_MARKERS = {"", "ALL YEARS", "ALL YEAR", "COMMON", "MIXED", "TBC", "TBD", "N/A", "NA"}


@dataclass(slots=True)
class Template2ValidationResult:
    """Validation outcome for the saved submission-ready Template 2 workbook."""

    ready: bool
    summary: dict[str, object] = field(default_factory=dict)
    programme_rows: list[dict[str, object]] = field(default_factory=list)
    required_field_rows: list[dict[str, object]] = field(default_factory=list)
    fixed_accuracy_rows: list[dict[str, object]] = field(default_factory=list)
    source_reconciliation_rows: list[dict[str, object]] = field(default_factory=list)
    duplicate_rows: list[dict[str, object]] = field(default_factory=list)
    invalid_rows: list[dict[str, object]] = field(default_factory=list)
    missing_programme_year_rows: list[dict[str, object]] = field(default_factory=list)
    identity_audit_rows: list[dict[str, object]] = field(default_factory=list)
    incomplete_programme_year_rows: list[dict[str, object]] = field(default_factory=list)
    saved_workbook_rows: list[dict[str, object]] = field(default_factory=list)


def normalise_template2_year(value: object) -> str:
    """Return canonical Y<number> text, or blank for unknown/mixed years."""
    text = _clean_text(value)
    if text in UNKNOWN_YEAR_MARKERS:
        return ""
    if re.search(r"(?:YR|YEAR|Y)?\s*[1-9]\s*/\s*[1-9]", text):
        return ""
    if re.fullmatch(r"[1-9]", text):
        return f"Y{text}"
    match = re.search(r"\b(?:YEAR|YR|Y)\s*([1-9])\b", text)
    return f"Y{match.group(1)}" if match else ""


def normalise_template2_programme_year(value: object, fallback_programme: object = "") -> str:
    """Return <PROGRAMME>/Y<number>, or blank when the year is unknown."""
    programme_year = canonical_programme_year(value)
    if programme_year:
        return programme_year
    year = normalise_template2_year(value)
    if year and fallback_programme:
        return canonical_programme_year(f"{fallback_programme}/{year}") or ""
    return ""


def saved_row_programme_year(row: dict[str, object]) -> str:
    """Return canonical programme-year from one saved Template 2 row."""
    for field, value in _saved_identity_candidates(row):
        programme_year = normalise_template2_programme_year(value)
        if programme_year:
            return programme_year
    return ""


def _saved_identity_candidates(row: dict[str, object]) -> list[tuple[str, object]]:
    """Return saved-row identity candidates in evidence order."""
    candidates: list[tuple[str, object]] = []
    for field in ("Programme/Year", "Group"):
        if row.get(field) not in (None, ""):
            candidates.append((field, row.get(field)))
    hostkey = row.get("Activity Hostkey")
    if hostkey not in (None, ""):
        text = str(hostkey)
        candidates.append(("Activity Hostkey", text.split("/", 1)[1] if "/" in text else text))
    return candidates


def _clean_text(value: object) -> str:
    """Return normalised uppercase text for conservative Template 2 parsing."""
    return re.sub(r"\s+", " ", str(value or "").strip().upper())


def _clean_programme(value: object) -> str:
    """Return a compact programme code without guessing a missing year."""
    text = _clean_text(value)
    text = re.sub(r"\b(?:YEAR|YR|Y)\s*[1-9]\b", "", text)
    text = text.strip(" /-_")
    if "/" in text:
        parts = [part.strip(" /-_") for part in text.split("/") if part.strip(" /-_")]
        for part in reversed(parts):
            if not normalise_template2_year(part):
                text = part
                break
    matches = re.findall(r"[A-Z][A-Z0-9]*(?:\s*\+\s*[A-Z][A-Z0-9]*)*", text)
    return matches[-1].replace(" ", "") if matches else ""


def _programme_before_year(value: object) -> str:
    """Return the programme text immediately before an explicit year token."""
    text = _clean_text(value)
    match = re.search(r"\b(?:YEAR|YR|Y)\s*[1-9]\b", text)
    if not match:
        return ""
    return _clean_programme(text[: match.start()])


def export_all_valid_scheduled_schedule(
    assignments: list[Assignment],
    output_path: Path,
    template2_path: Path,
    enable_remark_interpretation: bool = True,
    rooms: list[Room] | None = None,
) -> None:
    """Export every row-level valid scheduled assignment to Template 2."""
    export_schedule(
        submission_assignments(assignments, rooms=rooms),
        output_path,
        template2_path=template2_path,
        enable_remark_interpretation=enable_remark_interpretation,
        aggregate_teaching_weeks=False,
    )


def _is_submission_assignment(assignment: Assignment) -> bool:
    """Return True for complete scheduled rows eligible for submission."""
    return assignment.room is not None and assignment.timeslot is not None and not assignment.hard_violations


def _has_required_submission_values(assignment: Assignment) -> bool:
    """Return True when required Template 2 values can be populated."""
    course = assignment.course
    return all(
        [
            course.module_code,
            course.activity,
            course.group_ids or course.prog_yr,
            course.class_size > 0,
            assignment.room is not None and assignment.room.room_id,
            course.staff_names or course.staff_ids,
            assignment.timeslot is not None,
            course.duration_hrs > 0,
        ]
    )


def submission_assignments(
    assignments: list[Assignment],
    complete_programmes: set[str] | None = None,
    rooms: list[Room] | None = None,
) -> list[Assignment]:
    """Return scheduled assignments only, excluding unresolved placeholders."""
    valid_room_ids = _valid_rooms(rooms or []) if rooms is not None else None
    rows: list[Assignment] = []
    for assignment in assignments:
        if not _is_submission_assignment(assignment):
            continue
        if not _has_required_submission_values(assignment):
            continue
        programme = canonical_programme_year_from_source(assignment.course.prog_yr, assignment.course.source_file)
        if complete_programmes is not None and programme not in complete_programmes:
            continue
        if valid_room_ids is not None and assignment.room is not None and assignment.room.room_id not in valid_room_ids:
            continue
        rows.append(assignment)
    return rows


def export_submission_ready_schedule(
    assignments: list[Assignment],
    output_path: Path,
    template2_path: Path,
    enable_remark_interpretation: bool = True,
    complete_programmes: set[str] | None = None,
    rooms: list[Room] | None = None,
) -> None:
    """Export a Template 2 workbook containing only complete scheduled rows."""
    export_schedule(
        submission_assignments(assignments, complete_programmes=complete_programmes, rooms=rooms),
        output_path,
        template2_path=template2_path,
        enable_remark_interpretation=enable_remark_interpretation,
        aggregate_teaching_weeks=True,
    )


def _headers(workbook_path: Path) -> list[object]:
    """Return Timetable sheet headers."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if "Timetable" not in workbook.sheetnames:
            return []
        return [cell.value for cell in workbook["Timetable"][1]]
    finally:
        workbook.close()


def _row_dicts(workbook_path: Path) -> list[dict[str, object]]:
    """Return Timetable rows as dictionaries."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if "Timetable" not in workbook.sheetnames:
            return []
        sheet = workbook["Timetable"]
        headers = [cell.value for cell in sheet[1]]
        rows: list[dict[str, object]] = []
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value not in (None, "") for value in row):
                continue
            values = {str(header): row[position] for position, header in enumerate(headers) if header is not None}
            values["_row"] = row_index
            rows.append(values)
        return rows
    finally:
        workbook.close()


def _duration_valid(row: dict[str, object]) -> bool:
    """Return True when Template 2 start/end/duration values agree."""
    try:
        start = str(row.get("Start") or "").zfill(4)
        end = str(row.get("End") or "").zfill(4)
        duration = float(row.get("Duration") or 0)
        start_minutes = int(start[:2]) * 60 + int(start[2:])
        end_minutes = int(end[:2]) * 60 + int(end[2:])
    except (TypeError, ValueError):
        return False
    return end_minutes > start_minutes and abs((end_minutes - start_minutes) / 60 - duration) < 0.01


def _valid_rooms(rooms: list[Room]) -> set[str]:
    """Return valid room IDs for submission validation."""
    return {room.room_id for room in rooms}


def _required_field_validation(rows: list[dict[str, object]], rooms: list[Room]) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    """Validate required row fields and return field rows plus invalid rows."""
    valid_rooms = _valid_rooms(rooms)
    field_rows: list[dict[str, object]] = []
    invalid_rows: list[dict[str, object]] = []
    for row in rows:
        row_number = row["_row"]
        row_issues: list[str] = []
        for column in REQUIRED_SUBMISSION_COLUMNS:
            ok = row.get(column) not in (None, "")
            field_rows.append({"Row": row_number, "Field": column, "Status": "PASS" if ok else "FAIL", "Value": row.get(column)})
            if not ok:
                row_issues.append(f"Missing {column}")
        if not _duration_valid(row):
            row_issues.append("Duration does not match start/end")
        room = row.get("Room1")
        if room not in valid_rooms:
            row_issues.append(f"Room '{room}' is not in venue data")
        if row_issues:
            invalid_rows.append({"Row": row_number, "Module": row.get("Module"), "Issues": "; ".join(row_issues)})
    return field_rows, invalid_rows


def _duplicate_rows(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    """Return duplicate output occurrence rows."""
    keys = [
        (
            row.get("Module"),
            row.get("Class Type"),
            row.get("Group"),
            row.get("Day"),
            row.get("Start"),
            row.get("Tri Week"),
            row.get("Room1"),
        )
        for row in rows
    ]
    counts = Counter(keys)
    return [
        {"Module": key[0], "Class Type": key[1], "Group": key[2], "Day": key[3], "Start": key[4], "Tri Week": key[5], "Room1": key[6], "Count": count}
        for key, count in counts.items()
        if count > 1
    ]


def _invalid_row_numbers(invalid_rows: list[dict[str, object]]) -> set[int]:
    """Return row numbers already rejected by row-level validation."""
    numbers: set[int] = set()
    for row in invalid_rows:
        try:
            numbers.add(int(row.get("Row") or 0))
        except (TypeError, ValueError):
            continue
    return numbers


def _saved_programme_year_counts(
    rows: list[dict[str, object]],
    invalid_rows: list[dict[str, object]],
) -> Counter[str]:
    """Count programme-years from actual saved Template 2 rows only."""
    invalid_numbers = _invalid_row_numbers(invalid_rows)
    counts: Counter[str] = Counter()
    for row in rows:
        try:
            row_number = int(row.get("_row") or 0)
        except (TypeError, ValueError):
            row_number = 0
        if row_number in invalid_numbers:
            continue
        programme_year = saved_row_programme_year(row)
        if programme_year:
            counts[programme_year] += 1
    return counts


def _saved_workbook_verification_rows(
    rows: list[dict[str, object]],
    invalid_rows: list[dict[str, object]],
) -> list[dict[str, object]]:
    """Return actual saved workbook rows with canonical identity evidence."""
    invalid_numbers = _invalid_row_numbers(invalid_rows)
    evidence: list[dict[str, object]] = []
    for row in rows:
        row_number = int(row.get("_row") or 0)
        candidates = _saved_identity_candidates(row)
        raw = next((str(value) for _, value in candidates if value not in (None, "")), "")
        identity = identify_programme_year(raw)
        canonical = saved_row_programme_year(row)
        evidence.append(
            {
                "Saved Row": row_number,
                "Module": row.get("Module"),
                "Class Type": row.get("Class Type"),
                "Raw identity": raw,
                "Canonical programme-year": canonical,
                "Normalisation rule": identity.rule,
                "Identity status": "confident" if canonical else identity.status,
                "Row validation status": "FAIL" if row_number in invalid_numbers else "PASS",
            }
        )
    return evidence


def _template2_assignment_id(assignment: Assignment) -> tuple[object, ...]:
    """Return a row identity matching the Template 2 saved row fields."""
    row = assignment_to_row(assignment)
    return (
        row.get("Module"),
        row.get("Class Type"),
        row.get("Group"),
        row.get("Day"),
        row.get("Start"),
        row.get("Tri Week"),
        row.get("Room1"),
    )


def _template2_row_id(row: dict[str, object]) -> tuple[object, ...]:
    """Return a saved-row identity for source-to-output reconciliation."""
    return (
        row.get("Module"),
        row.get("Class Type"),
        row.get("Group"),
        row.get("Day"),
        row.get("Start"),
        str(row.get("Tri Week") or ""),
        row.get("Room1"),
    )


def _source_reference(course: Course) -> dict[str, object]:
    """Return source workbook references for an audit row."""
    return {
        "Source Workbook": course.source_file,
        "Source Sheet": course.source_sheet,
        "Source Row": course.source_row,
    }


def _assignment_exclusion_row(
    assignment: Assignment,
    rule: str,
    reason: str,
    missing_field: str = "",
) -> dict[str, object]:
    """Return one Template 2 exclusion audit row."""
    course = assignment.course
    return {
        "Programme-Year": programme_year_report_value_from_source(course.prog_yr, course.source_file),
        "Module": course.module_code,
        "Activity": course.activity,
        "Scheduled Row ID": "/".join(str(part or "") for part in _template2_assignment_id(assignment)),
        "Exclusion Rule": rule,
        "Exclusion Reason": reason,
        "Missing Field": missing_field,
        **_source_reference(course),
    }


def _missing_required_fields(assignment: Assignment) -> list[str]:
    """Return Template 2 required values missing from a scheduled assignment."""
    course = assignment.course
    missing: list[str] = []
    if not course.module_code:
        missing.append("Module")
    if not course.activity:
        missing.append("Class Type")
    if not (course.group_ids or course.prog_yr):
        missing.append("Group")
    if course.class_size <= 0:
        missing.append("Class Size")
    if assignment.room is None or not assignment.room.room_id:
        missing.append("Room1")
    if not (course.staff_names or course.staff_ids):
        missing.append("Staff1")
    if assignment.timeslot is None:
        missing.extend(["Day", "Start", "End", "Tri Week"])
    if course.duration_hrs <= 0:
        missing.append("Duration")
    return missing


def build_template2_exclusion_audit_rows(
    assignments: list[Assignment],
    complete_programmes: set[str] | None = None,
    rooms: list[Room] | None = None,
) -> list[dict[str, object]]:
    """Return row-level reasons assignments are absent from strict Template 2."""
    valid_room_ids = _valid_rooms(rooms or []) if rooms is not None else None
    rows: list[dict[str, object]] = []
    for assignment in assignments:
        programme = canonical_programme_year_from_source(assignment.course.prog_yr, assignment.course.source_file)
        if not _is_submission_assignment(assignment):
            reason = "; ".join(assignment.hard_violations) if assignment.hard_violations else "Assignment is unscheduled."
            rows.append(_assignment_exclusion_row(assignment, "unscheduled-or-hard-invalid", reason))
            continue
        missing = _missing_required_fields(assignment)
        if missing:
            rows.append(
                _assignment_exclusion_row(
                    assignment,
                    "missing-required-template2-value",
                    "One or more required Template 2 fields cannot be populated.",
                    ", ".join(missing),
                )
            )
            continue
        if valid_room_ids is not None and assignment.room is not None and assignment.room.room_id not in valid_room_ids:
            rows.append(
                _assignment_exclusion_row(
                    assignment,
                    "invalid-room-mapping",
                    f"Room '{assignment.room.room_id}' is not in the venue data.",
                    "Room1",
                )
            )
            continue
        if complete_programmes is not None and programme not in complete_programmes:
            rows.append(
                _assignment_exclusion_row(
                    assignment,
                    "incomplete-programme-year-filter",
                    "Valid scheduled row retained in all-valid output but excluded from strict submission-ready workbook.",
                )
            )
    return rows


def export_template2_exclusion_audit(
    assignments: list[Assignment],
    output_path: Path,
    complete_programmes: set[str] | None = None,
    rooms: list[Room] | None = None,
) -> None:
    """Export row-level Template 2 exclusion reasons."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    rows = build_template2_exclusion_audit_rows(assignments, complete_programmes=complete_programmes, rooms=rooms)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame(rows).to_excel(writer, sheet_name="Exclusion Audit", index=False)


def _programme_coverage(
    source_requirements: list[Course],
    assignments: list[Assignment],
    submission_rows: list[dict[str, object]],
    invalid_rows: list[dict[str, object]] | None = None,
    all_valid_rows: list[dict[str, object]] | None = None,
    quarantined_requirements: list[object] | None = None,
) -> list[dict[str, object]]:
    """Build programme-year schedule coverage rows."""
    required: Counter[str] = Counter()
    scheduled: Counter[str] = Counter()
    fixed_counts: Counter[str] = Counter()
    hard_counts: Counter[str] = Counter()
    quarantined_counts: Counter[str] = Counter()
    raw_identities: dict[str, set[str]] = defaultdict(set)
    input_requirements: Counter[str] = Counter()
    for demand in build_requirement_demands(source_requirements, assignments):
        programme = programme_year_report_value_from_source(demand.course.prog_yr, demand.course.source_file)
        raw_identities[programme].add(str(demand.course.prog_yr))
        input_requirements[programme] += 1
        required[programme] += demand.required_week_count
        scheduled[programme] += demand.scheduled_week_count
    for assignment in assignments:
        programme = programme_year_report_value_from_source(assignment.course.prog_yr, assignment.course.source_file)
        raw_identities[programme].add(str(assignment.course.prog_yr))
        if assignment.is_fixed:
            fixed_counts[programme] += 1
        if assignment.room is not None and assignment.timeslot is not None:
            hard_counts[programme] += len(assignment.hard_violations)
    for item in quarantined_requirements or []:
        raw_programme = getattr(item, "programme_year", "")
        programme = programme_year_report_value_from_source(raw_programme, getattr(item, "source_file", ""))
        raw_identities[programme].add(str(raw_programme))
        quarantined_counts[programme] += int(getattr(item, "affected_occurrences", 0) or 0)
    submission_count = _saved_programme_year_counts(submission_rows, invalid_rows or [])
    all_valid_count = _saved_programme_year_counts(all_valid_rows or [], [])
    for programme in submission_count:
        raw_identities[programme].add(programme)
    for programme in all_valid_count:
        raw_identities[programme].add(programme)

    rows: list[dict[str, object]] = []
    for programme in sorted(set(required) | set(submission_count) | set(all_valid_count) | set(quarantined_counts)):
        required_count = required[programme]
        scheduled_count = scheduled[programme]
        quarantined_count = quarantined_counts[programme]
        search_failures = max(required_count - scheduled_count - quarantined_count, 0)
        hard_count = hard_counts[programme]
        canonical = canonical_programme_year(programme)
        complete = (
            bool(canonical)
            and required_count > 0
            and quarantined_count == 0
            and search_failures == 0
            and scheduled_count == required_count
            and hard_count == 0
        )
        saved_rows = submission_count[programme]
        submission_ready = complete and saved_rows > 0
        failure_reasons: list[str] = []
        if not canonical:
            failure_reasons.append("Ambiguous or unknown programme-year identity")
        if quarantined_count:
            failure_reasons.append("Quarantined required occurrences remain")
        if search_failures:
            failure_reasons.append("Scheduler search-failure occurrences remain")
        if scheduled_count != required_count:
            failure_reasons.append("Scheduled occurrences do not equal recorded required occurrences")
        if hard_count:
            failure_reasons.append("Scheduled hard violations exist")
        if not saved_rows:
            failure_reasons.append("No valid rows in saved strict submission workbook")
        rows.append(
            {
                "Raw identities observed": "; ".join(sorted(raw_identities[programme])),
                "Normalised Programme/Year": canonical or programme,
                "Canonical programme-year": canonical,
                "Required Assignments": input_requirements[programme],
                "Required Teaching Occurrences": required_count,
                "Recorded occurrences": required_count,
                "Quarantined occurrences": quarantined_count,
                "Schedulable occurrences": max(required_count - quarantined_count, 0),
                "Fixed Assignments": fixed_counts[programme],
                "Scheduled Assignments": scheduled_count,
                "Scheduled occurrences": scheduled_count,
                "Search failures": search_failures,
                "Hard violations": hard_count,
                "Unscheduled Assignments": max(required_count - scheduled_count, 0),
                "All-valid rows": all_valid_count[programme],
                "Submission Rows": saved_rows,
                "Strict saved rows": saved_rows,
                "Actual Saved Submission Rows": saved_rows,
                "Completion Percentage": (scheduled_count / required_count * 100) if required_count else 0,
                "Complete Schedule Status": "PASS" if complete else "FAIL",
                "Included In Submission": "Yes" if saved_rows else "No",
                "Submission-Ready Status": "PASS" if submission_ready else "FAIL",
                "Counts Toward Minimum 20": "Yes" if submission_ready else "No",
                "Exclusion Reason": "" if submission_ready else "; ".join(failure_reasons),
            }
        )
    return rows


def _row_level_reconciliation(
    assignments: list[Assignment],
    submission_rows: list[dict[str, object]],
    invalid_rows: list[dict[str, object]],
    complete_programmes: set[str] | None = None,
) -> list[dict[str, object]]:
    """Map scheduled assignment rows to actual saved Template 2 rows."""
    saved_row_numbers = _invalid_row_numbers(invalid_rows)
    saved_index: dict[tuple[object, ...], list[dict[str, object]]] = defaultdict(list)
    for row in submission_rows:
        if int(row.get("_row") or 0) not in saved_row_numbers:
            saved_index[_template2_row_id(row)].append(row)

    rows: list[dict[str, object]] = []
    for assignment in assignments:
        course = assignment.course
        row_id = _template2_assignment_id(assignment)
        saved_matches = saved_index.get(row_id, [])
        programme = programme_year_report_value_from_source(course.prog_yr, course.source_file)
        if saved_matches:
            status = "Saved in submission workbook"
            reason = ""
            saved_row = saved_matches[0].get("_row")
        elif not _is_submission_assignment(assignment):
            status = "Excluded"
            reason = "; ".join(assignment.hard_violations) if assignment.hard_violations else "Assignment is unscheduled."
            saved_row = ""
        elif complete_programmes is not None and programme not in complete_programmes:
            status = "Excluded"
            reason = "Programme-year is incomplete; row remains available in all-valid output."
            saved_row = ""
        else:
            status = "Missing from saved workbook"
            reason = "No matching saved Template 2 row found."
            saved_row = ""
        rows.append(
            {
                "Programme-Year": programme,
                "Module": course.module_code,
                "Activity": course.activity,
                "Scheduled Row ID": "/".join(str(part or "") for part in row_id),
                "Saved Timetable Row": saved_row,
                "Status": status,
                "Reason": reason,
                **_source_reference(course),
            }
        )
    return rows


def _missing_programme_year_rows(programme_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    """Return scheduled programme-years absent from the saved submission workbook."""
    rows: list[dict[str, object]] = []
    for row in programme_rows:
        if int(row.get("Scheduled Assignments") or 0) <= 0:
            continue
        if int(row.get("Submission Rows") or 0) > 0:
            continue
        rows.append(
            {
                "Programme-Year": row.get("Canonical programme-year") or row.get("Normalised Programme/Year"),
                "Scheduled Assignments": row.get("Scheduled Assignments"),
                "Unscheduled Assignments": row.get("Unscheduled Assignments"),
                "Exclusion Reason": row.get("Exclusion Reason"),
            }
        )
    return rows


def _identity_audit_rows(values: list[object]) -> list[dict[str, object]]:
    """Return raw-to-canonical programme-year normalisation evidence."""
    rows: list[dict[str, object]] = []
    seen: set[tuple[str, str, str]] = set()
    for value in values:
        identity = identify_programme_year(value)
        key = (str(value or ""), identity.canonical, identity.status)
        if key in seen:
            continue
        seen.add(key)
        rows.append(
            {
                "Raw value": str(value or ""),
                "Canonical value": identity.canonical,
                "Normalisation rule": identity.rule,
                "Confidence/status": identity.status,
            }
        )
    return rows


def _incomplete_programme_rows(programme_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    """Return programme-years failing the authoritative completeness gate."""
    return [
        row
        for row in programme_rows
        if row.get("Complete Schedule Status") != "PASS" or row.get("Submission-Ready Status") != "PASS"
    ]


def _fixed_accuracy(assignments: list[Assignment]) -> list[dict[str, object]]:
    """Return fixed placement accuracy rows from the exported assignments."""
    rows: list[dict[str, object]] = []
    for assignment in assignments:
        if not assignment.is_fixed:
            continue
        rows.append(
            {
                "Fixed Source": assignment.fixed_source,
                "Module": assignment.course.module_code,
                "Week": assignment.timeslot.week if assignment.timeslot else "",
                "Day": assignment.timeslot.day if assignment.timeslot else "",
                "Start": assignment.timeslot.start_time if assignment.timeslot else "",
                "Duration": assignment.course.duration_hrs,
                "Room": assignment.room.room_id if assignment.room else "",
                "Status": "PASS" if _is_submission_assignment(assignment) else "FAIL",
            }
        )
    return rows


def validate_template2_submission(
    workbook_path: Path,
    source_requirements: list[Course],
    fixed_sessions: list[FixedSession],
    assignments: list[Assignment],
    rooms: list[Room],
    template2_path: Path,
    all_valid_workbook_path: Path | None = None,
    quarantined_requirements: list[object] | None = None,
) -> Template2ValidationResult:
    """Validate the actual saved submission-ready Template 2 workbook."""
    headers = _headers(workbook_path)
    template_headers = _headers(template2_path)
    rows = _row_dicts(workbook_path)
    all_valid_rows = _row_dicts(all_valid_workbook_path) if all_valid_workbook_path is not None else []
    field_rows, invalid_rows = _required_field_validation(rows, rooms)
    duplicates = _duplicate_rows(rows)
    programme_rows = _programme_coverage(
        source_requirements,
        assignments,
        rows,
        invalid_rows,
        all_valid_rows=all_valid_rows,
        quarantined_requirements=quarantined_requirements,
    )
    demand = build_demand_metrics(source_requirements, assignments, input_course_records=len(source_requirements))
    represented_programmes = {
        str(row.get("Canonical programme-year"))
        for row in programme_rows
        if row.get("Canonical programme-year") and int(row.get("Strict saved rows") or 0) > 0
    }
    all_valid_programmes = {
        str(row.get("Canonical programme-year"))
        for row in programme_rows
        if row.get("Canonical programme-year") and int(row.get("All-valid rows") or 0) > 0
    }
    complete_programmes = {
        str(row.get("Canonical programme-year"))
        for row in programme_rows
        if row.get("Complete Schedule Status") == "PASS" and row.get("Canonical programme-year")
    }
    submission_ready_programmes = {
        str(row.get("Canonical programme-year"))
        for row in programme_rows
        if row.get("Submission-Ready Status") == "PASS" and row.get("Canonical programme-year")
    }
    saved_programme_year_counts = _saved_programme_year_counts(rows, invalid_rows)
    actual_saved_programmes = len(saved_programme_year_counts)
    submission_ready_count = len(submission_ready_programmes)
    minimum_status = "PASS" if submission_ready_count >= REQUIRED_MIN_PROGRAMME_YEARS else "FAIL"
    missing_columns = [column for column in REQUIRED_SUBMISSION_COLUMNS if column not in headers]
    extra_columns = [column for column in headers if column not in template_headers]
    ready = not missing_columns and not extra_columns and not invalid_rows and not duplicates and minimum_status == "PASS"
    reconciliation = _row_level_reconciliation(assignments, rows, invalid_rows, complete_programmes)
    missing_programmes = _missing_programme_year_rows(programme_rows)
    identity_values: list[object] = []
    identity_values.extend(course.prog_yr for course in source_requirements)
    identity_values.extend(assignment.course.prog_yr for assignment in assignments)
    identity_values.extend(getattr(item, "programme_year", "") for item in quarantined_requirements or [])
    for row in [*rows, *all_valid_rows]:
        identity_values.extend(value for _, value in _saved_identity_candidates(row))
    summary = {
        "fixed source rows": len(fixed_sessions),
        "valid fixed assignments": sum(1 for item in assignments if item.is_fixed),
        "fixed teaching occurrences": sum(1 for item in assignments if item.is_fixed and item.timeslot is not None),
        "fixed assignments anchored": sum(1 for item in assignments if item.is_fixed and _is_submission_assignment(item)),
        "fixed-source conflicts": sum(1 for item in assignments if item.is_fixed and item.hard_violations),
        "non-fixed assignments": sum(1 for item in assignments if not item.is_fixed),
        "total required assignments": len(source_requirements),
        "total required teaching occurrences": demand.required_teaching_occurrences,
        "total scheduled occurrences": demand.scheduled_teaching_occurrences,
        "total unscheduled occurrences": demand.unscheduled_teaching_occurrences,
        "Template 2 output rows": len(rows),
        "Actual saved Template 2 rows": len(rows),
        "All-valid Template 2 rows": len(all_valid_rows),
        "rows with missing required fields": sum(1 for row in invalid_rows if "Missing" in row["Issues"]),
        "rows with mapping errors": len(invalid_rows),
        "distinct programme-year schedules": actual_saved_programmes,
        "programme-years represented in submission workbook": len(represented_programmes),
        "programme-years represented in all-valid workbook": len(all_valid_programmes),
        "actual saved programme-year schedules": actual_saved_programmes,
        "complete programme-year schedules": len(complete_programmes),
        "submission-ready programme-year schedules": submission_ready_count,
        "qualifying submission-ready programme-years": submission_ready_count,
        "required minimum programme-year schedules": REQUIRED_MIN_PROGRAMME_YEARS,
        "minimum programme-year status": minimum_status,
        "Template 2 readiness status": "PASS" if ready else "FAIL",
        "missing columns": ", ".join(missing_columns),
        "extra accidental columns": ", ".join(str(column) for column in extra_columns),
    }
    return Template2ValidationResult(
        ready=ready,
        summary=summary,
        programme_rows=programme_rows,
        required_field_rows=field_rows,
        fixed_accuracy_rows=_fixed_accuracy(assignments),
        source_reconciliation_rows=reconciliation,
        duplicate_rows=duplicates,
        invalid_rows=invalid_rows,
        missing_programme_year_rows=missing_programmes,
        identity_audit_rows=_identity_audit_rows(identity_values),
        incomplete_programme_year_rows=_incomplete_programme_rows(programme_rows),
        saved_workbook_rows=_saved_workbook_verification_rows(rows, invalid_rows),
    )


def export_template2_validation_report(result: Template2ValidationResult, output_path: Path) -> None:
    """Export the Template 2 submission validation workbook."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame([{"Metric": key, "Value": value} for key, value in result.summary.items()]).to_excel(
            writer,
            sheet_name="Summary",
            index=False,
        )
        pd.DataFrame(result.programme_rows).to_excel(writer, sheet_name="Programme Schedule Coverage", index=False)
        pd.DataFrame(result.fixed_accuracy_rows).to_excel(writer, sheet_name="Fixed Session Accuracy", index=False)
        pd.DataFrame(result.required_field_rows).to_excel(writer, sheet_name="Required Field Validation", index=False)
        pd.DataFrame(result.source_reconciliation_rows).to_excel(writer, sheet_name="Source-to-Output Reconciliation", index=False)
        pd.DataFrame(result.duplicate_rows).to_excel(writer, sheet_name="Duplicate Check", index=False)
        pd.DataFrame(result.invalid_rows).to_excel(writer, sheet_name="Invalid Rows", index=False)
        pd.DataFrame(result.missing_programme_year_rows).to_excel(writer, sheet_name="Missing Programme-Years", index=False)
        pd.DataFrame(result.identity_audit_rows).to_excel(writer, sheet_name="Identity Normalisation Audit", index=False)
        pd.DataFrame(result.incomplete_programme_year_rows).to_excel(writer, sheet_name="Incomplete Programme-Years", index=False)
        pd.DataFrame(result.saved_workbook_rows).to_excel(writer, sheet_name="Saved Workbook Verification", index=False)
        pd.DataFrame(
            [
                {
                    "Check": "Submission readiness",
                    "Status": "PASS" if result.ready else "FAIL",
                    "Notes": "Requires complete programme-years, valid saved rows, valid mappings, no duplicates, and at least 20 qualifying submission-ready programme-years.",
                }
            ]
        ).to_excel(writer, sheet_name="Submission Readiness", index=False)


def export_template2_programme_year_reconciliation(result: Template2ValidationResult, output_path: Path) -> None:
    """Export programme-year and row-level Template 2 reconciliation evidence."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    summary_keys = [
        "programme-years represented in all-valid workbook",
        "programme-years represented in submission workbook",
        "complete programme-year schedules",
        "submission-ready programme-year schedules",
        "qualifying submission-ready programme-years",
        "required minimum programme-year schedules",
        "minimum programme-year status",
        "total required teaching occurrences",
        "total scheduled occurrences",
        "total unscheduled occurrences",
        "Template 2 output rows",
        "All-valid Template 2 rows",
        "actual saved programme-year schedules",
        "Template 2 readiness status",
    ]
    summary_values = getattr(result, "summary", {})
    summary = [{"Metric": key, "Value": summary_values.get(key, "")} for key in summary_keys]
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame(summary).to_excel(writer, sheet_name="Summary", index=False)
        pd.DataFrame(getattr(result, "programme_rows", [])).to_excel(writer, sheet_name="Programme-Year Reconciliation", index=False)
        pd.DataFrame(getattr(result, "missing_programme_year_rows", [])).to_excel(writer, sheet_name="Missing Programme-Years", index=False)
        pd.DataFrame(getattr(result, "source_reconciliation_rows", [])).to_excel(writer, sheet_name="Row-Level Reconciliation", index=False)
        pd.DataFrame(getattr(result, "identity_audit_rows", [])).to_excel(writer, sheet_name="Identity Normalisation Audit", index=False)
        pd.DataFrame(getattr(result, "incomplete_programme_year_rows", [])).to_excel(writer, sheet_name="Incomplete Programme-Years", index=False)
        pd.DataFrame(getattr(result, "saved_workbook_rows", [])).to_excel(writer, sheet_name="Saved Workbook Verification", index=False)
