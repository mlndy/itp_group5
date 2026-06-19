"""Export preflight and post-run stakeholder reports."""

from __future__ import annotations

import re
from collections import Counter, defaultdict
from copy import deepcopy
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from config import BLOCKED_START_TIMES, DEFAULT_TEMPLATE2_FILE, LATEST_END_HOUR, SOFT_CONSTRAINT_WEIGHTS, VALID_DAYS, VALID_START_TIMES
from data.models import Assignment, Course, Room, TimeSlot
from engine.constraint_checker import (
    annotate_schedule_violations,
    assignment_end_hour,
    is_online_course,
    occupied_start_times,
    soft_violation_breakdown,
    time_to_hour,
    weighted_soft_score,
)
from engine.demand_metrics import DemandMetrics, build_demand_metrics, requirement_demand_lookup, requirement_key
from engine.remarks_interpreter import (
    RemarkEnforcement,
    RemarkHandlingStatus,
    RemarkRequirements,
    assignment_room_ids,
    assignment_rooms,
    assignment_satisfies_interpretation,
    course_remark_requirements,
    hard_enforceable_interpretations,
    is_hard_enforceable,
)
from engine.resource_audit import ResourceAudit, audit_resources
from generator.scheduler import get_candidate_rooms, schedulable_weeks

PREFLIGHT_COLUMNS = ["severity", "entity_type", "entity_id", "issue", "recommendation"]
VIOLATION_COLUMNS = ["Module", "Activity", "Programme/Year", "Week", "Day", "Start", "Room", "Violation"]
UNSCHEDULED_ANALYSIS_COLUMNS = [
    "Original Reason",
    "Reason Category",
    "Programme/Year",
    "Module Code",
    "Activity",
    "Class Size",
    "Class Size Band",
    "Delivery Mode",
    "Duration",
    "Teaching Week",
    "Common Module",
    "Candidate Limit",
    "Source File",
    "Required Week Count",
    "Scheduled Week Count",
    "Unscheduled Week Count",
    "Base Unscheduled Reason",
    "Remark Unscheduled Reason",
]
UNSCHEDULED_BREAKDOWN_COLUMNS = ["Breakdown", "Value", "Count"]
RESIDUAL_F2F_COLUMNS = [
    "Programme/Year",
    "Year",
    "Module Code",
    "Activity",
    "Class Size",
    "Duration",
    "Required Teaching Weeks",
    "Schedulable Teaching Weeks",
    "Scheduled Weeks",
    "Unscheduled Weeks",
    "Common Module",
    "Compatible Physical Room Count",
    "Smallest Suitable Room",
    "Largest Suitable Room",
    "Failed Reason",
    "Candidate Limit",
    "Feasible Start Windows Before Clash Checking",
    "Residual Classification",
    "Classification Evidence",
    "Source File",
]
PROGRAMME_COLUMNS = [
    "Programme/Year",
    "Source File",
    "DSC Indicator",
    "Assignments",
    "Scheduled Assignments",
    "Unscheduled Assignments",
    "Hard Violations on Scheduled Assignments",
]
VALIDATION_COLUMNS = ["Check", "Value", "Status", "Notes"]
METADATA_COLUMNS = ["Setting", "Value"]
OPTIMISATION_COLUMNS = ["Metric", "Value"]
STAKEHOLDER_PROGRAMME_COLUMNS = ["Programme/Year", "Week", "Day", "Start", "End", "Module", "Activity", "Room", "Delivery Mode", "Tutor"]
STAKEHOLDER_TUTOR_COLUMNS = [
    "Tutor",
    "Week",
    "Day",
    "Start",
    "End",
    "Module",
    "Activity",
    "Programme/Year",
    "Location",
    "Delivery Mode",
    "Idle Gap Since Previous",
    "Online/F2F Transition",
]
STAKEHOLDER_ROOM_COLUMNS = [
    "Room",
    "Week",
    "Day",
    "Start",
    "End",
    "Module",
    "Activity",
    "Programme/Year",
    "Class Size",
    "Room Capacity",
    "Utilisation Percentage",
]
EXCEPTION_QUEUE_COLUMNS = [
    "Programme/Year",
    "Module Code",
    "Activity",
    "Class Size",
    "Required Weeks",
    "Missing Weeks",
    "Original Reason",
    "Classification",
    "Compatible Physical Room Count",
    "Recommended Operational Action",
    "Review Status",
    "Source File",
]
REMARKS_INTERPRETATION_COLUMNS = [
    "Source Workbook",
    "Source Sheet",
    "Source Row",
    "Programme/Year",
    "Module",
    "Activity",
    "Raw Remark",
    "Detected Rule",
    "Extracted Parameters",
    "Enforcement",
    "Confidence",
    "Hard Enforceable",
    "Applied Status",
    "Assigned Rooms",
    "Selected Delivery Mode",
    "Explanation",
    "Review Reason",
]
SPECIAL_REQUEST_REVIEW_COLUMNS = [
    "Programme",
    "Module",
    "Activity",
    "Teaching Weeks",
    "Special Request",
    "What the System Understood",
    "Confidence",
    "Requirement Type",
    "How It Was Handled",
    "Scheduled?",
    "Assigned Day",
    "Assigned Time",
    "Assigned Room(s)",
    "Selected Delivery Mode",
    "Needs Manual Review",
    "Why Review Is Needed",
    "Recommended Action",
    "Review Status",
    "Review Notes",
]
TEMPLATE2_REQUIRED_COLUMNS = [
    "Module",
    "Class Type",
    "Group",
    "Day",
    "Start",
    "End",
    "Class Size",
    "Room1",
    "Staff1",
    "Staff2",
    "Tri Week",
    "Activity Type",
    "Duration",
    "Location Hostkey",
    "Remark",
]


def _metric_rows(values: dict[str, object]) -> pd.DataFrame:
    """Return metric/value rows in insertion order."""
    return pd.DataFrame([{"Metric": key, "Value": value} for key, value in values.items()])


def export_preflight_report(issues: list[dict[str, str]], output_path: Path) -> None:
    """Export preflight issues to an Excel workbook."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    rows = issues or [
        {
            "severity": "info",
            "entity_type": "run",
            "entity_id": "preflight",
            "issue": "No preflight issues found",
            "recommendation": "Proceed with scheduling.",
        }
    ]
    pd.DataFrame(rows, columns=PREFLIGHT_COLUMNS).to_excel(output_path, sheet_name="Preflight Issues", index=False)


def _snapshot(assignments: list[Assignment]) -> list[Assignment]:
    """Return an annotated copy while preserving original unscheduled reasons."""
    original_unscheduled_reasons = {
        id(assignment): list(assignment.hard_violations)
        for assignment in assignments
        if assignment.room is None or assignment.timeslot is None
    }
    copied = deepcopy(assignments)
    annotate_schedule_violations(copied)
    for original, copied_assignment in zip(assignments, copied, strict=False):
        reasons = original_unscheduled_reasons.get(id(original))
        if reasons:
            copied_assignment.hard_violations = reasons
    return copied


def _demand_summary_rows(demand_metrics: DemandMetrics | None) -> dict[str, object]:
    """Return demand metrics for Summary when available."""
    if demand_metrics is None:
        return {}
    return {
        "input_course_records": demand_metrics.input_course_records,
        "consolidated_course_requirements": demand_metrics.consolidated_course_requirements,
        "required_teaching_occurrences": demand_metrics.required_teaching_occurrences,
        "scheduled_teaching_occurrences": demand_metrics.scheduled_teaching_occurrences,
        "unscheduled_teaching_occurrences": demand_metrics.unscheduled_teaching_occurrences,
        "coverage_rate_percent": demand_metrics.coverage_rate_percent,
        "courses_fully_scheduled": demand_metrics.courses_fully_scheduled,
        "courses_partially_scheduled": demand_metrics.courses_partially_scheduled,
        "courses_fully_unscheduled": demand_metrics.courses_fully_unscheduled,
    }


def build_run_summary(assignments: list[Assignment], demand_metrics: DemandMetrics | None = None) -> dict[str, object]:
    """Build headline metrics for a completed run."""
    snapshot = _snapshot(assignments)
    total = len(snapshot)
    scheduled = [item for item in snapshot if item.room is not None and item.timeslot is not None]
    unscheduled_items = [item for item in snapshot if item.room is None or item.timeslot is None]
    unscheduled = total - len(scheduled)
    scheduled_hard = sum(len(item.hard_violations) for item in scheduled)
    unscheduled_hard = sum(len(item.hard_violations) for item in unscheduled_items)
    all_hard = sum(len(item.hard_violations) for item in snapshot)
    all_soft = sum(len(item.soft_violations) for item in snapshot)
    summary = {
        "assignments": total,
        "scheduled_assignments": len(scheduled),
        "unscheduled_assignments": unscheduled,
        "hard_violations_on_scheduled_assignments": scheduled_hard,
        "hard_violations_from_unscheduled_feasibility_failures": unscheduled_hard,
        "hard_violations_all_assignments": all_hard,
        "soft_violations": all_soft,
        "feasibility_rate": len(scheduled) / total if total else 0,
    }
    summary.update(_demand_summary_rows(demand_metrics))
    return summary


def _summary_df(assignments: list[Assignment], demand_metrics: DemandMetrics | None = None) -> pd.DataFrame:
    """Return summary metrics as rows."""
    summary = build_run_summary(assignments, demand_metrics=demand_metrics)
    labels = {
        "assignments": "Assignments",
        "scheduled_assignments": "Scheduled assignments",
        "unscheduled_assignments": "Unscheduled assignments",
        "hard_violations_on_scheduled_assignments": "Hard violations on scheduled assignments",
        "hard_violations_from_unscheduled_feasibility_failures": "Hard violations from unscheduled feasibility failures",
        "hard_violations_all_assignments": "Hard violations on all assignments",
        "soft_violations": "Soft violations",
        "feasibility_rate": "Feasibility rate",
        "input_course_records": "Input course records",
        "consolidated_course_requirements": "Consolidated course requirements",
        "required_teaching_occurrences": "Required teaching occurrences",
        "scheduled_teaching_occurrences": "Scheduled teaching occurrences",
        "unscheduled_teaching_occurrences": "Unscheduled teaching occurrences",
        "coverage_rate_percent": "Coverage rate percent",
        "courses_fully_scheduled": "Fully scheduled course requirements",
        "courses_partially_scheduled": "Partially scheduled course requirements",
        "courses_fully_unscheduled": "Fully unscheduled course requirements",
    }
    return pd.DataFrame([{"Metric": labels[key], "Value": value} for key, value in summary.items()])


def _violations_df(assignments: list[Assignment], violation_type: str) -> pd.DataFrame:
    """Return hard or soft violations as spreadsheet rows."""
    rows: list[dict[str, object]] = []
    for assignment in assignments:
        violations = assignment.hard_violations if violation_type == "hard" else assignment.soft_violations
        for violation in violations:
            rows.append(
                {
                    "Module": assignment.course.module_code,
                    "Activity": assignment.course.activity,
                    "Programme/Year": assignment.course.prog_yr,
                    "Week": assignment.timeslot.week if assignment.timeslot else None,
                    "Day": assignment.timeslot.day if assignment.timeslot else None,
                    "Start": assignment.timeslot.start_time if assignment.timeslot else None,
                    "Room": assignment_room_ids(assignment),
                    "Violation": violation,
                }
            )
    return pd.DataFrame(rows, columns=VIOLATION_COLUMNS)


def _unscheduled_reasons_df(assignments: list[Assignment]) -> pd.DataFrame:
    """Return counts of reasons unscheduled assignments remain unresolved."""
    reasons: Counter[str] = Counter()
    for assignment in assignments:
        if assignment.room is not None and assignment.timeslot is not None:
            continue
        if assignment.hard_violations:
            reasons.update(assignment.hard_violations)
        else:
            reasons["Unscheduled without recorded reason"] += 1
    return pd.DataFrame([{"Reason": reason, "Count": count} for reason, count in reasons.most_common()], columns=["Reason", "Count"])


def categorise_unscheduled_reason(assignment: Assignment, reason: str) -> str:
    """Map one unscheduled reason to a stakeholder-facing category."""
    text = reason.lower()
    if "max candidate pattern limit" in text:
        return "Candidate pattern limit"
    if "capacity" in text or "room large enough" in text:
        return "Room capacity"
    if "compatible room" in text or "delivery mode" in text or "virtual room" in text:
        return "No compatible room"
    if "staff clash" in text or "tutor" in text:
        return "Tutor clash"
    if "student group clash" in text or "group conflict" in text:
        return "Student-group clash"
    if "blocked" in text or "time window" in text or "no schedulable" in text or "valid teaching week" in text:
        return "Blocked or unavailable timeslot"
    if "feasible slot for week" in text or "weekly room/day/start pattern" in text:
        return "No complete multi-week placement"
    if assignment.course.is_common_module:
        return "Common-module constraint"
    return "Other / unknown"


def _class_size_band(class_size: int) -> str:
    """Return a compact class-size band for unscheduled breakdowns."""
    if class_size <= 30:
        return "1-30"
    if class_size <= 60:
        return "31-60"
    if class_size <= 100:
        return "61-100"
    if class_size <= 150:
        return "101-150"
    return "151+"


def _failed_week(reason: str) -> int | None:
    """Extract a failed teaching week from a reason when present."""
    match = re.search(r"week\s+(\d+)", reason, flags=re.IGNORECASE)
    return int(match.group(1)) if match else None


def _unscheduled_analysis_df(assignments: list[Assignment], demand_courses: list[Course] | None = None) -> pd.DataFrame:
    """Return detailed unscheduled bottleneck rows without replacing original reasons."""
    demand_lookup = requirement_demand_lookup(demand_courses, assignments) if demand_courses is not None else {}
    rows: list[dict[str, object]] = []
    for assignment in assignments:
        if assignment.room is not None and assignment.timeslot is not None:
            continue
        demand = demand_lookup.get(requirement_key(assignment.course))
        reasons = assignment.hard_violations or ["Unscheduled without recorded reason"]
        for reason in reasons:
            rows.append(
                {
                    "Original Reason": reason,
                    "Reason Category": categorise_unscheduled_reason(assignment, reason),
                    "Programme/Year": assignment.course.prog_yr,
                    "Module Code": assignment.course.module_code,
                    "Activity": assignment.course.activity,
                    "Class Size": assignment.course.class_size,
                    "Class Size Band": _class_size_band(assignment.course.class_size),
                    "Delivery Mode": assignment.selected_delivery_mode or assignment.course.delivery_mode,
                    "Duration": assignment.course.duration_hrs,
                    "Teaching Week": _failed_week(reason),
                    "Common Module": "Yes" if assignment.course.is_common_module else "No",
                    "Candidate Limit": "Yes" if "max candidate pattern limit" in reason.lower() else "No",
                    "Source File": assignment.course.source_file,
                    "Required Week Count": demand.required_week_count if demand else None,
                    "Scheduled Week Count": demand.scheduled_week_count if demand else None,
                    "Unscheduled Week Count": demand.unscheduled_week_count if demand else None,
                    "Base Unscheduled Reason": assignment.base_unscheduled_reason,
                    "Remark Unscheduled Reason": assignment.remark_unscheduled_reason,
                }
            )
    return pd.DataFrame(rows, columns=UNSCHEDULED_ANALYSIS_COLUMNS)


def _unscheduled_breakdown_df(analysis_df: pd.DataFrame) -> pd.DataFrame:
    """Return grouped counts for unscheduled bottleneck dimensions."""
    if analysis_df.empty:
        return pd.DataFrame(columns=UNSCHEDULED_BREAKDOWN_COLUMNS)

    breakdowns = {
        "Reason Category": "Reason Category",
        "Programme/Year": "Programme/Year",
        "Activity": "Activity",
        "Delivery Mode": "Delivery Mode",
        "Class Size Band": "Class Size Band",
        "Duration": "Duration",
        "Common Module": "Common Module",
    }
    rows: list[dict[str, object]] = []
    for label, column in breakdowns.items():
        counts = analysis_df[column].fillna("Unknown").astype(str).value_counts()
        for value, count in counts.items():
            rows.append({"Breakdown": label, "Value": value, "Count": int(count)})
    return pd.DataFrame(rows, columns=UNSCHEDULED_BREAKDOWN_COLUMNS)


def _year_from_programme(programme: str) -> str:
    """Extract a compact year label from programme text when present."""
    match = re.search(r"(?:year|yr)\s*([0-9]+)", programme, flags=re.IGNORECASE)
    return f"Year {match.group(1)}" if match else ""


def _format_weeks(weeks: list[int] | set[int]) -> str:
    """Return teaching weeks as compact comma-separated text."""
    return ", ".join(str(week) for week in sorted(weeks))


def _scheduled_weeks_by_key(assignments: list[Assignment]) -> dict[object, set[int]]:
    """Return scheduled weeks by consolidated requirement key."""
    scheduled: dict[object, set[int]] = defaultdict(set)
    for assignment in assignments:
        if assignment.room is None or assignment.timeslot is None:
            continue
        scheduled[requirement_key(assignment.course)].add(assignment.timeslot.week)
    return scheduled


def _feasible_start_window_count(course: Course, compatible_rooms: list[Room]) -> int:
    """Count room/week/day/start windows before occupancy clash checks."""
    count = 0
    for room in compatible_rooms:
        for week in schedulable_weeks(course.teaching_weeks):
            for day in VALID_DAYS:
                for start_time in VALID_START_TIMES:
                    candidate = Assignment(course=course, room=room, timeslot=TimeSlot(day, start_time, week))
                    if time_to_hour(start_time) + course.duration_hrs > LATEST_END_HOUR:
                        continue
                    if occupied_start_times(candidate) & BLOCKED_START_TIMES.get(day, set()):
                        continue
                    count += 1
    return count


def _classify_residual_f2f(
    assignment: Assignment,
    reasons: list[str],
    compatible_rooms: list[Room],
    schedulable_week_count: int,
) -> tuple[str, str]:
    """Classify one residual F2F requirement for stakeholder review."""
    reason_text = " | ".join(reasons).lower()
    if "max candidate pattern limit" in reason_text:
        return "Search limitation", "Candidate-pattern safety cap stopped search before all possibilities were explored."
    if schedulable_week_count == 0 or "no schedulable teaching weeks" in reason_text:
        return "Calendar restriction", "All requested teaching weeks are blocked by academic-calendar rules."
    if not compatible_rooms:
        return "Physical room scarcity", "No loaded physical room has enough capacity for this class size."
    if "weekly room/day/start pattern" in reason_text:
        return "Recurring-pattern infeasibility", "No single room/day/start pattern remained feasible across all required weeks."
    if "staff clash" in reason_text or "tutor" in reason_text or "student group" in reason_text:
        return "Tutor or student-group conflict", "Remaining otherwise valid slots clash with tutor or cohort occupancy."
    if "blocked" in reason_text or "time window" in reason_text:
        return "Calendar restriction", "Institutional blocked-time or valid-hour rules remove the requested window."
    return "Input issue", "Residual reason does not match a schedulable resource or search category."


def _residual_f2f_analysis_df(
    assignments: list[Assignment],
    rooms: list[Room] | None = None,
) -> pd.DataFrame:
    """Return detailed residual F2F classifications for unresolved requirements."""
    physical_rooms = rooms or []
    scheduled_by_key = _scheduled_weeks_by_key(assignments)
    grouped: dict[object, dict[str, object]] = {}
    for assignment in assignments:
        if assignment.room is not None and assignment.timeslot is not None:
            continue
        delivery_mode = assignment.selected_delivery_mode or assignment.course.delivery_mode
        if "online" in delivery_mode.lower():
            continue
        key = requirement_key(assignment.course)
        row = grouped.setdefault(
            key,
            {
                "assignment": assignment,
                "reasons": [],
                "candidate_limit": False,
                "failed_weeks": set(),
            },
        )
        reasons = assignment.hard_violations or ["Unscheduled without recorded reason"]
        row["reasons"].extend(reasons)  # type: ignore[union-attr]
        if any("max candidate pattern limit" in reason.lower() for reason in reasons):
            row["candidate_limit"] = True
        for reason in reasons:
            week = _failed_week(reason)
            if week is not None:
                row["failed_weeks"].add(week)  # type: ignore[union-attr]

    rows: list[dict[str, object]] = []
    for key, data in grouped.items():
        assignment = data["assignment"]
        if not isinstance(assignment, Assignment):
            continue
        course = assignment.course
        reasons = list(dict.fromkeys(data["reasons"]))  # type: ignore[arg-type]
        compatible_rooms = sorted(
            [room for room in get_candidate_rooms(course, physical_rooms) if room.room_type == "physical"],
            key=lambda room: (room.capacity, room.room_id),
        )
        scheduled_weeks = scheduled_by_key.get(key, set())
        schedulable = set(schedulable_weeks(course.teaching_weeks))
        failed_weeks = set(data["failed_weeks"])  # type: ignore[arg-type]
        unscheduled_weeks = failed_weeks or (schedulable - scheduled_weeks)
        classification, evidence = _classify_residual_f2f(
            assignment,
            reasons,
            compatible_rooms,
            len(schedulable),
        )
        rows.append(
            {
                "Programme/Year": course.prog_yr,
                "Year": _year_from_programme(course.prog_yr),
                "Module Code": course.module_code,
                "Activity": course.activity,
                "Class Size": course.class_size,
                "Duration": course.duration_hrs,
                "Required Teaching Weeks": _format_weeks(course.teaching_weeks),
                "Schedulable Teaching Weeks": _format_weeks(schedulable),
                "Scheduled Weeks": _format_weeks(scheduled_weeks),
                "Unscheduled Weeks": _format_weeks(unscheduled_weeks),
                "Common Module": "Yes" if course.is_common_module else "No",
                "Compatible Physical Room Count": len(compatible_rooms),
                "Smallest Suitable Room": compatible_rooms[0].room_id if compatible_rooms else "",
                "Largest Suitable Room": compatible_rooms[-1].room_id if compatible_rooms else "",
                "Failed Reason": " | ".join(reasons),
                "Candidate Limit": "Yes" if data["candidate_limit"] else "No",
                "Feasible Start Windows Before Clash Checking": _feasible_start_window_count(course, compatible_rooms),
                "Residual Classification": classification,
                "Classification Evidence": evidence,
                "Source File": course.source_file,
            }
        )
    return pd.DataFrame(rows, columns=RESIDUAL_F2F_COLUMNS)


def _room_utilisation_df(assignments: list[Assignment]) -> pd.DataFrame:
    """Return scheduled room usage and enrolment utilisation."""
    usage: dict[str, dict[str, object]] = defaultdict(lambda: {"Scheduled Assignments": 0, "Total Enrolment": 0, "Total Capacity": 0})
    for assignment in assignments:
        if assignment.room is None or assignment.timeslot is None:
            continue
        for room in assignment_rooms(assignment):
            row = usage[room.room_id]
            row["Scheduled Assignments"] = int(row["Scheduled Assignments"]) + 1
            row["Total Enrolment"] = int(row["Total Enrolment"]) + assignment.course.class_size
            row["Total Capacity"] = int(row["Total Capacity"]) + room.capacity

    rows: list[dict[str, object]] = []
    for room_id, row in sorted(usage.items()):
        total_capacity = int(row["Total Capacity"])
        utilisation = int(row["Total Enrolment"]) / total_capacity if total_capacity else 0
        rows.append({"Room": room_id, **row, "Average Utilisation": utilisation})
    return pd.DataFrame(rows, columns=["Room", "Scheduled Assignments", "Total Enrolment", "Total Capacity", "Average Utilisation"])


def _resource_audit_df(resource_audit: ResourceAudit | None) -> pd.DataFrame:
    """Return resource audit metrics for spreadsheet export."""
    if resource_audit is None:
        return _metric_rows({"Resource audit": "Not available"})
    return _metric_rows(
        {
            "Total raw room rows": resource_audit.total_raw_room_rows,
            "Total loaded rooms": resource_audit.total_loaded_rooms,
            "Loaded physical room count": resource_audit.physical_room_count,
            "Loaded virtual room count": resource_audit.virtual_room_count,
            "Virtual room IDs": ", ".join(room.room_id for room in resource_audit.virtual_rooms),
            "Virtual room policy": resource_audit.virtual_room_policy,
            "Duplicate room-ID count": resource_audit.duplicate_room_id_count,
            "Skipped or invalid room rows": resource_audit.skipped_or_invalid_room_rows,
            "Online course requirement count": resource_audit.online_course_requirements,
            "Required online teaching occurrences": resource_audit.required_online_teaching_occurrences,
            "Scheduled online teaching occurrences": resource_audit.scheduled_online_teaching_occurrences,
            "Unscheduled online teaching occurrences": resource_audit.unscheduled_online_teaching_occurrences,
            "Online coverage rate percent": resource_audit.online_coverage_rate_percent,
            "Virtual room policy note": resource_audit.exclusivity_note,
        }
    )


def _virtual_room_detail_df(resource_audit: ResourceAudit | None) -> pd.DataFrame:
    """Return loaded virtual-room details and online demand peaks."""
    rows: list[dict[str, object]] = []
    if resource_audit is not None:
        for room in resource_audit.virtual_rooms:
            rows.append(
                {
                    "Section": "Virtual Room",
                    "Room ID": room.room_id,
                    "Capacity": room.capacity,
                    "Resource Type": room.resource_type,
                    "Recording": room.recording,
                    "Programme/Year": None,
                    "Week": None,
                    "Required Online Occurrences": None,
                }
            )
        for row in resource_audit.peak_online_demand_by_week[:25]:
            rows.append(
                {
                    "Section": "Peak Online Demand",
                    "Room ID": None,
                    "Capacity": None,
                    "Resource Type": None,
                    "Recording": None,
                    **row,
                }
            )
    return pd.DataFrame(
        rows,
        columns=[
            "Section",
            "Room ID",
            "Capacity",
            "Resource Type",
            "Recording",
            "Programme/Year",
            "Week",
            "Required Online Occurrences",
        ],
    )


def _is_dsc_assignment(assignment: Assignment) -> bool:
    """Return True when an assignment appears to belong to DSC input or programme data."""
    course = assignment.course
    text = " ".join([course.module_code, course.prog_yr, course.source_file]).upper()
    return "DSC" in text


def _programme_breakdown_df(assignments: list[Assignment]) -> pd.DataFrame:
    """Return scheduled and unscheduled counts by programme/source."""
    grouped: dict[tuple[str, str], dict[str, object]] = defaultdict(
        lambda: {
            "Assignments": 0,
            "Scheduled Assignments": 0,
            "Unscheduled Assignments": 0,
            "Hard Violations on Scheduled Assignments": 0,
            "DSC Indicator": "No",
        }
    )
    for assignment in assignments:
        key = (assignment.course.prog_yr, assignment.course.source_file)
        row = grouped[key]
        row["Assignments"] = int(row["Assignments"]) + 1
        if _is_dsc_assignment(assignment):
            row["DSC Indicator"] = "Yes"
        if assignment.room is not None and assignment.timeslot is not None:
            row["Scheduled Assignments"] = int(row["Scheduled Assignments"]) + 1
            row["Hard Violations on Scheduled Assignments"] = int(row["Hard Violations on Scheduled Assignments"]) + len(assignment.hard_violations)
        else:
            row["Unscheduled Assignments"] = int(row["Unscheduled Assignments"]) + 1

    rows = [
        {"Programme/Year": programme, "Source File": source_file, **values}
        for (programme, source_file), values in sorted(grouped.items())
    ]
    return pd.DataFrame(rows, columns=PROGRAMME_COLUMNS)


def _validation_checks_df(
    assignments: list[Assignment],
    generated_at: str,
    demand_metrics: DemandMetrics | None = None,
    optimisation_summary: dict[str, object] | None = None,
    resource_audit: ResourceAudit | None = None,
) -> pd.DataFrame:
    """Return validation checks for Engineering final evidence."""
    summary = build_run_summary(assignments, demand_metrics=demand_metrics)
    programme_df = _programme_breakdown_df(assignments)
    reasons_df = _unscheduled_reasons_df(assignments)
    total = int(summary["assignments"])
    scheduled = int(summary["scheduled_assignments"])
    unscheduled = int(summary["unscheduled_assignments"])
    scheduled_hard = int(summary["hard_violations_on_scheduled_assignments"])
    total_check = scheduled + unscheduled
    has_dsc = bool((programme_df["DSC Indicator"] == "Yes").any()) if not programme_df.empty else False
    reasons_exist = not reasons_df.empty
    unscheduled_visible = unscheduled == 0 or reasons_exist
    optimisation = optimisation_summary or {}
    demand_unchanged = optimisation.get("required_teaching_occurrences_before") == optimisation.get("required_teaching_occurrences_after")
    coverage_unchanged = optimisation.get("coverage_unchanged_status")
    hard_after = optimisation.get("hard_violations_after")
    soft_not_worsened = optimisation.get("soft_score_not_worsened_status")
    online_preserved = optimisation.get("online_coverage_preserved_status")
    online_coverage_value = (
        ""
        if resource_audit is None
        else f"{resource_audit.scheduled_online_teaching_occurrences} / {resource_audit.required_online_teaching_occurrences}"
    )
    rows = [
        {"Check": "Generated timestamp", "Value": generated_at, "Status": "INFO", "Notes": "Report generation time."},
        {"Check": "Total assignments", "Value": total, "Status": "INFO", "Notes": "All scheduled and unscheduled assignment rows in this run."},
        {"Check": "Scheduled assignments", "Value": scheduled, "Status": "INFO", "Notes": "Assignments with both room and timeslot."},
        {"Check": "Unscheduled assignments", "Value": unscheduled, "Status": "INFO", "Notes": "Assignments intentionally left unresolved rather than forced invalid."},
        {
            "Check": "Scheduled + unscheduled total check",
            "Value": total_check,
            "Status": "PASS" if total_check == total else "FAIL",
            "Notes": "Must equal total assignments.",
        },
        {
            "Check": "Demand occurrence consistency",
            "Value": (
                ""
                if demand_metrics is None
                else demand_metrics.scheduled_teaching_occurrences + demand_metrics.unscheduled_teaching_occurrences
            ),
            "Status": "INFO" if demand_metrics is None else ("PASS" if demand_metrics.is_consistent else "FAIL"),
            "Notes": "Required teaching occurrences must equal scheduled plus unscheduled occurrences.",
        },
        {
            "Check": "Stable demand metric status",
            "Value": "" if demand_metrics is None else demand_metrics.required_teaching_occurrences,
            "Status": "INFO" if demand_metrics is None else "PASS",
            "Notes": "Required teaching occurrences are calculated from consolidated input requirements, not room availability.",
        },
        {
            "Check": "Hard violations on scheduled assignments",
            "Value": scheduled_hard,
            "Status": "PASS" if scheduled_hard == 0 else "FAIL",
            "Notes": "Main hard-constraint safety metric.",
        },
        {
            "Check": "Hard-constraint safety status",
            "Value": "0 scheduled hard violations" if scheduled_hard == 0 else f"{scheduled_hard} scheduled hard violations",
            "Status": "PASS" if scheduled_hard == 0 else "FAIL",
            "Notes": "Scheduled timetable entries must be hard-feasible.",
        },
        {
            "Check": "DSC inclusion status",
            "Value": "DSC rows found" if has_dsc else "No DSC rows found",
            "Status": "PASS" if has_dsc else "FAIL",
            "Notes": "Programme Breakdown must contain DSC evidence for Engineering scope.",
        },
        {
            "Check": "Unscheduled visibility status",
            "Value": f"{unscheduled} unscheduled assignments",
            "Status": "PASS" if unscheduled_visible else "FAIL",
            "Notes": "Unscheduled assignments must remain visible with reasons when present.",
        },
        {
            "Check": "Teaching demand unchanged after optimisation",
            "Value": (
                ""
                if not optimisation
                else f"{optimisation.get('required_teaching_occurrences_before')} -> {optimisation.get('required_teaching_occurrences_after')}"
            ),
            "Status": "INFO" if not optimisation else ("PASS" if demand_unchanged else "FAIL"),
            "Notes": "Required teaching occurrence demand must not change during optimisation.",
        },
        {
            "Check": "Scheduled coverage unchanged after optimisation",
            "Value": (
                ""
                if not optimisation
                else f"{optimisation.get('scheduled_teaching_occurrences_before')} -> {optimisation.get('scheduled_teaching_occurrences_after')}"
            ),
            "Status": "INFO" if not optimisation else str(coverage_unchanged or "INFO"),
            "Notes": "Optimisation may improve quality only; it must not change scheduled coverage.",
        },
        {
            "Check": "Hard violations after optimisation",
            "Value": "" if hard_after is None else hard_after,
            "Status": "INFO" if hard_after is None else ("PASS" if int(hard_after) == 0 else "FAIL"),
            "Notes": "Optimised scheduled timetable entries must remain hard-feasible.",
        },
        {
            "Check": "Soft score not worsened",
            "Value": (
                ""
                if not optimisation
                else f"{optimisation.get('soft_violations_before')} -> {optimisation.get('soft_violations_after')}"
            ),
            "Status": "INFO" if not optimisation else str(soft_not_worsened or "INFO"),
            "Notes": "A worse soft score must not replace the baseline schedule.",
        },
        {
            "Check": "Online coverage preserved",
            "Value": online_coverage_value,
            "Status": "INFO" if not optimisation else str(online_preserved or "INFO"),
            "Notes": "Shared online-room policy and online scheduled occurrence coverage must remain stable.",
        },
        {
            "Check": "Result total comparability note",
            "Value": "Compare scheduled counts only when total assignment pool and run metadata match.",
            "Status": "INFO",
            "Notes": "Retry and candidate settings can change week-level unscheduled placeholder counts.",
        },
    ]
    return pd.DataFrame(rows, columns=VALIDATION_COLUMNS)


def _optimisation_summary_df(optimisation_summary: dict[str, object] | None) -> pd.DataFrame:
    """Return optimisation acceptance metrics as spreadsheet rows."""
    summary = optimisation_summary or {"optimisation_enabled": "No", "status": "Skipped"}
    labels = {
        "optimisation_enabled": "Optimisation enabled",
        "status": "Status",
        "stop_reason": "Stop reason",
        "requested_max_iterations": "Requested maximum iterations",
        "time_limit_seconds": "Time limit seconds",
        "patience": "Patience",
        "iterations_completed": "Iterations completed",
        "runtime_seconds": "Runtime seconds",
        "scheduled_teaching_occurrences_before": "Scheduled teaching occurrences before optimisation",
        "scheduled_teaching_occurrences_after": "Scheduled teaching occurrences after optimisation",
        "required_teaching_occurrences_before": "Required teaching occurrences before optimisation",
        "required_teaching_occurrences_after": "Required teaching occurrences after optimisation",
        "online_scheduled_occurrences_before": "Online scheduled occurrences before optimisation",
        "online_scheduled_occurrences_after": "Online scheduled occurrences after optimisation",
        "online_required_occurrences_before": "Online required occurrences before optimisation",
        "online_required_occurrences_after": "Online required occurrences after optimisation",
        "hard_violations_before": "Hard violations before optimisation",
        "hard_violations_after": "Hard violations after optimisation",
        "soft_violations_before": "Soft violations before optimisation",
        "soft_violations_after": "Soft violations after optimisation",
        "weighted_soft_score_before": "Weighted soft score before optimisation",
        "weighted_soft_score_after": "Weighted soft score after optimisation",
        "absolute_soft_violation_improvement": "Absolute soft-violation improvement",
        "percentage_soft_violation_improvement": "Percentage soft-violation improvement",
        "absolute_weighted_soft_score_improvement": "Absolute weighted soft-score improvement",
        "percentage_weighted_soft_score_improvement": "Percentage weighted soft-score improvement",
        "coverage_unchanged_status": "Coverage unchanged status",
        "hard_safety_status": "Hard-safety status",
        "soft_score_not_worsened_status": "Soft score not worsened status",
        "online_coverage_preserved_status": "Online coverage preserved status",
    }
    return pd.DataFrame(
        [{"Metric": labels.get(key, key), "Value": value} for key, value in summary.items()],
        columns=OPTIMISATION_COLUMNS,
    )


def _metadata_df(metadata: dict[str, object] | None, generated_at: str) -> pd.DataFrame:
    """Return run metadata as setting/value rows."""
    rows = [{"Setting": "generated_at", "Value": generated_at}]
    for key, value in (metadata or {}).items():
        rows.append({"Setting": key, "Value": value})
    return pd.DataFrame(rows, columns=METADATA_COLUMNS)


def _is_scheduled_assignment(assignment: Assignment) -> bool:
    """Return True when an assignment has a room and timeslot."""
    return assignment.room is not None and assignment.timeslot is not None


def _end_time_text(assignment: Assignment) -> str:
    """Return the assignment end time as HH:MM text."""
    if assignment.timeslot is None:
        return ""
    return f"{assignment_end_hour(assignment):02d}:00"


def _assignment_tutors(assignment: Assignment) -> list[str]:
    """Return display tutor values for stakeholder views."""
    tutors = assignment.course.staff_names or assignment.course.staff_ids
    return [tutor for tutor in tutors if tutor] or ["Unassigned"]


def _programme_timetable_df(assignments: list[Assignment]) -> pd.DataFrame:
    """Return a readable weekly timetable by programme/year."""
    rows: list[dict[str, object]] = []
    for assignment in assignments:
        if not _is_scheduled_assignment(assignment):
            continue
        rows.append(
            {
                "Programme/Year": assignment.course.prog_yr,
                "Week": assignment.timeslot.week,
                "Day": assignment.timeslot.day,
                "Start": assignment.timeslot.start_time,
                "End": _end_time_text(assignment),
                "Module": assignment.course.module_code,
                "Activity": assignment.course.activity,
                "Room": assignment_room_ids(assignment),
                "Delivery Mode": assignment.selected_delivery_mode or assignment.course.delivery_mode,
                "Tutor": ", ".join(_assignment_tutors(assignment)),
            }
        )
    return pd.DataFrame(rows, columns=STAKEHOLDER_PROGRAMME_COLUMNS).sort_values(
        by=["Programme/Year", "Week", "Day", "Start", "Module"],
        kind="stable",
        ignore_index=True,
    )


def _tutor_timetable_df(assignments: list[Assignment]) -> pd.DataFrame:
    """Return tutor timetable rows with idle-gap and transition notes."""
    tutor_rows: list[dict[str, object]] = []
    grouped: dict[tuple[str, int, str], list[Assignment]] = defaultdict(list)
    for assignment in assignments:
        if not _is_scheduled_assignment(assignment):
            continue
        for tutor in _assignment_tutors(assignment):
            grouped[(tutor, assignment.timeslot.week, assignment.timeslot.day)].append(assignment)

    for (tutor, week, day), items in sorted(grouped.items()):
        ordered = sorted(items, key=lambda item: (time_to_hour(item.timeslot.start_time), item.course.module_code))
        previous: Assignment | None = None
        for assignment in ordered:
            idle_gap = ""
            transition = "No"
            if previous is not None:
                gap = time_to_hour(assignment.timeslot.start_time) - assignment_end_hour(previous)
                idle_gap = max(gap, 0)
                transition = "Yes" if is_online_course(previous.course) != is_online_course(assignment.course) else "No"
            tutor_rows.append(
                {
                    "Tutor": tutor,
                    "Week": week,
                    "Day": day,
                    "Start": assignment.timeslot.start_time,
                    "End": _end_time_text(assignment),
                    "Module": assignment.course.module_code,
                    "Activity": assignment.course.activity,
                    "Programme/Year": assignment.course.prog_yr,
                    "Location": assignment_room_ids(assignment),
                    "Delivery Mode": assignment.selected_delivery_mode or assignment.course.delivery_mode,
                    "Idle Gap Since Previous": idle_gap,
                    "Online/F2F Transition": transition,
                }
            )
            previous = assignment
    return pd.DataFrame(tutor_rows, columns=STAKEHOLDER_TUTOR_COLUMNS)


def _room_timetable_df(assignments: list[Assignment]) -> pd.DataFrame:
    """Return physical-room occupancy rows for stakeholder review."""
    rows: list[dict[str, object]] = []
    for assignment in assignments:
        if not _is_scheduled_assignment(assignment):
            continue
        for room in assignment_rooms(assignment):
            if room.room_type == "virtual":
                continue
            utilisation = assignment.course.class_size / room.capacity if room.capacity else 0
            rows.append(
                {
                    "Room": room.room_id,
                    "Week": assignment.timeslot.week,
                    "Day": assignment.timeslot.day,
                    "Start": assignment.timeslot.start_time,
                    "End": _end_time_text(assignment),
                    "Module": assignment.course.module_code,
                    "Activity": assignment.course.activity,
                    "Programme/Year": assignment.course.prog_yr,
                    "Class Size": assignment.course.class_size,
                    "Room Capacity": room.capacity,
                    "Utilisation Percentage": round(utilisation * 100, 2),
                }
            )
    return pd.DataFrame(rows, columns=STAKEHOLDER_ROOM_COLUMNS).sort_values(
        by=["Room", "Week", "Day", "Start", "Module"],
        kind="stable",
        ignore_index=True,
    )


def _recommended_action(classification: str, compatible_room_count: int) -> str:
    """Return a non-automatic operational action for one exception."""
    if classification == "Physical room scarcity" or compatible_room_count == 0:
        return "Review large-room availability or approved delivery arrangement."
    if classification == "Recurring-pattern infeasibility":
        return "Manual timetabling-team review of recurring pattern options."
    if classification == "Calendar restriction":
        return "Revise approved timeslot or verify blocked-week inputs."
    if classification == "Tutor or student-group conflict":
        return "Review tutor or cohort availability with programme lead."
    if classification == "Search limitation":
        return "Rerun with a higher candidate limit if demo time allows."
    return "Verify input data and review manually."


def _exception_queue_df(assignments: list[Assignment], rooms: list[Room] | None = None) -> pd.DataFrame:
    """Return unresolved requirements as an operational exception queue."""
    rows: list[dict[str, object]] = []
    for assignment in assignments:
        if _is_scheduled_assignment(assignment):
            continue
        compatible_rooms = [
            room for room in get_candidate_rooms(assignment.course, rooms or []) if room.room_type == "physical"
        ]
        reasons = assignment.hard_violations or ["Unscheduled without recorded reason"]
        classification, _ = _classify_residual_f2f(
            assignment,
            reasons,
            compatible_rooms,
            len(schedulable_weeks(assignment.course.teaching_weeks)),
        )
        missing_weeks = sorted({_failed_week(reason) for reason in reasons if _failed_week(reason) is not None})
        rows.append(
            {
                "Programme/Year": assignment.course.prog_yr,
                "Module Code": assignment.course.module_code,
                "Activity": assignment.course.activity,
                "Class Size": assignment.course.class_size,
                "Required Weeks": _format_weeks(assignment.course.teaching_weeks),
                "Missing Weeks": _format_weeks([week for week in missing_weeks if week is not None]),
                "Original Reason": " | ".join(reasons),
                "Classification": classification,
                "Compatible Physical Room Count": len(compatible_rooms),
                "Recommended Operational Action": _recommended_action(classification, len(compatible_rooms)),
                "Review Status": "Open",
                "Source File": assignment.course.source_file,
            }
        )
    return pd.DataFrame(rows, columns=EXCEPTION_QUEUE_COLUMNS)


def _remarks_result_text(assignment: Assignment) -> str:
    """Return a concise timetable result for a special request."""
    if not _is_scheduled_assignment(assignment):
        reasons = " | ".join(assignment.hard_violations) or "Unscheduled without recorded reason"
        return f"Unscheduled: {reasons}"
    return f"Scheduled {assignment.timeslot.day} {assignment.timeslot.start_time}, rooms {assignment_room_ids(assignment)}"


HANDLING_STATUS_LABELS = {
    RemarkHandlingStatus.AUTOMATICALLY_APPLIED: "Applied automatically",
    RemarkHandlingStatus.PREFERENCE_CONSIDERED: "Considered as a preference",
    RemarkHandlingStatus.SCHEDULED_NEEDS_CONFIRMATION: "Scheduled, confirmation required",
    RemarkHandlingStatus.UNSCHEDULED_DUE_TO_REQUEST: "Could not schedule because of explicit request",
    RemarkHandlingStatus.UNSUPPORTED_NON_BLOCKING: "Not automatically interpreted",
    RemarkHandlingStatus.NO_SCHEDULING_ACTION: "No timetable action needed",
}


def _course_review_key(course: Course) -> tuple[object, ...]:
    """Return a stable course-level key for special-request review."""
    return (
        course.source_file,
        course.source_row,
        course.module_code,
        course.activity,
        course.prog_yr,
        tuple(course.group_ids),
        tuple(course.teaching_weeks),
    )


def _group_remark_assignments(assignments: list[Assignment]) -> list[list[Assignment]]:
    """Group assignment rows by remarked course requirement."""
    grouped: dict[tuple[object, ...], list[Assignment]] = {}
    for assignment in assignments:
        if not str(assignment.course.remarks or "").strip():
            continue
        grouped.setdefault(_course_review_key(assignment.course), []).append(assignment)
    return list(grouped.values())


def _interpretation_confidence_text(requirements: RemarkRequirements) -> str:
    """Return a compact confidence summary for one remark."""
    order = {value: index for index, value in enumerate(["high", "medium", "low"])}
    values = sorted({item.confidence.value for item in requirements.interpretations}, key=lambda value: order.get(value, 99))
    return "; ".join(values)


def _interpretation_type_text(requirements: RemarkRequirements) -> str:
    """Return plain requirement-type labels for one remark."""
    labels: list[str] = []
    hard_items = hard_enforceable_interpretations(requirements)
    if hard_items:
        labels.append("Explicit hard requirement")
    if any(item.enforcement == RemarkEnforcement.SOFT for item in requirements.interpretations):
        labels.append("Preference")
    if any(item.enforcement == RemarkEnforcement.FALLBACK for item in requirements.interpretations):
        labels.append("Fallback option")
    if any(item.enforcement == RemarkEnforcement.REVIEW for item in requirements.interpretations):
        labels.append("Manual review")
    return "; ".join(labels) or "No scheduling action"


def _has_informational_remark(requirements: RemarkRequirements) -> bool:
    """Return True when a remark is explicitly informational."""
    return any(bool(item.parameters.get("informational")) for item in requirements.interpretations)


def _has_unsupported_only(requirements: RemarkRequirements) -> bool:
    """Return True when no supported scheduling pattern was detected."""
    return bool(requirements.interpretations) and all(
        item.rule_name == "unsupported_remark" for item in requirements.interpretations
    )


def _has_explicit_remark_failure(assignments: list[Assignment]) -> bool:
    """Return True when an unscheduled row carries remark-specific failure evidence."""
    for assignment in assignments:
        if _is_scheduled_assignment(assignment):
            continue
        if assignment.remark_unscheduled_reason:
            return True
        if any("explicit" in reason.casefold() for reason in assignment.hard_violations):
            return True
    return False


def _hard_items_satisfied(assignments: list[Assignment], requirements: RemarkRequirements) -> bool:
    """Return True when all scheduled rows satisfy every hard-enforceable interpretation."""
    hard_items = hard_enforceable_interpretations(requirements)
    scheduled = [assignment for assignment in assignments if _is_scheduled_assignment(assignment)]
    if not hard_items or not scheduled:
        return False
    return all(
        all(assignment_satisfies_interpretation(assignment, interpretation) for interpretation in hard_items)
        for assignment in scheduled
    )


def _remark_handling_status(assignments: list[Assignment]) -> RemarkHandlingStatus:
    """Return one primary handling status for a remarked course."""
    course = assignments[0].course
    requirements = course_remark_requirements(course)
    scheduled = any(_is_scheduled_assignment(assignment) for assignment in assignments)
    if _has_informational_remark(requirements) and not requirements.needs_manual_review:
        return RemarkHandlingStatus.NO_SCHEDULING_ACTION
    if _has_explicit_remark_failure(assignments):
        return RemarkHandlingStatus.UNSCHEDULED_DUE_TO_REQUEST
    if requirements.needs_manual_review:
        if _has_unsupported_only(requirements):
            return RemarkHandlingStatus.UNSUPPORTED_NON_BLOCKING
        if scheduled:
            return RemarkHandlingStatus.SCHEDULED_NEEDS_CONFIRMATION
        return RemarkHandlingStatus.UNSUPPORTED_NON_BLOCKING
    if any(item.enforcement == RemarkEnforcement.SOFT for item in requirements.interpretations):
        return RemarkHandlingStatus.PREFERENCE_CONSIDERED
    if _hard_items_satisfied(assignments, requirements):
        return RemarkHandlingStatus.AUTOMATICALLY_APPLIED
    if _has_unsupported_only(requirements):
        return RemarkHandlingStatus.UNSUPPORTED_NON_BLOCKING
    return RemarkHandlingStatus.NO_SCHEDULING_ACTION


def _scheduled_text(course: Course, assignments: list[Assignment]) -> str:
    """Return Yes, Partial, or No for course-level scheduled coverage."""
    scheduled_weeks = {assignment.timeslot.week for assignment in assignments if assignment.timeslot is not None}
    if not scheduled_weeks:
        return "No"
    required_weeks = set(schedulable_weeks(course.teaching_weeks))
    if required_weeks and required_weeks <= scheduled_weeks:
        return "Yes"
    return "Partial"


def _assigned_values(assignments: list[Assignment], attribute: str) -> str:
    """Return unique scheduled values for day or start time."""
    values: list[str] = []
    for assignment in assignments:
        if assignment.timeslot is None:
            continue
        value = str(getattr(assignment.timeslot, attribute))
        if value not in values:
            values.append(value)
    return ", ".join(values)


def _assigned_rooms_text(assignments: list[Assignment]) -> str:
    """Return unique room groups across scheduled rows."""
    values: list[str] = []
    for assignment in assignments:
        if not _is_scheduled_assignment(assignment):
            continue
        value = assignment_room_ids(assignment)
        if value and value not in values:
            values.append(value)
    return "; ".join(values)


def _selected_delivery_text(assignments: list[Assignment]) -> str:
    """Return unique selected delivery modes for scheduled rows."""
    values: list[str] = []
    for assignment in assignments:
        if not _is_scheduled_assignment(assignment):
            continue
        value = assignment.selected_delivery_mode or assignment.course.delivery_mode
        if value and value not in values:
            values.append(value)
    return "; ".join(values)


def _review_needed(status: RemarkHandlingStatus, requirements: RemarkRequirements) -> bool:
    """Return True when staff review remains useful."""
    return status in {
        RemarkHandlingStatus.SCHEDULED_NEEDS_CONFIRMATION,
        RemarkHandlingStatus.UNSCHEDULED_DUE_TO_REQUEST,
        RemarkHandlingStatus.UNSUPPORTED_NON_BLOCKING,
    } or bool(requirements.needs_manual_review)


def _why_review_needed(assignments: list[Assignment], requirements: RemarkRequirements) -> str:
    """Return concise review evidence for a special request."""
    remark_reasons = [
        assignment.remark_unscheduled_reason
        for assignment in assignments
        if assignment.remark_unscheduled_reason
    ]
    if remark_reasons:
        return "; ".join(dict.fromkeys(remark_reasons))
    if requirements.review_reason:
        return requirements.review_reason
    return ""


def _recommended_special_request_action(status: RemarkHandlingStatus) -> str:
    """Return the next stakeholder action for one handling status."""
    if status == RemarkHandlingStatus.AUTOMATICALLY_APPLIED:
        return "No action required unless operational context changes."
    if status == RemarkHandlingStatus.PREFERENCE_CONSIDERED:
        return "Review only if the preference is operationally critical."
    if status == RemarkHandlingStatus.SCHEDULED_NEEDS_CONFIRMATION:
        return "Confirm the special-request details with programme staff."
    if status == RemarkHandlingStatus.UNSCHEDULED_DUE_TO_REQUEST:
        return "Review the explicit request, room supply, or approved exception."
    if status == RemarkHandlingStatus.UNSUPPORTED_NON_BLOCKING:
        return "Review the raw remark and decide whether manual timetabling action is needed."
    return "No timetable action required."


def _remarks_interpretation_df(assignments: list[Assignment]) -> pd.DataFrame:
    """Return row-level explainability for interpreted scheduling remarks."""
    rows: list[dict[str, object]] = []
    seen: set[tuple[object, ...]] = set()
    for assignment in assignments:
        requirements = course_remark_requirements(assignment.course)
        if not assignment.course.remarks and not requirements.interpretations:
            continue
        for interpretation in requirements.interpretations:
            key = (
                assignment.course.source_file,
                assignment.course.source_row,
                assignment.course.module_code,
                assignment.course.activity,
                interpretation.rule_name,
                str(interpretation.parameters),
                assignment.timeslot.week if assignment.timeslot else None,
            )
            if key in seen:
                continue
            seen.add(key)
            applied = assignment_satisfies_interpretation(assignment, interpretation)
            if interpretation.enforcement == RemarkEnforcement.REVIEW:
                applied_status = "Manual review"
            elif applied:
                applied_status = "Applied"
            elif _is_scheduled_assignment(assignment):
                applied_status = "Not applied"
            else:
                applied_status = "Unscheduled"
            rows.append(
                {
                    "Source Workbook": assignment.course.source_file,
                    "Source Sheet": assignment.course.source_sheet,
                    "Source Row": assignment.course.source_row,
                    "Programme/Year": assignment.course.prog_yr,
                    "Module": assignment.course.module_code,
                    "Activity": assignment.course.activity,
                    "Raw Remark": interpretation.raw_text,
                    "Detected Rule": interpretation.rule_name,
                    "Extracted Parameters": str(interpretation.parameters),
                    "Enforcement": interpretation.enforcement.value,
                    "Confidence": interpretation.confidence.value,
                    "Hard Enforceable": "Yes" if is_hard_enforceable(interpretation) else "No",
                    "Applied Status": applied_status,
                    "Assigned Rooms": assignment_room_ids(assignment),
                    "Selected Delivery Mode": assignment.selected_delivery_mode or assignment.course.delivery_mode,
                    "Explanation": interpretation.explanation,
                    "Review Reason": requirements.review_reason,
                }
            )
    return pd.DataFrame(rows, columns=REMARKS_INTERPRETATION_COLUMNS)


def _special_requests_review_df(assignments: list[Assignment]) -> pd.DataFrame:
    """Return plain-language special-request review rows."""
    rows: list[dict[str, object]] = []
    for grouped_assignments in _group_remark_assignments(assignments):
        assignment = grouped_assignments[0]
        course = assignment.course
        requirements = course_remark_requirements(course)
        status = _remark_handling_status(grouped_assignments)
        needs_review = _review_needed(status, requirements)
        understood = "; ".join(item.explanation for item in requirements.interpretations)
        rows.append(
            {
                "Programme": course.prog_yr,
                "Module": course.module_code,
                "Activity": course.activity,
                "Teaching Weeks": _format_weeks(course.teaching_weeks),
                "Special Request": course.remarks,
                "What the System Understood": understood or "No supported pattern detected.",
                "Confidence": _interpretation_confidence_text(requirements),
                "Requirement Type": _interpretation_type_text(requirements),
                "How It Was Handled": HANDLING_STATUS_LABELS[status],
                "Scheduled?": _scheduled_text(course, grouped_assignments),
                "Assigned Day": _assigned_values(grouped_assignments, "day"),
                "Assigned Time": _assigned_values(grouped_assignments, "start_time"),
                "Assigned Room(s)": _assigned_rooms_text(grouped_assignments),
                "Selected Delivery Mode": _selected_delivery_text(grouped_assignments),
                "Needs Manual Review": "Yes" if needs_review else "No",
                "Why Review Is Needed": _why_review_needed(grouped_assignments, requirements),
                "Recommended Action": _recommended_special_request_action(status),
                "Review Status": "Open" if needs_review else "Closed",
                "Review Notes": "",
            }
        )
    return pd.DataFrame(rows, columns=SPECIAL_REQUEST_REVIEW_COLUMNS)


def _special_requests_summary_df(assignments: list[Assignment]) -> pd.DataFrame:
    """Return headline counts for special-request handling."""
    review_df = _special_requests_review_df(assignments)
    if review_df.empty:
        return _metric_rows(
            {
                "Total non-empty remarks": 0,
                "Applied automatically": 0,
                "Preferences considered": 0,
                "Scheduled needing confirmation": 0,
                "Unscheduled due to explicit request": 0,
                "Unsupported non-blocking": 0,
                "No scheduling action required": 0,
            }
        )
    handling_counts = review_df["How It Was Handled"].value_counts()
    return _metric_rows(
        {
            "Total non-empty remarks": len(review_df),
            "Applied automatically": int(handling_counts.get(HANDLING_STATUS_LABELS[RemarkHandlingStatus.AUTOMATICALLY_APPLIED], 0)),
            "Preferences considered": int(handling_counts.get(HANDLING_STATUS_LABELS[RemarkHandlingStatus.PREFERENCE_CONSIDERED], 0)),
            "Scheduled needing confirmation": int(handling_counts.get(HANDLING_STATUS_LABELS[RemarkHandlingStatus.SCHEDULED_NEEDS_CONFIRMATION], 0)),
            "Unscheduled due to explicit request": int(handling_counts.get(HANDLING_STATUS_LABELS[RemarkHandlingStatus.UNSCHEDULED_DUE_TO_REQUEST], 0)),
            "Unsupported non-blocking": int(handling_counts.get(HANDLING_STATUS_LABELS[RemarkHandlingStatus.UNSUPPORTED_NON_BLOCKING], 0)),
            "No scheduling action required": int(handling_counts.get(HANDLING_STATUS_LABELS[RemarkHandlingStatus.NO_SCHEDULING_ACTION], 0)),
        }
    )


def export_stakeholder_views(assignments: list[Assignment], rooms: list[Room], output_path: Path) -> None:
    """Export separate stakeholder timetable and exception views."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    snapshot = _snapshot(assignments)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        _programme_timetable_df(snapshot).to_excel(writer, sheet_name="Programme Timetable", index=False)
        _tutor_timetable_df(snapshot).to_excel(writer, sheet_name="Tutor Timetable", index=False)
        _room_timetable_df(snapshot).to_excel(writer, sheet_name="Room Timetable", index=False)
        _exception_queue_df(snapshot, rooms=rooms).to_excel(writer, sheet_name="Exception Queue", index=False)
        _special_requests_review_df(snapshot).to_excel(writer, sheet_name="Special Requests Review", index=False)
        _special_requests_summary_df(snapshot).to_excel(writer, sheet_name="Special Requests Summary", index=False)


def _soft_weights_df() -> pd.DataFrame:
    """Return configured soft-constraint weights as rows."""
    return pd.DataFrame(
        [{"Soft Rule": rule, "Weight": weight} for rule, weight in SOFT_CONSTRAINT_WEIGHTS.items()],
        columns=["Soft Rule", "Weight"],
    )


def _template1_field_status(courses: list[Course]) -> tuple[str, str]:
    """Validate scheduler-required fields on consolidated Template 1 records."""
    required = ["module_code", "activity", "prog_yr", "class_size", "delivery_mode", "teaching_weeks", "duration_hrs"]
    missing = 0
    for course in courses:
        for field_name in required:
            value = getattr(course, field_name)
            if value in ("", [], None) or (isinstance(value, int) and value <= 0 and field_name in {"class_size", "duration_hrs"}):
                missing += 1
    if missing:
        return "FAIL", f"{missing} required field value(s) missing or invalid."
    return "PASS", "All consolidated records expose scheduler-required fields."


def _template2_status(timetable_path: Path | None, template2_path: Path) -> tuple[str, str]:
    """Validate Template 2 workbook sheet and required columns."""
    target = timetable_path if timetable_path and timetable_path.exists() else template2_path
    if not target.exists():
        return "FAIL", f"Template 2 workbook not found: {target}"
    workbook = load_workbook(target, read_only=True, data_only=True)
    try:
        if "Timetable" not in workbook.sheetnames:
            return "FAIL", "Timetable sheet is missing."
        headers = [cell.value for cell in workbook["Timetable"][1]]
        missing = [column for column in TEMPLATE2_REQUIRED_COLUMNS if column not in headers]
        if missing:
            return "FAIL", f"Missing Template 2 columns: {', '.join(missing)}"
        return "PASS", "Template 2 Timetable sheet retains expected columns."
    finally:
        workbook.close()


def _traceability_status(assignments: list[Assignment], scheduled: bool) -> tuple[str, str]:
    """Validate source traceability for scheduled or unresolved rows."""
    target = [item for item in assignments if _is_scheduled_assignment(item) == scheduled]
    if not target:
        return "PASS", "No rows in this category."
    missing = [
        item
        for item in target
        if not item.course.module_code or not item.course.activity or not item.course.prog_yr or not item.course.source_file
    ]
    if missing:
        return "FAIL", f"{len(missing)} row(s) lack module/activity/programme/source traceability."
    return "PASS", "Rows retain module, activity, programme/year, and source workbook traceability."


def _template_validation_df(courses: list[Course], assignments: list[Assignment], timetable_path: Path | None, template2_path: Path) -> pd.DataFrame:
    """Return Template 1, Template 2, and traceability validation rows."""
    template1_status, template1_notes = _template1_field_status(courses)
    template2_status, template2_notes = _template2_status(timetable_path, template2_path)
    scheduled_status, scheduled_notes = _traceability_status(assignments, scheduled=True)
    unresolved_status, unresolved_notes = _traceability_status(assignments, scheduled=False)
    rows = [
        {"Check": "Consolidated Template 1 scheduler fields present", "Status": template1_status, "Notes": template1_notes},
        {"Check": "Template 2 output structure retained", "Status": template2_status, "Notes": template2_notes},
        {"Check": "Scheduled row traceability", "Status": scheduled_status, "Notes": scheduled_notes},
        {"Check": "Unresolved row traceability", "Status": unresolved_status, "Notes": unresolved_notes},
    ]
    return pd.DataFrame(rows, columns=["Check", "Status", "Notes"])


def _traceability_df(assignments: list[Assignment]) -> pd.DataFrame:
    """Return row-level schedule traceability evidence."""
    rows: list[dict[str, object]] = []
    for assignment in assignments:
        rows.append(
            {
                "Status": "Scheduled" if _is_scheduled_assignment(assignment) else "Unscheduled",
                "Source File": assignment.course.source_file,
                "Programme/Year": assignment.course.prog_yr,
                "Module Code": assignment.course.module_code,
                "Activity": assignment.course.activity,
                "Teaching Weeks": _format_weeks(assignment.course.teaching_weeks),
                "Scheduled Week": assignment.timeslot.week if assignment.timeslot else "",
                "Day": assignment.timeslot.day if assignment.timeslot else "",
                "Start": assignment.timeslot.start_time if assignment.timeslot else "",
                "Room": assignment_room_ids(assignment),
                "Original Reason": " | ".join(assignment.hard_violations),
            }
        )
    return pd.DataFrame(
        rows,
        columns=[
            "Status",
            "Source File",
            "Programme/Year",
            "Module Code",
            "Activity",
            "Teaching Weeks",
            "Scheduled Week",
            "Day",
            "Start",
            "Room",
            "Original Reason",
        ],
    )


def export_run_manifest(
    courses: list[Course],
    assignments: list[Assignment],
    output_path: Path,
    metadata: dict[str, object] | None = None,
    rooms: list[Room] | None = None,
    output_files: dict[str, Path] | None = None,
    template2_path: Path = DEFAULT_TEMPLATE2_FILE,
) -> None:
    """Export reproducibility, validation, and traceability evidence."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    snapshot = _snapshot(assignments)
    demand = build_demand_metrics(courses, snapshot, input_course_records=len(courses))
    summary = build_run_summary(snapshot, demand_metrics=demand)
    template_validation = _template_validation_df(courses, snapshot, (output_files or {}).get("timetable"), template2_path)
    validation_status = "PASS" if demand.is_consistent and summary["hard_violations_on_scheduled_assignments"] == 0 else "FAIL"
    generated_at = datetime.now().isoformat(timespec="seconds")
    manifest_rows = [
        {"Setting": "run_timestamp", "Value": generated_at},
        {"Setting": "scope", "Value": (metadata or {}).get("scope", "")},
        {"Setting": "Input role", "Value": "Template 1 - Scheduling Requirements"},
        {"Setting": "Input workbook", "Value": (metadata or {}).get("input_workbook", "")},
        {"Setting": "Output-template role", "Value": "Template 2 - Proposed Timetable"},
        {"Setting": "Output template", "Value": str(template2_path)},
        {"Setting": "Generated timetable", "Value": str((output_files or {}).get("timetable", ""))},
        {"Setting": "input_files", "Value": "; ".join(sorted({course.source_file for course in courses if course.source_file}))},
        {"Setting": "input_record_count", "Value": len(courses)},
        {"Setting": "consolidated_requirement_count", "Value": demand.consolidated_course_requirements},
        {"Setting": "required_teaching_occurrences", "Value": demand.required_teaching_occurrences},
        {"Setting": "scheduled_teaching_occurrences", "Value": demand.scheduled_teaching_occurrences},
        {"Setting": "unscheduled_teaching_occurrences", "Value": demand.unscheduled_teaching_occurrences},
        {"Setting": "coverage_rate_percent", "Value": demand.coverage_rate_percent},
        {"Setting": "hard_violations_on_scheduled_assignments", "Value": summary["hard_violations_on_scheduled_assignments"]},
        {"Setting": "weighted_soft_score", "Value": weighted_soft_score(snapshot)},
        {"Setting": "validation_status", "Value": validation_status},
        {"Setting": "loaded_rooms", "Value": len(rooms or [])},
    ]
    for key, value in (metadata or {}).items():
        if key.startswith("optimisation") or key in {"max_iterations", "skip_optimisation"}:
            manifest_rows.append({"Setting": key, "Value": value})
    for name, path in (output_files or {}).items():
        manifest_rows.append({"Setting": f"output_file_{name}", "Value": str(path)})

    before_after = pd.DataFrame(
        [{"Soft Rule": rule, "Count": count} for rule, count in soft_violation_breakdown(snapshot).items()],
        columns=["Soft Rule", "Count"],
    )
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame(manifest_rows, columns=["Setting", "Value"]).to_excel(writer, sheet_name="Run Manifest", index=False)
        _soft_weights_df().to_excel(writer, sheet_name="Soft Constraint Weights", index=False)
        before_after.to_excel(writer, sheet_name="Soft Rule Baseline", index=False)
        template_validation.to_excel(writer, sheet_name="Template Validation", index=False)
        _traceability_df(snapshot).to_excel(writer, sheet_name="Traceability", index=False)


def export_run_summary(
    assignments: list[Assignment],
    output_path: Path,
    metadata: dict[str, object] | None = None,
    demand_courses: list[Course] | None = None,
    input_course_records: int | None = None,
    rooms: list[Room] | None = None,
    room_source_path: Path | None = None,
    optimisation_summary: dict[str, object] | None = None,
) -> None:
    """Export a stakeholder-friendly run summary workbook."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    snapshot = _snapshot(assignments)
    generated_at = datetime.now().isoformat(timespec="seconds")
    demand_metrics = (
        build_demand_metrics(demand_courses, snapshot, input_course_records=input_course_records)
        if demand_courses is not None
        else None
    )
    unscheduled_analysis = _unscheduled_analysis_df(snapshot, demand_courses=demand_courses)
    resource_audit = (
        audit_resources(demand_courses, rooms, snapshot, room_source_path=room_source_path)
        if demand_courses is not None and rooms is not None
        else None
    )
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        _summary_df(snapshot, demand_metrics=demand_metrics).to_excel(writer, sheet_name="Summary", index=False)
        _violations_df(snapshot, "hard").to_excel(writer, sheet_name="Hard Violations", index=False)
        _violations_df(snapshot, "soft").to_excel(writer, sheet_name="Soft Violations", index=False)
        _unscheduled_reasons_df(snapshot).to_excel(writer, sheet_name="Unscheduled Reasons", index=False)
        unscheduled_analysis.to_excel(writer, sheet_name="Unscheduled Analysis", index=False)
        _unscheduled_breakdown_df(unscheduled_analysis).to_excel(writer, sheet_name="Unscheduled Breakdown", index=False)
        _residual_f2f_analysis_df(snapshot, rooms=rooms).to_excel(writer, sheet_name="Residual F2F Analysis", index=False)
        _room_utilisation_df(snapshot).to_excel(writer, sheet_name="Room Utilisation", index=False)
        _resource_audit_df(resource_audit).to_excel(writer, sheet_name="Resource Audit", index=False)
        _virtual_room_detail_df(resource_audit).to_excel(writer, sheet_name="Virtual Room Detail", index=False)
        _programme_breakdown_df(snapshot).to_excel(writer, sheet_name="Programme Breakdown", index=False)
        _remarks_interpretation_df(snapshot).to_excel(writer, sheet_name="Remarks Interpretation", index=False)
        _validation_checks_df(
            snapshot,
            generated_at,
            demand_metrics=demand_metrics,
            optimisation_summary=optimisation_summary,
            resource_audit=resource_audit,
        ).to_excel(writer, sheet_name="Validation Checks", index=False)
        _optimisation_summary_df(optimisation_summary).to_excel(writer, sheet_name="Optimisation Summary", index=False)
        _metadata_df(metadata, generated_at).to_excel(writer, sheet_name="Run Metadata", index=False)
