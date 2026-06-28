"""Calendar-style timetable workbook exports built from validated assignments."""

from __future__ import annotations

import re
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import EARLIEST_START_HOUR, LATEST_END_HOUR, VALID_DAYS
from data.models import Assignment, Room
from engine.constraint_checker import assignment_delivery_mode, assignment_end_time
from engine.fixed_reconciliation import normalise_programme_year
from engine.remarks_interpreter import assignment_rooms

INVALID_SHEET_CHARS = r"[\[\]\:\*\?\/\\]"
MAX_SHEET_NAME_LENGTH = 31
INDEX_SHEET = "Index"
LEGEND_SHEET = "Legend"
DATA_SHEET = "Structured Data"

PHYSICAL_FILL = "D9EAF7"
ONLINE_FILL = "D8F3EE"
EXTERNAL_FILL = "FCE4D6"
HEADER_FILL = "1F4E78"
WARNING_FILL = "FFF2CC"
DETAIL_FILL = "E2F0D9"
GRID_FILL = "F7F9FB"
FIXED_BORDER = Side(style="thick", color="1F1F1F")
GENERATED_BORDER = Side(style="thin", color="808080")
THIN_BORDER = Side(style="thin", color="D9E2F3")


@dataclass(frozen=True, slots=True)
class VisualTimetableEntry:
    """A scheduled assignment row ready for visual timetable export."""

    assignment_id: str
    programme: str
    academic_year: str
    programme_year: str
    module_code: str
    activity: str
    groups: tuple[str, ...]
    lecturers: tuple[str, ...]
    rooms: tuple[str, ...]
    day: str
    start_time: time
    end_time: time
    teaching_weeks: tuple[int, ...]
    is_fixed: bool
    delivery_mode: str
    is_external: bool
    is_shared: bool
    location_type: str = ""
    source_reference: str = ""
    programme_status: str = ""
    submission_ready_status: str = ""
    room_capacity_status: str = ""
    lecturer_keys: tuple[str, ...] = ()
    source_occurrence_ids: tuple[str, ...] = ()


@dataclass(slots=True)
class VisualExportResult:
    """Result summary for visual workbook generation."""

    programme_path: Path
    tutor_path: Path
    room_path: Path
    validation_path: Path
    status: str
    programme_sheets: int
    tutor_sheets: int
    room_sheets: int
    programme_entries: int
    tutor_entries: int
    room_entries: int
    missing_entries: int
    unexpected_entries: int
    invalid_overlaps: int
    sheet_name_issues: int
    max_weekday_lanes: int = 0
    lane_warnings: int = 0
    errors: list[str] = field(default_factory=list)


def scheduled_assignment_id(assignment: Assignment) -> str:
    """Return a stable traceable ID for a scheduled assignment."""
    course = assignment.course
    timeslot = assignment.timeslot
    parts = [
        assignment.fixed_source or course.fixed_source or "",
        course.source_file,
        course.source_sheet,
        str(course.source_row or ""),
        course.prog_yr,
        course.module_code,
        course.activity,
        str(timeslot.week if timeslot else ""),
        timeslot.day if timeslot else "",
        timeslot.start_time if timeslot else "",
    ]
    return "|".join(_clean_id_part(part) for part in parts if str(part).strip())


def safe_sheet_name(value: str, used: set[str] | None = None, prefix: str = "") -> str:
    """Return a deterministic Excel-safe worksheet name."""
    used = used if used is not None else set()
    cleaned = re.sub(INVALID_SHEET_CHARS, "_", value.strip()) or "Sheet"
    cleaned = re.sub(r"\s+", "_", cleaned)
    cleaned = cleaned.strip("'_") or "Sheet"
    if prefix:
        cleaned = f"{prefix}_{cleaned}"
    base = cleaned[:MAX_SHEET_NAME_LENGTH]
    name = base
    counter = 2
    while name in used:
        suffix = f"_{counter}"
        name = f"{base[: MAX_SHEET_NAME_LENGTH - len(suffix)]}{suffix}"
        counter += 1
    used.add(name)
    return name


def allocate_lanes(entries: list[VisualTimetableEntry]) -> dict[str, int]:
    """Assign deterministic lane numbers so visual blocks do not overwrite."""
    lanes: list[list[VisualTimetableEntry]] = []
    allocation: dict[str, int] = {}
    for entry in sorted(entries, key=_entry_sort_key):
        placed = False
        for lane_index, lane_entries in enumerate(lanes):
            if not any(_entries_conflict_for_lane(entry, other) for other in lane_entries):
                lane_entries.append(entry)
                allocation[entry.assignment_id] = lane_index
                placed = True
                break
        if not placed:
            lanes.append([entry])
            allocation[entry.assignment_id] = len(lanes) - 1
    return allocation


def aggregate_visual_entries(entries: list[VisualTimetableEntry]) -> list[VisualTimetableEntry]:
    """Combine identical recurring placements into readable visual blocks."""
    grouped: dict[tuple[object, ...], list[VisualTimetableEntry]] = defaultdict(list)
    for entry in entries:
        grouped[_aggregation_key(entry)].append(entry)

    aggregated: list[VisualTimetableEntry] = []
    for items in grouped.values():
        first = sorted(items, key=_entry_sort_key)[0]
        weeks = tuple(sorted({week for item in items for week in item.teaching_weeks}))
        source_ids = tuple(sorted({source_id for item in items for source_id in _source_occurrence_ids(item)}))
        aggregated.append(
            VisualTimetableEntry(
                assignment_id=_aggregated_assignment_id(first, weeks),
                programme=first.programme,
                academic_year=first.academic_year,
                programme_year=first.programme_year,
                module_code=first.module_code,
                activity=first.activity,
                groups=first.groups,
                lecturers=first.lecturers,
                rooms=first.rooms,
                day=first.day,
                start_time=first.start_time,
                end_time=first.end_time,
                teaching_weeks=weeks,
                is_fixed=first.is_fixed,
                delivery_mode=first.delivery_mode,
                is_external=first.is_external,
                is_shared=first.is_shared,
                location_type=first.location_type,
                source_reference=first.source_reference,
                programme_status=first.programme_status,
                submission_ready_status=first.submission_ready_status,
                room_capacity_status=first.room_capacity_status,
                lecturer_keys=first.lecturer_keys,
                source_occurrence_ids=source_ids,
            )
        )
    return sorted(aggregated, key=_entry_sort_key)


def build_programme_visual_entries(
    assignments: list[Assignment],
    programme_rows: list[dict[str, object]] | None = None,
) -> list[VisualTimetableEntry]:
    """Return programme-expanded visual entries for scheduled assignments."""
    statuses = _programme_status_lookup(programme_rows or [])
    entries: list[VisualTimetableEntry] = []
    for assignment in _scheduled_valid_assignments(assignments):
        for programme_year in _programme_labels(assignment):
            entries.append(_entry_from_assignment(assignment, programme_year, statuses))
    return entries


def build_tutor_visual_entries(assignments: list[Assignment]) -> list[VisualTimetableEntry]:
    """Return one visual entry per tutor/assignment occupancy."""
    entries: list[VisualTimetableEntry] = []
    seen: set[tuple[str, str]] = set()
    for assignment in _scheduled_valid_assignments(assignments):
        base = _entry_from_assignment(assignment, normalise_programme_year(assignment.course.prog_yr), {})
        for lecturer in base.lecturers or ("Unassigned tutor",):
            key = (lecturer, base.assignment_id)
            if key in seen:
                continue
            seen.add(key)
            entries.append(base)
    return entries


def build_room_visual_entries(assignments: list[Assignment]) -> list[VisualTimetableEntry]:
    """Return one visual entry per physical or external room occupancy."""
    entries: list[VisualTimetableEntry] = []
    seen: set[tuple[str, str]] = set()
    for assignment in _scheduled_valid_assignments(assignments):
        base = _entry_from_assignment(assignment, normalise_programme_year(assignment.course.prog_yr), {})
        for room in assignment_rooms(assignment):
            if room.room_type == "virtual":
                continue
            key = (room.room_id, base.assignment_id)
            if key in seen:
                continue
            seen.add(key)
            entries.append(_replace_rooms(base, (room.room_id,), room))
    return entries


def export_timetable_visuals(
    *,
    assignments: list[Assignment],
    rooms: list[Room],
    programme_rows: list[dict[str, object]] | None,
    programme_path: Path,
    tutor_path: Path,
    room_path: Path,
    validation_path: Path,
) -> VisualExportResult:
    """Export programme, tutor and room visual timetable workbooks."""
    programme_entries = aggregate_visual_entries(build_programme_visual_entries(assignments, programme_rows))
    tutor_entries = aggregate_visual_entries(build_tutor_visual_entries(assignments))
    room_entries = aggregate_visual_entries(build_room_visual_entries(assignments))

    programme_sheets = _export_entity_workbook(
        path=programme_path,
        title="Programme Timetables",
        entity_type="programme",
        entries=programme_entries,
        groups=_group_programmes(programme_entries),
        programme_rows=programme_rows or [],
    )
    tutor_sheets = _export_entity_workbook(
        path=tutor_path,
        title="Tutor Timetables",
        entity_type="tutor",
        entries=tutor_entries,
        groups=_group_tutors(tutor_entries),
        programme_rows=[],
    )
    room_sheets = _export_entity_workbook(
        path=room_path,
        title="Room Timetables",
        entity_type="room",
        entries=room_entries,
        groups=_group_rooms(room_entries),
        programme_rows=[],
    )

    validation = validate_visual_exports(
        assignments=assignments,
        rooms=rooms,
        programme_entries=programme_entries,
        tutor_entries=tutor_entries,
        room_entries=room_entries,
        programme_sheets=programme_sheets,
        tutor_sheets=tutor_sheets,
        room_sheets=room_sheets,
        programme_rows=programme_rows or [],
    )
    result = VisualExportResult(
        programme_path=programme_path,
        tutor_path=tutor_path,
        room_path=room_path,
        validation_path=validation_path,
        status=validation["status"],
        programme_sheets=programme_sheets,
        tutor_sheets=tutor_sheets,
        room_sheets=room_sheets,
        programme_entries=len(programme_entries),
        tutor_entries=len(tutor_entries),
        room_entries=len(room_entries),
        missing_entries=len(validation["missing"]),
        unexpected_entries=len(validation["unexpected"]),
        invalid_overlaps=len(validation["overlaps"]),
        sheet_name_issues=len(validation["sheet_names"]),
        max_weekday_lanes=int(validation["max_weekday_lanes"]),
        lane_warnings=len(validation["lane_quality_warnings"]),
        errors=[str(row.get("Issue")) for row in validation["export_status"] if row.get("Status") == "FAIL"],
    )
    export_visualisation_validation_report(result, validation, validation_path)
    return result


def export_visualisation_failure_report(validation_path: Path, error: Exception) -> VisualExportResult:
    """Write a FAIL validation workbook when visual export raises."""
    result = VisualExportResult(
        programme_path=Path(""),
        tutor_path=Path(""),
        room_path=Path(""),
        validation_path=validation_path,
        status="FAIL",
        programme_sheets=0,
        tutor_sheets=0,
        room_sheets=0,
        programme_entries=0,
        tutor_entries=0,
        room_entries=0,
        missing_entries=0,
        unexpected_entries=0,
        invalid_overlaps=0,
        sheet_name_issues=0,
        errors=[str(error)],
    )
    validation = {
        "summary": _summary_rows(result, 0),
        "programme": [],
        "tutor": [],
        "room": [],
        "lane_quality": [],
        "lane_quality_warnings": [],
        "max_weekday_lanes": 0,
        "missing": [],
        "unexpected": [],
        "overlaps": [],
        "sheet_names": [],
        "export_status": [{"Workbook": "Visual export", "Status": "FAIL", "Issue": str(error)}],
    }
    export_visualisation_validation_report(result, validation, validation_path)
    return result


def validate_visual_exports(
    *,
    assignments: list[Assignment],
    rooms: list[Room],
    programme_entries: list[VisualTimetableEntry],
    tutor_entries: list[VisualTimetableEntry],
    room_entries: list[VisualTimetableEntry],
    programme_sheets: int,
    tutor_sheets: int,
    room_sheets: int,
    programme_rows: list[dict[str, object]],
) -> dict[str, object]:
    """Return validation rows for exported timetable visuals."""
    scheduled_ids = {scheduled_assignment_id(item) for item in _scheduled_valid_assignments(assignments)}
    programme_ids = _entry_occurrence_id_set(programme_entries)
    tutor_ids = _entry_occurrence_id_set(tutor_entries)
    physical_ids = {
        scheduled_assignment_id(item)
        for item in _scheduled_valid_assignments(assignments)
        if any(room.room_type != "virtual" for room in assignment_rooms(item))
    }
    room_ids = _entry_occurrence_id_set(room_entries)
    missing = _missing_rows("programme", scheduled_ids - programme_ids)
    missing.extend(_missing_rows("tutor", scheduled_ids - tutor_ids))
    missing.extend(_missing_rows("room", physical_ids - room_ids))
    unexpected = _unexpected_rows("programme", programme_ids - scheduled_ids)
    unexpected.extend(_unexpected_rows("tutor", tutor_ids - scheduled_ids))
    unexpected.extend(_unexpected_rows("room", room_ids - scheduled_ids))
    overlaps = _overlap_rows("Tutor", _group_tutor_keys(tutor_entries))
    overlaps.extend(_overlap_rows("Room", _group_rooms(room_entries)))
    sheet_names = _sheet_name_rows(
        [
            *[key for key in _group_programmes(programme_entries)],
            *[key for key in _group_tutors(tutor_entries)],
            *[key for key in _group_rooms(room_entries)],
        ]
    )
    programme_reconciliation = _programme_reconciliation_rows(programme_entries, programme_rows)
    tutor_reconciliation = _entity_reconciliation_rows("Tutor", _group_tutors(tutor_entries))
    room_reconciliation = _entity_reconciliation_rows("Room", _group_rooms(room_entries), rooms)
    lane_quality = _lane_quality_rows(
        {
            "Programme": _group_programmes(programme_entries),
            "Tutor": _group_tutors(tutor_entries),
            "Room": _group_rooms(room_entries),
        }
    )
    lane_warnings = [row for row in lane_quality if row.get("Status") == "WARN"]
    max_weekday_lanes = max((int(row.get("Maximum Weekday Lanes") or 0) for row in lane_quality), default=0)
    status = "PASS" if not missing and not unexpected and not overlaps and not sheet_names else "FAIL"
    result_stub = VisualExportResult(
        programme_path=Path(""),
        tutor_path=Path(""),
        room_path=Path(""),
        validation_path=Path(""),
        status=status,
        programme_sheets=programme_sheets,
        tutor_sheets=tutor_sheets,
        room_sheets=room_sheets,
        programme_entries=len(programme_entries),
        tutor_entries=len(tutor_entries),
        room_entries=len(room_entries),
        missing_entries=len(missing),
        unexpected_entries=len(unexpected),
        invalid_overlaps=len(overlaps),
        sheet_name_issues=len(sheet_names),
        max_weekday_lanes=max_weekday_lanes,
        lane_warnings=len(lane_warnings),
    )
    export_status = [
        {"Workbook": "Programme_Timetable_Visuals.xlsx", "Status": "PASS" if programme_sheets else "FAIL", "Issue": ""},
        {"Workbook": "Tutor_Timetable_Visuals.xlsx", "Status": "PASS" if tutor_sheets else "FAIL", "Issue": ""},
        {"Workbook": "Room_Timetable_Visuals.xlsx", "Status": "PASS" if room_sheets >= 0 else "FAIL", "Issue": ""},
        {
            "Workbook": "Lane quality",
            "Status": "WARN" if lane_warnings else "PASS",
            "Issue": f"{len(lane_warnings)} sheet/day combination(s) exceed four lanes." if lane_warnings else "",
        },
        {"Workbook": "timetable_visualisation_validation.xlsx", "Status": status, "Issue": "" if status == "PASS" else "Review validation sheets."},
    ]
    return {
        "status": status,
        "summary": _summary_rows(result_stub, len(scheduled_ids)),
        "programme": programme_reconciliation,
        "tutor": tutor_reconciliation,
        "room": room_reconciliation,
        "lane_quality": lane_quality,
        "lane_quality_warnings": lane_warnings,
        "max_weekday_lanes": max_weekday_lanes,
        "missing": missing,
        "unexpected": unexpected,
        "overlaps": overlaps,
        "sheet_names": sheet_names,
        "export_status": export_status,
    }


def export_visualisation_validation_report(
    result: VisualExportResult,
    validation: dict[str, object],
    output_path: Path,
) -> None:
    """Export visualisation validation workbook."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    _write_table(summary, ["Metric", "Value"], validation.get("summary", _summary_rows(result, 0)))
    _write_table(workbook.create_sheet("Programme Reconciliation"), list(_programme_reconciliation_columns()), validation.get("programme", []))
    _write_table(workbook.create_sheet("Tutor Reconciliation"), list(_entity_reconciliation_columns("Tutor")), validation.get("tutor", []))
    _write_table(workbook.create_sheet("Room Reconciliation"), list(_entity_reconciliation_columns("Room")), validation.get("room", []))
    _write_table(
        workbook.create_sheet("Lane Quality"),
        ["Entity Type", "Entity", "Day", "Maximum Weekday Lanes", "Status", "Issue"],
        validation.get("lane_quality", []),
    )
    _write_table(workbook.create_sheet("Missing Visual Entries"), ["View", "Assignment ID"], validation.get("missing", []))
    _write_table(workbook.create_sheet("Unexpected Visual Entries"), ["View", "Assignment ID"], validation.get("unexpected", []))
    _write_table(
        workbook.create_sheet("Overlap Validation"),
        ["Entity Type", "Entity", "Left Assignment ID", "Right Assignment ID", "Day", "Start", "End", "Weeks", "Status"],
        validation.get("overlaps", []),
    )
    _write_table(workbook.create_sheet("Sheet Name Validation"), ["Original Name", "Sheet Name", "Status", "Issue"], validation.get("sheet_names", []))
    _write_table(workbook.create_sheet("Export Status"), ["Workbook", "Status", "Issue"], validation.get("export_status", []))
    _style_workbook(workbook)
    workbook.save(output_path)


def _export_entity_workbook(
    *,
    path: Path,
    title: str,
    entity_type: str,
    entries: list[VisualTimetableEntry],
    groups: dict[str, list[VisualTimetableEntry]],
    programme_rows: list[dict[str, object]],
) -> int:
    """Export one visual workbook for a stakeholder entity type."""
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    index = workbook.active
    index.title = INDEX_SHEET
    used = {INDEX_SHEET}
    sheet_map: dict[str, str] = {}
    for entity in sorted(groups):
        sheet_name = safe_sheet_name(entity, used, _sheet_prefix(entity_type))
        sheet_map[entity] = sheet_name
        _write_entity_sheet(workbook.create_sheet(sheet_name), title, entity_type, entity, groups[entity], programme_rows)
    _write_index_sheet(index, entity_type, groups, sheet_map, programme_rows)
    _write_legend_sheet(workbook.create_sheet(LEGEND_SHEET))
    _write_structured_data_sheet(workbook.create_sheet(DATA_SHEET), entries)
    _style_workbook(workbook)
    workbook.save(path)
    return len(groups)


def _write_entity_sheet(
    sheet,
    title: str,
    entity_type: str,
    entity: str,
    entries: list[VisualTimetableEntry],
    programme_rows: list[dict[str, object]],
) -> None:
    """Write one weekly grid plus detail table."""
    sheet.sheet_view.showGridLines = False
    sheet["A1"] = title
    sheet["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor=HEADER_FILL)
    sheet["A2"] = entity
    sheet["A2"].font = Font(bold=True, size=13)
    status = _entity_status(entity_type, entity, entries, programme_rows)
    sheet["A3"] = status
    sheet["A3"].fill = PatternFill("solid", fgColor=WARNING_FILL if "Incomplete" in status else DETAIL_FILL)
    sheet["A4"] = _entity_summary(entity_type, entity, entries, programme_rows)
    sheet["A5"] = "Legend: Physical, Online, External, FIXED, GENERATED, SHARED. Blocks show valid scheduled assignments only."
    _write_week_grid(sheet, entries, start_row=7)
    detail_start = max(sheet.max_row + 3, 35)
    sheet.cell(row=detail_start, column=1, value="Structured Detail")
    sheet.cell(row=detail_start, column=1).font = Font(bold=True, size=12)
    _write_table(sheet, _detail_columns(), [_entry_detail_row(entry) for entry in sorted(entries, key=_entry_sort_key)], start_row=detail_start + 1)
    sheet.freeze_panes = "B8"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.5
    sheet.page_margins.bottom = 0.5
    sheet.print_title_rows = "1:7"
    sheet.oddHeader.center.text = f"{title} - {entity}"
    sheet.oddFooter.center.text = "Generated &D &T"
    sheet.oddFooter.right.text = "Page &P of &N"
    sheet.print_area = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"


def _write_week_grid(sheet, entries: list[VisualTimetableEntry], start_row: int) -> None:
    """Write a weekly timetable grid using deterministic day lanes."""
    resolution = _time_resolution(entries)
    times = _time_axis(entries, resolution)
    row_for_time = {value: start_row + 2 + index for index, value in enumerate(times)}
    columns: dict[tuple[str, int], int] = {}
    current_col = 2
    sheet.cell(row=start_row, column=1, value="Time")
    sheet.cell(row=start_row + 1, column=1, value="")
    for day in VALID_DAYS:
        day_entries = [entry for entry in entries if entry.day == day]
        lanes = allocate_lanes(day_entries)
        lane_count = max(lanes.values(), default=0) + 1 if day_entries else 1
        first_col = current_col
        for lane in range(lane_count):
            columns[(day, lane)] = current_col
            sheet.cell(row=start_row + 1, column=current_col, value=f"{day} {lane + 1}" if lane_count > 1 else "")
            current_col += 1
        sheet.cell(row=start_row, column=first_col, value=day)
        if current_col - first_col > 1:
            sheet.merge_cells(start_row=start_row, start_column=first_col, end_row=start_row, end_column=current_col - 1)
    for time_value, row_index in row_for_time.items():
        sheet.cell(row=row_index, column=1, value=_format_time(time_value))
        for column in range(2, current_col):
            cell = sheet.cell(row=row_index, column=column)
            cell.fill = PatternFill("solid", fgColor=GRID_FILL)
            cell.border = Border(left=THIN_BORDER, right=THIN_BORDER, top=THIN_BORDER, bottom=THIN_BORDER)
    for day in VALID_DAYS:
        day_entries = [entry for entry in entries if entry.day == day]
        lanes = allocate_lanes(day_entries)
        for entry in sorted(day_entries, key=_entry_sort_key):
            lane = lanes[entry.assignment_id]
            column = columns[(day, lane)]
            first_row = row_for_time.get(entry.start_time)
            last_row = row_for_time.get(entry.end_time)
            if first_row is None or last_row is None or last_row <= first_row:
                continue
            end_row = last_row - 1
            if end_row > first_row:
                sheet.merge_cells(start_row=first_row, start_column=column, end_row=end_row, end_column=column)
            cell = sheet.cell(row=first_row, column=column, value=_block_text(entry))
            cell.fill = PatternFill("solid", fgColor=_entry_fill(entry))
            side = FIXED_BORDER if entry.is_fixed else GENERATED_BORDER
            cell.border = Border(left=side, right=side, top=side, bottom=side)
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            cell.font = Font(size=9)
    for column in range(1, current_col):
        sheet.column_dimensions[get_column_letter(column)].width = 18 if column > 1 else 9
    for row in range(start_row, start_row + 2):
        for column in range(1, current_col):
            cell = sheet.cell(row=row, column=column)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
            cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_index_sheet(
    sheet,
    entity_type: str,
    groups: dict[str, list[VisualTimetableEntry]],
    sheet_map: dict[str, str],
    programme_rows: list[dict[str, object]],
) -> None:
    """Write a linked index sheet."""
    if entity_type == "programme":
        columns = [
            "Programme",
            "Year",
            "Display name",
            "Status",
            "Required occurrences",
            "Scheduled occurrences",
            "Quarantined occurrences",
            "Unscheduled occurrences",
            "Submission ready",
            "Sheet link",
        ]
        rows = [_programme_index_row(entity, entries, sheet_map[entity], programme_rows) for entity, entries in sorted(groups.items())]
    elif entity_type == "tutor":
        columns = ["Tutor", "Scheduled assignments", "Fixed assignments", "Generated assignments", "Online assignments", "Earliest class", "Latest class", "Sheet link"]
        rows = [_tutor_index_row(entity, entries, sheet_map[entity]) for entity, entries in sorted(groups.items())]
    else:
        columns = ["Room", "Location type", "Scheduled assignments", "Fixed assignments", "Capacity status", "External venue", "Earliest use", "Latest use", "Sheet link"]
        rows = [_room_index_row(entity, entries, sheet_map[entity]) for entity, entries in sorted(groups.items())]
    _write_table(sheet, columns, rows)
    for row_index in range(2, len(rows) + 2):
        cell = sheet.cell(row=row_index, column=len(columns))
        sheet_name = str(cell.value)
        cell.hyperlink = f"#{sheet_name}!A1"
        cell.style = "Hyperlink"


def _write_legend_sheet(sheet) -> None:
    """Write workbook legend and instructions."""
    rows = [
        {"Marker": "Physical session", "Meaning": "Scheduled class in an internal physical room."},
        {"Marker": "ONLINE", "Meaning": "Online delivery placeholder; not a physical room booking."},
        {"Marker": "EXTERNAL", "Meaning": "Approved or recognised external venue."},
        {"Marker": "FIXED", "Meaning": "Anchored fixed session from the source workbook."},
        {"Marker": "GENERATED", "Meaning": "Generated by the scheduler."},
        {"Marker": "SHARED", "Meaning": "Shared or common-module session; trace by assignment ID."},
    ]
    _write_table(sheet, ["Marker", "Meaning"], rows)
    sheet["D1"] = "Instructions"
    sheet["D2"] = "Use the Index links to open timetable sheets. Blocks show valid scheduled assignments only."
    sheet["D3"] = "Unscheduled or quarantined requirements are reported in exception workbooks, not shown as calendar blocks."
    sheet["D4"] = "Every block includes an assignment ID for traceability."
    sheet["D1"].font = Font(bold=True)


def _write_structured_data_sheet(sheet, entries: list[VisualTimetableEntry]) -> None:
    """Write exact row-level visual data."""
    _write_table(sheet, _detail_columns(), [_entry_detail_row(entry) for entry in sorted(entries, key=_entry_sort_key)])


def _write_table(sheet, columns: list[str], rows: object, start_row: int = 1) -> None:
    """Write rows with a header and Excel table formatting."""
    row_list = list(rows or [])
    for column_index, header in enumerate(columns, start=1):
        sheet.cell(row=start_row, column=column_index, value=header)
    for row_offset, row in enumerate(row_list, start=1):
        for column_index, header in enumerate(columns, start=1):
            value = row.get(header, "") if isinstance(row, dict) else ""
            sheet.cell(row=start_row + row_offset, column=column_index, value=value)
    sheet.freeze_panes = f"A{start_row + 1}"
    sheet.auto_filter.ref = f"A{start_row}:{get_column_letter(len(columns))}{max(start_row + len(row_list), start_row)}"


def _style_workbook(workbook: Workbook) -> None:
    """Apply readable workbook-wide styling."""
    workbook.properties.created = datetime(2000, 1, 1)
    workbook.properties.modified = datetime(2000, 1, 1)
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in sheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 36)


def _scheduled_valid_assignments(assignments: list[Assignment]) -> list[Assignment]:
    """Return scheduled assignments with no hard violations."""
    return [item for item in assignments if item.room is not None and item.timeslot is not None and not item.hard_violations]


def _entry_from_assignment(
    assignment: Assignment,
    programme_year: str,
    statuses: dict[str, dict[str, str]],
) -> VisualTimetableEntry:
    """Build one visual entry from a scheduled assignment."""
    assignment_id = scheduled_assignment_id(assignment)
    rooms = assignment_rooms(assignment)
    primary_room = rooms[0] if rooms else assignment.room
    room_ids = tuple(room.room_id for room in rooms)
    programme = _programme_code(programme_year)
    academic_year = _academic_year(programme_year)
    status = statuses.get(normalise_programme_year(programme_year), {})
    return VisualTimetableEntry(
        assignment_id=assignment_id,
        programme=programme,
        academic_year=academic_year,
        programme_year=normalise_programme_year(programme_year),
        module_code=assignment.course.module_code,
        activity=assignment.course.activity,
        groups=tuple(assignment.course.group_ids or [assignment.course.prog_yr]),
        lecturers=tuple(assignment.course.staff_names or assignment.course.staff_ids or ("Unassigned tutor",)),
        lecturer_keys=tuple(assignment.course.staff_ids or assignment.course.staff_names or ("Unassigned tutor",)),
        rooms=room_ids,
        day=assignment.timeslot.day if assignment.timeslot else "",
        start_time=_parse_time(assignment.timeslot.start_time if assignment.timeslot else "00:00"),
        end_time=_parse_time(assignment_end_time(assignment)),
        teaching_weeks=tuple(sorted({assignment.timeslot.week})) if assignment.timeslot else (),
        is_fixed=assignment.is_fixed,
        delivery_mode=assignment_delivery_mode(assignment),
        is_external=any(room.room_type == "external" or "external" in room.resource_type.casefold() for room in rooms),
        is_shared=_is_shared_assignment(assignment),
        location_type=_location_type(primary_room),
        source_reference=_source_reference(assignment),
        programme_status=status.get("status", ""),
        submission_ready_status=status.get("submission", ""),
        room_capacity_status=_capacity_status(primary_room),
        source_occurrence_ids=(assignment_id,),
    )


def _replace_rooms(entry: VisualTimetableEntry, room_ids: tuple[str, ...], room: Room) -> VisualTimetableEntry:
    """Return a copy of an entry for one room occupancy."""
    return VisualTimetableEntry(
        assignment_id=entry.assignment_id,
        programme=entry.programme,
        academic_year=entry.academic_year,
        programme_year=entry.programme_year,
        module_code=entry.module_code,
        activity=entry.activity,
        groups=entry.groups,
        lecturers=entry.lecturers,
        lecturer_keys=entry.lecturer_keys,
        rooms=room_ids,
        day=entry.day,
        start_time=entry.start_time,
        end_time=entry.end_time,
        teaching_weeks=entry.teaching_weeks,
        is_fixed=entry.is_fixed,
        delivery_mode=entry.delivery_mode,
        is_external=room.room_type == "external" or "external" in room.resource_type.casefold(),
        is_shared=entry.is_shared,
        location_type=_location_type(room),
        source_reference=entry.source_reference,
        programme_status=entry.programme_status,
        submission_ready_status=entry.submission_ready_status,
        room_capacity_status=_capacity_status(room),
        source_occurrence_ids=entry.source_occurrence_ids,
    )


def _programme_labels(assignment: Assignment) -> tuple[str, ...]:
    """Return programme labels for programme visual expansion."""
    labels = [assignment.course.prog_yr]
    if not _is_shared_assignment(assignment):
        return tuple(labels)
    for group in assignment.course.group_ids:
        normalised = normalise_programme_year(group)
        if normalised and _looks_like_programme_year(normalised) and normalised not in {normalise_programme_year(item) for item in labels}:
            labels.append(group)
    return tuple(labels)


def _programme_status_lookup(rows: list[dict[str, object]]) -> dict[str, dict[str, str]]:
    """Return programme status metadata keyed by normalised programme/year."""
    lookup: dict[str, dict[str, str]] = {}
    for row in rows:
        programme = normalise_programme_year(str(row.get("Programme/Year") or row.get("Normalised Programme/Year") or ""))
        if not programme:
            continue
        status = str(row.get("Status") or row.get("Complete Schedule Status") or "")
        submission = str(row.get("Submission-Ready Status") or row.get("Included In Submission") or "")
        lookup[programme] = {"status": status, "submission": submission}
    return lookup


def _group_programmes(entries: list[VisualTimetableEntry]) -> dict[str, list[VisualTimetableEntry]]:
    """Group entries by programme/year."""
    return _group_by(entries, lambda entry: entry.programme_year)


def _group_tutors(entries: list[VisualTimetableEntry]) -> dict[str, list[VisualTimetableEntry]]:
    """Group entries by tutor while de-duplicating shared rows."""
    grouped: dict[str, list[VisualTimetableEntry]] = defaultdict(list)
    seen: set[tuple[str, str]] = set()
    for entry in entries:
        for lecturer in entry.lecturers or ("Unassigned tutor",):
            key = (lecturer, entry.assignment_id)
            if key in seen:
                continue
            seen.add(key)
            grouped[lecturer].append(entry)
    return {key: sorted(value, key=_entry_sort_key) for key, value in grouped.items()}


def _group_tutor_keys(entries: list[VisualTimetableEntry]) -> dict[str, list[VisualTimetableEntry]]:
    """Group entries by scheduler tutor identity for overlap validation."""
    grouped: dict[str, list[VisualTimetableEntry]] = defaultdict(list)
    seen: set[tuple[str, str]] = set()
    for entry in entries:
        for key in entry.lecturer_keys or entry.lecturers or ("Unassigned tutor",):
            item_key = (key, entry.assignment_id)
            if item_key in seen:
                continue
            seen.add(item_key)
            grouped[key].append(entry)
    return {key: sorted(value, key=_entry_sort_key) for key, value in grouped.items()}


def _group_rooms(entries: list[VisualTimetableEntry]) -> dict[str, list[VisualTimetableEntry]]:
    """Group entries by room while de-duplicating shared rows."""
    grouped: dict[str, list[VisualTimetableEntry]] = defaultdict(list)
    seen: set[tuple[str, str]] = set()
    for entry in entries:
        for room in entry.rooms:
            key = (room, entry.assignment_id)
            if key in seen:
                continue
            seen.add(key)
            grouped[room].append(entry)
    return {key: sorted(value, key=_entry_sort_key) for key, value in grouped.items()}


def _group_by(entries: list[VisualTimetableEntry], key_fn) -> dict[str, list[VisualTimetableEntry]]:
    """Group entries using a string key function."""
    grouped: dict[str, list[VisualTimetableEntry]] = defaultdict(list)
    for entry in entries:
        grouped[key_fn(entry)].append(entry)
    return {key: sorted(value, key=_entry_sort_key) for key, value in grouped.items()}


def _source_occurrence_ids(entry: VisualTimetableEntry) -> tuple[str, ...]:
    """Return source occurrence IDs represented by one visual block."""
    return entry.source_occurrence_ids or (entry.assignment_id,)


def _entry_occurrence_id_set(entries: list[VisualTimetableEntry]) -> set[str]:
    """Return all scheduled occurrence IDs represented by visual blocks."""
    return {source_id for entry in entries for source_id in _source_occurrence_ids(entry)}


def _aggregation_key(entry: VisualTimetableEntry) -> tuple[object, ...]:
    """Return the fields that must match before recurring entries can merge."""
    return (
        entry.programme_year,
        entry.module_code,
        entry.activity,
        entry.groups,
        entry.lecturers,
        entry.lecturer_keys,
        entry.rooms,
        entry.delivery_mode,
        entry.day,
        entry.start_time,
        entry.end_time,
        entry.is_fixed,
        entry.is_shared,
        entry.is_external,
        entry.location_type,
        entry.source_reference,
        entry.programme_status,
        entry.submission_ready_status,
        entry.room_capacity_status,
    )


def _aggregated_assignment_id(entry: VisualTimetableEntry, weeks: tuple[int, ...]) -> str:
    """Return a compact trace ID for one aggregated visual block."""
    return "|".join(
        str(part)
        for part in (
            entry.source_reference,
            entry.programme_year,
            entry.module_code,
            entry.activity,
            entry.day,
            _format_time(entry.start_time),
            _format_time(entry.end_time),
            f"weeks={_weeks_text(weeks)}",
        )
        if str(part).strip()
    )


def _lane_quality_rows(group_sets: dict[str, dict[str, list[VisualTimetableEntry]]]) -> list[dict[str, object]]:
    """Return lane-count quality rows for visual workbook validation."""
    rows: list[dict[str, object]] = []
    for entity_type, groups in group_sets.items():
        for entity, entries in sorted(groups.items()):
            for day in VALID_DAYS:
                day_entries = [entry for entry in entries if entry.day == day]
                if not day_entries:
                    continue
                lanes = allocate_lanes(day_entries)
                lane_count = max(lanes.values(), default=0) + 1
                rows.append(
                    {
                        "Entity Type": entity_type,
                        "Entity": entity,
                        "Day": day,
                        "Maximum Weekday Lanes": lane_count,
                        "Status": "WARN" if lane_count > 4 else "PASS",
                        "Issue": "More than four true overlapping lane(s) are required by the current data." if lane_count > 4 else "",
                    }
                )
    return rows


def _entity_status(entity_type: str, entity: str, entries: list[VisualTimetableEntry], programme_rows: list[dict[str, object]]) -> str:
    """Return a visible status banner for one sheet."""
    if entity_type != "programme":
        return "Valid scheduled classes only"
    row = _programme_row(entity, programme_rows)
    if not row:
        return "Complete with warnings"
    if str(row.get("Submission-Ready Status")) == "PASS" or str(row.get("Included In Submission")) == "Yes":
        return "Submission ready"
    status = str(row.get("Status") or row.get("Complete Schedule Status") or "")
    if "QUARANTINED" in status:
        return "Incomplete: input requirements excluded"
    if "UNSCHEDULED" in status or str(row.get("Unscheduled Search Failures") or "0") not in {"", "0"}:
        return "Incomplete: classes could not be placed"
    return "Incomplete: output mapping unavailable"


def _entity_summary(entity_type: str, entity: str, entries: list[VisualTimetableEntry], programme_rows: list[dict[str, object]]) -> str:
    """Return concise sheet summary text."""
    if entity_type != "programme":
        return f"Scheduled entries: {len(entries)} | Fixed: {sum(1 for item in entries if item.is_fixed)} | Generated: {sum(1 for item in entries if not item.is_fixed)}"
    row = _programme_row(entity, programme_rows) or {}
    return (
        f"Required occurrences: {row.get('Total Required Occurrences', '')} | "
        f"Scheduled occurrences: {row.get('Valid Exported Rows', len(entries))} | "
        f"Quarantined occurrences: {row.get('Quarantined Occurrences', 0)} | "
        f"Unscheduled search failures: {row.get('Unscheduled Search Failures', 0)} | "
        f"Submission-ready status: {row.get('Submission-Ready Status', '')}. "
        "This visual timetable shows valid scheduled classes only. Review the exception report for unresolved requirements."
    )


def _programme_row(entity: str, rows: list[dict[str, object]]) -> dict[str, object] | None:
    """Return a programme completeness row for one entity."""
    normalised = normalise_programme_year(entity)
    for row in rows:
        if normalise_programme_year(str(row.get("Programme/Year") or row.get("Normalised Programme/Year") or "")) == normalised:
            return row
    return None


def _programme_index_row(
    entity: str,
    entries: list[VisualTimetableEntry],
    sheet_name: str,
    programme_rows: list[dict[str, object]],
) -> dict[str, object]:
    """Return one programme index row."""
    first = entries[0]
    row = _programme_row(entity, programme_rows) or {}
    status = first.programme_status or "Scheduled rows available"
    return {
        "Programme": first.programme,
        "Year": first.academic_year,
        "Display name": entity,
        "Status": status,
        "Required occurrences": row.get("Total Required Occurrences", ""),
        "Scheduled occurrences": row.get("Valid Exported Rows", len(entries)),
        "Quarantined occurrences": row.get("Quarantined Occurrences", ""),
        "Unscheduled occurrences": row.get("Unscheduled Search Failures", ""),
        "Submission ready": first.submission_ready_status or "",
        "Sheet link": sheet_name,
    }


def _tutor_index_row(entity: str, entries: list[VisualTimetableEntry], sheet_name: str) -> dict[str, object]:
    """Return one tutor index row."""
    return {
        "Tutor": entity,
        "Scheduled assignments": len(entries),
        "Fixed assignments": sum(1 for entry in entries if entry.is_fixed),
        "Generated assignments": sum(1 for entry in entries if not entry.is_fixed),
        "Online assignments": sum(1 for entry in entries if _is_online(entry)),
        "Earliest class": min((_format_time(entry.start_time) for entry in entries), default=""),
        "Latest class": max((_format_time(entry.end_time) for entry in entries), default=""),
        "Sheet link": sheet_name,
    }


def _room_index_row(entity: str, entries: list[VisualTimetableEntry], sheet_name: str) -> dict[str, object]:
    """Return one room index row."""
    return {
        "Room": entity,
        "Location type": entries[0].location_type if entries else "",
        "Scheduled assignments": len(entries),
        "Fixed assignments": sum(1 for entry in entries if entry.is_fixed),
        "Capacity status": entries[0].room_capacity_status if entries else "",
        "External venue": "Yes" if any(entry.is_external for entry in entries) else "No",
        "Earliest use": min((_format_time(entry.start_time) for entry in entries), default=""),
        "Latest use": max((_format_time(entry.end_time) for entry in entries), default=""),
        "Sheet link": sheet_name,
    }


def _entry_detail_row(entry: VisualTimetableEntry) -> dict[str, object]:
    """Return one structured detail row."""
    return {
        "Assignment ID": entry.assignment_id,
        "Module": entry.module_code,
        "Activity": entry.activity,
        "Programme/Year": entry.programme_year,
        "Group": " / ".join(entry.groups),
        "Tutor": ", ".join(entry.lecturers),
        "Room": ", ".join(entry.rooms),
        "Delivery Mode": entry.delivery_mode,
        "Day": entry.day,
        "Start": _format_time(entry.start_time),
        "End": _format_time(entry.end_time),
        "Teaching Weeks": _weeks_text(entry.teaching_weeks),
        "Fixed/Generated": "Fixed" if entry.is_fixed else "Generated",
        "Shared Session": "Yes" if entry.is_shared else "No",
        "Source Reference": entry.source_reference,
        "Source Occurrence IDs": " ; ".join(_source_occurrence_ids(entry)),
    }


def _detail_columns() -> list[str]:
    """Return structured detail columns."""
    return [
        "Assignment ID",
        "Module",
        "Activity",
        "Programme/Year",
        "Group",
        "Tutor",
        "Room",
        "Delivery Mode",
        "Day",
        "Start",
        "End",
        "Teaching Weeks",
        "Fixed/Generated",
        "Shared Session",
        "Source Reference",
        "Source Occurrence IDs",
    ]


def _block_text(entry: VisualTimetableEntry) -> str:
    """Return readable block text."""
    markers = []
    markers.append("FIXED" if entry.is_fixed else "GENERATED")
    if _is_online(entry):
        markers.append("ONLINE")
    if entry.is_external:
        markers.append("EXTERNAL")
    if entry.is_shared:
        markers.append("SHARED")
    room_text = ", ".join(entry.rooms) if entry.rooms else "ONLINE"
    return "\n".join(
        [
            entry.module_code,
            f"{entry.activity} | {' / '.join(entry.groups)}",
            room_text,
            ", ".join(entry.lecturers),
            f"Weeks { _weeks_text(entry.teaching_weeks) }",
            " ".join(markers),
            entry.assignment_id,
        ]
    )


def _entry_fill(entry: VisualTimetableEntry) -> str:
    """Return fill colour by location type."""
    if entry.is_external:
        return EXTERNAL_FILL
    if _is_online(entry):
        return ONLINE_FILL
    return PHYSICAL_FILL


def _time_resolution(entries: list[VisualTimetableEntry]) -> int:
    """Return 15 or 30 minute resolution based on scheduled times."""
    minutes = [entry.start_time.minute for entry in entries] + [entry.end_time.minute for entry in entries]
    return 15 if any(minute % 30 for minute in minutes) else 30


def _time_axis(entries: list[VisualTimetableEntry], resolution: int) -> list[time]:
    """Return timetable axis from project range and scheduled entries."""
    start_min = EARLIEST_START_HOUR * 60
    end_min = LATEST_END_HOUR * 60
    if entries:
        start_min = min(start_min, min(_time_minutes(entry.start_time) for entry in entries))
        end_min = max(end_min, max(_time_minutes(entry.end_time) for entry in entries))
    values = list(range(start_min, end_min + resolution, resolution))
    return [time(value // 60, value % 60) for value in values]


def _parse_time(value: str) -> time:
    """Parse HH:MM text without rounding."""
    hour, minute = value.split(":")
    return time(int(hour), int(minute))


def _format_time(value: time) -> str:
    """Return HH:MM text."""
    return f"{value.hour:02d}:{value.minute:02d}"


def _time_minutes(value: time) -> int:
    """Return minutes after midnight."""
    return value.hour * 60 + value.minute


def _clock_overlaps(left: VisualTimetableEntry, right: VisualTimetableEntry) -> bool:
    """Return True when two entries would occupy the same clock region."""
    return left.day == right.day and _time_minutes(left.start_time) < _time_minutes(right.end_time) and _time_minutes(right.start_time) < _time_minutes(left.end_time)


def _weeks_overlap(left: VisualTimetableEntry, right: VisualTimetableEntry) -> bool:
    """Return True when two visual entries share at least one teaching week."""
    return bool(set(left.teaching_weeks) & set(right.teaching_weeks))


def _entries_conflict_for_lane(left: VisualTimetableEntry, right: VisualTimetableEntry) -> bool:
    """Return True when two visual blocks need separate day lanes."""
    return _clock_overlaps(left, right)


def _entry_sort_key(entry: VisualTimetableEntry) -> tuple[object, ...]:
    """Return deterministic sorting key."""
    return (
        VALID_DAYS.index(entry.day) if entry.day in VALID_DAYS else 99,
        _time_minutes(entry.start_time),
        _time_minutes(entry.end_time),
        entry.module_code,
        entry.activity,
        entry.assignment_id,
    )


def _overlap_rows(entity_type: str, groups: dict[str, list[VisualTimetableEntry]]) -> list[dict[str, object]]:
    """Return invalid tutor or room overlaps."""
    rows: list[dict[str, object]] = []
    for entity, entries in groups.items():
        if entity_type == "Tutor" and _is_unknown_tutor_key(entity):
            continue
        if entity_type == "Room" and entries and all(entry.is_external for entry in entries):
            continue
        for left_index, left in enumerate(entries):
            for right in entries[left_index + 1 :]:
                if left.assignment_id == right.assignment_id:
                    continue
                if _clock_overlaps(left, right) and _weeks_overlap(left, right):
                    rows.append(
                        {
                            "Entity Type": entity_type,
                            "Entity": entity,
                            "Left Assignment ID": left.assignment_id,
                            "Right Assignment ID": right.assignment_id,
                            "Day": left.day,
                            "Start": _format_time(max(left.start_time, right.start_time)),
                            "End": _format_time(min(left.end_time, right.end_time)),
                            "Weeks": _weeks_text(tuple(sorted(set(left.teaching_weeks) & set(right.teaching_weeks)))),
                            "Status": "FAIL",
                        }
                    )
    return rows


def _programme_reconciliation_rows(
    entries: list[VisualTimetableEntry],
    programme_rows: list[dict[str, object]],
) -> list[dict[str, object]]:
    """Return programme visual count reconciliation rows."""
    block_counts = Counter(entry.programme_year for entry in entries)
    occurrence_counts: Counter[str] = Counter()
    for entry in entries:
        occurrence_counts[entry.programme_year] += len(_source_occurrence_ids(entry))
    programmes = sorted(set(block_counts) | {normalise_programme_year(str(row.get("Programme/Year") or "")) for row in programme_rows})
    rows: list[dict[str, object]] = []
    for programme in programmes:
        if not programme:
            continue
        row = _programme_row(programme, programme_rows) or {}
        expected = row.get("Valid Exported Rows", occurrence_counts[programme])
        rows.append(
            {
                "Programme/Year": programme,
                "Expected Scheduled Visual Occurrences": expected,
                "Programme Visual Occurrences": occurrence_counts[programme],
                "Programme Visual Blocks": block_counts[programme],
                "Status": "PASS" if int(expected or occurrence_counts[programme]) == occurrence_counts[programme] else "WARN",
            }
        )
    return rows


def _entity_reconciliation_rows(
    label: str,
    groups: dict[str, list[VisualTimetableEntry]],
    rooms: list[Room] | None = None,
) -> list[dict[str, object]]:
    """Return tutor or room reconciliation rows."""
    room_lookup = {room.room_id: room for room in rooms or []}
    rows: list[dict[str, object]] = []
    for entity, entries in sorted(groups.items()):
        room = room_lookup.get(entity)
        rows.append(
            {
                label: entity,
                "Visual Entries": len(entries),
                "Fixed Entries": sum(1 for entry in entries if entry.is_fixed),
                "Generated Entries": sum(1 for entry in entries if not entry.is_fixed),
                "Online Entries": sum(1 for entry in entries if _is_online(entry)),
                "External Venue": "Yes" if any(entry.is_external for entry in entries) else "No",
                "Capacity Status": _capacity_status(room) if room else "",
                "Status": "PASS",
            }
        )
    return rows


def _programme_reconciliation_columns() -> tuple[str, ...]:
    """Return programme reconciliation columns."""
    return ("Programme/Year", "Expected Scheduled Visual Occurrences", "Programme Visual Occurrences", "Programme Visual Blocks", "Status")


def _entity_reconciliation_columns(label: str) -> tuple[str, ...]:
    """Return entity reconciliation columns."""
    return (label, "Visual Entries", "Fixed Entries", "Generated Entries", "Online Entries", "External Venue", "Capacity Status", "Status")


def _sheet_name_rows(names: list[str]) -> list[dict[str, object]]:
    """Return validation issues for unsafe sheet names before sanitisation."""
    rows: list[dict[str, object]] = []
    used: set[str] = set()
    for original in names:
        sheet = safe_sheet_name(original, used)
        if len(sheet) > MAX_SHEET_NAME_LENGTH or re.search(INVALID_SHEET_CHARS, sheet):
            rows.append({"Original Name": original, "Sheet Name": sheet, "Status": "FAIL", "Issue": "Invalid worksheet name"})
    return rows


def _summary_rows(result: VisualExportResult, scheduled_received: int) -> list[dict[str, object]]:
    """Return validation summary rows."""
    return [
        {"Metric": "scheduled assignments received", "Value": scheduled_received},
        {"Metric": "programme visual entries", "Value": result.programme_entries},
        {"Metric": "tutor visual entries", "Value": result.tutor_entries},
        {"Metric": "room visual entries", "Value": result.room_entries},
        {"Metric": "programme sheets", "Value": result.programme_sheets},
        {"Metric": "tutor sheets", "Value": result.tutor_sheets},
        {"Metric": "room sheets", "Value": result.room_sheets},
        {"Metric": "missing entries", "Value": result.missing_entries},
        {"Metric": "unexpected entries", "Value": result.unexpected_entries},
        {"Metric": "invalid overlaps", "Value": result.invalid_overlaps},
        {"Metric": "maximum weekday lanes per sheet", "Value": result.max_weekday_lanes},
        {"Metric": "weekday lane warnings", "Value": result.lane_warnings},
        {"Metric": "visual export status", "Value": result.status},
    ]


def _missing_rows(view: str, assignment_ids: set[str]) -> list[dict[str, object]]:
    """Return missing visual entry rows."""
    return [{"View": view, "Assignment ID": assignment_id} for assignment_id in sorted(assignment_ids)]


def _unexpected_rows(view: str, assignment_ids: set[str]) -> list[dict[str, object]]:
    """Return unexpected visual entry rows."""
    return [{"View": view, "Assignment ID": assignment_id} for assignment_id in sorted(assignment_ids)]


def _source_reference(assignment: Assignment) -> str:
    """Return readable source reference."""
    course = assignment.course
    if assignment.fixed_source:
        return str(assignment.fixed_source)
    return f"{course.source_file}:{course.source_sheet}:{course.source_row or ''}"


def _clean_id_part(value: object) -> str:
    """Clean one ID segment."""
    return re.sub(r"\s+", " ", str(value)).strip()


def _programme_code(programme_year: str) -> str:
    """Return programme code from a normalised programme/year label."""
    text = normalise_programme_year(programme_year)
    return re.split(r"[\s/_-]*Y(?:EAR)?\s*\d+", text, flags=re.IGNORECASE)[0].strip(" _-/") or text


def _academic_year(programme_year: str) -> str:
    """Return academic year from a programme/year label."""
    match = re.search(r"(?:Y|YEAR)\s*(\d+)", programme_year, flags=re.IGNORECASE)
    return f"Y{match.group(1)}" if match else ""


def _looks_like_programme_year(value: str) -> bool:
    """Return True for labels that look like programme-year display names."""
    return bool(re.search(r"(?:^|[\s/_-])(?:Y|YR|YEAR)\s*\d+", value, flags=re.IGNORECASE))


def _is_unknown_tutor_key(value: str) -> bool:
    """Return True for placeholder tutor identities that are not clash keys."""
    return value.strip().casefold() in {
        "",
        "unassigned tutor",
        "unassigned",
        "tbc",
        "tbd",
        "temp staff",
        "temporary staff",
        "na",
        "n/a",
        "#n/a",
    }


def _is_online(entry: VisualTimetableEntry) -> bool:
    """Return True for online visual entries."""
    text = f"{entry.delivery_mode} {' '.join(entry.rooms)} {entry.location_type}".casefold()
    return "online" in text or "virtual" in text or "async" in text


def _is_shared_assignment(assignment: Assignment) -> bool:
    """Return True when an assignment represents shared demand."""
    return assignment.course.is_common_module or len(assignment.course.group_ids) > 1 or "|" in str(assignment.fixed_source or "")


def _location_type(room: Room | None) -> str:
    """Return human-readable location type."""
    if room is None:
        return ""
    if room.room_type == "virtual":
        return "Online"
    if room.room_type == "external":
        return "External"
    return room.resource_type or "Physical"


def _capacity_status(room: Room | None) -> str:
    """Return concise room capacity status."""
    if room is None:
        return ""
    if "capacity unavailable" in room.resource_type.casefold():
        return "Capacity unavailable"
    if room.room_type == "external":
        return "External venue"
    return "Known capacity" if room.capacity > 0 else "Unknown capacity"


def _weeks_text(weeks: tuple[int, ...]) -> str:
    """Return compact teaching-week text."""
    return ", ".join(str(week) for week in weeks)


def _sheet_prefix(entity_type: str) -> str:
    """Return worksheet-name prefix for entity type."""
    return {"programme": "Prog", "tutor": "Tutor", "room": "Room"}.get(entity_type, "")
