"""Load structured fixed-session requirements from the Engineering lab workbook."""

from __future__ import annotations

import math
import re
from dataclasses import dataclass, field
from datetime import datetime, time
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook

from data.models import FixedSession

METADATA_SHEETS = {"mod list", "uni wide mod", "sheet1", "sample"}
FIXED_AUDIT_COLUMNS = [
    "source workbook",
    "source sheet",
    "source row",
    "programme/year",
    "module code",
    "group",
    "group size",
    "day",
    "start time",
    "duration",
    "teaching weeks",
    "location",
    "staff",
    "candidate status",
    "loader status",
    "severity",
    "issue",
    "intended treatment",
    "reconciliation status",
]
ISSUE_COLUMNS = ["severity", "workbook", "sheet", "row", "field", "entered value", "problem", "how to correct it"]


@dataclass(slots=True)
class FixedSessionLoaderReport:
    """Diagnostics produced while loading fixed-session rows."""

    workbook_path: str
    authoritative_sheets: list[str] = field(default_factory=list)
    ignored_sheets: list[str] = field(default_factory=list)
    source_rows: int = 0
    duplicate_rows_removed: int = 0
    issues: list[dict[str, object]] = field(default_factory=list)
    audit_rows: list[dict[str, object]] = field(default_factory=list)

    @property
    def fixed_rows_loaded(self) -> int:
        """Return the number of valid non-duplicate fixed sessions loaded."""
        return sum(1 for row in self.audit_rows if row.get("loader status") == "loaded")

    @property
    def critical_errors(self) -> int:
        """Return the number of critical fixed-session loader issues."""
        return sum(1 for issue in self.issues if issue.get("severity") == "critical")

    @property
    def warnings(self) -> int:
        """Return the number of warning fixed-session loader issues."""
        return sum(1 for issue in self.issues if issue.get("severity") == "warning")


def _clean(value: object) -> str:
    """Return a stripped string for messy Excel values."""
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ").strip())


def _header_key(value: object) -> str:
    """Return a punctuation-tolerant header key."""
    return re.sub(r"[^0-9a-z]+", " ", _clean(value).casefold()).strip()


def _headers_from_row(values: Iterable[object]) -> dict[str, int]:
    """Return normalised headers mapped to one-based column indexes."""
    return {key: index for index, value in enumerate(values, start=1) if (key := _header_key(value))}


def _find_header_row(worksheet) -> tuple[int | None, dict[str, int]]:
    """Find the fixed-session header row in one worksheet."""
    required = {"module code", "group", "weeks", "location"}
    placement = {"day", "start time", "duration hr"}
    for row_index in range(1, min(worksheet.max_row, 12) + 1):
        headers = _headers_from_row(cell.value for cell in worksheet[row_index])
        if required <= set(headers) and placement <= set(headers):
            return row_index, headers
    return None, {}


def _get(row: tuple[object, ...], headers: dict[str, int], aliases: Iterable[str]) -> str:
    """Return the first matching cell value by header alias."""
    for alias in aliases:
        index = headers.get(_header_key(alias))
        if index is not None and index - 1 < len(row):
            return _clean(row[index - 1])
    return ""


def _has_payload(row: tuple[object, ...], headers: dict[str, int]) -> bool:
    """Return True when a row has real fixed-session data."""
    aliases = [
        "Prog/Yr",
        "Prog",
        "Yr",
        "Module code",
        "Group",
        "Group Size",
        "Day",
        "Start Time",
        "Duration (Hr)",
        "Weeks",
        "Location",
        "Staff 1",
        "Staff 2",
        "Staff 3",
        "Staff 4",
        "Staff 5",
        "Staff 6",
    ]
    return any(_get(row, headers, [alias]) for alias in aliases)


def normalise_fixed_day(value: object) -> str:
    """Normalise a fixed-session day name to the scheduler's full weekday."""
    text = _clean(value).casefold()
    mapping = {
        "mon": "Monday",
        "monday": "Monday",
        "tue": "Tuesday",
        "tues": "Tuesday",
        "tuesday": "Tuesday",
        "wed": "Wednesday",
        "wednesday": "Wednesday",
        "thu": "Thursday",
        "thur": "Thursday",
        "thurs": "Thursday",
        "thursday": "Thursday",
        "fri": "Friday",
        "friday": "Friday",
    }
    return mapping.get(text, "")


def _time_from_excel_fraction(value: float) -> str:
    """Convert an Excel time fraction to HH:MM."""
    total_minutes = int(round(value * 24 * 60))
    hour, minute = divmod(total_minutes, 60)
    return f"{hour:02d}:{minute:02d}"


def normalise_fixed_start_time(value: object) -> str:
    """Normalise mixed time formats such as 9am, 2.30pm and 14:00."""
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, datetime):
        return value.time().strftime("%H:%M")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if 0 < float(value) < 1:
            return _time_from_excel_fraction(float(value))
        if 0 <= float(value) <= 23:
            return f"{int(value):02d}:00"

    text = _clean(value).casefold().replace(" ", "")
    if not text:
        return ""
    text = text.replace(".", ":")
    meridiem = ""
    if text.endswith("am") or text.endswith("pm"):
        meridiem = text[-2:]
        text = text[:-2]
    if re.fullmatch(r"\d{3,4}", text) and ":" not in text:
        text = text.zfill(4)
        hour = int(text[:-2])
        minute = int(text[-2:])
    else:
        match = re.fullmatch(r"(\d{1,2})(?::(\d{2}))?", text)
        if not match:
            return ""
        hour = int(match.group(1))
        minute = int(match.group(2) or "0")
    if meridiem == "pm" and hour != 12:
        hour += 12
    if meridiem == "am" and hour == 12:
        hour = 0
    if not (0 <= hour <= 23 and 0 <= minute <= 59):
        return ""
    return f"{hour:02d}:{minute:02d}"


def parse_fixed_duration(value: object) -> float | None:
    """Parse a positive fixed-session duration in hours."""
    text = _clean(value)
    if not text:
        return None
    try:
        duration = float(text)
    except ValueError:
        return None
    return duration if duration > 0 else None


def parse_fixed_weeks(value: object) -> tuple[int, ...]:
    """Parse fixed-session teaching weeks from comma-separated values."""
    text = _clean(value)
    if not text:
        return ()
    weeks: set[int] = set()
    for part in re.split(r"[,;/\s]+", text):
        if not part:
            continue
        if "-" in part:
            start, end = part.split("-", 1)
            if start.isdigit() and end.isdigit():
                weeks.update(range(int(start), int(end) + 1))
        elif part.isdigit():
            weeks.add(int(part))
    return tuple(sorted(week for week in weeks if 1 <= week <= 20))


def parse_group_size(value: object) -> int | None:
    """Parse group size, leaving TBC and blanks as None."""
    text = _clean(value)
    if not text or text.casefold() in {"tbc", "all"}:
        return None
    match = re.search(r"\d+", text)
    return int(match.group()) if match else None


def split_locations(value: object) -> tuple[str, ...]:
    """Split multiple fixed locations while preserving readable names."""
    text = _clean(value)
    if not text:
        return ()
    codes = re.findall(r"\b[A-Z]\d-[0-9A-Z]{2}-[0-9A-Z]{2}(?:-[A-Z0-9]+)?\b", text.upper())
    if len(codes) > 1:
        return tuple(codes)
    if "," not in text:
        return (text,)
    parts = [part.strip(" ,") for part in text.split(",")]
    return tuple(part for part in parts if part)


def _collect_staff(row: tuple[object, ...], headers: dict[str, int]) -> tuple[str, ...]:
    """Collect staff names from Staff 1 to Staff 6."""
    names: list[str] = []
    for index in range(1, 7):
        name = _get(row, headers, [f"Staff {index}", f"Staff ID {index}"])
        if name and name not in names:
            names.append(name)
    return tuple(names)


def _resolve_programme_year(sheet_name: str, prog: str, year: str) -> str:
    """Return a stable programme/year label for fixed sessions."""
    text = _clean(prog or year)
    if not text:
        return ""
    if re.fullmatch(r"Y(?:ear|r)?\s*\d+", text, flags=re.IGNORECASE):
        return f"{sheet_name}/{text.upper().replace('YEAR', 'Y').replace('YR', 'Y')}"
    return text


def _issue(
    report: FixedSessionLoaderReport,
    *,
    severity: str,
    sheet: str,
    row: int,
    field: str,
    value: object,
    problem: str,
    recommendation: str,
) -> None:
    """Append one fixed-session loader issue."""
    report.issues.append(
        {
            "severity": severity,
            "workbook": report.workbook_path,
            "sheet": sheet,
            "row": row,
            "field": field,
            "entered value": _clean(value),
            "problem": problem,
            "how to correct it": recommendation,
        }
    )


def _audit_row(
    *,
    workbook: str,
    sheet: str,
    row: int,
    programme_year: str,
    module_code: str,
    group: str,
    group_size: int | None,
    day: str,
    start: str,
    duration: object,
    weeks: tuple[int, ...],
    locations: tuple[str, ...],
    staff: tuple[str, ...],
    candidate_status: str,
    loader_status: str,
    severity: str,
    issue: str,
    intended_treatment: str,
    reconciliation_status: str = "Not reconciled",
) -> dict[str, object]:
    """Build one audit row for the fixed-session workbook."""
    return {
        "source workbook": workbook,
        "source sheet": sheet,
        "source row": row,
        "programme/year": programme_year,
        "module code": module_code,
        "group": group,
        "group size": group_size,
        "day": day,
        "start time": start,
        "duration": duration,
        "teaching weeks": ", ".join(str(week) for week in weeks),
        "location": "; ".join(locations),
        "staff": "; ".join(staff),
        "candidate status": candidate_status,
        "loader status": loader_status,
        "severity": severity,
        "issue": issue,
        "intended treatment": intended_treatment,
        "reconciliation status": reconciliation_status,
    }


def _fingerprint(session: FixedSession) -> tuple[object, ...]:
    """Return a deterministic duplicate key for one fixed session."""
    return (
        session.programme_year.casefold(),
        session.module_code.casefold(),
        session.group_id.casefold(),
        session.day,
        session.start_time,
        session.duration_hours,
        session.teaching_weeks,
        tuple(location.casefold() for location in session.locations),
        tuple(staff.casefold() for staff in session.staff_names),
    )


def _candidate_sheets(workbook) -> tuple[list[str], list[str]]:
    """Return authoritative fixed-session sheets and ignored sheets."""
    sheet_headers: dict[str, int] = {}
    ignored: list[str] = []
    for worksheet in workbook.worksheets:
        if worksheet.title.casefold() in METADATA_SHEETS:
            ignored.append(worksheet.title)
            continue
        header_row, _headers = _find_header_row(worksheet)
        if header_row is None:
            ignored.append(worksheet.title)
        else:
            sheet_headers[worksheet.title] = header_row
    programme_sheets = [name for name in sheet_headers if name != "Module"]
    return (programme_sheets or list(sheet_headers), ignored)


def load_fixed_sessions(workbook_path: Path) -> tuple[list[FixedSession], FixedSessionLoaderReport]:
    """Load valid fixed-session rows from the official fixed-session workbook."""
    workbook_path = Path(workbook_path)
    report = FixedSessionLoaderReport(workbook_path=str(workbook_path))
    if not workbook_path.exists():
        _issue(
            report,
            severity="critical",
            sheet="",
            row=0,
            field="workbook",
            value=workbook_path,
            problem="Fixed-session workbook was not found.",
            recommendation="Restore the official Engineering fixed-session workbook.",
        )
        return [], report

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        authoritative_sheets, ignored_sheets = _candidate_sheets(workbook)
        report.authoritative_sheets = authoritative_sheets
        report.ignored_sheets = ignored_sheets
        sessions: list[FixedSession] = []
        seen: set[tuple[object, ...]] = set()
        for sheet_name in authoritative_sheets:
            worksheet = workbook[sheet_name]
            header_row, headers = _find_header_row(worksheet)
            if header_row is None:
                continue
            previous = {"prog": "", "year": "", "module": ""}
            for row_number, row in enumerate(worksheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                if not _has_payload(row, headers):
                    continue
                report.source_rows += 1
                prog = _get(row, headers, ["Prog/Yr", "Prog", "Programme"])
                year = _get(row, headers, ["Yr", "Year"])
                module_code = _get(row, headers, ["Module code", "Module"])
                if prog:
                    previous["prog"] = prog
                else:
                    prog = previous["prog"]
                if year:
                    previous["year"] = year
                else:
                    year = previous["year"]
                if module_code:
                    previous["module"] = module_code
                else:
                    module_code = previous["module"]

                programme_year = _resolve_programme_year(sheet_name, prog, year)
                group = _get(row, headers, ["Group"])
                group_size = parse_group_size(_get(row, headers, ["Group Size"]))
                day_raw = _get(row, headers, ["Day"])
                start_raw = _get(row, headers, ["Start Time"])
                duration_raw = _get(row, headers, ["Duration (Hr)", "Duration"])
                weeks_raw = _get(row, headers, ["Weeks"])
                locations = split_locations(_get(row, headers, ["Location"]))
                staff_names = _collect_staff(row, headers)
                day = normalise_fixed_day(day_raw)
                start = normalise_fixed_start_time(start_raw)
                duration = parse_fixed_duration(duration_raw)
                weeks = parse_fixed_weeks(weeks_raw)

                if not day_raw or not start_raw:
                    report.audit_rows.append(
                        _audit_row(
                            workbook=workbook_path.name,
                            sheet=sheet_name,
                            row=row_number,
                            programme_year=programme_year,
                            module_code=module_code,
                            group=group,
                            group_size=group_size,
                            day=day_raw,
                            start=start_raw,
                            duration=duration_raw,
                            weeks=weeks,
                            locations=locations,
                            staff=staff_names,
                            candidate_status="candidate",
                            loader_status="not loaded",
                            severity="warning",
                            issue="Row is not a complete fixed placement because day or start time is blank.",
                            intended_treatment="Treat as non-fixed or source-review row until fixed placement is completed.",
                        )
                    )
                    _issue(
                        report,
                        severity="warning",
                        sheet=sheet_name,
                        row=row_number,
                        field="day/start time",
                        value=f"{day_raw} / {start_raw}",
                        problem="Row was not loaded as a fixed session because it lacks a complete fixed placement.",
                        recommendation="Confirm whether this row is non-fixed demand or complete the fixed day and start time.",
                    )
                    continue

                errors: list[str] = []
                if not programme_year:
                    errors.append("programme/year")
                if not module_code:
                    errors.append("module code")
                if not day:
                    errors.append("day")
                if not start:
                    errors.append("start time")
                if duration is None:
                    errors.append("duration")
                if not weeks:
                    errors.append("weeks")
                if not locations:
                    errors.append("location")
                for field_name in errors:
                    _issue(
                        report,
                        severity="critical",
                        sheet=sheet_name,
                        row=row_number,
                        field=field_name,
                        value="",
                        problem=f"Invalid or missing fixed-session {field_name}.",
                        recommendation="Correct the source fixed-session row before generating the timetable.",
                    )
                if errors:
                    report.audit_rows.append(
                        _audit_row(
                            workbook=workbook_path.name,
                            sheet=sheet_name,
                            row=row_number,
                            programme_year=programme_year,
                            module_code=module_code,
                            group=group,
                            group_size=group_size,
                            day=day_raw,
                            start=start_raw,
                            duration=duration_raw,
                            weeks=weeks,
                            locations=locations,
                            staff=staff_names,
                            candidate_status="candidate",
                            loader_status="invalid",
                            severity="critical",
                            issue=f"Missing or invalid fields: {', '.join(errors)}",
                            intended_treatment="Block generation until the source row is corrected or excluded by reconciliation.",
                        )
                    )
                    continue

                session = FixedSession(
                    programme_year=programme_year,
                    module_code=module_code.upper(),
                    group_id=group,
                    group_size=group_size,
                    day=day,
                    start_time=start,
                    duration_hours=duration,
                    teaching_weeks=weeks,
                    locations=locations,
                    staff_ids=staff_names,
                    staff_names=staff_names,
                    source_file=workbook_path.name,
                    source_sheet=sheet_name,
                    source_row=row_number,
                )
                key = _fingerprint(session)
                if key in seen:
                    report.duplicate_rows_removed += 1
                    report.audit_rows.append(
                        _audit_row(
                            workbook=workbook_path.name,
                            sheet=sheet_name,
                            row=row_number,
                            programme_year=programme_year,
                            module_code=module_code,
                            group=group,
                            group_size=group_size,
                            day=day,
                            start=start,
                            duration=duration,
                            weeks=weeks,
                            locations=locations,
                            staff=staff_names,
                            candidate_status="candidate",
                            loader_status="duplicate removed",
                            severity="info",
                            issue="Duplicate fixed-session row was not loaded twice.",
                            intended_treatment="Exclude duplicate from demand.",
                        )
                    )
                    continue
                seen.add(key)
                sessions.append(session)
                report.audit_rows.append(
                    _audit_row(
                        workbook=workbook_path.name,
                        sheet=sheet_name,
                        row=row_number,
                        programme_year=programme_year,
                        module_code=module_code,
                        group=group,
                        group_size=group_size,
                        day=day,
                        start=start,
                        duration=duration,
                        weeks=weeks,
                        locations=locations,
                        staff=staff_names,
                        candidate_status="candidate",
                        loader_status="loaded",
                        severity="info",
                        issue="",
                        intended_treatment="Anchor as fixed session if reconciliation remains unambiguous.",
                    )
                )
        return sessions, report
    finally:
        workbook.close()


def export_fixed_sessions_audit(report: FixedSessionLoaderReport, output_path: Path) -> None:
    """Export fixed-session source audit and loader issues."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    summary = pd.DataFrame(
        [
            {"Metric": "Workbook", "Value": report.workbook_path},
            {"Metric": "Authoritative sheets", "Value": ", ".join(report.authoritative_sheets)},
            {"Metric": "Ignored sheets", "Value": ", ".join(report.ignored_sheets)},
            {"Metric": "Source rows inspected", "Value": report.source_rows},
            {"Metric": "Fixed rows loaded", "Value": report.fixed_rows_loaded},
            {"Metric": "Duplicate rows removed", "Value": report.duplicate_rows_removed},
            {"Metric": "Critical errors", "Value": report.critical_errors},
            {"Metric": "Warnings", "Value": report.warnings},
        ]
    )
    audit = pd.DataFrame(report.audit_rows, columns=FIXED_AUDIT_COLUMNS)
    issues = pd.DataFrame(report.issues, columns=ISSUE_COLUMNS)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        audit.to_excel(writer, sheet_name="Fixed Sessions", index=False)
        issues.to_excel(writer, sheet_name="Issues", index=False)
