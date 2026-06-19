"""Load SIT timetabling input files into dataclasses."""

from __future__ import annotations

import csv
import math
import re
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Iterable, Optional

import pandas as pd
from openpyxl import load_workbook

from config import (
    ACTIVITY_DURATION_HOURS,
    DEFAULT_UNKNOWN_ROOM_CAPACITY,
    TERM_WEEKS,
    VIRTUAL_ROOM_CAPACITY,
    VIRTUAL_ROOM_ID,
)
from data.models import Course, Room
from engine.remarks_interpreter import interpret_remarks


COMMON_COLUMNS = {
    "prog_yr": ["Prog/Yr", "Prog Yr", "Programme/Year", "Programme Year"],
    "class_size": ["Class Size", "Size", "Enrolment"],
    "module_code": ["Module Code", "Module", "Course Code", "Module code"],
    "activity": ["Activity", "Class Type", "Activity Type"],
    "delivery_mode": ["Delivery Mode", "Mode"],
    "teaching_weeks": ["Teaching Weeks", "Tri Week", "Weeks"],
    "staff_1": ["Staff ID 1", "Staff1", "SIS Staff ID"],
    "staff_2": ["Staff ID 2", "Staff2", "SIS Staff ID.1"],
    "staff_3": ["Staff ID 3", "Staff3", "SIS Staff ID.2"],
    "staff_4": ["Staff ID 4", "Staff4", "SIS Staff ID.3"],
    "staff_name_1": ["Staff 1", "Staff1", "Staff"],
    "staff_name_2": ["Staff 2", "Staff2"],
    "staff_name_3": ["Staff 3", "Staff3"],
    "staff_name_4": ["Staff 4", "Staff4"],
    "remarks": ["Remarks", "Remark"],
}

TEMPLATE2_OUTPUT_HEADERS = {
    "module",
    "class type",
    "template",
    "group",
    "day",
    "start",
    "end",
    "room1",
    "staff1",
    "tri week",
    "activity hostkey",
}

TEMPLATE1_LAB_REQUIREMENT_HEADERS = {
    "prog yr",
    "module code",
    "group",
    "group size",
    "weeks",
    "staff 1",
}

TEMPLATE1_VALIDATION_MESSAGE = (
    "This workbook does not match the consolidated schedule format. "
    "Please select the consolidated schedule."
)
TEMPLATE2_INPUT_MESSAGE = "This workbook appears to be a generated timetable.\nPlease select the consolidated schedule."


class WorkbookRole(Enum):
    """Known workbook roles for timetabling inputs and outputs."""

    TEMPLATE1_REQUIREMENTS = "template1_requirements"
    TEMPLATE2_TIMETABLE = "template2_timetable"
    UNKNOWN = "unknown"


class ConsolidatedScheduleValidationError(ValueError):
    """Raised when a selected workbook is not valid consolidated Template 1 input."""


@dataclass(slots=True)
class WorkbookDiagnostic:
    """Diagnostic information for one workbook load attempt."""

    file_path: str
    sheet_name: str
    status: str
    reason: str
    missing_columns: list[str] = field(default_factory=list)
    rows_parsed: int = 0
    rows_skipped: int = 0


@dataclass(slots=True)
class LoaderReport:
    """Collection of workbook diagnostics from a loading run."""

    workbooks: list[WorkbookDiagnostic] = field(default_factory=list)

    def add(self, diagnostic: WorkbookDiagnostic) -> None:
        """Store one workbook diagnostic."""
        self.workbooks.append(diagnostic)

    @property
    def skipped_workbooks(self) -> int:
        """Return the number of fully skipped workbooks."""
        return sum(1 for item in self.workbooks if item.status == "skipped")

    @property
    def partial_workbooks(self) -> int:
        """Return the number of partially parsed workbooks."""
        return sum(1 for item in self.workbooks if item.status == "partial")

    @property
    def parsed_workbooks(self) -> int:
        """Return the number of fully parsed workbooks."""
        return sum(1 for item in self.workbooks if item.status == "parsed")

    def diagnostics_dataframe(self) -> pd.DataFrame:
        """Return workbook diagnostics as a DataFrame."""
        rows = [
            {
                "File Path": item.file_path,
                "Sheet Name": item.sheet_name,
                "Status": item.status,
                "Reason": item.reason,
                "Missing Columns": ", ".join(item.missing_columns),
                "Rows Parsed": item.rows_parsed,
                "Rows Skipped": item.rows_skipped,
            }
            for item in self.workbooks
        ]
        columns = ["File Path", "Sheet Name", "Status", "Reason", "Missing Columns", "Rows Parsed", "Rows Skipped"]
        return pd.DataFrame(rows, columns=columns)

    def summary_dataframe(self) -> pd.DataFrame:
        """Return a compact summary DataFrame."""
        total_rows = sum(item.rows_parsed for item in self.workbooks)
        skipped_rows = sum(item.rows_skipped for item in self.workbooks)
        return pd.DataFrame(
            [
                {"Metric": "Workbooks scanned", "Value": len(self.workbooks)},
                {"Metric": "Workbooks parsed", "Value": self.parsed_workbooks},
                {"Metric": "Workbooks partially parsed", "Value": self.partial_workbooks},
                {"Metric": "Workbooks skipped", "Value": self.skipped_workbooks},
                {"Metric": "Parsed course rows", "Value": total_rows},
                {"Metric": "Skipped course rows", "Value": skipped_rows},
            ]
        )


def _normalise_header(value: object) -> str:
    """Return a case-insensitive, punctuation-tolerant header key."""
    text = str(value or "").replace("\xa0", " ").strip().casefold()
    return re.sub(r"[^0-9a-z]+", " ", text).strip()


def _headers_from_row(values: Iterable[object]) -> set[str]:
    """Return normalised, non-empty headers from a worksheet row."""
    return {header for header in (_normalise_header(value) for value in values) if header}


def _workbook_header_rows(path: Path, max_rows: int = 10) -> list[tuple[str, set[str]]]:
    """Return candidate header rows from every worksheet in a workbook."""
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ConsolidatedScheduleValidationError(f"Unable to open workbook: {exc}") from exc

    rows: list[tuple[str, set[str]]] = []
    try:
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(min_row=1, max_row=max_rows, values_only=True):
                headers = _headers_from_row(row)
                if headers:
                    rows.append((worksheet.title, headers))
    finally:
        workbook.close()
    return rows


def _workbook_appears_template2_output(path: str | Path) -> bool:
    """Return True when workbook headers match proposed Template 2 output."""
    path = Path(path)
    for sheet_name, headers in _workbook_header_rows(path):
        if _normalise_header(sheet_name) == "timetable" and len(headers & TEMPLATE2_OUTPUT_HEADERS) >= 7:
            return True
        if len(headers & TEMPLATE2_OUTPUT_HEADERS) >= 9:
            return True
    return False


def _find_consolidated_schedule_sheet(path: Path) -> str | None:
    """Return the first worksheet containing required Template 1 fields."""
    required_fields = set(_required_column_names())
    for sheet_name, headers in _workbook_header_rows(path):
        present = {
            field_name
            for field_name in required_fields
            if any(_normalise_header(candidate) in headers for candidate in COMMON_COLUMNS[field_name])
        }
        if present == required_fields:
            return sheet_name
    return None


def _find_template1_like_sheet(path: Path) -> str | None:
    """Return a worksheet that has Template 1-style requirement headers."""
    for sheet_name, headers in _workbook_header_rows(path):
        if _normalise_header(sheet_name) != "module":
            continue
        if TEMPLATE1_LAB_REQUIREMENT_HEADERS <= headers:
            return sheet_name
    return None


def detect_workbook_role(path: str | Path) -> WorkbookRole:
    """Detect workbook role from worksheet structure and headers."""
    workbook_path = Path(path)
    if _workbook_appears_template2_output(workbook_path):
        return WorkbookRole.TEMPLATE2_TIMETABLE
    if _find_consolidated_schedule_sheet(workbook_path) is not None or _find_template1_like_sheet(workbook_path) is not None:
        return WorkbookRole.TEMPLATE1_REQUIREMENTS
    return WorkbookRole.UNKNOWN


def workbook_appears_template2_output(path: str | Path) -> bool:
    """Return True when workbook headers match proposed Template 2 output."""
    return detect_workbook_role(path) == WorkbookRole.TEMPLATE2_TIMETABLE


def normalise_delivery_mode(value: object) -> str:
    """Normalise common delivery-mode text while preserving recognisable detail."""
    text = _clean_text(value, default="f2f")
    key = text.casefold().strip()
    if key in {"f2f", "face to face", "face-to-face", "physical", "in person", "in-person", "campus"}:
        return "f2f"
    if "async" in key:
        return "Online - Asynchronous"
    if "online" in key or "virtual" in key or "e learning" in key or "elearning" in key:
        return "Online - Synchronous"
    return text


def validate_consolidated_schedule_structure(path: str | Path) -> None:
    """Validate that a workbook is consolidated Template 1 scheduling input."""
    workbook_path = Path(path)
    if not workbook_path.exists() or not workbook_path.is_file():
        raise ConsolidatedScheduleValidationError("Select a valid Excel workbook")
    if workbook_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ConsolidatedScheduleValidationError("Select a valid Excel workbook")
    role = detect_workbook_role(workbook_path)
    if role == WorkbookRole.TEMPLATE2_TIMETABLE:
        raise ConsolidatedScheduleValidationError(TEMPLATE2_INPUT_MESSAGE)
    if role != WorkbookRole.TEMPLATE1_REQUIREMENTS:
        raise ConsolidatedScheduleValidationError(TEMPLATE1_VALIDATION_MESSAGE)


def _find_header_row(path: Path, sheet_name: str) -> int:
    """Find the row index containing the Module sheet headers."""
    preview = pd.read_excel(path, sheet_name=sheet_name, header=None, nrows=20, engine="openpyxl")
    module_markers = {"module", "module code", "course code"}
    for idx, row in preview.iterrows():
        values = {_normalise_header(value) for value in row.tolist()}
        if values & module_markers and "activity" in values:
            return int(idx)
    return 0


def _get_column(df: pd.DataFrame, candidates: Iterable[str]) -> Optional[str]:
    """Return the first matching column from candidate names."""
    normalised = {_normalise_header(col): col for col in df.columns}
    for candidate in candidates:
        key = _normalise_header(candidate)
        if key in normalised:
            return normalised[key]
    return None


def _clean_text(value: object, default: str = "") -> str:
    """Return a stripped string, treating NaN as empty."""
    if value is None:
        return default
    if isinstance(value, float) and math.isnan(value):
        return default
    text = str(value).strip()
    return text if text and text.lower() != "nan" else default


def _clean_int(value: object, default: int = 0) -> int:
    """Return an integer from messy Excel values."""
    if value is None:
        return default
    if isinstance(value, float) and math.isnan(value):
        return default
    text = str(value).strip()
    if not text:
        return default
    match = re.search(r"\d+", text)
    return int(match.group()) if match else default


def parse_teaching_weeks(value: object) -> list[int]:
    """Parse teaching weeks from values like '1,2,3' or '1-6,8,9'."""
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return TERM_WEEKS.copy()
    if isinstance(value, int):
        return [value]
    if isinstance(value, float) and value.is_integer():
        return [int(value)]

    text = str(value).strip()
    if not text:
        return TERM_WEEKS.copy()

    weeks: set[int] = set()
    for part in re.split(r"[,;/\s]+", text):
        if not part:
            continue
        if "-" in part:
            start_text, end_text = part.split("-", 1)
            if start_text.isdigit() and end_text.isdigit():
                weeks.update(range(int(start_text), int(end_text) + 1))
        elif part.isdigit():
            weeks.add(int(part))

    parsed = sorted(week for week in weeks if 1 <= week <= 20)
    return parsed or TERM_WEEKS.copy()


def infer_week_pattern(weeks: list[int]) -> str:
    """Infer ALL, ODD, EVEN, or CUSTOM from teaching weeks."""
    if not weeks:
        return "ALL"
    if all(week % 2 == 1 for week in weeks):
        return "ODD"
    if all(week % 2 == 0 for week in weeks):
        return "EVEN"
    if set(weeks) == set(TERM_WEEKS):
        return "ALL"
    return "CUSTOM"


def infer_duration_hrs(activity: str) -> int:
    """Infer a default duration from the activity name."""
    activity_key = activity.strip().lower()
    for keyword, duration in ACTIVITY_DURATION_HOURS.items():
        if keyword in activity_key:
            return duration
    return 2


def _collect_staff_ids(row: pd.Series, df: pd.DataFrame) -> list[str]:
    """Collect non-empty staff IDs from a row."""
    ids: list[str] = []
    for key in ["staff_1", "staff_2", "staff_3", "staff_4"]:
        col = _get_column(df, COMMON_COLUMNS[key])
        if not col:
            continue
        staff_id = _clean_text(row.get(col))
        if staff_id and staff_id not in ids:
            ids.append(staff_id)
    return ids


def _collect_staff_names(row: pd.Series, df: pd.DataFrame) -> list[str]:
    """Collect non-empty staff names from a row."""
    names: list[str] = []
    for key in ["staff_name_1", "staff_name_2", "staff_name_3", "staff_name_4"]:
        col = _get_column(df, COMMON_COLUMNS[key])
        if not col:
            continue
        name = _clean_text(row.get(col))
        if name and name not in names:
            names.append(name)
    return names


def load_common_modules(path: str | Path) -> set[str]:
    """Load common module codes from the common modules CSV."""
    path = Path(path)
    if not path.exists():
        return set()

    modules: set[str] = set()
    with path.open(newline="", encoding="utf-8-sig") as file:
        reader = csv.DictReader(file)
        for row in reader:
            module_text = _clean_text(row.get("Module"))
            for module in re.split(r"[/,;&]+", module_text):
                clean = module.strip().upper()
                if clean:
                    modules.add(clean)
    return modules


def _resolve_sheet_name(path: Path, preferred_sheet_name: str) -> tuple[Optional[str], str]:
    """Resolve a workbook sheet name without guessing aggressively."""
    try:
        workbook = load_workbook(path, read_only=True)
    except Exception as exc:  # pragma: no cover - exercised via load failure tests
        return None, f"Unable to open workbook: {exc}"

    try:
        sheet_lookup = {_normalise_header(sheet_name): sheet_name for sheet_name in workbook.sheetnames}
        preferred_key = _normalise_header(preferred_sheet_name)
        if preferred_key in sheet_lookup:
            return sheet_lookup[preferred_key], ""
        if preferred_sheet_name in workbook.sheetnames:
            return preferred_sheet_name, ""
        return None, f"Sheet '{preferred_sheet_name}' not found"
    finally:
        workbook.close()


def _required_column_names() -> dict[str, str]:
    """Return human-friendly labels for the required loader fields."""
    return {
        "prog_yr": " / ".join(COMMON_COLUMNS["prog_yr"]),
        "class_size": " / ".join(COMMON_COLUMNS["class_size"]),
        "module_code": " / ".join(COMMON_COLUMNS["module_code"]),
        "activity": " / ".join(COMMON_COLUMNS["activity"]),
        "delivery_mode": " / ".join(COMMON_COLUMNS["delivery_mode"]),
        "teaching_weeks": " / ".join(COMMON_COLUMNS["teaching_weeks"]),
    }


def _missing_required_columns(df: pd.DataFrame) -> list[str]:
    """List any missing required semantic columns."""
    field_names = ["prog_yr", "class_size", "module_code", "activity", "delivery_mode", "teaching_weeks"]
    missing: list[str] = []
    for field_name in field_names:
        if _get_column(df, COMMON_COLUMNS[field_name]) is None:
            missing.append(_required_column_names()[field_name])
    return missing


def _build_diagnostic(
    *,
    path: Path,
    sheet_name: str,
    status: str,
    reason: str,
    missing_columns: Optional[list[str]] = None,
    rows_parsed: int = 0,
    rows_skipped: int = 0,
) -> WorkbookDiagnostic:
    """Create a workbook diagnostic entry."""
    return WorkbookDiagnostic(
        file_path=str(path),
        sheet_name=sheet_name,
        status=status,
        reason=reason,
        missing_columns=missing_columns or [],
        rows_parsed=rows_parsed,
        rows_skipped=rows_skipped,
    )


def load_courses_from_requirements(
    path: str | Path,
    common_modules: Optional[set[str]] = None,
    sheet_name: str = "Module",
) -> tuple[list[Course], WorkbookDiagnostic]:
    """Load course activities from a requirements workbook."""
    path = Path(path)
    common_modules = common_modules or set()

    resolved_sheet, sheet_reason = _resolve_sheet_name(path, sheet_name)
    if resolved_sheet is None:
        diagnostic = _build_diagnostic(
            path=path,
            sheet_name=sheet_name,
            status="skipped",
            reason=sheet_reason,
        )
        return [], diagnostic

    try:
        header_row = _find_header_row(path, resolved_sheet)
        df = pd.read_excel(path, sheet_name=resolved_sheet, header=header_row, engine="openpyxl")
        df = df.loc[:, ~df.columns.astype(str).str.startswith("Unnamed")]
        df.columns = [str(col).strip() for col in df.columns]
    except Exception as exc:
        diagnostic = _build_diagnostic(
            path=path,
            sheet_name=resolved_sheet,
            status="skipped",
            reason=f"Unable to read sheet: {exc}",
        )
        return [], diagnostic

    missing_columns = _missing_required_columns(df)
    if missing_columns:
        diagnostic = _build_diagnostic(
            path=path,
            sheet_name=resolved_sheet,
            status="skipped",
            reason="Missing required columns",
            missing_columns=missing_columns,
        )
        return [], diagnostic

    prog_col = _get_column(df, COMMON_COLUMNS["prog_yr"])
    size_col = _get_column(df, COMMON_COLUMNS["class_size"])
    module_col = _get_column(df, COMMON_COLUMNS["module_code"])
    activity_col = _get_column(df, COMMON_COLUMNS["activity"])
    mode_col = _get_column(df, COMMON_COLUMNS["delivery_mode"])
    weeks_col = _get_column(df, COMMON_COLUMNS["teaching_weeks"])
    remarks_col = _get_column(df, COMMON_COLUMNS["remarks"])

    # The required-column check above guarantees these are present.
    assert prog_col is not None
    assert size_col is not None
    assert module_col is not None
    assert activity_col is not None
    assert mode_col is not None
    assert weeks_col is not None

    for col in [prog_col, size_col, module_col]:
        df[col] = df[col].ffill()

    courses: list[Course] = []
    skipped_rows = 0
    for row_index, row in df.iterrows():
        module_code = _clean_text(row.get(module_col)).upper()
        activity = _clean_text(row.get(activity_col))
        prog_yr = _clean_text(row.get(prog_col))
        delivery_mode = normalise_delivery_mode(row.get(mode_col))

        if not module_code or not activity or not prog_yr:
            skipped_rows += 1
            continue
        if module_code.lower() in {"module code", "nan"}:
            skipped_rows += 1
            continue

        weeks = parse_teaching_weeks(row.get(weeks_col))
        class_size = _clean_int(row.get(size_col), default=1)
        staff_ids = _collect_staff_ids(row, df)
        staff_names = _collect_staff_names(row, df)
        remarks = _clean_text(row.get(remarks_col)) if remarks_col else ""
        source_row = header_row + int(row_index) + 2

        courses.append(
            Course(
                module_code=module_code,
                activity=activity,
                prog_yr=prog_yr,
                class_size=class_size,
                delivery_mode=delivery_mode,
                teaching_weeks=weeks,
                week_pattern=infer_week_pattern(weeks),
                staff_ids=staff_ids,
                duration_hrs=infer_duration_hrs(activity),
                is_common_module=module_code in common_modules,
                staff_names=staff_names,
                remarks=remarks,
                source_file=path.name,
                group_ids=[prog_yr],
                source_sheet=resolved_sheet,
                source_row=source_row,
                remark_requirements=interpret_remarks(remarks),
            )
        )

    if not courses:
        diagnostic = _build_diagnostic(
            path=path,
            sheet_name=resolved_sheet,
            status="skipped",
            reason="No valid course rows were found",
            rows_parsed=0,
            rows_skipped=skipped_rows,
        )
        return [], diagnostic

    status = "parsed" if skipped_rows == 0 else "partial"
    reason = "Workbook parsed successfully" if skipped_rows == 0 else f"Parsed with {skipped_rows} skipped row(s)"
    diagnostic = _build_diagnostic(
        path=path,
        sheet_name=resolved_sheet,
        status=status,
        reason=reason,
        rows_parsed=len(courses),
        rows_skipped=skipped_rows,
    )
    return courses, diagnostic


def load_consolidated_schedule_with_report(
    workbook_path: str | Path,
    common_modules: Optional[set[str]] = None,
) -> tuple[list[Course], WorkbookDiagnostic]:
    """Load one consolidated Template 1 workbook into course requirements."""
    path = Path(workbook_path)
    validate_consolidated_schedule_structure(path)
    sheet_name = _find_consolidated_schedule_sheet(path)
    if sheet_name is None:
        raise ConsolidatedScheduleValidationError(TEMPLATE1_VALIDATION_MESSAGE)

    courses, diagnostic = load_courses_from_requirements(path, common_modules=common_modules, sheet_name=sheet_name)
    if diagnostic.reason == "Missing required columns":
        raise ConsolidatedScheduleValidationError(TEMPLATE1_VALIDATION_MESSAGE)
    if not courses:
        raise ConsolidatedScheduleValidationError("The selected workbook contains no scheduling records")
    return courses, diagnostic


def load_consolidated_schedule(workbook_path: str | Path, common_modules: Optional[set[str]] = None) -> list[Course]:
    """Load one consolidated Template 1 workbook into existing Course objects."""
    courses, _diagnostic = load_consolidated_schedule_with_report(workbook_path, common_modules=common_modules)
    return courses


def load_courses_from_folder(folder: str | Path, common_modules: Optional[set[str]] = None) -> tuple[list[Course], LoaderReport]:
    """Load courses from all requirement workbooks in a folder."""
    courses: list[Course] = []
    report = LoaderReport()
    for path in sorted(Path(folder).glob("*.xlsx")):
        if path.name.startswith("~$"):
            report.add(
                _build_diagnostic(
                    path=path,
                    sheet_name="Unknown",
                    status="skipped",
                    reason="Temporary Excel lock file was ignored",
                )
            )
            continue
        workbook_courses, diagnostic = load_courses_from_requirements(path, common_modules=common_modules)
        courses.extend(workbook_courses)
        report.add(diagnostic)
    return courses, report


def _parse_capacity(value: object) -> int:
    """Parse capacity, using a safe high default for blank teaching rooms."""
    parsed = _clean_int(value, default=0)
    return parsed if parsed > 0 else DEFAULT_UNKNOWN_ROOM_CAPACITY


def load_rooms_from_csv(path: str | Path) -> list[Room]:
    """Load physical rooms from the venue CSV and add one virtual room."""
    path = Path(path)
    rooms: list[Room] = [
        Room(
            room_id=VIRTUAL_ROOM_ID,
            capacity=VIRTUAL_ROOM_CAPACITY,
            room_type="virtual",
            resource_type="Virtual Room",
            recording="Yes",
        )
    ]

    with path.open(newline="", encoding="cp1252") as file:
        reader = csv.DictReader(file)
        for row in reader:
            room_id = _clean_text(row.get("Location Name"))
            if not room_id:
                continue
            rooms.append(
                Room(
                    room_id=room_id,
                    capacity=_parse_capacity(row.get("Capacity")),
                    room_type="physical",
                    resource_type=_clean_text(row.get("Resource Type")),
                    recording=_clean_text(row.get("Recording?")),
                )
            )
    return rooms


def export_loader_report(report: LoaderReport, output_path: str | Path) -> None:
    """Export loader diagnostics to an Excel report."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    summary_df = report.summary_dataframe()
    diagnostics_df = report.diagnostics_dataframe()

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        diagnostics_df.to_excel(writer, sheet_name="Diagnostics", index=False)
