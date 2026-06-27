"""Controlled supervisor resolution workbook support."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ALLOWED_DECISIONS = {
    "CONFIRM_SHARED_SESSION",
    "CONFIRM_SEPARATE_SESSIONS",
    "CONFIRM_ROOM_ALIAS",
    "CONFIRM_EXTERNAL_VENUE",
    "PROVIDE_TEACHING_WEEKS",
    "CONFIRM_PROGRAMME_GROUP",
    "REMOVE_DUPLICATE_SOURCE_ROW",
    "SOURCE_DATA_REQUIRES_CORRECTION",
}
DECISIONS_REQUIRING_VALUE = {
    "CONFIRM_ROOM_ALIAS",
    "CONFIRM_EXTERNAL_VENUE",
    "PROVIDE_TEACHING_WEEKS",
    "CONFIRM_PROGRAMME_GROUP",
}
RESOLUTION_COLUMNS = ["Query ID", "Decision", "Approved Value", "Reason", "Approved By", "Approval Date", "Notes"]


@dataclass(slots=True)
class ResolutionDecision:
    """One approved supervisor decision."""

    query_id: str
    decision: str
    approved_value: str
    reason: str
    approved_by: str
    approval_date: str
    notes: str = ""


@dataclass(slots=True)
class ResolutionLoadResult:
    """Validated resolution workbook result."""

    decisions: list[ResolutionDecision] = field(default_factory=list)
    errors: list[dict[str, object]] = field(default_factory=list)

    @property
    def valid(self) -> bool:
        """Return True when all populated decisions are valid."""
        return not self.errors


def export_resolution_template(output_path: Path, query_ids: list[str] | None = None) -> None:
    """Create a blank supervisor resolution template workbook."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    rows = [{"Query ID": query_id, "Decision": "", "Approved Value": "", "Reason": "", "Approved By": "", "Approval Date": "", "Notes": ""} for query_id in (query_ids or [])]
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame(rows, columns=RESOLUTION_COLUMNS).to_excel(writer, sheet_name="Resolution Decisions", index=False)
        pd.DataFrame({"Supported Decisions": sorted(ALLOWED_DECISIONS)}).to_excel(writer, sheet_name="Decision Types", index=False)


def load_resolution_workbook(path: Path, valid_query_ids: set[str]) -> ResolutionLoadResult:
    """Load and validate a filled supervisor resolution workbook."""
    path = Path(path)
    result = ResolutionLoadResult()
    if not path.exists():
        return result
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if "Resolution Decisions" not in workbook.sheetnames:
            result.errors.append({"Row": "", "Query ID": "", "Error": "Missing Resolution Decisions sheet"})
            return result
        worksheet = workbook["Resolution Decisions"]
        headers = [str(cell.value or "") for cell in worksheet[1]]
        seen: set[str] = set()
        for row_index, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            row = {headers[index]: str(value or "").strip() for index, value in enumerate(values) if index < len(headers)}
            if not any(row.get(column, "") for column in RESOLUTION_COLUMNS):
                continue
            decision = row.get("Decision", "").upper()
            query_id = row.get("Query ID", "")
            if query_id and not decision and not any(row.get(column, "") for column in RESOLUTION_COLUMNS if column != "Query ID"):
                continue
            errors = _validate_resolution_row(row, decision, query_id, valid_query_ids, seen)
            if errors:
                result.errors.extend({"Row": row_index, "Query ID": query_id, "Error": error} for error in errors)
                continue
            seen.add(query_id)
            result.decisions.append(
                ResolutionDecision(
                    query_id=query_id,
                    decision=decision,
                    approved_value=row.get("Approved Value", ""),
                    reason=row.get("Reason", ""),
                    approved_by=row.get("Approved By", ""),
                    approval_date=row.get("Approval Date", ""),
                    notes=row.get("Notes", ""),
                )
            )
    finally:
        workbook.close()
    return result


def _validate_resolution_row(
    row: dict[str, str],
    decision: str,
    query_id: str,
    valid_query_ids: set[str],
    seen: set[str],
) -> list[str]:
    """Return validation errors for one resolution row."""
    errors: list[str] = []
    if query_id not in valid_query_ids:
        errors.append("Unknown query ID")
    if query_id in seen:
        errors.append("Duplicate query ID decision")
    if decision not in ALLOWED_DECISIONS:
        errors.append("Unsupported decision")
    if decision in DECISIONS_REQUIRING_VALUE and not row.get("Approved Value", ""):
        errors.append("Approved Value is required for this decision")
    if decision == "PROVIDE_TEACHING_WEEKS" and row.get("Approved Value", "") and not _valid_week_expression(row["Approved Value"]):
        errors.append("Approved teaching weeks are invalid")
    if not row.get("Reason", ""):
        errors.append("Reason is required")
    if not row.get("Approved By", ""):
        errors.append("Approved By is required")
    return errors


def _valid_week_expression(value: str) -> bool:
    """Return True when a teaching-week expression is parseable and in range."""
    text = str(value or "").replace(" ", "")
    if not text:
        return False
    weeks: set[int] = set()
    for part in text.split(","):
        if not part:
            return False
        if "-" in part:
            start, end = part.split("-", 1)
            if not start.isdigit() or not end.isdigit():
                return False
            weeks.update(range(int(start), int(end) + 1))
        elif part.isdigit():
            weeks.add(int(part))
        else:
            return False
    return bool(weeks) and all(1 <= week <= 20 for week in weeks)


def export_resolution_audit(result: ResolutionLoadResult, output_path: Path) -> None:
    """Export a traceable resolution audit workbook."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    decisions = [decision.__dict__ for decision in result.decisions]
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame(
            [
                {"Metric": "Valid", "Value": "Yes" if result.valid else "No"},
                {"Metric": "Decisions accepted", "Value": len(result.decisions)},
                {"Metric": "Validation errors", "Value": len(result.errors)},
            ]
        ).to_excel(writer, sheet_name="Summary", index=False)
        pd.DataFrame(decisions).to_excel(writer, sheet_name="Accepted Decisions", index=False)
        pd.DataFrame(result.errors).to_excel(writer, sheet_name="Validation Errors", index=False)
