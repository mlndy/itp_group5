"""Authoritative fixed-session location evidence and classification."""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from config import DEFAULT_TEMPLATE2_FILE, DEFAULT_UNKNOWN_ROOM_CAPACITY
from data.models import Room


@dataclass(frozen=True, slots=True)
class LocationEvidence:
    """Evidence-backed location classification."""

    original_value: str
    normalised_value: str
    candidate_venue_code: str
    authoritative_evidence_source: str
    capacity: int | None
    room_type: str
    host_key: str
    confidence: str
    treatment: str
    blocking_status: str


def normalise_location_code(value: str) -> str:
    """Return a punctuation-normalised institutional location code."""
    text = str(value or "").upper()
    code = _extract_code(text)
    if code:
        return code
    return re.sub(r"[^A-Z0-9]+", "-", text).strip("-")


def _extract_code(value: str) -> str:
    """Extract venue codes including hyphen or space separated variants."""
    text = str(value or "").upper()
    match = re.search(r"\b([A-Z]\d)[-\s]?([0-9A-Z]{2})[-\s]?([0-9A-Z]{2})(?:[-\s]?([A-Z0-9]+))?\b", text)
    if not match:
        return ""
    parts = [part for part in match.groups() if part]
    return "-".join(parts)


def _room_indexes(rooms: list[Room]) -> tuple[dict[str, Room], dict[str, Room]]:
    """Return exact and normalised room lookup indexes."""
    exact = {room.room_id.casefold(): room for room in rooms}
    normalised = {normalise_location_code(room.room_id).casefold(): room for room in rooms}
    return exact, normalised


def load_template2_location_rows(template2_path: Path = DEFAULT_TEMPLATE2_FILE) -> list[dict[str, object]]:
    """Load Template 2 Location support rows."""
    template2_path = Path(template2_path)
    if not template2_path.exists():
        return []
    workbook = load_workbook(template2_path, read_only=True, data_only=True)
    try:
        if "Location" not in workbook.sheetnames:
            return []
        worksheet = workbook["Location"]
        headers = [str(cell.value or "") for cell in worksheet[1]]
        rows: list[dict[str, object]] = []
        for row_index, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value not in (None, "") for value in values):
                continue
            item = {headers[index]: value for index, value in enumerate(values) if index < len(headers)}
            item["_source row"] = row_index
            rows.append(item)
        return rows
    finally:
        workbook.close()


def _template2_index(template2_rows: list[dict[str, object]]) -> dict[str, dict[str, object]]:
    """Return Template 2 locations indexed by name, host key and normalised code."""
    index: dict[str, dict[str, object]] = {}
    for row in template2_rows:
        for field in ["Name", "Host Key"]:
            value = str(row.get(field) or "").strip()
            if not value:
                continue
            index[value.casefold()] = row
            index[normalise_location_code(value).casefold()] = row
    return index


def classify_location(
    value: str,
    rooms: list[Room],
    template2_path: Path = DEFAULT_TEMPLATE2_FILE,
) -> LocationEvidence:
    """Classify one fixed-session location using authoritative project files."""
    text = str(value or "").strip()
    normalised = normalise_location_code(text)
    exact_rooms, normalised_rooms = _room_indexes(rooms)
    template_rows = load_template2_location_rows(template2_path)
    template_index = _template2_index(template_rows)

    room = exact_rooms.get(text.casefold()) or normalised_rooms.get(normalised.casefold())
    if room is None and normalised:
        for candidate in rooms:
            if candidate.room_id.casefold().startswith(f"{normalised.casefold()}-"):
                room = candidate
                break
    if room is not None:
        return LocationEvidence(
            original_value=text,
            normalised_value=normalised,
            candidate_venue_code=room.room_id,
            authoritative_evidence_source="Venue Information(Campus Court).csv",
            capacity=room.capacity,
            room_type=room.room_type,
            host_key=room.room_id,
            confidence="exact" if room.room_id.casefold() == text.casefold() else "alias",
            treatment="exact internal venue fully validated" if room.room_id.casefold() == text.casefold() else "official venue alias safely resolved",
            blocking_status="non-blocking",
        )

    template_row = template_index.get(text.casefold()) or template_index.get(normalised.casefold())
    if template_row is None and "external" in text.casefold():
        for row in template_rows:
            name = str(row.get("Name") or "")
            host_key = str(row.get("Host Key") or "")
            if (name and name.casefold() in text.casefold()) or (host_key and host_key.casefold() in text.casefold()):
                template_row = row
                break
    if template_row is not None:
        host_key = str(template_row.get("Host Key") or template_row.get("Name") or text)
        capacity = _parse_capacity(template_row.get("Capacity"))
        treatment = "recognised external venue" if "external" in host_key.casefold() else "recognised institutional venue missing capacity data"
        return LocationEvidence(
            original_value=text,
            normalised_value=normalised,
            candidate_venue_code=host_key,
            authoritative_evidence_source=f"{Path(template2_path).name}:Location row {template_row.get('_source row')}",
            capacity=capacity,
            room_type="external" if "external" in host_key.casefold() else "physical",
            host_key=host_key,
            confidence="template2-supported",
            treatment=treatment,
            blocking_status="non-blocking" if treatment == "recognised external venue" or capacity is not None else "warning",
        )

    return LocationEvidence(
        original_value=text,
        normalised_value=normalised,
        candidate_venue_code=normalised,
        authoritative_evidence_source="No authoritative match found",
        capacity=None,
        room_type="unknown",
        host_key="",
        confidence="none",
        treatment="unknown venue",
        blocking_status="blocking",
    )


def _parse_capacity(value: object) -> int | None:
    """Parse capacity from a support sheet value."""
    try:
        if value in (None, ""):
            return None
        return int(float(str(value)))
    except ValueError:
        return None


def room_from_location_evidence(evidence: LocationEvidence) -> Room | None:
    """Create a placeholder room for non-blocking recognised non-CSV venues."""
    if evidence.blocking_status == "blocking":
        return None
    if evidence.room_type == "external":
        return Room(evidence.host_key or evidence.original_value, evidence.capacity or DEFAULT_UNKNOWN_ROOM_CAPACITY, "external", "External Venue")
    if evidence.treatment == "recognised institutional venue missing capacity data":
        return Room(
            evidence.host_key or evidence.candidate_venue_code,
            evidence.capacity or DEFAULT_UNKNOWN_ROOM_CAPACITY,
            "physical",
            "Recognised Venue (capacity unavailable)",
        )
    return None


def export_location_mapping_evidence(
    mapping_issues: list[dict[str, object]],
    location_rows: list[dict[str, object]],
    output_path: Path,
) -> None:
    """Export the location mapping evidence workbook."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    rows = location_rows
    unresolved = [row for row in rows if row.get("blocking status") == "blocking"]
    exact = [row for row in rows if row.get("treatment") == "exact internal venue fully validated"]
    alias = [row for row in rows if row.get("treatment") == "official venue alias safely resolved"]
    external = [row for row in rows if row.get("treatment") == "recognised external venue"]
    missing = [row for row in rows if row.get("treatment") == "recognised institutional venue missing capacity data"]
    manual = [row for row in rows if row.get("blocking status") == "blocking"]
    summary = [
        {"Metric": "Original location mapping issues", "Value": len(mapping_issues)},
        {"Metric": "Unique locations reviewed", "Value": len({row.get("original value") for row in rows})},
        {"Metric": "Exact matches", "Value": len(exact)},
        {"Metric": "Alias matches", "Value": len(alias)},
        {"Metric": "Recognised external venues", "Value": len(external)},
        {"Metric": "Missing venue records", "Value": len(missing)},
        {"Metric": "Manual clarification required", "Value": len(manual)},
    ]
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame(summary).to_excel(writer, sheet_name="Summary", index=False)
        pd.DataFrame(unresolved).to_excel(writer, sheet_name="Unresolved Locations", index=False)
        pd.DataFrame(exact).to_excel(writer, sheet_name="Exact Matches", index=False)
        pd.DataFrame(alias).to_excel(writer, sheet_name="Alias Matches", index=False)
        pd.DataFrame(external).to_excel(writer, sheet_name="External Venues", index=False)
        pd.DataFrame(missing).to_excel(writer, sheet_name="Missing Venue Records", index=False)
        pd.DataFrame(manual).to_excel(writer, sheet_name="Manual Clarification Required", index=False)
