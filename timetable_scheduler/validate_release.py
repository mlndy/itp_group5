"""Validate existing v1.1 release Excel artefacts without regenerating them."""

from __future__ import annotations

import argparse
import ast
import hashlib
import sys
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from config import (
    BASE_DIR,
    DEFAULT_FIXED_SESSION_INTEGRITY_FILE,
    DEFAULT_GUARDED_GENERATION_REPORT_FILE,
    DEFAULT_PROGRAMME_VISUALS_FILE,
    DEFAULT_ROOM_VISUALS_FILE,
    DEFAULT_RUN_MANIFEST_FILE,
    DEFAULT_RUN_SUMMARY_FILE,
    DEFAULT_STAKEHOLDER_VIEWS_FILE,
    DEFAULT_TEMPLATE2_PROGRAMME_YEAR_RECONCILIATION_FILE,
    DEFAULT_TEMPLATE2_SUBMISSION_FILE,
    DEFAULT_TEMPLATE2_SUBMISSION_VALIDATION_FILE,
    DEFAULT_TIMETABLE_VISUALISATION_VALIDATION_FILE,
    DEFAULT_TUTOR_VISUALS_FILE,
    OUTPUT_DIR,
)
from output.submission_validator import saved_row_programme_year

CURRENT_TIMETABLE_FILENAME = "Proposed_Timetable.xlsx"
LEGACY_TIMETABLE_FILENAME = "final_timetable_engineering_cluster.xlsx"
DEFAULT_TIMETABLE_FILE = OUTPUT_DIR / LEGACY_TIMETABLE_FILENAME
RUNS_DIR = OUTPUT_DIR / "runs"

EXPECTED_MIN_TESTS = 314
EXPECTED_TOTAL_TEACHING_OCCURRENCES = 3562
EXPECTED_SCHEDULABLE_OCCURRENCES = 3323
EXPECTED_QUARANTINED_OCCURRENCES = 239
EXPECTED_SCHEDULED_OCCURRENCES = 3214
EXPECTED_SEARCH_FAILURE_OCCURRENCES = 109
EXPECTED_SCHEDULED_HARD_VIOLATIONS = 0
EXPECTED_TEMPLATE2_SUBMISSION_ROWS = 212
EXPECTED_ALL_VALID_TEMPLATE2_ROWS = 2980
MIN_SUBMISSION_READY_PROGRAMME_YEARS = 20
EXPECTED_FIXED_SOURCE_ROWS = 250
EXPECTED_FIXED_SOURCE_OCCURRENCES = 834
EXPECTED_ANCHORED_FIXED_SOURCE_OCCURRENCES = 595
EXPECTED_QUARANTINED_FIXED_SOURCE_OCCURRENCES = 239
OFFICIAL_TEMPLATE2_SHEET_COUNT = 16
OFFICIAL_TIMETABLE_COLUMN_COUNT = 31
OFFICIAL_TEMPLATE2_SHEETS = [
    "Timetable",
    "Course Code",
    "Location",
    "Staff",
    "Group",
    "LocationGroup",
    "Zone",
    "Sheet4",
    "Sheet1",
    "Time",
    "Day",
    "Sheet2",
    "Sheet3",
    "Class Type",
    "Template",
    "StaffGroup",
]
OFFICIAL_TIMETABLE_COLUMNS = [
    "Module",
    "Class Type",
    "Template",
    "Group",
    "Day",
    "Start",
    "End",
    "Class Size",
    "Sector",
    "RoomGrouping",
    "Room1",
    "Room2",
    "StaffGrouping",
    "Staff1",
    "Staff2",
    "Tri Week",
    "Recording Mode",
    "Remark",
    "FMTS Tri Start Week",
    "Activity Hostkey",
    "SIS Module Code",
    "Term",
    "Activity Type",
    "Duration",
    "Staff Suitability ID",
    "SIS Staff ID",
    "SIS Staff ID",
    "Zone Hoskey",
    "Location Suitability ID",
    "Location Hostkey",
    "Location Hostkey",
]

REQUIRED_RUN_SUMMARY_SHEETS = [
    "Summary",
    "Validation Checks",
    "Run Metadata",
    "Programme Breakdown",
    "Resource Audit",
    "Virtual Room Detail",
    "Unscheduled Analysis",
    "Unscheduled Breakdown",
    "Residual F2F Analysis",
    "Optimisation Summary",
]
REQUIRED_GUARDED_SHEETS = [
    "Summary",
    "Quarantined Requirements",
    "Unscheduled Search Failures",
    "Programme Completeness",
    "Submission Exclusions",
]
REQUIRED_TEMPLATE2_VALIDATION_SHEETS = [
    "Summary",
    "Programme Schedule Coverage",
    "Required Field Validation",
    "Source-to-Output Reconciliation",
    "Invalid Rows",
    "Missing Programme-Years",
    "Identity Normalisation Audit",
    "Incomplete Programme-Years",
    "Saved Workbook Verification",
    "Submission Readiness",
]
REQUIRED_VISUAL_VALIDATION_SHEETS = [
    "Summary",
    "Programme Reconciliation",
    "Tutor Reconciliation",
    "Room Reconciliation",
    "Missing Visual Entries",
    "Unexpected Visual Entries",
    "Overlap Validation",
    "Export Status",
]
REQUIRED_FIXED_INTEGRITY_SHEETS = [
    "Summary",
    "Fixed Source Integrity",
    "Integrity Issues",
    "Quarantined Fixed Sources",
]
REQUIRED_TIMETABLE_SHEETS = ["Timetable"]
REQUIRED_TEMPLATE2_RECONCILIATION_SHEETS = [
    "Summary",
    "Programme-Year Reconciliation",
    "Missing Programme-Years",
    "Row-Level Reconciliation",
    "Identity Normalisation Audit",
    "Incomplete Programme-Years",
    "Saved Workbook Verification",
]
REQUIRED_TIMETABLE_COLUMNS = [
    "Module",
    "Class Type",
    "Group",
    "Day",
    "Start",
    "End",
    "Class Size",
    "Room1",
    "Staff1",
    "Staff2",
    "Tri Week",
    "Activity Type",
    "Duration",
    "Location Hostkey",
    "Remark",
]
RUN_DIR_FILE_MAP = {
    "run_summary": "run_summary.xlsx",
    "stakeholder_views": "stakeholder_views.xlsx",
    "run_manifest": "run_manifest.xlsx",
    "guarded_report": "guarded_generation_report.xlsx",
    "template2_validation": "template2_submission_validation.xlsx",
    "template2_submission": "Template2_Submission_Ready.xlsx",
    "template2_reconciliation": "template2_programme_year_reconciliation.xlsx",
    "visual_validation": "timetable_visualisation_validation.xlsx",
    "fixed_integrity": "fixed_session_integrity_validation.xlsx",
    "programme_visuals": "Programme_Timetable_Visuals.xlsx",
    "tutor_visuals": "Tutor_Timetable_Visuals.xlsx",
    "room_visuals": "Room_Timetable_Visuals.xlsx",
}
TIMETABLE_CANDIDATE_FILENAMES = (CURRENT_TIMETABLE_FILENAME, LEGACY_TIMETABLE_FILENAME)
REQUIRED_STAKEHOLDER_SHEETS = ["Programme Timetable", "Tutor Timetable", "Room Timetable", "Exception Queue"]
REQUIRED_MANIFEST_SHEETS = [
    "Run Manifest",
    "Soft Constraint Weights",
    "Soft Rule Baseline",
    "Template Validation",
    "Traceability",
]


@dataclass(slots=True)
class ReleaseValidationResult:
    """Release validation outcome and failure details."""

    passed: bool
    failures: list[str] = field(default_factory=list)


@dataclass(frozen=True, slots=True)
class ReleaseEvidencePaths:
    """Locations of generated evidence files used by release validation."""

    run_summary: Path = DEFAULT_RUN_SUMMARY_FILE
    timetable: Path = DEFAULT_TIMETABLE_FILE
    stakeholder_views: Path = DEFAULT_STAKEHOLDER_VIEWS_FILE
    run_manifest: Path = DEFAULT_RUN_MANIFEST_FILE
    guarded_report: Path = DEFAULT_GUARDED_GENERATION_REPORT_FILE
    template2_validation: Path = DEFAULT_TEMPLATE2_SUBMISSION_VALIDATION_FILE
    template2_submission: Path = DEFAULT_TEMPLATE2_SUBMISSION_FILE
    template2_reconciliation: Path = DEFAULT_TEMPLATE2_PROGRAMME_YEAR_RECONCILIATION_FILE
    visual_validation: Path = DEFAULT_TIMETABLE_VISUALISATION_VALIDATION_FILE
    fixed_integrity: Path = DEFAULT_FIXED_SESSION_INTEGRITY_FILE
    programme_visuals: Path = DEFAULT_PROGRAMME_VISUALS_FILE
    tutor_visuals: Path = DEFAULT_TUTOR_VISUALS_FILE
    room_visuals: Path = DEFAULT_ROOM_VISUALS_FILE
    test_root: Path = BASE_DIR / "tests"
    evidence_location: Path | None = None
    run_id: str = "current-output-paths"


@dataclass(frozen=True, slots=True)
class EvidenceResolution:
    """Resolved release evidence location and file paths."""

    paths: ReleaseEvidencePaths
    message: str


def _file_digest(path: Path) -> str:
    """Return a stable digest for workbook ambiguity checks."""
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _missing_run_files(run_dir: Path) -> list[str]:
    """Return required run-folder evidence files missing from a run directory."""
    missing = [filename for filename in RUN_DIR_FILE_MAP.values() if not (run_dir / filename).exists()]
    if not any((run_dir / filename).exists() for filename in TIMETABLE_CANDIDATE_FILENAMES):
        missing.append(f"{CURRENT_TIMETABLE_FILENAME} or {LEGACY_TIMETABLE_FILENAME}")
    return missing


def resolve_run_timetable_path(run_dir: Path) -> Path:
    """Resolve current or legacy timetable workbook path for one run folder."""
    current = run_dir / CURRENT_TIMETABLE_FILENAME
    legacy = run_dir / LEGACY_TIMETABLE_FILENAME
    current_exists = current.exists()
    legacy_exists = legacy.exists()
    if current_exists and legacy_exists:
        if _file_digest(current) != _file_digest(legacy):
            raise ValueError(
                "Run directory contains ambiguous mixed timetable evidence: "
                f"{CURRENT_TIMETABLE_FILENAME} and {LEGACY_TIMETABLE_FILENAME} both exist but differ"
            )
        return current
    if current_exists:
        return current
    if legacy_exists:
        return legacy
    raise FileNotFoundError(
        "Run directory is missing the proposed timetable workbook: expected "
        f"{CURRENT_TIMETABLE_FILENAME} or {LEGACY_TIMETABLE_FILENAME}"
    )


def _paths_from_run_dir(run_dir: Path, test_root: Path = BASE_DIR / "tests") -> ReleaseEvidencePaths:
    """Return evidence paths rooted in one isolated run directory."""
    run_dir = run_dir.resolve()
    timetable_path = resolve_run_timetable_path(run_dir)
    return ReleaseEvidencePaths(
        run_summary=run_dir / RUN_DIR_FILE_MAP["run_summary"],
        timetable=timetable_path,
        stakeholder_views=run_dir / RUN_DIR_FILE_MAP["stakeholder_views"],
        run_manifest=run_dir / RUN_DIR_FILE_MAP["run_manifest"],
        guarded_report=run_dir / RUN_DIR_FILE_MAP["guarded_report"],
        template2_validation=run_dir / RUN_DIR_FILE_MAP["template2_validation"],
        template2_submission=run_dir / RUN_DIR_FILE_MAP["template2_submission"],
        template2_reconciliation=run_dir / RUN_DIR_FILE_MAP["template2_reconciliation"],
        visual_validation=run_dir / RUN_DIR_FILE_MAP["visual_validation"],
        fixed_integrity=run_dir / RUN_DIR_FILE_MAP["fixed_integrity"],
        programme_visuals=run_dir / RUN_DIR_FILE_MAP["programme_visuals"],
        tutor_visuals=run_dir / RUN_DIR_FILE_MAP["tutor_visuals"],
        room_visuals=run_dir / RUN_DIR_FILE_MAP["room_visuals"],
        test_root=test_root,
        evidence_location=run_dir,
        run_id=run_dir.name,
    )


def _run_dir_is_complete(run_dir: Path) -> bool:
    """Return True when the folder contains the required release evidence files."""
    return run_dir.is_dir() and not _missing_run_files(run_dir)


def latest_completed_run_dir(runs_dir: Path | None = None) -> Path | None:
    """Return the newest complete isolated run folder, if one exists."""
    runs_dir = runs_dir or RUNS_DIR
    if not runs_dir.exists():
        return None
    candidates = sorted(
        (path for path in runs_dir.iterdir() if path.is_dir()),
        key=lambda path: (path.stat().st_mtime, path.name),
        reverse=True,
    )
    for path in candidates:
        if _missing_run_files(path):
            continue
        resolve_run_timetable_path(path)
        return path
    return None


def resolve_evidence_paths(
    *,
    run_dir: Path | None = None,
    test_root: Path = BASE_DIR / "tests",
    explicit_paths: ReleaseEvidencePaths | None = None,
) -> EvidenceResolution:
    """Resolve one non-mixed evidence source for release validation."""
    if run_dir is not None and explicit_paths is not None:
        raise ValueError("--run-dir cannot be combined with individual evidence paths")
    if run_dir is not None:
        resolved = run_dir.resolve()
        missing = _missing_run_files(resolved)
        if missing:
            raise FileNotFoundError(f"Run directory is incomplete: {resolved}; missing: {', '.join(missing)}")
        paths = _paths_from_run_dir(resolved, test_root)
        return EvidenceResolution(
            paths,
            f"Resolved run ID: {resolved.name}\n"
            f"Resolved evidence location: {resolved}\n"
            f"Resolved timetable filename: {paths.timetable.name}",
        )
    if explicit_paths is not None:
        return EvidenceResolution(
            explicit_paths,
            f"Resolved run ID: explicit-paths\n"
            f"Resolved evidence location: individual CLI paths\n"
            f"Resolved timetable filename: {explicit_paths.timetable.name}",
        )
    latest = latest_completed_run_dir()
    if latest is not None:
        paths = _paths_from_run_dir(latest, test_root)
        return EvidenceResolution(
            paths,
            f"Resolved run ID: {latest.name}\n"
            f"Resolved evidence location: {latest.resolve()}\n"
            f"Resolved timetable filename: {paths.timetable.name}",
        )
    raise FileNotFoundError(f"No complete release evidence run folder found in {RUNS_DIR}")


def _load_workbook(path: Path, failures: list[str]) -> Workbook | None:
    """Open one workbook and record a failure if it cannot be read."""
    if not path.exists():
        failures.append(f"Missing workbook: {path}")
        return None
    try:
        return load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - exact openpyxl errors vary
        failures.append(f"Could not open workbook {path}: {exc}")
        return None


def _sheet_dict(workbook: Workbook, sheet_name: str) -> dict[object, object]:
    """Return first-column/second-column rows as a dictionary."""
    worksheet = workbook[sheet_name]
    return {
        row[0]: row[1]
        for row in worksheet.iter_rows(min_row=2, values_only=True)
        if row and row[0] is not None
    }


def _rows_as_dicts(workbook: Workbook, sheet_name: str) -> list[dict[str, object]]:
    """Return worksheet rows as dictionaries keyed by header."""
    worksheet = workbook[sheet_name]
    try:
        headers = [cell.value for cell in worksheet[1]]
    except IndexError:
        return []
    if not any(header is not None for header in headers):
        return []
    return [
        {str(header): row[index] for index, header in enumerate(headers) if header is not None}
        for row in worksheet.iter_rows(min_row=2, values_only=True)
        if row and any(value not in (None, "") for value in row)
    ]


def _validation_statuses(workbook: Workbook) -> dict[object, object]:
    """Return Validation Checks statuses keyed by check name."""
    worksheet = workbook["Validation Checks"]
    return {
        row[0]: row[2]
        for row in worksheet.iter_rows(min_row=2, values_only=True)
        if row and row[0] is not None
    }


def _check_required_sheets(workbook: Workbook, required: list[str], workbook_name: str, failures: list[str]) -> None:
    """Check required sheets exist."""
    for sheet_name in required:
        if sheet_name not in workbook.sheetnames:
            failures.append(f"{workbook_name} missing required sheet: {sheet_name}")


def _check_test_suite_size(test_root: Path, failures: list[str]) -> None:
    """Check the normal test suite still contains the v1.1 minimum test count."""
    if not test_root.exists():
        failures.append(f"Missing test folder: {test_root}")
        return
    test_count = 0
    for path in sorted(test_root.glob("test_*.py")):
        try:
            tree = ast.parse(path.read_text(encoding="utf-8"))
        except SyntaxError as exc:
            failures.append(f"Could not parse test file {path}: {exc}")
            continue
        test_count += sum(
            isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)) and node.name.startswith("test_")
            for node in tree.body
        )
    if test_count < EXPECTED_MIN_TESTS:
        failures.append(f"Normal test suite contains {test_count} tests, expected at least {EXPECTED_MIN_TESTS}")


def _check_validation_statuses(workbook: Workbook, failures: list[str]) -> None:
    """Check key Validation Checks rows are PASS."""
    statuses = _validation_statuses(workbook)
    required_passes = {
        "Demand occurrence consistency": "demand consistency",
        "Hard-constraint safety status": "hard-constraint safety",
        "DSC inclusion status": "DSC inclusion",
    }
    for check, label in required_passes.items():
        if statuses.get(check) != "PASS":
            failures.append(f"Validation Checks does not show PASS for {label}")


def _check_summary_metrics(workbook: Workbook, failures: list[str]) -> None:
    """Check final all-demand and hard-safety metrics."""
    summary = _sheet_dict(workbook, "Summary")
    required = summary.get("Required teaching occurrences")
    scheduled = summary.get("Scheduled teaching occurrences")
    unscheduled = summary.get("Unscheduled teaching occurrences")
    scheduled_hard = summary.get("Hard violations on scheduled assignments")

    if required != EXPECTED_TOTAL_TEACHING_OCCURRENCES:
        failures.append(
            f"Required teaching occurrences is {required}, expected {EXPECTED_TOTAL_TEACHING_OCCURRENCES}"
        )
    if not isinstance(required, int) or not isinstance(scheduled, int) or not isinstance(unscheduled, int):
        failures.append("Teaching occurrence metrics must be numeric")
        return
    if required != scheduled + unscheduled:
        failures.append(f"Demand inconsistency: {required} != {scheduled} + {unscheduled}")
    if scheduled != EXPECTED_SCHEDULED_OCCURRENCES:
        failures.append(
            "Official Engineering release scheduled teaching occurrences is "
            f"{scheduled}, expected {EXPECTED_SCHEDULED_OCCURRENCES}"
        )
    if scheduled_hard != EXPECTED_SCHEDULED_HARD_VIOLATIONS:
        failures.append(f"Scheduled hard violations is {scheduled_hard}, expected 0")


def _check_programme_breakdown(workbook: Workbook, failures: list[str]) -> None:
    """Check Programme Breakdown contains DSC evidence."""
    worksheet = workbook["Programme Breakdown"]
    headers = [cell.value for cell in worksheet[1]]
    try:
        dsc_index = headers.index("DSC Indicator")
    except ValueError:
        failures.append("Programme Breakdown missing DSC Indicator column")
        return
    has_dsc = any(row[dsc_index] == "Yes" for row in worksheet.iter_rows(min_row=2, values_only=True))
    if not has_dsc:
        failures.append("Programme Breakdown does not contain DSC rows")


def _check_guarded_report(workbook: Workbook, failures: list[str]) -> None:
    """Check guarded-generation summary metrics for v1.1 evidence."""
    summary = _sheet_dict(workbook, "Summary")
    total = summary.get("Total teaching occurrences")
    schedulable = summary.get("Schedulable occurrences")
    quarantined = summary.get("Quarantined occurrences")
    scheduled = summary.get("Scheduled occurrences")
    search_failures = summary.get("Unscheduled search failures")
    scheduled_hard = summary.get("Scheduled hard violations")
    submission_ready = summary.get("Submission-ready programme-years")

    expected_values = {
        "Total teaching occurrences": (total, EXPECTED_TOTAL_TEACHING_OCCURRENCES),
        "Official Engineering release schedulable occurrences": (schedulable, EXPECTED_SCHEDULABLE_OCCURRENCES),
        "Official Engineering release quarantined occurrences": (quarantined, EXPECTED_QUARANTINED_OCCURRENCES),
        "Official Engineering release scheduled occurrences": (scheduled, EXPECTED_SCHEDULED_OCCURRENCES),
        "Official Engineering release search failures": (search_failures, EXPECTED_SEARCH_FAILURE_OCCURRENCES),
        "Scheduled hard violations": (scheduled_hard, EXPECTED_SCHEDULED_HARD_VIOLATIONS),
    }
    for label, (actual, expected) in expected_values.items():
        if actual != expected:
            failures.append(f"{label} is {actual}, expected {expected}")
    if isinstance(total, int) and isinstance(schedulable, int) and isinstance(quarantined, int):
        if total != schedulable + quarantined:
            failures.append(f"Recorded demand split is inconsistent: {total} != {schedulable} + {quarantined}")
    if isinstance(schedulable, int) and isinstance(scheduled, int) and isinstance(search_failures, int):
        if schedulable != scheduled + search_failures:
            failures.append(f"Schedulable split is inconsistent: {schedulable} != {scheduled} + {search_failures}")
    if not isinstance(submission_ready, int) or submission_ready < MIN_SUBMISSION_READY_PROGRAMME_YEARS:
        failures.append(
            f"Submission-ready programme-years is {submission_ready}, expected at least {MIN_SUBMISSION_READY_PROGRAMME_YEARS}"
        )


def _check_template2_validation(workbook: Workbook, failures: list[str]) -> None:
    """Check submission-ready Template 2 validation metrics."""
    summary = _sheet_dict(workbook, "Summary")
    missing = summary.get("rows with missing required fields")
    mapping_errors = summary.get("rows with mapping errors")
    saved_rows = summary.get("Actual saved Template 2 rows", summary.get("Template 2 output rows"))
    all_valid_rows = summary.get("All-valid Template 2 rows")
    submission_ready = summary.get("submission-ready programme-year schedules")
    qualifying = summary.get("qualifying submission-ready programme-years")
    minimum_status = summary.get("minimum programme-year status")
    status = summary.get("Template 2 readiness status")
    expected_values = {
        "rows with missing required fields": (missing, 0),
        "rows with mapping errors": (mapping_errors, 0),
        "Actual saved Template 2 rows": (saved_rows, EXPECTED_TEMPLATE2_SUBMISSION_ROWS),
        "All-valid Template 2 rows": (all_valid_rows, EXPECTED_ALL_VALID_TEMPLATE2_ROWS),
    }
    for label, (actual, expected) in expected_values.items():
        if actual != expected:
            failures.append(f"{label} is {actual}, expected {expected}")
    if submission_ready != qualifying:
        failures.append(
            f"submission-ready programme-year schedules ({submission_ready}) does not match qualifying count ({qualifying})"
        )
    if not isinstance(submission_ready, int) or submission_ready < MIN_SUBMISSION_READY_PROGRAMME_YEARS:
        failures.append(
            f"Fewer than {MIN_SUBMISSION_READY_PROGRAMME_YEARS} qualifying programme-years: {submission_ready}"
        )
    if minimum_status != "PASS":
        failures.append(f"minimum programme-year status is {minimum_status}, expected PASS")
    if status != "PASS":
        failures.append(f"Template 2 readiness status is {status}, expected PASS")
    readiness = workbook["Submission Readiness"]
    statuses = [row[1] for row in readiness.iter_rows(min_row=2, values_only=True) if row and row[0]]
    if not statuses or any(status != "PASS" for status in statuses):
        failures.append("Submission Readiness sheet does not show PASS")
    coverage = _rows_as_dicts(workbook, "Programme Schedule Coverage")
    for row in coverage:
        if row.get("Submission-Ready Status") == "PASS" and row.get("Complete Schedule Status") != "PASS":
            failures.append(f"{row.get('Canonical programme-year')} is submission-ready but incomplete")
        if row.get("Counts Toward Minimum 20") == "Yes" and not row.get("Canonical programme-year"):
            failures.append("Ambiguous programme-year identity counted toward minimum")
    invalid_rows = _rows_as_dicts(workbook, "Invalid Rows")
    if invalid_rows:
        failures.append(f"Template 2 validation reports {len(invalid_rows)} invalid saved row(s), expected 0")


def _saved_template2_row_metrics(workbook: Workbook) -> tuple[int, set[str]]:
    """Return populated Timetable rows and distinct canonical programme-years."""
    worksheet = workbook["Timetable"]
    headers = [cell.value for cell in worksheet[1]]
    row_count = 0
    programme_years: set[str] = set()
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in values):
            continue
        row_count += 1
        row = {str(header): values[index] for index, header in enumerate(headers) if header is not None}
        programme_year = saved_row_programme_year(row)
        if programme_year:
            programme_years.add(programme_year)
    return row_count, programme_years


def _qualifying_programme_years(workbook: Workbook) -> set[str]:
    """Return canonical programme-years counted by reconciliation evidence."""
    if "Programme-Year Reconciliation" not in workbook.sheetnames:
        return set()
    rows = _rows_as_dicts(workbook, "Programme-Year Reconciliation")
    return {
        str(row.get("Canonical programme-year"))
        for row in rows
        if row.get("Counts Toward Minimum 20") == "Yes" and row.get("Canonical programme-year")
    }


def _check_saved_template2_submission(
    workbook: Workbook,
    validation_workbook: Workbook | None,
    reconciliation_workbook: Workbook | None,
    failures: list[str],
) -> None:
    """Check the saved strict Template 2 workbook against validation evidence."""
    if workbook.sheetnames != OFFICIAL_TEMPLATE2_SHEETS:
        failures.append("Template2_Submission_Ready.xlsx does not preserve the 16 official worksheets in order")
    if len(workbook.sheetnames) != OFFICIAL_TEMPLATE2_SHEET_COUNT:
        failures.append(
            f"Template2_Submission_Ready.xlsx has {len(workbook.sheetnames)} sheets, expected {OFFICIAL_TEMPLATE2_SHEET_COUNT}"
        )
    _check_required_sheets(workbook, ["Timetable"], "Template2_Submission_Ready.xlsx", failures)
    if "Timetable" not in workbook.sheetnames:
        return
    headers = [cell.value for cell in workbook["Timetable"][1]]
    if headers != OFFICIAL_TIMETABLE_COLUMNS:
        failures.append("Template2_Submission_Ready.xlsx Timetable columns do not match the official 31-column structure")
    if len(headers) != OFFICIAL_TIMETABLE_COLUMN_COUNT:
        failures.append(f"Template2_Submission_Ready.xlsx Timetable has {len(headers)} columns, expected 31")
    row_count, programme_years = _saved_template2_row_metrics(workbook)
    if row_count != EXPECTED_TEMPLATE2_SUBMISSION_ROWS:
        failures.append(f"Saved Template 2 row count is {row_count}, expected {EXPECTED_TEMPLATE2_SUBMISSION_ROWS}")
    if validation_workbook is None or "Summary" not in validation_workbook.sheetnames:
        return
    summary = _sheet_dict(validation_workbook, "Summary")
    expected_rows = summary.get("Actual saved Template 2 rows", summary.get("Template 2 output rows"))
    expected_programmes = summary.get(
        "programme-years represented in submission workbook",
        summary.get("actual saved programme-year schedules"),
    )
    if expected_rows != row_count:
        failures.append(f"Saved Template 2 row count is {row_count}, validation report says {expected_rows}")
    if expected_programmes != len(programme_years):
        failures.append(
            "Saved Template 2 represented programme-years is "
            f"{len(programme_years)}, validation report says {expected_programmes}"
        )
    if reconciliation_workbook is not None:
        qualifying = _qualifying_programme_years(reconciliation_workbook)
        if programme_years != qualifying:
            failures.append("Saved Template 2 programme-years do not match reconciliation qualifying programme-years")
        if len(qualifying) < MIN_SUBMISSION_READY_PROGRAMME_YEARS:
            failures.append(
                f"Reconciliation qualifying programme-years is {len(qualifying)}, "
                f"expected at least {MIN_SUBMISSION_READY_PROGRAMME_YEARS}"
            )


def _check_template2_reconciliation(workbook: Workbook, validation_workbook: Workbook | None, failures: list[str]) -> None:
    """Check programme-year reconciliation workbook exists and agrees with readiness summary."""
    _check_required_sheets(
        workbook,
        REQUIRED_TEMPLATE2_RECONCILIATION_SHEETS,
        "template2_programme_year_reconciliation.xlsx",
        failures,
    )
    if validation_workbook is None or "Summary" not in workbook.sheetnames or "Summary" not in validation_workbook.sheetnames:
        return
    reconciliation_summary = _sheet_dict(workbook, "Summary")
    validation_summary = _sheet_dict(validation_workbook, "Summary")
    if reconciliation_summary.get("Template 2 readiness status") != "PASS":
        failures.append("Template 2 reconciliation readiness status is not PASS")
    if reconciliation_summary.get("minimum programme-year status") != "PASS":
        failures.append("Template 2 reconciliation minimum programme-year status is not PASS")
    qualifying = reconciliation_summary.get("qualifying submission-ready programme-years")
    if not isinstance(qualifying, int) or qualifying < MIN_SUBMISSION_READY_PROGRAMME_YEARS:
        failures.append(
            f"Template 2 reconciliation qualifying programme-years is {qualifying}, "
            f"expected at least {MIN_SUBMISSION_READY_PROGRAMME_YEARS}"
        )
    if reconciliation_summary.get("actual saved programme-year schedules") != qualifying:
        failures.append("Template 2 reconciliation saved programme-year count does not match qualifying count")
    if reconciliation_summary.get("Template 2 output rows") != EXPECTED_TEMPLATE2_SUBMISSION_ROWS:
        failures.append(
            f"Template 2 reconciliation output rows is {reconciliation_summary.get('Template 2 output rows')}, "
            f"expected {EXPECTED_TEMPLATE2_SUBMISSION_ROWS}"
        )
    if reconciliation_summary.get("All-valid Template 2 rows") != EXPECTED_ALL_VALID_TEMPLATE2_ROWS:
        failures.append(
            f"Template 2 reconciliation all-valid rows is {reconciliation_summary.get('All-valid Template 2 rows')}, "
            f"expected {EXPECTED_ALL_VALID_TEMPLATE2_ROWS}"
        )
    for metric in [
        "submission-ready programme-year schedules",
        "qualifying submission-ready programme-years",
        "minimum programme-year status",
        "Template 2 readiness status",
    ]:
        if reconciliation_summary.get(metric) != validation_summary.get(metric):
            failures.append(
                f"Template 2 reconciliation {metric} is {reconciliation_summary.get(metric)}, "
                f"validation report says {validation_summary.get(metric)}"
            )


def _check_visual_validation(workbook: Workbook, failures: list[str]) -> None:
    """Check visual timetable reconciliation metrics."""
    summary = _sheet_dict(workbook, "Summary")
    expected_values = {
        "scheduled assignments received": (summary.get("scheduled assignments received"), EXPECTED_SCHEDULED_OCCURRENCES),
        "missing entries": (summary.get("missing entries"), 0),
        "unexpected entries": (summary.get("unexpected entries"), 0),
        "invalid overlaps": (summary.get("invalid overlaps"), 0),
        "visual export status": (summary.get("visual export status"), "PASS"),
    }
    for label, (actual, expected) in expected_values.items():
        if actual != expected:
            failures.append(f"{label} is {actual}, expected {expected}")


def _check_fixed_integrity(workbook: Workbook, failures: list[str]) -> None:
    """Check fixed-session integrity evidence."""
    summary = _sheet_dict(workbook, "Summary")
    expected_values = {
        "fixed source rows": (summary.get("fixed source rows"), EXPECTED_FIXED_SOURCE_ROWS),
        "expected fixed teaching occurrences": (summary.get("expected fixed teaching occurrences"), EXPECTED_FIXED_SOURCE_OCCURRENCES),
        "anchored fixed teaching occurrences": (summary.get("anchored fixed teaching occurrences"), EXPECTED_ANCHORED_FIXED_SOURCE_OCCURRENCES),
        "quarantined fixed teaching occurrences": (summary.get("quarantined fixed teaching occurrences"), EXPECTED_QUARANTINED_FIXED_SOURCE_OCCURRENCES),
        "missing fixed teaching occurrences": (summary.get("missing fixed teaching occurrences"), 0),
        "placement mismatches": (summary.get("placement mismatches"), 0),
        "scheduled hard violations on fixed assignments": (summary.get("scheduled hard violations on fixed assignments"), 0),
        "fixed-session integrity status": (summary.get("fixed-session integrity status"), "PASS"),
    }
    for label, (actual, expected) in expected_values.items():
        if actual != expected:
            failures.append(f"{label} is {actual}, expected {expected}")


def _check_timetable_workbook(workbook: Workbook, failures: list[str]) -> None:
    """Check proposed timetable workbook structure and row count."""
    _check_required_sheets(workbook, REQUIRED_TIMETABLE_SHEETS, "Timetable workbook", failures)
    if "Timetable" not in workbook.sheetnames:
        return
    sheet = workbook["Timetable"]
    headers = [cell.value for cell in sheet[1]]
    for column in REQUIRED_TIMETABLE_COLUMNS:
        if column not in headers:
            failures.append(f"Timetable sheet missing required column: {column}")


def _check_stakeholder_views(workbook: Workbook, failures: list[str]) -> None:
    """Check stakeholder workbook contains the required operational views."""
    _check_required_sheets(workbook, REQUIRED_STAKEHOLDER_SHEETS, "stakeholder_views.xlsx", failures)
    if "Exception Queue" in workbook.sheetnames:
        headers = [cell.value for cell in workbook["Exception Queue"][1]]
        for column in ["Original Reason", "Classification", "Recommended Operational Action", "Review Status"]:
            if column not in headers:
                failures.append(f"Exception Queue missing required column: {column}")


def _check_run_manifest(workbook: Workbook, failures: list[str]) -> None:
    """Check run manifest contains acceptance validation evidence."""
    _check_required_sheets(workbook, REQUIRED_MANIFEST_SHEETS, "run_manifest.xlsx", failures)
    if "Run Manifest" in workbook.sheetnames:
        manifest = _sheet_dict(workbook, "Run Manifest")
        if manifest.get("validation_status") != "PASS":
            failures.append("Run Manifest validation_status is not PASS")
        if manifest.get("required_teaching_occurrences") != EXPECTED_TOTAL_TEACHING_OCCURRENCES:
            failures.append("Run Manifest required teaching occurrences do not match v1.1 evidence")
    if "Template Validation" in workbook.sheetnames:
        worksheet = workbook["Template Validation"]
        statuses = [row[1] for row in worksheet.iter_rows(min_row=2, values_only=True) if row and row[0]]
        if any(status != "PASS" for status in statuses):
            failures.append("Template Validation contains non-PASS status")


def validate_release(
    run_summary_path: Path = DEFAULT_RUN_SUMMARY_FILE,
    timetable_path: Path = DEFAULT_TIMETABLE_FILE,
    stakeholder_views_path: Path = DEFAULT_STAKEHOLDER_VIEWS_FILE,
    run_manifest_path: Path = DEFAULT_RUN_MANIFEST_FILE,
    guarded_report_path: Path = DEFAULT_GUARDED_GENERATION_REPORT_FILE,
    template2_validation_path: Path = DEFAULT_TEMPLATE2_SUBMISSION_VALIDATION_FILE,
    template2_submission_path: Path = DEFAULT_TEMPLATE2_SUBMISSION_FILE,
    template2_reconciliation_path: Path = DEFAULT_TEMPLATE2_PROGRAMME_YEAR_RECONCILIATION_FILE,
    visual_validation_path: Path = DEFAULT_TIMETABLE_VISUALISATION_VALIDATION_FILE,
    fixed_integrity_path: Path = DEFAULT_FIXED_SESSION_INTEGRITY_FILE,
    programme_visuals_path: Path = DEFAULT_PROGRAMME_VISUALS_FILE,
    tutor_visuals_path: Path = DEFAULT_TUTOR_VISUALS_FILE,
    room_visuals_path: Path = DEFAULT_ROOM_VISUALS_FILE,
    test_root: Path = BASE_DIR / "tests",
) -> ReleaseValidationResult:
    """Validate existing release artefacts and return the result."""
    failures: list[str] = []
    _check_test_suite_size(test_root, failures)
    run_summary = _load_workbook(run_summary_path, failures)
    timetable = _load_workbook(timetable_path, failures)
    stakeholder_views = _load_workbook(stakeholder_views_path, failures)
    run_manifest = _load_workbook(run_manifest_path, failures)
    guarded_report = _load_workbook(guarded_report_path, failures)
    template2_validation = _load_workbook(template2_validation_path, failures)
    template2_submission = _load_workbook(template2_submission_path, failures)
    template2_reconciliation = _load_workbook(template2_reconciliation_path, failures)
    visual_validation = _load_workbook(visual_validation_path, failures)
    fixed_integrity = _load_workbook(fixed_integrity_path, failures)
    _load_workbook(programme_visuals_path, failures)
    _load_workbook(tutor_visuals_path, failures)
    _load_workbook(room_visuals_path, failures)

    if run_summary is not None:
        _check_required_sheets(run_summary, REQUIRED_RUN_SUMMARY_SHEETS, "run_summary.xlsx", failures)
        if all(sheet in run_summary.sheetnames for sheet in REQUIRED_RUN_SUMMARY_SHEETS):
            _check_validation_statuses(run_summary, failures)
            _check_summary_metrics(run_summary, failures)
            _check_programme_breakdown(run_summary, failures)

    if guarded_report is not None:
        _check_required_sheets(guarded_report, REQUIRED_GUARDED_SHEETS, "guarded_generation_report.xlsx", failures)
        if "Summary" in guarded_report.sheetnames:
            _check_guarded_report(guarded_report, failures)

    if template2_validation is not None:
        _check_required_sheets(
            template2_validation,
            REQUIRED_TEMPLATE2_VALIDATION_SHEETS,
            "template2_submission_validation.xlsx",
            failures,
        )
        if "Summary" in template2_validation.sheetnames and "Submission Readiness" in template2_validation.sheetnames:
            _check_template2_validation(template2_validation, failures)

    if template2_submission is not None:
        _check_saved_template2_submission(template2_submission, template2_validation, template2_reconciliation, failures)

    if template2_reconciliation is not None:
        _check_template2_reconciliation(template2_reconciliation, template2_validation, failures)

    if visual_validation is not None:
        _check_required_sheets(
            visual_validation,
            REQUIRED_VISUAL_VALIDATION_SHEETS,
            "timetable_visualisation_validation.xlsx",
            failures,
        )
        if "Summary" in visual_validation.sheetnames:
            _check_visual_validation(visual_validation, failures)

    if fixed_integrity is not None:
        _check_required_sheets(
            fixed_integrity,
            REQUIRED_FIXED_INTEGRITY_SHEETS,
            "fixed_session_integrity_validation.xlsx",
            failures,
        )
        if "Summary" in fixed_integrity.sheetnames:
            _check_fixed_integrity(fixed_integrity, failures)

    if timetable is not None:
        _check_timetable_workbook(timetable, failures)
    if stakeholder_views is not None:
        _check_stakeholder_views(stakeholder_views, failures)
    if run_manifest is not None:
        _check_run_manifest(run_manifest, failures)

    return ReleaseValidationResult(passed=not failures, failures=failures)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    """Parse release-validator command-line arguments."""
    parser = argparse.ArgumentParser(description="Validate generated final-release Excel artefacts")
    parser.add_argument("--run-dir", type=Path, default=None)
    parser.add_argument("--run-summary", type=Path, default=None)
    parser.add_argument("--timetable", type=Path, default=None)
    parser.add_argument("--stakeholder-views", type=Path, default=None)
    parser.add_argument("--run-manifest", type=Path, default=None)
    parser.add_argument("--guarded-report", type=Path, default=None)
    parser.add_argument("--template2-validation", type=Path, default=None)
    parser.add_argument("--template2-submission", type=Path, default=None)
    parser.add_argument(
        "--template2-reconciliation",
        type=Path,
        default=None,
    )
    parser.add_argument("--visual-validation", type=Path, default=None)
    parser.add_argument("--fixed-session-integrity", type=Path, default=None)
    parser.add_argument("--programme-visuals", type=Path, default=None)
    parser.add_argument("--tutor-visuals", type=Path, default=None)
    parser.add_argument("--room-visuals", type=Path, default=None)
    parser.add_argument("--test-root", type=Path, default=BASE_DIR / "tests")
    return parser.parse_args(argv)


def _explicit_paths_from_args(args: argparse.Namespace) -> ReleaseEvidencePaths | None:
    """Return explicit evidence paths when any individual path option is used."""
    values = {
        "run_summary": args.run_summary,
        "timetable": args.timetable,
        "stakeholder_views": args.stakeholder_views,
        "run_manifest": args.run_manifest,
        "guarded_report": args.guarded_report,
        "template2_validation": args.template2_validation,
        "template2_submission": args.template2_submission,
        "template2_reconciliation": args.template2_reconciliation,
        "visual_validation": args.visual_validation,
        "fixed_integrity": args.fixed_session_integrity,
        "programme_visuals": args.programme_visuals,
        "tutor_visuals": args.tutor_visuals,
        "room_visuals": args.room_visuals,
    }
    if not any(value is not None for value in values.values()):
        return None
    return ReleaseEvidencePaths(
        run_summary=values["run_summary"] or DEFAULT_RUN_SUMMARY_FILE,
        timetable=values["timetable"] or DEFAULT_TIMETABLE_FILE,
        stakeholder_views=values["stakeholder_views"] or DEFAULT_STAKEHOLDER_VIEWS_FILE,
        run_manifest=values["run_manifest"] or DEFAULT_RUN_MANIFEST_FILE,
        guarded_report=values["guarded_report"] or DEFAULT_GUARDED_GENERATION_REPORT_FILE,
        template2_validation=values["template2_validation"] or DEFAULT_TEMPLATE2_SUBMISSION_VALIDATION_FILE,
        template2_submission=values["template2_submission"] or DEFAULT_TEMPLATE2_SUBMISSION_FILE,
        template2_reconciliation=values["template2_reconciliation"] or DEFAULT_TEMPLATE2_PROGRAMME_YEAR_RECONCILIATION_FILE,
        visual_validation=values["visual_validation"] or DEFAULT_TIMETABLE_VISUALISATION_VALIDATION_FILE,
        fixed_integrity=values["fixed_integrity"] or DEFAULT_FIXED_SESSION_INTEGRITY_FILE,
        programme_visuals=values["programme_visuals"] or DEFAULT_PROGRAMME_VISUALS_FILE,
        tutor_visuals=values["tutor_visuals"] or DEFAULT_TUTOR_VISUALS_FILE,
        room_visuals=values["room_visuals"] or DEFAULT_ROOM_VISUALS_FILE,
        test_root=args.test_root,
        evidence_location=None,
        run_id="explicit-paths",
    )


def _metric(path: Path, sheet_name: str, key: str) -> object:
    """Read one summary-style metric from a workbook."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return _sheet_dict(workbook, sheet_name).get(key)
    finally:
        workbook.close()


def _release_metric_lines(paths: ReleaseEvidencePaths) -> list[str]:
    """Return concise final release evidence lines for CLI output."""
    return [
        f"Required occurrences: {_metric(paths.guarded_report, 'Summary', 'Total teaching occurrences')}",
        f"Scheduled occurrences: {_metric(paths.guarded_report, 'Summary', 'Scheduled occurrences')}",
        f"Quarantined occurrences: {_metric(paths.guarded_report, 'Summary', 'Quarantined occurrences')}",
        f"Search failures: {_metric(paths.guarded_report, 'Summary', 'Unscheduled search failures')}",
        f"Scheduled hard violations: {_metric(paths.guarded_report, 'Summary', 'Scheduled hard violations')}",
        f"Fixed-session integrity: {_metric(paths.fixed_integrity, 'Summary', 'fixed-session integrity status')}",
        f"Qualifying programme-years: {_metric(paths.template2_reconciliation, 'Summary', 'qualifying submission-ready programme-years')}",
        f"Template 2 readiness: {_metric(paths.template2_reconciliation, 'Summary', 'Template 2 readiness status')}",
        f"Visual validation: {_metric(paths.visual_validation, 'Summary', 'visual export status')}",
    ]


def main(argv: list[str] | None = None) -> int:
    """Run release validation and print a short result."""
    args = parse_args(argv)
    try:
        resolution = resolve_evidence_paths(
            run_dir=args.run_dir,
            test_root=args.test_root,
            explicit_paths=_explicit_paths_from_args(args),
        )
    except (FileNotFoundError, ValueError) as exc:
        print(f"FINAL RELEASE VALIDATION: FAIL")
        print(f"- {exc}")
        return 1
    paths = resolution.paths
    print(resolution.message)
    result = validate_release(
        paths.run_summary,
        paths.timetable,
        paths.stakeholder_views,
        paths.run_manifest,
        paths.guarded_report,
        paths.template2_validation,
        paths.template2_submission,
        paths.template2_reconciliation,
        paths.visual_validation,
        paths.fixed_integrity,
        paths.programme_visuals,
        paths.tutor_visuals,
        paths.room_visuals,
        paths.test_root,
    )
    if result.passed:
        for line in _release_metric_lines(paths):
            print(line)
        print("FINAL RELEASE VALIDATION: PASS")
        return 0

    print("FINAL RELEASE VALIDATION: FAIL")
    for failure in result.failures:
        print(f"- {failure}")
    return 1


if __name__ == "__main__":
    sys.exit(main())
