"""Reusable orchestration service for CLI and desktop UI callers."""

from __future__ import annotations

import zipfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from threading import Event
from time import perf_counter
from types import SimpleNamespace
from typing import Callable, Literal
from uuid import uuid4

from openpyxl import load_workbook

from config import (
    DEFAULT_COMMON_MODULE_FILE,
    DEFAULT_COURSE_FILE,
    DEFAULT_ENGINEERING_FOLDER,
    DEFAULT_FIXED_RECONCILIATION_FILE,
    DEFAULT_FIXED_CONFLICT_TRIAGE_FILE,
    DEFAULT_FIXED_RESOLUTION_AUDIT_FILE,
    DEFAULT_FIXED_RESOLUTION_TEMPLATE_FILE,
    DEFAULT_FIXED_ROOT_CAUSE_FILE,
    DEFAULT_FIXED_SESSION_FILE,
    DEFAULT_FIXED_SESSIONS_AUDIT_FILE,
    DEFAULT_GUARDED_GENERATION_REPORT_FILE,
    DEFAULT_INPUT_READINESS_REPORT_FILE,
    DEFAULT_LOCATION_MAPPING_EVIDENCE_FILE,
    DEFAULT_LOADER_REPORT_FILE,
    DEFAULT_PROGRAMME_VISUALS_FILE,
    DEFAULT_PREFLIGHT_REPORT_FILE,
    DEFAULT_REMARKS_AUDIT_FILE,
    DEFAULT_ROOM_FILE,
    DEFAULT_ROOM_VISUALS_FILE,
    DEFAULT_RUN_MANIFEST_FILE,
    DEFAULT_RUN_SUMMARY_FILE,
    DEFAULT_STAKEHOLDER_VIEWS_FILE,
    DEFAULT_SUPERVISOR_CLARIFICATION_PACK_FILE,
    DEFAULT_SUPERVISOR_FIXED_QUERIES_FILE,
    DEFAULT_TEMPLATE2_SUBMISSION_FILE,
    DEFAULT_TEMPLATE2_SUBMISSION_VALIDATION_FILE,
    DEFAULT_TEMPLATE2_FILE,
    DEFAULT_TIMETABLE_VISUALISATION_VALIDATION_FILE,
    DEFAULT_TUTOR_VISUALS_FILE,
    DEFAULT_UNSCHEDULED_DIAGNOSTICS_FILE,
    OUTPUT_DIR,
)
from data.loader import (
    LoaderReport,
    export_loader_report,
    load_consolidated_schedule_with_report,
    load_common_modules,
    load_courses_from_folder,
    load_courses_from_requirements,
    load_rooms_from_csv,
)
from data.fixed_sessions import export_fixed_sessions_audit, load_fixed_sessions
from data.models import Course
from engine.constraint_checker import annotate_schedule_violations, count_soft_violations
from engine.demand_metrics import build_demand_metrics
from engine.fixed_reconciliation import (
    adjusted_courses_after_exact_matches,
    export_fixed_reconciliation_report,
    reconcile_fixed_sessions,
)
from engine.fixed_scope import filter_fixed_sessions_to_selected_scope
from engine.fixed_issue_analysis import export_fixed_issue_workbooks
from engine.guarded_generation import (
    build_guarded_generation_state,
    build_programme_completeness_rows,
    complete_programme_set,
    export_guarded_generation_report,
    quarantined_requirement_courses,
)
from engine.input_readiness import build_input_readiness_result, export_input_readiness_report
from engine.preflight_validator import run_preflight_checks
from engine.remarks_interpreter import export_remarks_audit
from engine.resource_audit import audit_resources
from engine.unscheduled_diagnostics import diagnose_unscheduled_assignments, export_unscheduled_diagnostics
from generator.scheduler import generate_schedule
from generator.fixed_scheduler import create_fixed_assignments, validate_fixed_assignments
from main import (
    _completed_optimisation_summary,
    _count_current_hard_violations,
    _count_scheduled_hard_violations,
    _count_weighted_soft_score,
    _fixed_requirement_courses,
    _restore_unscheduled_reasons,
    _skipped_optimisation_summary,
    _snapshot_unscheduled_reasons,
    export_outputs,
)
from optimiser.local_search import optimise_schedule_with_stats
from output.report_exporter import export_preflight_report, export_run_manifest, export_run_summary, export_stakeholder_views
from output.submission_validator import (
    export_submission_ready_schedule,
    export_template2_validation_report,
    validate_template2_submission,
)
from output.timetable_visualizer import export_timetable_visuals, export_visualisation_failure_report

ProgressCallback = Callable[[str], None]


@dataclass(slots=True)
class PipelineOptions:
    """Runtime options for a timetable pipeline run."""

    scope: str = "eng"
    consolidated_schedule_path: Path | None = None
    # Compatibility alias: when supplied, input_path means consolidated Template 1 requirements input.
    input_path: Path | None = None
    input_mode: Literal["selected_workbook", "selected_folder", "default_engineering_dataset"] | None = None
    room_source_path: Path | None = None
    common_module_path: Path = DEFAULT_COMMON_MODULE_FILE
    template2_output_template_path: Path = DEFAULT_TEMPLATE2_FILE
    # Compatibility alias for callers created before explicit workbook roles were added.
    room_path: Path | None = None
    run_optimisation: bool = False
    max_iterations: int = 5
    optimisation_time_limit: float | None = None
    optimisation_patience: int | None = None
    max_candidate_patterns: int | None = 300
    max_retry_assignments: int | None = 50
    skip_unscheduled_diagnostics: bool = True
    max_diagnostic_assignments: int | None = None
    progress_interval: int = 25
    skip_preflight: bool = False
    audit_demand_metrics: bool = True
    enable_remark_interpretation: bool = True
    run_id: str | None = None
    output_dir: Path | None = None


@dataclass(slots=True)
class PipelineResult:
    """Headline result values returned to the desktop UI."""

    required_occurrences: int
    scheduled_occurrences: int
    unscheduled_occurrences: int
    coverage_percent: float
    scheduled_hard_violations: int
    online_required: int
    online_scheduled: int
    dsc_included: bool
    validation_passed: bool
    optimisation_status: str
    output_paths: dict[str, Path]
    run_id: str = ""
    selected_recorded_occurrences: int = 0
    selected_quarantined_occurrences: int = 0
    selected_schedulable_occurrences: int = 0
    selected_scheduled_occurrences: int = 0
    selected_search_failures: int = 0
    selected_schedulable_coverage_percent: float = 0.0
    selected_total_coverage_percent: float = 0.0


class PipelineCancelled(RuntimeError):
    """Raised when a cooperative UI cancellation request is observed."""


def _emit(callback: ProgressCallback | None, message: str) -> None:
    """Send one progress message to the caller."""
    if callback is not None:
        callback(message)


def _check_cancel(cancel_event: Event | None) -> None:
    """Raise if the caller requested cancellation."""
    if cancel_event is not None and cancel_event.is_set():
        raise PipelineCancelled("Run cancelled between pipeline stages.")


def _consolidated_schedule_path(options: PipelineOptions) -> Path | None:
    """Return selected Template 1 path, including the compatibility alias."""
    return options.consolidated_schedule_path or options.input_path


def _room_source_path(options: PipelineOptions) -> Path:
    """Return the configured room source path, including the compatibility alias."""
    return options.room_source_path or options.room_path or DEFAULT_ROOM_FILE


def _input_mode(options: PipelineOptions) -> str:
    """Return the explicit input-source mode for a pipeline run."""
    if options.input_mode:
        return options.input_mode
    path = _consolidated_schedule_path(options)
    if path is None:
        return "default_engineering_dataset"
    if path.is_dir():
        return "selected_folder"
    return "selected_workbook"


def _load_courses_from_path(
    path: Path | None,
    scope: str,
    common_modules: set[str],
    input_mode: str = "default_engineering_dataset",
) -> tuple[list[Course], LoaderReport]:
    """Load courses from selected Template 1 input or existing CLI defaults."""
    if path is None:
        if scope == "eng" and DEFAULT_ENGINEERING_FOLDER.exists():
            return load_courses_from_folder(DEFAULT_ENGINEERING_FOLDER, common_modules=common_modules)
        courses, workbook_report = load_courses_from_requirements(DEFAULT_COURSE_FILE, common_modules=common_modules)
        report = LoaderReport()
        report.add(workbook_report)
        return courses, report

    if path.is_dir():
        return load_courses_from_folder(path, common_modules=common_modules)
    courses, workbook_report = load_consolidated_schedule_with_report(
        path,
        common_modules=common_modules,
        strict_teaching_week_dates=input_mode == "selected_workbook",
    )
    report = LoaderReport()
    report.add(workbook_report)
    return courses, report


def _new_run_id() -> str:
    """Return a compact run identifier for isolated UI outputs."""
    return f"{datetime.now():%Y%m%d_%H%M%S}_{uuid4().hex[:6]}"


def _run_output_dir(options: PipelineOptions, run_id: str) -> Path:
    """Return the isolated output directory for one pipeline run."""
    return options.output_dir or OUTPUT_DIR / "runs" / run_id


def _run_paths(run_dir: Path) -> dict[str, Path]:
    """Return all workbook paths for one isolated pipeline run."""
    return {
        "loader_report": run_dir / "loader_report.xlsx",
        "preflight_report": run_dir / "preflight_report.xlsx",
        "fixed_sessions_audit": run_dir / "fixed_sessions_audit.xlsx",
        "fixed_reconciliation": run_dir / "fixed_nonfixed_reconciliation.xlsx",
        "input_readiness_report": run_dir / "input_readiness_report.xlsx",
        "fixed_issue_root_cause_analysis": run_dir / "fixed_issue_root_cause_analysis.xlsx",
        "fixed_conflict_triage": run_dir / "fixed_conflict_triage.xlsx",
        "supervisor_fixed_session_queries": run_dir / "supervisor_fixed_session_queries.xlsx",
        "location_mapping_evidence": run_dir / "location_mapping_evidence.xlsx",
        "supervisor_clarification_pack": run_dir / "Supervisor_Fixed_Session_Clarification_Pack.xlsx",
        "fixed_resolution_template": run_dir / "fixed_session_resolution_template.xlsx",
        "fixed_resolution_audit": run_dir / "fixed_session_resolution_audit.xlsx",
        "run_summary": run_dir / "run_summary.xlsx",
        "stakeholder_views": run_dir / "stakeholder_views.xlsx",
        "remarks_audit": run_dir / "remarks_audit.xlsx",
        "unscheduled_diagnostics": run_dir / "unscheduled_diagnostics.xlsx",
        "proposed_timetable": run_dir / "Proposed_Timetable.xlsx",
        "violations": run_dir / "violation_report.xlsx",
        "submission_ready_timetable": run_dir / "Template2_Submission_Ready.xlsx",
        "template2_submission_validation": run_dir / "template2_submission_validation.xlsx",
        "guarded_generation_report": run_dir / "guarded_generation_report.xlsx",
        "programme_visuals": run_dir / "Programme_Timetable_Visuals.xlsx",
        "tutor_visuals": run_dir / "Tutor_Timetable_Visuals.xlsx",
        "room_visuals": run_dir / "Room_Timetable_Visuals.xlsx",
        "visualisation_validation": run_dir / "timetable_visualisation_validation.xlsx",
        "run_manifest": run_dir / "run_manifest.xlsx",
    }


def _workbook_reopens(path: Path) -> bool:
    """Return True when an XLSX file is a valid ZIP and openpyxl can reopen it."""
    if not path.exists() or path.stat().st_size <= 0:
        return False
    try:
        with zipfile.ZipFile(path) as archive:
            if archive.testzip() is not None:
                return False
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            return bool(workbook.sheetnames)
        finally:
            workbook.close()
    except Exception:
        return False


def _outputs_validate(output_paths: dict[str, Path]) -> bool:
    """Return True when reported workbook outputs can be reopened."""
    workbook_keys = [
        "proposed_timetable",
        "submission_ready_timetable",
        "run_summary",
        "run_manifest",
        "guarded_generation_report",
        "template2_submission_validation",
        "programme_visuals",
        "tutor_visuals",
        "room_visuals",
        "visualisation_validation",
    ]
    return all(
        _workbook_reopens(output_paths[key])
        for key in workbook_keys
        if key in output_paths
    )


def _args_namespace(options: PipelineOptions) -> SimpleNamespace:
    """Return an argparse-like namespace for existing summary helpers."""
    return SimpleNamespace(
        scope=options.scope,
        skip_optimisation=not options.run_optimisation,
        max_iterations=options.max_iterations,
        optimisation_time_limit=options.optimisation_time_limit,
        optimisation_patience=options.optimisation_patience,
        max_candidate_patterns=options.max_candidate_patterns,
        max_retry_assignments=options.max_retry_assignments,
        skip_unscheduled_diagnostics=options.skip_unscheduled_diagnostics,
        max_diagnostic_assignments=options.max_diagnostic_assignments,
        progress_interval=options.progress_interval,
        audit_demand_metrics=options.audit_demand_metrics,
        disable_remark_interpretation=not options.enable_remark_interpretation,
    )


def _metadata(options: PipelineOptions, run_id: str, run_dir: Path) -> dict[str, object]:
    """Return run metadata for report workbooks."""
    args = _args_namespace(options)
    return {
        "run_id": run_id,
        "run_output_dir": str(run_dir),
        "scope": args.scope,
        "input_mode": _input_mode(options),
        "skip_optimisation": args.skip_optimisation,
        "max_iterations": args.max_iterations,
        "optimisation_time_limit": args.optimisation_time_limit,
        "optimisation_patience": args.optimisation_patience,
        "max_candidate_patterns": args.max_candidate_patterns,
        "max_retry_assignments": args.max_retry_assignments,
        "skip_unscheduled_diagnostics": args.skip_unscheduled_diagnostics,
        "max_diagnostic_assignments": args.max_diagnostic_assignments,
        "progress_interval": args.progress_interval,
        "audit_demand_metrics": args.audit_demand_metrics,
        "remark_interpretation_enabled": not args.disable_remark_interpretation,
        "input_workbook": str(_consolidated_schedule_path(options) or DEFAULT_ENGINEERING_FOLDER),
        "consolidated_schedule_path": str(_consolidated_schedule_path(options) or DEFAULT_ENGINEERING_FOLDER),
        "output_template": str(options.template2_output_template_path),
        "template2_output_template_path": str(options.template2_output_template_path),
        "room_source_path": str(_room_source_path(options)),
        "common_module_path": str(options.common_module_path),
    }


def _dsc_included(courses: list[Course]) -> bool:
    """Return True when DSC appears in Engineering input data."""
    return any("DSC" in " ".join([course.module_code, course.prog_yr, course.source_file]).upper() for course in courses)


def run_timetable_pipeline(
    options: PipelineOptions,
    progress_callback: ProgressCallback | None = None,
    cancel_event: Event | None = None,
) -> PipelineResult:
    """Run the existing timetable pipeline and return UI-friendly metrics."""
    run_id = options.run_id or _new_run_id()
    run_dir = _run_output_dir(options, run_id)
    run_dir.mkdir(parents=True, exist_ok=True)
    paths = _run_paths(run_dir)
    input_mode = _input_mode(options)

    _emit(progress_callback, "Loading input")
    common_modules = load_common_modules(options.common_module_path)
    courses, loader_report = _load_courses_from_path(
        _consolidated_schedule_path(options),
        options.scope,
        common_modules,
        input_mode,
    )
    rooms = load_rooms_from_csv(_room_source_path(options))
    export_loader_report(loader_report, paths["loader_report"])
    _check_cancel(cancel_event)

    if options.skip_preflight:
        _emit(progress_callback, "Skipping preflight checks")
    else:
        _emit(progress_callback, "Running preflight checks")
        preflight_issues = run_preflight_checks(courses, rooms)
        export_preflight_report(preflight_issues, paths["preflight_report"])
    _check_cancel(cancel_event)

    fixed_sessions = []
    fixed_assignments: list = []
    fixed_conflict_issues: list[dict[str, object]] = []
    guarded_state = None
    schedule_courses = courses
    demand_courses = courses
    if options.scope == "eng" and DEFAULT_FIXED_SESSION_FILE.exists():
        _emit(progress_callback, "Checking fixed-session requirements")
        fixed_sessions, fixed_loader_report = load_fixed_sessions(DEFAULT_FIXED_SESSION_FILE)
        if input_mode != "default_engineering_dataset":
            fixed_sessions, fixed_loader_report, _fixed_scope_rows = filter_fixed_sessions_to_selected_scope(
                fixed_sessions,
                fixed_loader_report,
                courses,
            )
        export_fixed_sessions_audit(fixed_loader_report, paths["fixed_sessions_audit"])
        reconciliation_report = reconcile_fixed_sessions(fixed_sessions, courses, fixed_loader_report)
        export_fixed_reconciliation_report(reconciliation_report, paths["fixed_reconciliation"])
        fixed_assignments, fixed_mapping_issues = create_fixed_assignments(fixed_sessions, rooms)
        fixed_conflict_issues = validate_fixed_assignments(fixed_assignments)
        guarded_state = build_guarded_generation_state(
            courses=courses,
            rooms_loaded=len(rooms),
            fixed_sessions=fixed_sessions,
            fixed_loader_report=fixed_loader_report,
            reconciliation_report=reconciliation_report,
            fixed_assignments=fixed_assignments,
            fixed_assignment_issues=fixed_mapping_issues + fixed_conflict_issues,
        )
        readiness = build_input_readiness_result(
            fixed_loader_report=fixed_loader_report,
            reconciliation_report=reconciliation_report,
            fixed_assignment_issues=fixed_mapping_issues + fixed_conflict_issues,
            loader_report=loader_report,
            global_errors=guarded_state.global_errors,
            quarantined_requirements=guarded_state.quarantined_requirements,
        )
        export_input_readiness_report(readiness, paths["input_readiness_report"])
        export_fixed_issue_workbooks(
            fixed_sessions=fixed_sessions,
            courses=courses,
            assignments=fixed_assignments,
            rooms=rooms,
            loader_report=fixed_loader_report,
            reconciliation_report=reconciliation_report,
            mapping_issues=fixed_mapping_issues,
            conflict_issues=fixed_conflict_issues,
            root_cause_path=paths["fixed_issue_root_cause_analysis"],
            conflict_triage_path=paths["fixed_conflict_triage"],
            supervisor_queries_path=paths["supervisor_fixed_session_queries"],
            location_evidence_path=paths["location_mapping_evidence"],
            supervisor_pack_path=paths["supervisor_clarification_pack"],
            resolution_template_path=paths["fixed_resolution_template"],
            resolution_audit_path=paths["fixed_resolution_audit"],
        )
        if not readiness.ready:
            raise ValueError(readiness.message)
        fixed_assignments = guarded_state.anchored_fixed_assignments
        schedule_courses = adjusted_courses_after_exact_matches(courses, reconciliation_report)
        demand_courses = schedule_courses + _fixed_requirement_courses(fixed_assignments) + quarantined_requirement_courses(guarded_state.quarantined_requirements)
    _check_cancel(cancel_event)

    _emit(progress_callback, "Generating timetable")

    def _course_progress(position: int, total: int, course: Course) -> None:
        _emit(progress_callback, f"Generating timetable: {position}/{total} {course.module_code} {course.activity}")

    initial_schedule = generate_schedule(
        schedule_courses,
        rooms,
        initial_assignments=fixed_assignments,
        progress_callback=_course_progress,
        progress_interval=options.progress_interval,
        max_retry_assignments=options.max_retry_assignments,
        max_candidate_patterns=options.max_candidate_patterns,
        enable_remark_interpretation=options.enable_remark_interpretation,
    )
    initial_reasons = _snapshot_unscheduled_reasons(initial_schedule)
    annotate_schedule_violations(
        initial_schedule,
        enable_remark_interpretation=options.enable_remark_interpretation,
    )
    initial_soft = count_soft_violations(
        initial_schedule,
        enable_remark_interpretation=options.enable_remark_interpretation,
    )
    initial_soft_score = _count_weighted_soft_score(
        initial_schedule,
        options.enable_remark_interpretation,
    )
    _restore_unscheduled_reasons(initial_schedule, initial_reasons)

    args = _args_namespace(options)
    final_schedule = initial_schedule
    optimisation_metrics = _skipped_optimisation_summary(args, initial_soft, initial_soft_score, demand_courses, initial_schedule, rooms)
    if options.run_optimisation:
        _check_cancel(cancel_event)
        _emit(progress_callback, "Running optimiser")
        started = perf_counter()
        result = optimise_schedule_with_stats(
            initial_schedule,
            rooms,
            max_iterations=options.max_iterations,
            time_limit_seconds=options.optimisation_time_limit,
            patience=options.optimisation_patience,
            enable_remark_interpretation=options.enable_remark_interpretation,
        )
        runtime_seconds = perf_counter() - started
        final_schedule = result.assignments
        final_reasons = _snapshot_unscheduled_reasons(final_schedule)
        final_soft = count_soft_violations(
            final_schedule,
            enable_remark_interpretation=options.enable_remark_interpretation,
        )
        final_soft_score = _count_weighted_soft_score(
            final_schedule,
            options.enable_remark_interpretation,
        )
        _restore_unscheduled_reasons(final_schedule, final_reasons)
        optimisation_metrics = _completed_optimisation_summary(
            args,
            initial_schedule,
            final_schedule,
            demand_courses,
            rooms,
            initial_soft,
            final_soft,
            initial_soft_score,
            final_soft_score,
            runtime_seconds,
            result.iterations_completed,
            result.status,
            result.stop_reason,
        )
    _check_cancel(cancel_event)

    _emit(progress_callback, "Generating stakeholder reports")
    metadata = _metadata(options, run_id, run_dir)
    export_run_summary(
        final_schedule,
        paths["run_summary"],
        metadata=metadata,
        demand_courses=demand_courses,
        input_course_records=len(demand_courses),
        rooms=rooms,
        room_source_path=_room_source_path(options),
        optimisation_summary=optimisation_metrics,
        enable_remark_interpretation=options.enable_remark_interpretation,
    )
    export_stakeholder_views(
        final_schedule,
        rooms,
        paths["stakeholder_views"],
        enable_remark_interpretation=options.enable_remark_interpretation,
    )
    export_remarks_audit(demand_courses, paths["remarks_audit"])

    if not options.skip_unscheduled_diagnostics:
        report = diagnose_unscheduled_assignments(
            final_schedule,
            rooms,
            max_diagnostic_assignments=options.max_diagnostic_assignments,
            enable_remark_interpretation=options.enable_remark_interpretation,
        )
        export_unscheduled_diagnostics(report, paths["unscheduled_diagnostics"])
    _check_cancel(cancel_event)

    _emit(progress_callback, "Exporting proposed timetable")
    output_paths = export_outputs(
        final_schedule,
        options.scope,
        template2_path=options.template2_output_template_path,
        enable_remark_interpretation=options.enable_remark_interpretation,
        output_dir=run_dir,
        timetable_filename="Proposed_Timetable.xlsx",
    )
    output_paths["proposed_timetable"] = paths["proposed_timetable"]
    output_paths["timetable"] = paths["proposed_timetable"]
    programme_rows: list[dict[str, object]] = []
    if options.scope == "eng" and DEFAULT_FIXED_SESSION_FILE.exists():
        programme_rows = build_programme_completeness_rows(
            demand_courses,
            final_schedule,
            guarded_state.quarantined_requirements if guarded_state is not None else [],
        )
        complete_programmes = complete_programme_set(programme_rows)
        export_submission_ready_schedule(
            final_schedule,
            paths["submission_ready_timetable"],
            template2_path=options.template2_output_template_path,
            enable_remark_interpretation=options.enable_remark_interpretation,
            complete_programmes=complete_programmes,
            rooms=rooms,
        )
        template2_validation = validate_template2_submission(
            paths["submission_ready_timetable"],
            demand_courses,
            fixed_sessions,
            final_schedule,
            rooms,
            options.template2_output_template_path,
        )
        export_template2_validation_report(template2_validation, paths["template2_submission_validation"])
        if guarded_state is not None:
            programme_rows = build_programme_completeness_rows(
                demand_courses,
                final_schedule,
                guarded_state.quarantined_requirements,
                submission_ready_programmes={
                    str(row.get("Normalised Programme/Year"))
                    for row in template2_validation.programme_rows
                    if row.get("Included In Submission") == "Yes"
                },
            )
            export_guarded_generation_report(
                output_path=paths["guarded_generation_report"],
                global_errors=guarded_state.global_errors,
                quarantined=guarded_state.quarantined_requirements,
                fixed_conflict_issues=fixed_conflict_issues,
                warnings=guarded_state.warning_issues,
                assignments=final_schedule,
                demand_courses=demand_courses,
                programme_rows=programme_rows,
                template2_summary=template2_validation.summary,
            )
        output_paths["submission_ready_timetable"] = paths["submission_ready_timetable"]
        output_paths["template2_submission_validation"] = paths["template2_submission_validation"]
        output_paths["guarded_generation_report"] = paths["guarded_generation_report"]
    output_paths.update(
        {
            "loader_report": paths["loader_report"],
            "run_summary": paths["run_summary"],
            "stakeholder_views": paths["stakeholder_views"],
            "remarks_audit": paths["remarks_audit"],
        }
    )
    if not options.skip_preflight:
        output_paths["preflight_report"] = paths["preflight_report"]
    if options.scope == "eng" and DEFAULT_FIXED_SESSION_FILE.exists():
        output_paths["fixed_sessions_audit"] = paths["fixed_sessions_audit"]
        output_paths["fixed_reconciliation"] = paths["fixed_reconciliation"]
        output_paths["input_readiness_report"] = paths["input_readiness_report"]
        output_paths["fixed_issue_root_cause_analysis"] = paths["fixed_issue_root_cause_analysis"]
        output_paths["fixed_conflict_triage"] = paths["fixed_conflict_triage"]
        output_paths["supervisor_fixed_session_queries"] = paths["supervisor_fixed_session_queries"]
        output_paths["location_mapping_evidence"] = paths["location_mapping_evidence"]
        output_paths["supervisor_clarification_pack"] = paths["supervisor_clarification_pack"]
        output_paths["fixed_resolution_template"] = paths["fixed_resolution_template"]
        output_paths["fixed_resolution_audit"] = paths["fixed_resolution_audit"]
        output_paths["guarded_generation_report"] = paths["guarded_generation_report"]
    if not options.skip_unscheduled_diagnostics:
        output_paths["unscheduled_diagnostics"] = paths["unscheduled_diagnostics"]

    _emit(progress_callback, "Exporting visual timetables")
    try:
        export_timetable_visuals(
            assignments=final_schedule,
            rooms=rooms,
            programme_rows=programme_rows,
            programme_path=paths["programme_visuals"],
            tutor_path=paths["tutor_visuals"],
            room_path=paths["room_visuals"],
            validation_path=paths["visualisation_validation"],
        )
    except Exception as exc:  # pragma: no cover - keeps official timetable outputs intact
        export_visualisation_failure_report(paths["visualisation_validation"], exc)
    output_paths.update(
        {
            "programme_visuals": paths["programme_visuals"],
            "tutor_visuals": paths["tutor_visuals"],
            "room_visuals": paths["room_visuals"],
            "visualisation_validation": paths["visualisation_validation"],
        }
    )

    _emit(progress_callback, "Validating outputs")
    export_run_manifest(
        demand_courses,
        final_schedule,
        paths["run_manifest"],
        metadata=metadata,
        rooms=rooms,
        output_files=output_paths,
        template2_path=options.template2_output_template_path,
        enable_remark_interpretation=options.enable_remark_interpretation,
    )
    output_paths["run_manifest"] = paths["run_manifest"]
    output_paths["proposed_template2"] = paths["proposed_timetable"]
    output_paths["exception_queue"] = paths["stakeholder_views"]
    output_paths["output_folder"] = run_dir

    demand = build_demand_metrics(demand_courses, final_schedule, input_course_records=len(demand_courses))
    resource_audit = audit_resources(demand_courses, rooms, final_schedule, room_source_path=_room_source_path(options))
    validation_passed = _outputs_validate({key: Path(value) for key, value in output_paths.items()})
    selected_quarantined = (
        sum(item.affected_occurrences for item in guarded_state.quarantined_requirements)
        if guarded_state is not None
        else 0
    )
    selected_schedulable = max(demand.required_teaching_occurrences - selected_quarantined, 0)
    selected_search_failures = max(
        demand.required_teaching_occurrences - demand.scheduled_teaching_occurrences - selected_quarantined,
        0,
    )
    selected_schedulable_coverage = (
        demand.scheduled_teaching_occurrences / selected_schedulable * 100
        if selected_schedulable
        else 0.0
    )
    _emit(progress_callback, "Completed")
    return PipelineResult(
        required_occurrences=demand.required_teaching_occurrences,
        scheduled_occurrences=demand.scheduled_teaching_occurrences,
        unscheduled_occurrences=demand.unscheduled_teaching_occurrences,
        coverage_percent=demand.coverage_rate_percent,
        scheduled_hard_violations=_count_scheduled_hard_violations(final_schedule),
        online_required=resource_audit.required_online_teaching_occurrences,
        online_scheduled=resource_audit.scheduled_online_teaching_occurrences,
        dsc_included=_dsc_included(courses),
        validation_passed=validation_passed,
        optimisation_status=str(optimisation_metrics.get("status", "Skipped")),
        output_paths={key: Path(value) for key, value in output_paths.items()},
        run_id=run_id,
        selected_recorded_occurrences=demand.required_teaching_occurrences,
        selected_quarantined_occurrences=selected_quarantined,
        selected_schedulable_occurrences=selected_schedulable,
        selected_scheduled_occurrences=demand.scheduled_teaching_occurrences,
        selected_search_failures=selected_search_failures,
        selected_schedulable_coverage_percent=selected_schedulable_coverage,
        selected_total_coverage_percent=demand.coverage_rate_percent,
    )
