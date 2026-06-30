"""Export final release metrics from generated evidence workbooks."""

from __future__ import annotations

import argparse
import csv
import json
import re
import subprocess
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "final_verification"
DEFAULT_ZIP_PATH = PROJECT_ROOT / "dist" / "itp_group5_prototype_v1.1.0.zip"
DEFAULT_REMARKS_COMPARISON = PROJECT_ROOT / "timetable_scheduler" / "generated" / "remarks_coverage_comparison.xlsx"


def _sheet_dict(path: Path, sheet_name: str) -> dict[str, Any]:
    """Return first-column/second-column rows from one workbook sheet."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        return {
            str(row[0]): row[1]
            for row in sheet.iter_rows(min_row=2, values_only=True)
            if row and row[0] is not None
        }
    finally:
        workbook.close()


def _sheet_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    """Return sheet rows as dictionaries keyed by header."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        try:
            headers = [cell.value for cell in sheet[1]]
        except IndexError:
            return []
        return [
            {str(header): row[index] for index, header in enumerate(headers) if header is not None}
            for row in sheet.iter_rows(min_row=2, values_only=True)
            if row and any(value not in (None, "") for value in row)
        ]
    finally:
        workbook.close()


def _git_commit() -> str:
    """Return the current commit SHA when Git is available."""
    try:
        result = subprocess.run(
            ["git", "rev-parse", "HEAD"],
            cwd=PROJECT_ROOT,
            check=True,
            capture_output=True,
            text=True,
        )
    except Exception:
        return "UNKNOWN"
    return result.stdout.strip()


def _test_count(test_output_path: Path | None) -> int | None:
    """Extract pytest passed count from a captured output file."""
    if test_output_path is None or not test_output_path.exists():
        return None
    text = _read_text(test_output_path)
    matches = re.findall(r"(\d+)\s+passed", text)
    return int(matches[-1]) if matches else None


def _remarks_status(path: Path | None) -> str:
    """Return remarks attribution reconciliation status."""
    if path is None or not path.exists():
        return "NOT AVAILABLE"
    rows = _sheet_rows(path, "Attribution Reconciliation")
    for row in rows:
        if row.get("Metric") == "Attribution reconciliation":
            return str(row.get("Status") or "UNKNOWN")
    return "UNKNOWN"


def _visual_status(path: Path, workbook_name: str) -> str:
    """Return visual workbook export status from validation evidence."""
    rows = _sheet_rows(path, "Export Status")
    for row in rows:
        if row.get("Workbook") == workbook_name:
            return str(row.get("Status") or "UNKNOWN")
    return "UNKNOWN"


def _zip_file_count(path: Path | None) -> int | None:
    """Return the ZIP entry count."""
    if path is None or not path.exists():
        return None
    with zipfile.ZipFile(path) as archive:
        return len(archive.namelist())


def _read_hash(path: Path | None) -> str:
    """Read a one-line hash file when available."""
    if path is None or not path.exists():
        return ""
    return _read_text(path).lstrip("\ufeff").strip()


def _read_text(path: Path) -> str:
    """Read text files written by either Python or Windows PowerShell."""
    data = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-16"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def build_metrics(
    run_dir: Path,
    *,
    test_output: Path | None = None,
    zip_path: Path | None = DEFAULT_ZIP_PATH,
    zip_sha_path: Path | None = None,
    remarks_comparison: Path | None = DEFAULT_REMARKS_COMPARISON,
) -> dict[str, Any]:
    """Build final metrics from one isolated run directory."""
    guarded = _sheet_dict(run_dir / "guarded_generation_report.xlsx", "Summary")
    fixed = _sheet_dict(run_dir / "fixed_session_integrity_validation.xlsx", "Summary")
    template2 = _sheet_dict(run_dir / "template2_submission_validation.xlsx", "Summary")
    reconciliation = _sheet_dict(run_dir / "template2_programme_year_reconciliation.xlsx", "Summary")
    visual = _sheet_dict(run_dir / "timetable_visualisation_validation.xlsx", "Summary")

    required = int(guarded["Total teaching occurrences"])
    schedulable = int(guarded["Schedulable occurrences"])
    scheduled = int(guarded["Scheduled occurrences"])
    quarantined = int(guarded["Quarantined occurrences"])
    search_failures = int(guarded["Unscheduled search failures"])
    unscheduled = required - scheduled
    zip_file_count = _zip_file_count(zip_path)
    zip_size = zip_path.stat().st_size if zip_path is not None and zip_path.exists() else None

    return {
        "run_id": run_dir.name,
        "commit_sha": _git_commit(),
        "required_occurrences": required,
        "schedulable_occurrences": schedulable,
        "scheduled_occurrences": scheduled,
        "quarantined_occurrences": quarantined,
        "search_failures": search_failures,
        "unscheduled_occurrences": unscheduled,
        "schedulable_demand_coverage_percent": scheduled / schedulable * 100 if schedulable else 0.0,
        "total_demand_coverage_percent": scheduled / required * 100 if required else 0.0,
        "scheduled_hard_violations": int(guarded["Scheduled hard violations"]),
        "fixed_source_occurrences": int(fixed["expected fixed teaching occurrences"]),
        "anchored_fixed_occurrences": int(fixed["anchored fixed teaching occurrences"]),
        "quarantined_fixed_occurrences": int(fixed["quarantined fixed teaching occurrences"]),
        "fixed_placement_mismatches": int(fixed["placement mismatches"]),
        "fixed_session_integrity": str(fixed["fixed-session integrity status"]),
        "strict_template2_rows": int(template2["Actual saved Template 2 rows"]),
        "all_valid_template2_rows": int(template2["All-valid Template 2 rows"]),
        "invalid_template2_rows": len(_sheet_rows(run_dir / "template2_submission_validation.xlsx", "Invalid Rows")),
        "qualifying_programme_years": int(reconciliation["qualifying submission-ready programme-years"]),
        "minimum_required_programme_years": int(reconciliation["required minimum programme-year schedules"]),
        "template2_readiness": str(reconciliation["Template 2 readiness status"]),
        "programme_visual_status": _visual_status(run_dir / "timetable_visualisation_validation.xlsx", "Programme_Timetable_Visuals.xlsx"),
        "tutor_visual_status": _visual_status(run_dir / "timetable_visualisation_validation.xlsx", "Tutor_Timetable_Visuals.xlsx"),
        "room_visual_status": _visual_status(run_dir / "timetable_visualisation_validation.xlsx", "Room_Timetable_Visuals.xlsx"),
        "visual_validation": str(visual["visual export status"]),
        "remarks_attribution_status": _remarks_status(remarks_comparison),
        "test_count": _test_count(test_output),
        "zip_filename": zip_path.name if zip_path is not None else "",
        "zip_file_count": zip_file_count,
        "zip_size": zip_size,
        "zip_sha256": _read_hash(zip_sha_path),
    }


def write_outputs(metrics: dict[str, Any], output_dir: Path) -> None:
    """Write JSON, CSV and Markdown final metric files."""
    output_dir.mkdir(parents=True, exist_ok=True)
    json_path = output_dir / "final_release_metrics.json"
    csv_path = output_dir / "final_release_metrics.csv"
    summary_path = output_dir / "final_release_summary.md"
    json_path.write_text(json.dumps(metrics, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["metric", "value"])
        for key, value in metrics.items():
            writer.writerow([key, value])
    lines = ["# Final Release Summary", ""]
    for key, value in metrics.items():
        lines.append(f"- {key}: {value}")
    summary_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(description="Export final release metrics from evidence workbooks")
    parser.add_argument("--run-dir", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--test-output", type=Path, default=None)
    parser.add_argument("--zip-path", type=Path, default=DEFAULT_ZIP_PATH)
    parser.add_argument("--zip-sha-path", type=Path, default=None)
    parser.add_argument("--remarks-comparison", type=Path, default=DEFAULT_REMARKS_COMPARISON)
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    """Export metrics and return a process status code."""
    args = parse_args(argv)
    metrics = build_metrics(
        args.run_dir.resolve(),
        test_output=args.test_output,
        zip_path=args.zip_path,
        zip_sha_path=args.zip_sha_path,
        remarks_comparison=args.remarks_comparison,
    )
    write_outputs(metrics, args.output_dir.resolve())
    print(f"Wrote final metrics to {args.output_dir.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
